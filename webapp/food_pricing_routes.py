from flask import Blueprint, render_template, request, redirect, url_for, session, flash, Response
from functools import wraps
import io
import math
import re
import sqlite3

try:
    from store_scope import perth_today
except Exception:  # pragma: no cover - fallback if run outside the app package
    from datetime import date as _date
    def perth_today():
        return _date.today()

food_pricing = Blueprint('food_pricing', __name__, url_prefix='/admin/food-pricing')
DB_PATH = None

DEFAULT_MARGIN_PCT = 50.0
DEFAULT_OVERHEAD_PCT = 50.0
FOOD_PRICING_SEED_VERSION = 7

# Ingredient → inventory item mappings where Item List has a usable unit cost.
# Values are inventory_items.id. Ambiguous or missing products are left blank so
# admin can enter a precise cost manually instead of hiding bad assumptions.
AUTO_COST_MAP = {
    'bean sprouts': 476,            # Bean Sprout Organic 400G
    'beef': 290,                    # Beef Eyeround Platinum / KG
    'grilled beef': 290,
    'meat': 290,
    'raw beef': 290,
    'raw beef slices': 290,
    'sliced beef': 290,
    'beef brisket slices': 273,     # Beef Brisket / KG
    'beef brisket': 273,
    'beef balls': 261,              # Golden/B Beef Ball 500G
    'beef rib': 316,                # Beef Rib / KG
    'slow-cooked beef rib': 316,
    'beef shank': 354,              # LOTTE Wagyu Shin Shank / KG
    'broken rice': 541,
    'bun bo hue noodles': 552,      # 400G x 30 carton
    'chicken': 369,                 # Chicken Breast Fillet S/OFF / KG
    'grilled chicken': 369,
    'grilled lemongrass chicken': 369,
    'grilled lemongrass beef': 290,
    'sliced chicken': 369,
    'sliced chicken meat': 369,
    'chilli': 484,                  # Chilli Green / KG
    'sliced chili': 484,
    'chinese sausage': 474,         # Pork Sausage 1KG
    'coriander': 487,               # Coriander / KG
    'dry noodles': 552,
    'egg': 254,                     # Golder Eggs 700G / dozen
    'eggs': 254,
    'fish sauce': 564,              # Fish Sauce Golden 725ML
    'fish sauce with chili': 564,
    'fresh vermicelli noodles': 552,
    'fried tofu': 256,
    'herbs': 487,
    'mayonnaise': 570,
    'mint': 493,                    # Mint / KG
    'onion': 478,
    'pho noodles': 550,             # Rice Stick Pho 400G
    'prawn': 515,                   # Tiger prawn / KG
    'rice': 541,                    # Broken Rice 20KG
    'rice noodles': 550,
    'rice paper': 549,              # Rice paper 340G
    'sausage': 474,
    'sliced onion': 478,            # Brown onion 20KG
    'soy sauce': 521,
    'steamed rice': 541,
    'sticky rice': 542,             # Glutinous Rice 10KG
    'stir-fried tofu': 256,
    'tangy sweet fish sauce': 564,
    'tangy sweet sauce': 564,
    'tofu': 256,                    # Fried Tofu 400G
    'tomato': 502,                  # Tomato / KG
    'vermicelli noodles': 552,
    'xl vermicelli noodles': 552,
}

AUTO_COST_MAP_BY_ITEM = {
    ('goi cuon', 'sauce'): 564,      # fish sauce based roll sauce
    ('banh mi', 'sauce'): 570,       # Kewpie mayonnaise
    ('banh mi heo quay', 'sauce'): 570,
    ('traditional pork bánh mì', 'sauce'): 570,
    ('roast pork bánh mì', 'sauce'): 570,
    ('grilled chicken bánh mì', 'mayonnaise'): 570,
    ('grilled pork bánh mì', 'mayonnaise'): 570,
    ('grilled beef bánh mì', 'mayonnaise'): 570,
}

ZERO_COST_INGREDIENTS = {'scallion oil', 'onion oil', 'fried shallots'}

INGREDIENT_ALIASES = {
    'beef brisket': ['beef brisket slices'],
    'crispy roast pork': ['roast pork'],
    'dry noodles': ['fresh vermicelli noodles'],
    'fried tofu': ['tofu'],
    'grilled beef': ['beef'],
    'grilled chicken': ['chicken'],
    'grilled lemongrass beef': ['grilled beef', 'beef'],
    'grilled lemongrass chicken': ['chicken'],
    'lettuce': ['salad'],
    'onion oil': ['scallion oil'],
    'pickled vegetables': ['pickles'],
    'pork': ['meat'],
    'pork ham': ['vietnamese pork loaf'],
    'pork hock': ['pork hock meat'],
    'raw beef slices': ['raw beef'],
    'rice noodles': ['pho noodles'],
    'slow-cooked beef rib': ['beef rib'],
    'steamed rice': ['rice'],
    'stir-fried tofu': ['tofu'],
    'tangy sweet fish sauce': ['fish sauce'],
    'tangy sweet sauce': ['fish sauce'],
    'vermicelli noodles': ['bun bo hue noodles'],
    'xl vermicelli noodles': ['bun bo hue noodles'],
}


FOOD_PRICING_SEED = [
    {
        'category': 'Pho Noodle Soup',
        'name': 'Raw Beef Pho',
        'aliases': ['Pho Tai'],
        'ingredients': [
            ('Rice noodles', 200, 'g', ''),
            ('Beef broth', 750, 'ml', ''),
            ('Raw beef slices', 110, 'g', ''),
            ('Herbs', 10, 'g', ''),
        ],
    },
    {
        'category': 'Pho Noodle Soup',
        'name': 'Raw Beef and Beef Balls Pho',
        'aliases': ['Pho Tai Bo Vien'],
        'ingredients': [
            ('Rice noodles', 200, 'g', ''),
            ('Beef broth', 750, 'ml', ''),
            ('Raw beef slices', 85, 'g', ''),
            ('Beef balls', 50, 'g', '5 pieces'),
            ('Herbs', 10, 'g', ''),
        ],
    },
    {
        'category': 'Pho Noodle Soup',
        'name': 'Beef Brisket Pho',
        'aliases': ['Pho Nam'],
        'ingredients': [
            ('Rice noodles', 200, 'g', ''),
            ('Beef broth', 750, 'ml', ''),
            ('Beef brisket', 90, 'g', '8 slices'),
            ('Herbs', 10, 'g', ''),
        ],
    },
    {
        'category': 'Pho Noodle Soup',
        'name': 'Slow-Cooked Beef Rib Pho',
        'aliases': ['Pho Suon Bo'],
        'ingredients': [
            ('Rice noodles', 200, 'g', ''),
            ('Beef broth', 750, 'ml', ''),
            ('Slow-cooked beef rib', 350, 'g', '1 small rib piece'),
            ('Herbs', 10, 'g', ''),
        ],
    },
    {
        'category': 'Pho Noodle Soup',
        'name': 'MCQ Special Pho',
        'aliases': ['Pho Dac Biet'],
        'ingredients': [
            ('Rice noodles', 200, 'g', ''),
            ('Beef broth', 750, 'ml', ''),
            ('Raw beef slices', 60, 'g', ''),
            ('Beef brisket', 30, 'g', '4 slices'),
            ('Beef balls', 40, 'g', '4 pieces'),
            ('Herbs', 10, 'g', ''),
        ],
    },
    {
        'category': 'Pho Noodle Soup',
        'name': 'Chicken Pho',
        'aliases': ['Pho Ga'],
        'ingredients': [
            ('Rice noodles', 200, 'g', ''),
            ('Beef broth', 750, 'ml', ''),
            ('Sliced chicken meat', 180, 'g', '1 piece'),
            ('Herbs', 10, 'g', ''),
        ],
    },
    {
        'category': 'Pho Noodle Soup',
        'name': 'Beef Balls Pho',
        'aliases': [],
        'ingredients': [
            ('Rice noodles', 200, 'g', ''),
            ('Beef broth', 750, 'ml', ''),
            ('Beef balls', 50, 'g', ''),
            ('Herbs', 10, 'g', ''),
        ],
    },
    {
        'category': 'Pho Noodle Soup',
        'name': 'Bun Bo Hue',
        'aliases': ['Bun Bo Hue (Beef Spicy Noodle)'],
        'ingredients': [
            ('Vermicelli noodles', 250, 'g', ''),
            ('Beef broth', 750, 'ml', ''),
            ('Beef shank', 30, 'g', ''),
            ('Pork ham', 30, 'g', ''),
            ('Pork hock', 40, 'g', ''),
            ('Raw beef slices', 50, 'g', ''),
            ('Satay', 20, 'g', '1 tea spoon'),
            ('Bun Bo Hue sauce', 20, 'g', '1 tea spoon'),
        ],
    },
    {
        'category': 'Cup Noodles',
        'name': 'Chicken Pho Cup',
        'aliases': [],
        'ingredients': [
            ('Rice noodles', 60, 'g', ''),
            ('Broth', 500, 'ml', ''),
            ('Sliced chicken', 70, 'g', ''),
            ('Coriander', 5, 'g', ''),
            ('Green onion', 5, 'g', ''),
            ('Sliced onion', 5, 'g', ''),
        ],
    },
    {
        'category': 'Cup Noodles',
        'name': 'Bun Bo Hue Cup',
        'aliases': [],
        'ingredients': [
            ('XL vermicelli noodles', 150, 'g', ''),
            ('Broth', 500, 'ml', ''),
            ('Sliced beef', 20, 'g', ''),
            ('Pork ham', 15, 'g', ''),
            ('Beef shank', 20, 'g', ''),
            ('Raw beef slices', 10, 'g', ''),
            ('Coriander', 5, 'g', ''),
            ('Green onion', 5, 'g', ''),
            ('Sliced onion', 5, 'g', ''),
        ],
    },
    {
        'category': 'Cup Noodles',
        'name': 'Beef Pho Cup',
        'aliases': ['Pho Cup'],
        'ingredients': [
            ('Rice noodles', 60, 'g', ''),
            ('Broth', 500, 'ml', ''),
            ('Sliced beef', 70, 'g', ''),
            ('Coriander', 5, 'g', ''),
            ('Green onion', 5, 'g', ''),
            ('Sliced onion', 5, 'g', ''),
        ],
    },
    {
        'category': 'Pho Combo',
        'name': 'Pho Combo',
        'aliases': [],
        'ingredients': [
            ('Any pho', None, 'serve', 'Combo component'),
            ('Coffee', None, 'serve', 'Combo component'),
            ('Juice', None, 'serve', 'Combo component'),
        ],
    },
    {
        'category': 'Dry Noodles',
        'name': 'Rice Vermicelli Salad',
        'aliases': [],
        'ingredients': [
            ('Fresh vermicelli noodles', 200, 'g', ''),
            ('Salad', 60, 'g', ''),
            ('Cucumber', 40, 'g', ''),
            ('Pickles', 40, 'g', ''),
            ('Fish sauce', 40, 'g', ''),
            ('Scallion oil', 5, 'g', 'Free of charge'),
            ('Fried shallots', 5, 'g', 'Free of charge'),
        ],
    },
    {
        'category': 'Dry Noodles',
        'name': 'Roast Pork Dry Noodles',
        'aliases': ['Bun Heo Quay'],
        'ingredients': [
            ('Dry noodles', 200, 'g', ''),
            ('Roast pork', 150, 'g', ''),
            ('Lettuce', 60, 'g', ''),
            ('Pickles', 40, 'g', ''),
            ('Cucumber', 40, 'g', ''),
            ('Mint', 10, 'g', ''),
            ('Fish sauce with chili', 40, 'g', ''),
            ('Onion oil', 5, 'g', 'Free of charge'),
            ('Fried shallots', 5, 'g', 'Free of charge'),
        ],
    },
    {
        'category': 'Dry Noodles',
        'name': 'Grilled Lemongrass Chicken Dry Noodles',
        'aliases': ['Bun Ga'],
        'ingredients': [
            ('Dry noodles', 200, 'g', ''),
            ('Grilled lemongrass chicken', 180, 'g', ''),
            ('Lettuce', 60, 'g', ''),
            ('Pickles', 40, 'g', ''),
            ('Cucumber', 40, 'g', ''),
            ('Mint', 10, 'g', ''),
            ('Fish sauce with chili', 40, 'g', ''),
            ('Onion oil', 5, 'g', 'Free of charge'),
            ('Fried shallots', 5, 'g', 'Free of charge'),
        ],
    },
    {
        'category': 'Dry Noodles',
        'name': 'Grilled Lemongrass Beef Dry Noodles',
        'aliases': [],
        'ingredients': [
            ('Dry noodles', 200, 'g', ''),
            ('Grilled lemongrass beef', 90, 'g', ''),
            ('Lettuce', 60, 'g', ''),
            ('Pickles', 40, 'g', ''),
            ('Cucumber', 40, 'g', ''),
            ('Mint', 10, 'g', ''),
            ('Fish sauce with chili', 40, 'g', ''),
            ('Onion oil', 5, 'g', 'Free of charge'),
            ('Fried shallots', 5, 'g', 'Free of charge'),
        ],
    },
    {
        'category': 'Dry Noodles',
        'name': 'Grilled Lemongrass Pork Dry Noodles',
        'aliases': [],
        'ingredients': [
            ('Dry noodles', 200, 'g', ''),
            ('Grilled lemongrass pork', 120, 'g', ''),
            ('Lettuce', 60, 'g', ''),
            ('Pickles', 40, 'g', ''),
            ('Cucumber', 40, 'g', ''),
            ('Mint', 10, 'g', ''),
            ('Fish sauce with chili', 40, 'g', ''),
            ('Onion oil', 5, 'g', 'Free of charge'),
            ('Fried shallots', 5, 'g', 'Free of charge'),
        ],
    },
    {
        'category': 'Dry Noodles',
        'name': 'Stir-Fried Tofu Dry Noodles',
        'aliases': ['Bun Dau Hu'],
        'ingredients': [
            ('Dry noodles', 200, 'g', ''),
            ('Stir-fried tofu', 120, 'g', ''),
            ('Lettuce', 60, 'g', ''),
            ('Pickles', 40, 'g', ''),
            ('Cucumber', 40, 'g', ''),
            ('Mint', 10, 'g', ''),
            ('Soy sauce', 40, 'g', ''),
            ('Onion oil', 5, 'g', 'Free of charge'),
            ('Fried shallots', 5, 'g', 'Free of charge'),
        ],
    },
    {
        'category': 'Rice Paper Rolls',
        'name': 'Chicken Rice Paper Roll',
        'aliases': [],
        'ingredients': [
        ('Rice paper', 15, 'g', ''),
        ('Vermicelli noodles', 45, 'g', ''),
        ('Grilled chicken', 70, 'g', ''),
        ('Lettuce', 40, 'g', ''),
        ('Sauce', 60, 'g', ''),
        ],
    },
    {
        'category': 'Rice Paper Rolls',
        'name': 'Prawn and Pork Rice Paper Roll',
        'aliases': ['Goi Cuon'],
        'ingredients': [
        ('Rice paper', 15, 'g', ''),
        ('Prawn', 20, 'g', ''),
        ('Pork', 25, 'g', ''),
        ('Vermicelli noodles', 45, 'g', ''),
        ('Lettuce', 40, 'g', ''),
        ('Sauce', 60, 'g', ''),
        ],
    },
    {
        'category': 'Rice Paper Rolls',
        'name': 'Grilled Beef Rice Paper Roll',
        'aliases': [],
        'ingredients': [
        ('Rice paper', 15, 'g', ''),
        ('Vermicelli noodles', 45, 'g', ''),
        ('Grilled beef', 90, 'g', ''),
        ('Lettuce', 40, 'g', ''),
        ('Sauce', 60, 'g', ''),
        ],
    },
    {
        'category': 'Sticky Rice',
        'name': 'Xoi',
        'aliases': [],
        'ingredients': [
        ('Egg', 30, 'g', ''),
        ('Vietnamese pork loaf', 30, 'g', ''),
        ('Chinese sausage', 20, 'g', ''),
        ('Sticky rice', 220, 'g', ''),
        ],
    },
    {
        'category': 'Mixed Juices',
        'name': 'Detox Juice',
        'aliases': [],
        'ingredients': [
        ('Celery', None, 'g', ''),
        ('Carrot', None, 'g', ''),
        ('Cucumber', None, 'g', ''),
        ('Apple', None, 'g', ''),
        ('Mint', None, 'g', ''),
        ],
    },
    {
        'category': 'Mixed Juices',
        'name': 'Immunity Juice',
        'aliases': [],
        'ingredients': [
        ('Honey', None, 'g', ''),
        ('Ginger', None, 'g', ''),
        ('Carrot', None, 'g', ''),
        ('Orange', None, 'g', ''),
        ('Apple', None, 'g', ''),
        ],
    },
    {
        'category': 'Mixed Juices',
        'name': 'Sweet Beets Juice',
        'aliases': [],
        'ingredients': [
        ('Beetroot', None, 'g', ''),
        ('Apple', None, 'g', ''),
        ('Orange', None, 'g', ''),
        ],
    },
    {
        'category': 'Mixed Juices',
        'name': 'Green Glow Juice',
        'aliases': [],
        'ingredients': [
        ('Kiwi', None, 'g', ''),
        ('Pineapple', None, 'g', ''),
        ('Cucumber', None, 'g', ''),
        ('Apple', None, 'g', ''),
        ],
    },
    {
        'category': 'Mixed Juices',
        'name': 'Tropical Juice',
        'aliases': [],
        'ingredients': [
        ('Apple', None, 'g', ''),
        ('Pineapple', None, 'g', ''),
        ('Orange', None, 'g', ''),
        ('Watermelon', None, 'g', ''),
        ('Watermelon juice', None, 'ml', ''),
        ],
    },
    {
        'category': 'Mixed Juices',
        'name': 'Sugarcane Juice',
        'aliases': [],
        'ingredients': [
        ('Fresh sugarcane juice', None, 'ml', ''),
        ],
    },
    {
        'category': 'Smoothies',
        'name': 'Avocado Smoothie',
        'aliases': [],
        'ingredients': [
        ('Milk', None, 'ml', ''),
        ('Avocado', None, 'g', ''),
        ('Ice', None, 'g', ''),
        ('Sugar', None, 'g', ''),
        ('Condensed milk', None, 'g', ''),
        ],
    },
    {
        'category': 'Smoothies',
        'name': 'Strawberry Smoothie',
        'aliases': [],
        'ingredients': [
        ('Milk', None, 'ml', ''),
        ('Strawberry', None, 'g', ''),
        ('Ice', None, 'g', ''),
        ('Sugar', None, 'g', ''),
        ('Condensed milk', None, 'g', ''),
        ],
    },
    {
        'category': 'Smoothies',
        'name': 'Mango Smoothie',
        'aliases': [],
        'ingredients': [
        ('Milk', None, 'ml', ''),
        ('Mango', None, 'g', ''),
        ('Ice', None, 'g', ''),
        ('Sugar', None, 'g', ''),
        ('Condensed milk', None, 'g', ''),
        ],
    },
    {
        'category': 'Smoothies',
        'name': 'Mixed Berry Smoothie',
        'aliases': [],
        'ingredients': [
        ('Milk', None, 'ml', ''),
        ('Mixed berry', None, 'g', ''),
        ('Ice', None, 'g', ''),
        ('Sugar', None, 'g', ''),
        ('Condensed milk', None, 'g', ''),
        ],
    },
    {
        'category': 'Smoothies',
        'name': 'Coconut Smoothie',
        'aliases': [],
        'ingredients': [
        ('Milk', None, 'ml', ''),
        ('Coconut', None, 'g', ''),
        ('Ice', None, 'g', ''),
        ('Sugar', None, 'g', ''),
        ('Condensed milk', None, 'g', ''),
        ],
    },
    {
        'category': 'Vietnamese Coffee',
        'name': 'Black Coffee',
        'aliases': [],
        'ingredients': [
        ('Vietnamese coffee', None, 'g', ''),
        ('Ice', None, 'g', ''),
        ],
    },
    {
        'category': 'Vietnamese Coffee',
        'name': 'Milk Coffee',
        'aliases': [],
        'ingredients': [
        ('Vietnamese coffee', None, 'g', ''),
        ('Condensed milk', None, 'g', ''),
        ('Milk', None, 'ml', ''),
        ('Ice', None, 'g', ''),
        ],
    },
    {
        'category': 'Lemonade Drinks',
        'name': 'Kiwi Lemonade',
        'aliases': [],
        'ingredients': [],
    },
    {
        'category': 'Lemonade Drinks',
        'name': 'Strawberry Lemonade',
        'aliases': [],
        'ingredients': [],
    },
    {
        'category': 'Lemonade Drinks',
        'name': 'Watermelon Lemonade',
        'aliases': [],
        'ingredients': [],
    },
    {
        'category': 'Lemonade Drinks',
        'name': 'Coconut Lemonade',
        'aliases': [],
        'ingredients': [],
    },
    {
        'category': 'Lemonade Drinks',
        'name': 'Pineapple Lemonade',
        'aliases': [],
        'ingredients': [],
    },
    {
        'category': 'Lemonade Drinks',
        'name': 'Aloe Vera Lemonade',
        'aliases': [],
        'ingredients': [],
    },
    {
        'category': 'Banh Mi',
        'name': 'Traditional Pork Bánh Mì',
        'aliases': ['Banh Mi'],
        'ingredients': [
        ('Bread', 70, 'g', ''),
        ('Pate', 20, 'g', ''),
        ('Butter', 20, 'g', ''),
        ('Vietnamese pork loaf', 30, 'g', ''),
        ('Grilled pork', 50, 'g', ''),
        ('Sauce', 20, 'g', ''),
        ('Pickled vegetables', 35, 'g', ''),
        ('Cucumber', 25, 'g', ''),
        ('Coriander', 10, 'g', ''),
        ('Sliced chili', 5, 'g', ''),
        ],
    },
    {
        'category': 'Banh Mi',
        'name': 'Roast Pork Bánh Mì',
        'aliases': ['Banh Mi Heo Quay'],
        'ingredients': [
        ('Bread', 70, 'g', ''),
        ('Crispy roast pork', 90, 'g', ''),
        ('Pate', 20, 'g', ''),
        ('Butter', 20, 'g', ''),
        ('Sauce', 20, 'g', ''),
        ('Pickled vegetables', 35, 'g', ''),
        ('Cucumber', 25, 'g', ''),
        ('Coriander', 10, 'g', ''),
        ('Sliced chili', 5, 'g', ''),
        ],
    },
    {
        'category': 'Banh Mi',
        'name': 'Grilled Chicken Bánh Mì',
        'aliases': [],
        'ingredients': [
        ('Bread', 70, 'g', ''),
        ('Grilled chicken', 50, 'g', ''),
        ('Butter', 20, 'g', ''),
        ('Mayonnaise', 20, 'g', ''),
        ('Pickled vegetables', 35, 'g', ''),
        ('Cucumber', 25, 'g', ''),
        ('Coriander', 10, 'g', ''),
        ('Sliced chili', 5, 'g', ''),
        ],
    },
    {
        'category': 'Banh Mi',
        'name': 'Grilled Pork Bánh Mì',
        'aliases': [],
        'ingredients': [
        ('Bread', 70, 'g', ''),
        ('Grilled pork', 50, 'g', ''),
        ('Butter', 20, 'g', ''),
        ('Mayonnaise', 20, 'g', ''),
        ('Pickled vegetables', 35, 'g', ''),
        ('Cucumber', 25, 'g', ''),
        ('Coriander', 10, 'g', ''),
        ('Sliced chili', 5, 'g', ''),
        ],
    },
    {
        'category': 'Banh Mi',
        'name': 'Grilled Beef Bánh Mì',
        'aliases': [],
        'ingredients': [
        ('Bread', 70, 'g', ''),
        ('Grilled beef', 50, 'g', ''),
        ('Butter', 20, 'g', ''),
        ('Mayonnaise', 20, 'g', ''),
        ('Pickled vegetables', 35, 'g', ''),
        ('Cucumber', 25, 'g', ''),
        ('Coriander', 10, 'g', ''),
        ('Sliced chili', 5, 'g', ''),
        ],
    },
    {
        'category': 'Hotplate',
        'name': 'Chicken Sizzling Hot Plate',
        'aliases': [],
        'ingredients': [
        ('Chicken', 180, 'g', ''),
        ('Pickles', 40, 'g', ''),
        ('Onion', 30, 'g', ''),
        ('Rice', 250, 'g', ''),
        ('Sauce', 40, 'g', ''),
        ('Egg', 55, 'g', 'Sunny side up egg'),
        ],
    },
    {
        'category': 'Hotplate',
        'name': 'Beef Sizzling Hot Plate',
        'aliases': [],
        'ingredients': [
        ('Grilled beef', 120, 'g', ''),
        ('Pickles', 40, 'g', ''),
        ('Onion', 30, 'g', ''),
        ('Rice', 250, 'g', ''),
        ('Sauce', 40, 'g', ''),
        ('Egg', 55, 'g', 'Sunny side up egg'),
        ],
    },
    {
        'category': 'Hotplate',
        'name': 'Pork Sizzling Hot Plate',
        'aliases': [],
        'ingredients': [
        ('Grilled pork', 120, 'g', ''),
        ('Pickles', 40, 'g', ''),
        ('Onion', 30, 'g', ''),
        ('Rice', 250, 'g', ''),
        ('Sauce', 40, 'g', ''),
        ('Egg', 55, 'g', 'Sunny side up egg'),
        ],
    },
    {
        'category': 'Hotplate',
        'name': 'Tofu Sizzling Hot Plate',
        'aliases': [],
        'ingredients': [
        ('Fried tofu', 120, 'g', ''),
        ('Pickles', 40, 'g', ''),
        ('Onion', 30, 'g', ''),
        ('Rice', 250, 'g', ''),
        ('Sauce', 40, 'g', ''),
        ('Egg', 55, 'g', 'Sunny side up egg'),
        ],
    },
    {
        'category': 'Hotplate',
        'name': 'MCQ Sizzling Beef with Bread',
        'aliases': ['Bo Ne'],
        'ingredients': [
        ('Eggs', 100, 'g', '2 eggs total'),
        ('Sausage', 50, 'g', ''),
        ('Beef', 50, 'g', ''),
        ('Pate', 30, 'g', ''),
        ('Pickles', 40, 'g', ''),
        ('Bread', 70, 'g', ''),
        ],
    },
    {
        'category': 'Hotplate',
        'name': 'Hotplate Egg Add-on',
        'aliases': [],
        'ingredients': [
        ('Egg', 55, 'g', '1 egg for hotplate dishes'),
        ],
    },
    {
        'category': 'Add-ons',
        'name': 'Extra Tai',
        'aliases': [],
        'ingredients': [
        ('Raw beef', 50, 'g', 'Current add-on price mentioned: $3'),
        ],
    },
    {
        'category': 'Add-ons',
        'name': 'Extra Nam',
        'aliases': [],
        'ingredients': [
        ('Beef brisket slices', 50, 'g', '2 slices'),
        ],
    },
    {
        'category': 'Dry Noodles & Salad',
        'name': 'Bun Bo Nuong Cha Gio',
        'aliases': [],
        'ingredients': [
        ('Fresh vermicelli noodles', 200, 'g', ''),
        ('Salad', 60, 'g', ''),
        ('Cucumber', 40, 'g', ''),
        ('Pickles', 40, 'g', ''),
        ('Grilled beef', 90, 'g', ''),
        ('Spring roll', 65, 'g', '1 piece'),
        ('Scallion oil', 5, 'g', 'Free of charge'),
        ('Fried shallots', 5, 'g', 'Free of charge'),
        ],
    },
    {
        'category': 'Rice Dishes',
        'name': 'Grilled Chicken Rice',
        'aliases': ['Com Ga'],
        'ingredients': [
        ('Steamed rice', 250, 'g', ''),
        ('Grilled chicken', 180, 'g', ''),
        ('Cucumber', 40, 'g', ''),
        ('Tomato', 30, 'g', ''),
        ('Pickles', 45, 'g', ''),
        ('Tangy sweet sauce', 40, 'g', ''),
        ],
    },
    {
        'category': 'Rice Dishes',
        'name': 'Roast Pork Rice',
        'aliases': ['Com Heo Quay'],
        'ingredients': [
        ('Steamed rice', 250, 'g', ''),
        ('Roast pork', 150, 'g', ''),
        ('Cucumber', 40, 'g', ''),
        ('Tomato', 30, 'g', ''),
        ('Pickles', 45, 'g', ''),
        ('Tangy sweet sauce', 40, 'g', ''),
        ],
    },
    {
        'category': 'Rice Dishes',
        'name': 'Grilled Beef Rice',
        'aliases': ['Com Bo'],
        'ingredients': [
        ('Steamed rice', 250, 'g', ''),
        ('Grilled beef', 120, 'g', ''),
        ('Cucumber', 40, 'g', ''),
        ('Tomato', 30, 'g', ''),
        ('Pickles', 45, 'g', ''),
        ('Tangy sweet sauce', 40, 'g', ''),
        ],
    },
    {
        'category': 'Rice Dishes',
        'name': 'Stir-Fried Tofu Rice',
        'aliases': [],
        'ingredients': [
        ('Steamed rice', 250, 'g', ''),
        ('Stir-fried tofu', 120, 'g', ''),
        ('Cucumber', 40, 'g', ''),
        ('Tomato', 30, 'g', ''),
        ('Pickles', 45, 'g', ''),
        ('Soy sauce', 40, 'g', ''),
        ],
    },
    {
        'category': 'Rice Dishes',
        'name': 'Com Suon',
        'aliases': [],
        'ingredients': [
        ('Rice', 250, 'g', ''),
        ('Cucumber', 40, 'g', ''),
        ('Tomato', 30, 'g', ''),
        ('Pickles', 45, 'g', ''),
        ('Fish sauce', 40, 'g', ''),
        ('Pork chop', 120, 'g', '1 piece'),
        ('Egg', 55, 'g', ''),
        ],
    },
    {
        'category': 'Rice Dishes',
        'name': 'Grilled Pork Chop with Broken Rice',
        'aliases': ['Com Tam Suon Bi Cha'],
        'ingredients': [
        ('Broken rice', 250, 'g', ''),
        ('Grilled pork chop', 120, 'g', '1 piece'),
        ('Shredded pork', 30, 'g', ''),
        ('Egg meatloaf', 50, 'g', ''),
        ('Cucumber', 40, 'g', ''),
        ('Tomato', 30, 'g', ''),
        ('Pickles', 45, 'g', ''),
        ('Tangy sweet fish sauce', 40, 'g', ''),
        ],
    },
]


# ── Recipe Book batch/prep recipes (from MCQ_Restaurant_Recipe_Book.docx) ──
# These are ADD-ON costing items appended below without altering any of the
# à-la-carte menu seed above. Each ingredient keeps the exact gram/ml amount
# documented in the recipe book; non-weighable amounts (bag, can, bottle, cup,
# tbsp, pcs) keep qty blank and record the original wording in Notes.
RECIPE_BOOK_CATEGORY = 'Recipe Book (Batch Prep)'
RECIPE_BOOK_SEED = [
    {
        'category': RECIPE_BOOK_CATEGORY,
        'name': 'Chicken Marinade (15kg Chicken)',
        'aliases': [],
        'ingredients': [
            ('Soy Sauce', 370, 'g', ''),
            ('Fish Sauce', 350, 'g', ''),
            ('Sugar', 550, 'g', ''),
            ('ABC Sauce', 400, 'g', ''),
            ('Oyster Sauce', 200, 'g', ''),
            ('Cooking Wine', 75, 'g', ''),
            ('Minced Lemongrass', None, 'bag', '1 bag'),
            ('Minced Garlic', 300, 'g', ''),
            ('Cooking Oil', 500, 'g', ''),
            ('Five Spice Powder', 35, 'g', ''),
        ],
    },
    {
        'category': RECIPE_BOOK_CATEGORY,
        'name': 'Stir Fried Sauce',
        'aliases': [],
        'ingredients': [
            ('Minced Garlic', 300, 'g', ''),
            ('Cooking Wine', 500, 'g', ''),
            ('Hoisin Sauce', 1000, 'g', ''),
            ('Soy Sauce', 2000, 'g', ''),
            ('Oyster Sauce', 750, 'g', ''),
            ('ABC Sauce', 750, 'g', ''),
            ('Sugar', 1000, 'g', ''),
            ('Potato Starch', None, 'bag', '1/4 bag (add after boiling)'),
        ],
    },
    {
        'category': RECIPE_BOOK_CATEGORY,
        'name': 'Roast Pork Seasoning',
        'aliases': [],
        'ingredients': [
            ('Garlic Powder', 1000, 'g', '1kg'),
            ('Onion Powder', 1000, 'g', '1kg'),
            ('MSG', 2000, 'g', '2kg'),
            ('Five Spice Powder', 1000, 'g', '1kg'),
            ('Salt', 8000, 'g', '8kg'),
        ],
    },
    {
        'category': RECIPE_BOOK_CATEGORY,
        'name': 'Broth Seasoning (Large Pot)',
        'aliases': [],
        'ingredients': [
            ('Cinnamon', 140, 'g', ''),
            ('Star Anise', 100, 'g', ''),
            ('Black Cardamom', 40, 'g', ''),
            ('Coriander Seed', 600, 'g', ''),
            ('Fennel Seed', 60, 'g', ''),
            ('Licorice', 10, 'g', ''),
            ('Cloves', None, 'pcs', '20 pcs'),
            ('Sugar', 2000, 'g', '2kg'),
            ('Salt', 1000, 'g', '1kg'),
            ('MSG', 600, 'g', ''),
            ('Fish Sauce', None, 'cup', '1/2 cup'),
        ],
    },
    {
        'category': RECIPE_BOOK_CATEGORY,
        'name': 'Broth Seasoning (Small Pot)',
        'aliases': [],
        'ingredients': [
            ('Cinnamon', 70, 'g', ''),
            ('Star Anise', 50, 'g', ''),
            ('Black Cardamom', 20, 'g', ''),
            ('Coriander Seed', 300, 'g', ''),
            ('Fennel Seed', 30, 'g', ''),
            ('Licorice', 5, 'g', ''),
            ('Cloves', None, 'pcs', '10 pcs'),
            ('Sugar', 1000, 'g', '1kg'),
            ('Salt', 500, 'g', ''),
            ('MSG', 300, 'g', ''),
            ('Fish Sauce', None, 'cup', '1/4 cup'),
        ],
    },
    {
        'category': RECIPE_BOOK_CATEGORY,
        'name': 'Com Tam Pork Marinade Sauce',
        'aliases': [],
        'ingredients': [
            ('Minced Lemongrass', None, 'bag', '1 bag'),
            ('Minced Garlic', None, 'bag', '1 bag'),
            ('Minced Spring Onion', None, 'bag', '1 bag'),
            ('Fish Sauce', 1400, 'ml', '1.4L'),
            ('ABC Sauce', 700, 'g', ''),
            ('Soy Sauce', None, 'bottle', '1 bottle'),
            ('Condensed Milk', None, 'can', '1 can'),
            ('Yellow Food Colour', None, 'bottle', '1/2 bottle'),
            ('MSG', None, 'tbsp', '3 tbsp'),
            ('Five Spice Powder', None, 'tbsp', '3 tbsp'),
            ('Cooking Oil', 1500, 'ml', '1.5L'),
        ],
    },
]

# Append (never mutate) so existing menu costing is untouched.
FOOD_PRICING_SEED = FOOD_PRICING_SEED + RECIPE_BOOK_SEED

FULL_MENU_ORDER = [
    'Raw Beef Pho',
    'Raw Beef and Beef Balls Pho',
    'Beef Brisket Pho',
    'Slow-Cooked Beef Rib Pho',
    'MCQ Special Pho',
    'Chicken Pho',
    'Beef Balls Pho',
    'Bun Bo Hue',
    'Beef Pho Cup',
    'Chicken Pho Cup',
    'Bun Bo Hue Cup',
    'Pho Combo',
    'Grilled Pork Chop with Broken Rice',
    'Roast Pork Rice',
    'Grilled Chicken Rice',
    'Grilled Beef Rice',
    'Stir-Fried Tofu Rice',
    'Chicken Sizzling Hot Plate',
    'Beef Sizzling Hot Plate',
    'Pork Sizzling Hot Plate',
    'Tofu Sizzling Hot Plate',
    'MCQ Sizzling Beef with Bread',
    'Roast Pork Dry Noodles',
    'Grilled Lemongrass Chicken Dry Noodles',
    'Grilled Lemongrass Beef Dry Noodles',
    'Grilled Lemongrass Pork Dry Noodles',
    'Stir-Fried Tofu Dry Noodles',
    'Roast Pork Bánh Mì',
    'Traditional Pork Bánh Mì',
    'Grilled Chicken Bánh Mì',
    'Grilled Pork Bánh Mì',
    'Grilled Beef Bánh Mì',
    'Chicken Rice Paper Roll',
    'Prawn and Pork Rice Paper Roll',
    'Grilled Beef Rice Paper Roll',
    'Detox Juice',
    'Immunity Juice',
    'Sweet Beets Juice',
    'Green Glow Juice',
    'Tropical Juice',
    'Sugarcane Juice',
    'Avocado Smoothie',
    'Strawberry Smoothie',
    'Mango Smoothie',
    'Mixed Berry Smoothie',
    'Coconut Smoothie',
    'Black Coffee',
    'Milk Coffee',
    'Kiwi Lemonade',
    'Strawberry Lemonade',
    'Watermelon Lemonade',
    'Coconut Lemonade',
    'Pineapple Lemonade',
    'Aloe Vera Lemonade',
    'Rice Vermicelli Salad',
    'Bun Bo Nuong Cha Gio',
    'Com Suon',
    'Hotplate Egg Add-on',
    'Extra Tai',
    'Extra Nam',
    'Xoi',
]

OBSOLETE_PRICING_ITEMS = {'Herb Toppings:'}


def _ordered_food_pricing_seed():
    order = {name: idx for idx, name in enumerate(FULL_MENU_ORDER)}
    return sorted(
        enumerate(FOOD_PRICING_SEED),
        key=lambda pair: (order.get(pair[1]['name'], len(order) + pair[0]), pair[0])
    )


def _get_db():
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    return conn


def _is_admin():
    return session.get('role') == 'admin'


def _admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login_page'))
        if not _is_admin():
            return render_template('access_denied.html'), 403
        return f(*args, **kwargs)
    return decorated


def _float_or_none(value):
    value = (value or '').strip() if isinstance(value, str) else value
    if value in ('', None):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _float_or_default(value, default=0.0):
    parsed = _float_or_none(value)
    return default if parsed is None else parsed


def _rounded_up(amount, step=0.5):
    if amount is None or amount <= 0:
        return None
    return math.ceil((amount / step) - 1e-9) * step


def _norm_name(value):
    return re.sub(r'\s+', ' ', (value or '').strip().lower())


def _unit_weight_grams(item):
    unit_size = (item.get('unit_size') or '').upper().replace(' ', '')
    original = (item.get('original') or '').upper()

    multi = re.search(r'(\d+(?:\.\d+)?)\s*(?:G|ML)?\s*[*X]\s*(\d+)', unit_size)
    if multi:
        return float(multi.group(1)) * int(multi.group(2))

    kg = re.fullmatch(r'(\d+(?:\.\d+)?)?KG', unit_size)
    if kg:
        return float(kg.group(1) or 1) * 1000

    grams = re.fullmatch(r'(\d+(?:\.\d+)?)G', unit_size)
    if grams:
        return float(grams.group(1))

    litres = re.fullmatch(r'(\d+(?:\.\d+)?)L', unit_size)
    if litres:
        return float(litres.group(1)) * 1000

    ml = re.fullmatch(r'(\d+(?:\.\d+)?)ML', unit_size)
    if ml:
        return float(ml.group(1))

    if unit_size == 'KG' or '/KG' in original:
        return 1000

    return None


def _lookup_inventory_cost(conn, item_name, ingredient_name, weight_qty):
    ing_key = _norm_name(ingredient_name)
    if ing_key in ZERO_COST_INGREDIENTS:
        return 0.0, 'Free of charge'

    item_key = _norm_name(item_name)
    inventory_id = AUTO_COST_MAP_BY_ITEM.get((item_key, ing_key)) or AUTO_COST_MAP.get(ing_key)
    if not inventory_id or weight_qty is None:
        return None, ''

    row = conn.execute('SELECT * FROM inventory_items WHERE id=? AND active=1', (inventory_id,)).fetchone()
    if not row:
        return None, ''
    inv = dict(row)
    total_weight = _unit_weight_grams(inv)
    cost_unit = float(inv.get('cost_unit') or 0)
    if not total_weight or not cost_unit:
        return None, ''

    cost = float(weight_qty or 0) * cost_unit / total_weight
    source = f"Auto: {inv['name']} ({inv.get('unit_size') or 'unit'} @ ${cost_unit:.2f})"
    return round(cost, 4), source


def autofill_food_pricing_costs(conn):
    rows = conn.execute(
        '''SELECT ing.id, ing.ingredient_name, ing.weight_qty, ing.cost_per_serve,
                  ing.cost_source, item.name AS item_name
           FROM food_pricing_ingredients ing
           JOIN food_pricing_items item ON item.id=ing.item_id
           WHERE ing.active=1 AND item.active=1'''
    ).fetchall()
    filled = 0
    preserved = 0
    skipped = []
    for row in rows:
        source = row['cost_source'] or ''
        if row['cost_per_serve'] is not None and not (
            source.startswith('Auto:') or source == 'Free of charge'
        ):
            preserved += 1
            continue
        cost, source = _lookup_inventory_cost(
            conn, row['item_name'], row['ingredient_name'], row['weight_qty']
        )
        if cost is None:
            skipped.append(row['ingredient_name'])
            continue
        conn.execute(
            '''UPDATE food_pricing_ingredients
               SET cost_per_serve=?, cost_source=?
               WHERE id=?''',
            (cost, source, row['id'])
        )
        filled += 1
    return filled, sorted(set(skipped)), preserved


def _calc_ingredient(row):
    ing = dict(row)
    cost = float(ing.get('cost_per_serve') or 0)
    overhead_pct = float(ing.get('overhead_pct') or 0)
    ing['overhead_amount'] = cost * overhead_pct / 100
    ing['adjusted_cost'] = cost + ing['overhead_amount']
    return ing


def _calc_item(item, ingredients):
    item = dict(item)
    ingredients = [_calc_ingredient(r) for r in ingredients]
    missing_cost_count = sum(1 for i in ingredients if i.get('cost_per_serve') is None)
    costed_count = len(ingredients) - missing_cost_count
    total_cost = sum(i['cost_per_serve'] or 0 for i in ingredients)
    total_overhead = sum(i['overhead_amount'] for i in ingredients)
    total_adjusted = sum(i['adjusted_cost'] for i in ingredients)
    margin_pct = float(item.get('margin_pct') or DEFAULT_MARGIN_PCT)
    exact_price = None
    rounded_price = None
    gross_profit = None
    achieved_margin = None
    if total_adjusted > 0 and margin_pct < 100:
        exact_price = total_adjusted / (1 - margin_pct / 100)
        rounded_price = _rounded_up(exact_price)
        if rounded_price:
            gross_profit = rounded_price - total_adjusted
            achieved_margin = gross_profit / rounded_price * 100
    item.update({
        'ingredients': ingredients,
        'ingredient_count': len(ingredients),
        'costed_count': costed_count,
        'missing_cost_count': missing_cost_count,
        'total_cost': total_cost,
        'total_overhead': total_overhead,
        'total_adjusted': total_adjusted,
        'exact_price': exact_price,
        'rounded_price': rounded_price,
        'gross_profit': gross_profit,
        'achieved_margin': achieved_margin,
    })
    return item


def _existing_items_by_name(conn):
    rows = conn.execute(
        '''SELECT id, name FROM food_pricing_items
           WHERE active=1
           ORDER BY id'''
    ).fetchall()
    return {_norm_name(row['name']): row for row in rows}


def _sync_seed_ingredients(conn, item_id, ingredients):
    existing = conn.execute(
        '''SELECT * FROM food_pricing_ingredients
           WHERE item_id=? AND active=1
           ORDER BY sort_order, id''',
        (item_id,)
    ).fetchall()
    by_name = {}
    for row in existing:
        by_name.setdefault(_norm_name(row['ingredient_name']), row)

    for ing_order, (ing_name, qty, unit, notes) in enumerate(ingredients):
        key = _norm_name(ing_name)
        aliases = [_norm_name(alias) for alias in INGREDIENT_ALIASES.get(key, [])]
        row = next((by_name.get(candidate) for candidate in [key] + aliases if by_name.get(candidate)), None)
        default_overhead = 0 if 'Free of charge' in notes else DEFAULT_OVERHEAD_PCT
        if row:
            overhead = 0 if 'Free of charge' in notes else row['overhead_pct']
            conn.execute(
                '''UPDATE food_pricing_ingredients
                   SET ingredient_name=?, weight_qty=?, weight_unit=?,
                       overhead_pct=?, sort_order=?, notes=?
                   WHERE id=?''',
                (ing_name, qty, unit, overhead, ing_order, notes, row['id'])
            )
            continue

        conn.execute(
            '''INSERT INTO food_pricing_ingredients
               (item_id, ingredient_name, weight_qty, weight_unit, cost_per_serve,
                overhead_pct, sort_order, notes)
               VALUES (?,?,?,?,?,?,?,?)''',
            (item_id, ing_name, qty, unit, None, default_overhead, ing_order, notes)
        )


def _sync_food_pricing_seed(conn):
    indexed = _existing_items_by_name(conn)
    for sort_order, (_, seed) in enumerate(_ordered_food_pricing_seed()):
        names = [seed['name']] + seed.get('aliases', [])
        row = next((indexed.get(_norm_name(name)) for name in names if indexed.get(_norm_name(name))), None)
        if row:
            item_id = row['id']
            conn.execute(
                '''UPDATE food_pricing_items
                   SET category=?, name=?, sort_order=?, updated_at=datetime('now','localtime')
                   WHERE id=?''',
                (seed['category'], seed['name'], sort_order, item_id)
            )
        else:
            cur = conn.execute(
                '''INSERT INTO food_pricing_items
                   (category, name, margin_pct, sort_order, notes)
                   VALUES (?,?,?,?,?)''',
                (seed['category'], seed['name'], DEFAULT_MARGIN_PCT, sort_order, '')
            )
            item_id = cur.lastrowid

        _sync_seed_ingredients(conn, item_id, seed['ingredients'])
        indexed = _existing_items_by_name(conn)


def _seed_food_pricing(conn):
    existing = conn.execute('SELECT COUNT(*) FROM food_pricing_items').fetchone()[0]
    state = conn.execute('SELECT version FROM food_pricing_seed_state WHERE id=1').fetchone()
    current_version = state['version'] if state else 0
    if OBSOLETE_PRICING_ITEMS:
        placeholders = ','.join('?' for _ in OBSOLETE_PRICING_ITEMS)
        conn.execute(
            f'''UPDATE food_pricing_items
                SET active=0, updated_at=datetime('now','localtime')
                WHERE name IN ({placeholders})''',
            tuple(OBSOLETE_PRICING_ITEMS)
        )
    if existing and current_version >= FOOD_PRICING_SEED_VERSION:
        return

    _sync_food_pricing_seed(conn)
    conn.execute(
        '''INSERT INTO food_pricing_seed_state (id, version)
           VALUES (1, ?)
           ON CONFLICT(id) DO UPDATE SET version=excluded.version''',
        (FOOD_PRICING_SEED_VERSION,)
    )


def init_food_pricing_tables(db_path):
    global DB_PATH
    DB_PATH = db_path
    with sqlite3.connect(db_path, timeout=30) as conn:
        conn.row_factory = sqlite3.Row
        conn.execute('''CREATE TABLE IF NOT EXISTS food_pricing_items (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            category    TEXT NOT NULL DEFAULT 'Menu',
            name        TEXT NOT NULL,
            margin_pct  REAL NOT NULL DEFAULT 50,
            sort_order  INTEGER NOT NULL DEFAULT 0,
            notes       TEXT NOT NULL DEFAULT '',
            active      INTEGER NOT NULL DEFAULT 1,
            created_at  TEXT DEFAULT (datetime('now','localtime')),
            updated_at  TEXT DEFAULT (datetime('now','localtime'))
        )''')
        conn.execute('''CREATE TABLE IF NOT EXISTS food_pricing_ingredients (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id         INTEGER NOT NULL REFERENCES food_pricing_items(id) ON DELETE CASCADE,
            ingredient_name TEXT NOT NULL,
            weight_qty      REAL,
            weight_unit     TEXT NOT NULL DEFAULT 'g',
            cost_per_serve  REAL,
            overhead_pct    REAL NOT NULL DEFAULT 50,
            sort_order      INTEGER NOT NULL DEFAULT 0,
            notes           TEXT NOT NULL DEFAULT '',
            active          INTEGER NOT NULL DEFAULT 1
        )''')
        conn.execute('''CREATE TABLE IF NOT EXISTS food_pricing_seed_state (
            id       INTEGER PRIMARY KEY CHECK(id=1),
            version  INTEGER NOT NULL DEFAULT 0
        )''')
        cols = [r[1] for r in conn.execute('PRAGMA table_info(food_pricing_ingredients)').fetchall()]
        if 'cost_source' not in cols:
            conn.execute("ALTER TABLE food_pricing_ingredients ADD COLUMN cost_source TEXT NOT NULL DEFAULT ''")
        if 'inventory_item_id' not in cols:
            conn.execute("ALTER TABLE food_pricing_ingredients ADD COLUMN inventory_item_id INTEGER")
        conn.execute('CREATE INDEX IF NOT EXISTS idx_food_pricing_item_category ON food_pricing_items(category, sort_order, name)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_food_pricing_ing_item ON food_pricing_ingredients(item_id, sort_order, id)')
        conn.execute('''DELETE FROM food_pricing_ingredients
                        WHERE item_id NOT IN (SELECT id FROM food_pricing_items)''')
        _seed_food_pricing(conn)
        autofill_food_pricing_costs(conn)


def _fetch_items(conn, category='', search=''):
    where = ['active=1']
    params = []
    if category and category != 'All':
        where.append('category=?')
        params.append(category)
    if search:
        where.append('(name LIKE ? OR category LIKE ? OR notes LIKE ?)')
        like = f'%{search}%'
        params.extend([like, like, like])
    rows = conn.execute(
        f'''SELECT * FROM food_pricing_items
            WHERE {' AND '.join(where)}
            ORDER BY sort_order, category, name''',
        params
    ).fetchall()
    items = []
    for row in rows:
        ing = conn.execute(
            '''SELECT * FROM food_pricing_ingredients
               WHERE item_id=? AND active=1
               ORDER BY sort_order, id''',
            (row['id'],)
        ).fetchall()
        items.append(_calc_item(row, ing))
    return items


@food_pricing.route('/')
@_admin_required
def food_pricing_home():
    category = request.args.get('category', 'All')
    search = request.args.get('q', '').strip()
    selected_id = request.args.get('item_id', type=int)

    with _get_db() as conn:
        categories = [r['category'] for r in conn.execute(
            '''SELECT category FROM food_pricing_items
               WHERE active=1
               GROUP BY category
               ORDER BY MIN(sort_order), category'''
        ).fetchall()]
        items = _fetch_items(conn, category, search)

    selected_item = None
    if items:
        selected_item = next((i for i in items if i['id'] == selected_id), None) or items[0]

    totals = {
        'items': len(items),
        'ingredients': sum(i['ingredient_count'] for i in items),
        'ready': sum(1 for i in items if i['ingredient_count'] and i['missing_cost_count'] == 0),
    }
    return render_template(
        'food_pricing.html',
        items=items,
        selected_item=selected_item,
        categories=categories,
        selected_category=category,
        search=search,
        totals=totals,
        default_margin_pct=DEFAULT_MARGIN_PCT,
        default_overhead_pct=DEFAULT_OVERHEAD_PCT,
    )


@food_pricing.route('/item/add', methods=['POST'])
@_admin_required
def add_item():
    name = request.form.get('name', '').strip()
    if not name:
        return redirect(url_for('food_pricing.food_pricing_home'))
    category = request.form.get('category', '').strip() or 'Menu'
    margin_pct = _float_or_default(request.form.get('margin_pct'), DEFAULT_MARGIN_PCT)
    notes = request.form.get('notes', '').strip()
    with _get_db() as conn:
        sort_order = conn.execute(
            'SELECT COALESCE(MAX(sort_order), -1) + 1 FROM food_pricing_items WHERE active=1'
        ).fetchone()[0]
        cur = conn.execute(
            '''INSERT INTO food_pricing_items
               (category, name, margin_pct, sort_order, notes)
               VALUES (?,?,?,?,?)''',
            (category, name, margin_pct, sort_order, notes)
        )
        item_id = cur.lastrowid
    return redirect(url_for('food_pricing.food_pricing_home', item_id=item_id, category=category))


@food_pricing.route('/item/<int:item_id>/update', methods=['POST'])
@_admin_required
def update_item(item_id):
    name = request.form.get('name', '').strip()
    if not name:
        return redirect(url_for('food_pricing.food_pricing_home', item_id=item_id))
    category = request.form.get('category', '').strip() or 'Menu'
    margin_pct = _float_or_default(request.form.get('margin_pct'), DEFAULT_MARGIN_PCT)
    sort_order = int(_float_or_default(request.form.get('sort_order'), 0))
    notes = request.form.get('notes', '').strip()
    with _get_db() as conn:
        conn.execute(
            '''UPDATE food_pricing_items
               SET category=?, name=?, margin_pct=?, sort_order=?, notes=?,
                   updated_at=datetime('now','localtime')
               WHERE id=?''',
            (category, name, margin_pct, sort_order, notes, item_id)
        )
    return redirect(url_for('food_pricing.food_pricing_home', item_id=item_id, category=category))


@food_pricing.route('/item/<int:item_id>/delete', methods=['POST'])
@_admin_required
def delete_item(item_id):
    with _get_db() as conn:
        conn.execute('DELETE FROM food_pricing_ingredients WHERE item_id=?', (item_id,))
        conn.execute('DELETE FROM food_pricing_items WHERE id=?', (item_id,))
    return redirect(url_for('food_pricing.food_pricing_home'))


@food_pricing.route('/ingredient/add', methods=['POST'])
@_admin_required
def add_ingredient():
    item_id = request.form.get('item_id', type=int)
    if not item_id:
        return redirect(url_for('food_pricing.food_pricing_home'))
    name = request.form.get('ingredient_name', '').strip()
    if not name:
        return redirect(url_for('food_pricing.food_pricing_home', item_id=item_id))
    weight_qty = _float_or_none(request.form.get('weight_qty'))
    weight_unit = request.form.get('weight_unit', '').strip() or 'g'
    cost_per_serve = _float_or_none(request.form.get('cost_per_serve'))
    overhead_pct = _float_or_default(request.form.get('overhead_pct'), DEFAULT_OVERHEAD_PCT)
    notes = request.form.get('notes', '').strip()
    with _get_db() as conn:
        sort_order = conn.execute(
            'SELECT COALESCE(MAX(sort_order), -1) + 1 FROM food_pricing_ingredients WHERE item_id=?',
            (item_id,)
        ).fetchone()[0]
        conn.execute(
            '''INSERT INTO food_pricing_ingredients
               (item_id, ingredient_name, weight_qty, weight_unit, cost_per_serve,
                overhead_pct, sort_order, notes)
               VALUES (?,?,?,?,?,?,?,?)''',
            (item_id, name, weight_qty, weight_unit, cost_per_serve, overhead_pct, sort_order, notes)
        )
        category = conn.execute('SELECT category FROM food_pricing_items WHERE id=?', (item_id,)).fetchone()
    return redirect(url_for('food_pricing.food_pricing_home', item_id=item_id, category=category['category'] if category else 'All'))


@food_pricing.route('/ingredient/<int:ingredient_id>/update', methods=['POST'])
@_admin_required
def update_ingredient(ingredient_id):
    item_id = request.form.get('item_id', type=int)
    name = request.form.get('ingredient_name', '').strip()
    if not item_id or not name:
        return redirect(url_for('food_pricing.food_pricing_home'))
    weight_qty = _float_or_none(request.form.get('weight_qty'))
    weight_unit = request.form.get('weight_unit', '').strip() or 'g'
    cost_per_serve = _float_or_none(request.form.get('cost_per_serve'))
    overhead_pct = _float_or_default(request.form.get('overhead_pct'), DEFAULT_OVERHEAD_PCT)
    sort_order = int(_float_or_default(request.form.get('sort_order'), 0))
    notes = request.form.get('notes', '').strip()
    with _get_db() as conn:
        conn.execute(
            '''UPDATE food_pricing_ingredients
               SET ingredient_name=?, weight_qty=?, weight_unit=?, cost_per_serve=?,
                   overhead_pct=?, sort_order=?, notes=?, cost_source=''
               WHERE id=?''',
            (name, weight_qty, weight_unit, cost_per_serve, overhead_pct,
             sort_order, notes, ingredient_id)
        )
        category = conn.execute('SELECT category FROM food_pricing_items WHERE id=?', (item_id,)).fetchone()
    return redirect(url_for('food_pricing.food_pricing_home', item_id=item_id, category=category['category'] if category else 'All'))


@food_pricing.route('/ingredient/<int:ingredient_id>/delete', methods=['POST'])
@_admin_required
def delete_ingredient(ingredient_id):
    item_id = request.form.get('item_id', type=int)
    with _get_db() as conn:
        if not item_id:
            row = conn.execute('SELECT item_id FROM food_pricing_ingredients WHERE id=?', (ingredient_id,)).fetchone()
            item_id = row['item_id'] if row else None
        conn.execute('DELETE FROM food_pricing_ingredients WHERE id=?', (ingredient_id,))
        category = conn.execute('SELECT category FROM food_pricing_items WHERE id=?', (item_id,)).fetchone() if item_id else None
    return redirect(url_for('food_pricing.food_pricing_home', item_id=item_id or '', category=category['category'] if category else 'All'))


@food_pricing.route('/autofill-costs', methods=['POST'])
@_admin_required
def autofill_costs():
    with _get_db() as conn:
        filled, skipped, preserved = autofill_food_pricing_costs(conn)
    flash(f'Auto-filled {filled} ingredient cost line(s) from Item List. Preserved {preserved} manual cost line(s). {len(skipped)} unmatched ingredient type(s) left for manual cost.', 'success')
    return redirect(url_for('food_pricing.food_pricing_home'))


# ── Excel export ──────────────────────────────────────────────────────────
# Builds a live, formula-linked workbook (not a static snapshot):
#   • "Cost Sheets" tab — one costing block per menu item. The per-1g/1ml unit
#     cost drives the portion cost, cost+ (labour), the totals and the sell
#     price, so editing any yellow input recalculates the whole sheet in Excel.
#   • "Menu Summary" tab — one row per item whose figures are =links back to the
#     matching Cost Sheets cells, so the summary updates the moment a detail cell
#     changes.
# Colour palette + banding follow the Item List export for a consistent look.
GREEN_DARK = '1B4332'
GREEN_MID = '2D6A4F'
GREEN_BAND = '1B3A2D'
GREEN_LIGHT = 'E8F5E9'
INPUT_FILL = 'FFF8E1'     # soft amber = editable input cell
CALC_FILL = 'F2F8F5'      # soft green = formula / calculated cell
HEADER_TXT = 'FFFFFF'
MONEY_FMT = '$#,##0.00'
UNIT_FMT = '$#,##0.0000'
WEIGHT_FMT = '#,##0.###'
PCT_FMT = '0.0'


@food_pricing.route('/export')
@_admin_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    category = request.args.get('category', 'All')
    search = request.args.get('q', '').strip()
    with _get_db() as conn:
        items = _fetch_items(conn, category, search)

    thin = Side(style='thin', color='D5E2DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor=GREEN_MID)
    band_fill = PatternFill('solid', fgColor=GREEN_BAND)
    input_fill = PatternFill('solid', fgColor=INPUT_FILL)
    calc_fill = PatternFill('solid', fgColor=CALC_FILL)
    title_fill = PatternFill('solid', fgColor=GREEN_DARK)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')
    left_mid = Alignment(horizontal='left', vertical='center', wrap_text=True)

    wb = Workbook()
    summary = wb.active
    summary.title = 'Menu Summary'
    sheet = wb.create_sheet('Cost Sheets')
    SHEET = "'Cost Sheets'"
    today_str = perth_today().strftime('%d %b %Y')

    # ═══ Cost Sheets tab ═══
    # A No | B Ingredient | C Weight | D Unit | E Cost/1g·1ml | F Portion Cost
    # G Labour% | H Cost+ | I Notes
    cs_headers = ['No', 'Ingredient', 'Weight', 'Unit', 'Cost / 1g·1ml ($)',
                  'Portion Cost ($)', 'Labour & Exp %', 'Cost + ($)', 'Notes']
    cs_widths = [5, 30, 10, 8, 16, 15, 13, 14, 34]
    for i, w in enumerate(cs_widths, 1):
        sheet.column_dimensions[get_column_letter(i)].width = w

    sheet.merge_cells('A1:I1')
    t = sheet['A1']
    t.value = f'MCQ Mirrabooka — Food Costing Sheets   (exported {today_str})'
    t.font = Font(bold=True, size=13, color=HEADER_TXT)
    t.fill = title_fill
    t.alignment = center
    sheet.row_dimensions[1].height = 26
    sheet.merge_cells('A2:I2')
    hint = sheet['A2']
    hint.value = ('Amber cells are inputs (weight, cost per 1g/1ml, labour %, margin %). '
                  'Green cells are live formulas — edit an input and Excel recalculates portion cost, cost+, totals and the sell price.')
    hint.font = Font(italic=True, size=9, color='5A6B62')
    hint.alignment = left_mid
    sheet.row_dimensions[2].height = 26

    refs = []           # per-item cell references for the summary tab
    cur_cat = None
    row = 4
    for item in items:
        if item['category'] != cur_cat:
            cur_cat = item['category']
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            c = sheet.cell(row=row, column=1, value=f'  {cur_cat}')
            c.font = Font(bold=True, color=HEADER_TXT, size=11)
            c.fill = band_fill
            c.alignment = Alignment(vertical='center')
            sheet.row_dimensions[row].height = 20
            row += 1

        # Item title
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        c = sheet.cell(row=row, column=1, value=f'  {item["name"]}')
        c.font = Font(bold=True, color=GREEN_DARK, size=11)
        c.fill = PatternFill('solid', fgColor=GREEN_LIGHT)
        c.alignment = Alignment(vertical='center')
        sheet.row_dimensions[row].height = 18
        row += 1

        # Column headers for this block
        for col, h in enumerate(cs_headers, 1):
            hc = sheet.cell(row=row, column=col, value=h)
            hc.font = Font(bold=True, color=HEADER_TXT, size=9)
            hc.fill = header_fill
            hc.alignment = center
            hc.border = border
        sheet.row_dimensions[row].height = 24
        row += 1

        first_ing = row
        ingredients = item['ingredients']
        if not ingredients:
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            c = sheet.cell(row=row, column=1, value='  (no ingredient lines yet)')
            c.font = Font(italic=True, color='9AA79F', size=9)
            c.alignment = Alignment(vertical='center')
            c.border = border
            last_ing = row
            row += 1
        else:
            for idx, ing in enumerate(ingredients, 1):
                r = row
                qty = ing.get('weight_qty')
                cost = ing.get('cost_per_serve')
                has_weight = qty not in (None, '') and float(qty) > 0
                unit_cost = (float(cost) / float(qty)) if (has_weight and cost not in (None, '')) else None

                sheet.cell(row=r, column=1, value=idx).alignment = center
                bcell = sheet.cell(row=r, column=2, value=ing.get('ingredient_name') or '')
                bcell.font = Font(bold=True, color='1B3A2D', size=9)
                bcell.alignment = left_mid
                # C weight (input)
                wc = sheet.cell(row=r, column=3, value=(float(qty) if has_weight else None))
                wc.number_format = WEIGHT_FMT
                wc.alignment = right
                if has_weight:
                    wc.fill = input_fill
                sheet.cell(row=r, column=4, value=ing.get('weight_unit') or '').alignment = center
                # E cost per 1g/1ml (input driver)
                ec = sheet.cell(row=r, column=5, value=unit_cost)
                ec.number_format = UNIT_FMT
                ec.alignment = right
                if has_weight:
                    ec.fill = input_fill
                # F portion cost — formula when weighted, else literal cost
                fc = sheet.cell(row=r, column=6)
                if has_weight:
                    fc.value = f'=IF(OR($C{r}="",$E{r}=""),"",$C{r}*$E{r})'
                    fc.fill = calc_fill
                else:
                    fc.value = (float(cost) if cost not in (None, '') else None)
                    if cost not in (None, ''):
                        fc.fill = input_fill
                fc.number_format = MONEY_FMT
                fc.alignment = right
                # G labour % (input)
                gc = sheet.cell(row=r, column=7, value=float(ing.get('overhead_pct') or 0))
                gc.number_format = PCT_FMT
                gc.alignment = right
                gc.fill = input_fill
                # H cost+ (formula)
                hc = sheet.cell(row=r, column=8, value=f'=IF($F{r}="","",$F{r}*(1+$G{r}/100))')
                hc.number_format = MONEY_FMT
                hc.alignment = right
                hc.fill = calc_fill
                # I notes
                nc = sheet.cell(row=r, column=9, value=ing.get('notes') or '')
                nc.font = Font(color='6A756F', size=9)
                nc.alignment = left_mid
                if ing.get('cost_source'):
                    nc.value = (nc.value + ('  ·  ' if nc.value else '') + ing['cost_source'])
                for col in range(1, 10):
                    sheet.cell(row=r, column=col).border = border
                sheet.row_dimensions[r].height = 16
                row += 1
            last_ing = row - 1

        # Subtotal row
        sr = row
        sheet.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=5)
        lab = sheet.cell(row=sr, column=1, value='TOTAL')
        lab.font = Font(bold=True, color=GREEN_DARK, size=9)
        lab.alignment = right
        fsum = sheet.cell(row=sr, column=6, value=f'=SUM(F{first_ing}:F{last_ing})')
        fsum.number_format = MONEY_FMT
        fsum.font = Font(bold=True, color='1565C0', size=9)
        fsum.alignment = right
        sheet.cell(row=sr, column=7, value='').alignment = right
        hsum = sheet.cell(row=sr, column=8, value=f'=SUM(H{first_ing}:H{last_ing})')
        hsum.number_format = MONEY_FMT
        hsum.font = Font(bold=True, color='00796B', size=9)
        hsum.alignment = right
        for col in range(1, 10):
            cc = sheet.cell(row=sr, column=col)
            cc.fill = PatternFill('solid', fgColor='F1F8F4')
            cc.border = border
        row += 1

        # Pricing block (Margin % → Exact → Rounded sell price)
        def _price_row(label, value, fmt, bold=False, is_input=False):
            nonlocal row
            rr = row
            sheet.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=7)
            lc = sheet.cell(row=rr, column=1, value=label)
            lc.font = Font(bold=bold, color=GREEN_DARK, size=9)
            lc.alignment = right
            vc = sheet.cell(row=rr, column=8, value=value)
            vc.number_format = fmt
            vc.alignment = right
            vc.font = Font(bold=bold, size=10, color=('C62828' if bold else '213029'))
            vc.fill = input_fill if is_input else calc_fill
            for col in range(1, 10):
                sheet.cell(row=rr, column=col).border = border
            row += 1
            return rr

        margin_r = _price_row('Margin Target %', float(item.get('margin_pct') or DEFAULT_MARGIN_PCT), PCT_FMT, is_input=True)
        exact_r = _price_row('Exact Sell Price',
                             f'=IF($H{sr}=0,"",IF($H{margin_r}>=100,"",$H{sr}/(1-$H{margin_r}/100)))',
                             MONEY_FMT)
        sell_r = _price_row('Sell Price (rounded to $0.50)',
                            f'=IF($H{exact_r}="","",CEILING($H{exact_r},0.5))',
                            MONEY_FMT, bold=True)

        refs.append({
            'category': item['category'],
            'name': item['name'],
            'count': item['ingredient_count'],
            'base': f'{SHEET}!F{sr}',
            'plus': f'{SHEET}!H{sr}',
            'margin': f'{SHEET}!H{margin_r}',
            'exact': f'{SHEET}!H{exact_r}',
            'sell': f'{SHEET}!H{sell_r}',
        })
        row += 1  # spacer between items

    sheet.freeze_panes = 'A4'

    # ═══ Menu Summary tab ═══
    # A # | B Menu Item | C Ingredients | D Base Cost | E Cost+ | F Margin%
    # G Exact | H Sell Price  (D–H are =links into Cost Sheets)
    sm_headers = ['#', 'Menu Item', 'Ingredients', 'Base Cost',
                  'Cost + (Labour)', 'Margin %', 'Exact Price', 'Sell Price']
    sm_widths = [5, 34, 12, 15, 16, 11, 14, 14]
    for i, w in enumerate(sm_widths, 1):
        summary.column_dimensions[get_column_letter(i)].width = w

    summary.merge_cells('A1:H1')
    t = summary['A1']
    t.value = f'MCQ Mirrabooka — Menu Pricing Summary   (exported {today_str})'
    t.font = Font(bold=True, size=13, color=HEADER_TXT)
    t.fill = title_fill
    t.alignment = center
    summary.row_dimensions[1].height = 26

    for col, h in enumerate(sm_headers, 1):
        c = summary.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color=HEADER_TXT, size=10)
        c.fill = header_fill
        c.alignment = center
        c.border = border
    summary.row_dimensions[2].height = 22

    cur_cat = None
    r = 3
    seq = 0
    for ref in refs:
        if ref['category'] != cur_cat:
            cur_cat = ref['category']
            summary.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            c = summary.cell(row=r, column=1, value=f'  {cur_cat}')
            c.font = Font(bold=True, color=HEADER_TXT, size=10)
            c.fill = band_fill
            c.alignment = Alignment(vertical='center')
            summary.row_dimensions[r].height = 18
            r += 1
            seq = 0
        seq += 1
        rowfill = PatternFill('solid', fgColor='FFFFFF') if seq % 2 == 0 else PatternFill('solid', fgColor='F7FCF9')
        summary.cell(row=r, column=1, value=seq).alignment = center
        nc = summary.cell(row=r, column=2, value=ref['name'])
        nc.font = Font(bold=True, color='1B3A2D', size=9)
        nc.alignment = left_mid
        summary.cell(row=r, column=3, value=ref['count']).alignment = center
        pairs = [
            (4, f'={ref["base"]}', MONEY_FMT, '1565C0'),
            (5, f'={ref["plus"]}', MONEY_FMT, '00796B'),
            (6, f'={ref["margin"]}', PCT_FMT, '5A6B62'),
            (7, f'={ref["exact"]}', MONEY_FMT, '5A6B62'),
            (8, f'={ref["sell"]}', MONEY_FMT, 'C62828'),
        ]
        for col, formula, fmt, color in pairs:
            c = summary.cell(row=r, column=col, value=formula)
            c.number_format = fmt
            c.alignment = right
            c.font = Font(bold=(col == 8), color=color, size=9)
        for col in range(1, 9):
            cc = summary.cell(row=r, column=col)
            cc.fill = rowfill
            cc.border = border
        summary.row_dimensions[r].height = 16
        r += 1

    summary.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    foot = summary.cell(row=r, column=1, value=f'  Total: {len(refs)} menu / recipe item(s) · figures link live to the Cost Sheets tab')
    foot.font = Font(bold=True, size=9, color=GREEN_DARK)
    foot.fill = PatternFill('solid', fgColor=GREEN_LIGHT)
    foot.alignment = Alignment(vertical='center')
    summary.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'MCQ_FoodPricing_{perth_today().isoformat()}.xlsx'
    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'}
    )
