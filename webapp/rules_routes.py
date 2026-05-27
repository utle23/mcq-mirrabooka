from flask import Blueprint, render_template, request, redirect, url_for, session, Response
import sqlite3, io, re, html as html_mod
from datetime import datetime
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

rules_bp = Blueprint('rules', __name__, url_prefix='/rules')
DB_PATH = None

RULES_SEED = [
    (1, 'Restaurant Departments',
     'The restaurant is divided into two main departments:\n\n'
     'FOH — Front of House\nManaged by Do Nguyen.\n\n'
     'BOH — Back of House\nManaged by Ms. Phung.\n\n'
     'Each staff member must follow the instructions of the person responsible for their department.'),

    (2, 'Daily Checklist and iPad Record',
     'All staff must complete their assigned checklist tasks every working day.\n\n'
     'Checklist tasks must be completed and recorded on the restaurant iPad.\n\n'
     'Staff must complete the checklist before the required deadline:\n'
     '- Morning checklist: must be completed before 10:30 AM.\n'
     '- Afternoon checklist: must be completed before 6:30 PM.\n\n'
     'Staff must not submit checklist records late unless there is a valid reason and the manager has been informed.\n\n'
     'The checklist must be completed honestly and accurately. Staff must not tick a task as done if it has not been completed.\n\n'
     'At the end of the day, a designated person will double-check and verify your checklist.'),

    (3, 'Temperature Record Policy',
     'Temperature records must be completed on the iPad by the responsible staff only.\n\n'
     'The following positions are responsible for temperature records:\n'
     '- Cashier\n'
     '- Banh Mi staff\n'
     '- Chef\n\n'
     'Temperature records must be completed before the required deadline:\n'
     '- Morning temperature record: before 10:30 AM.\n'
     '- Afternoon temperature record: before 6:30 PM.\n\n'
     'If the temperature is incorrect or unsafe, staff must report the issue immediately through the iPad and inform the manager.'),

    (4, 'iPad Issue Reporting',
     'The restaurant iPad must be used to report work-related issues.\n\n'
     'Staff must report issues through the iPad when there is a problem such as:\n'
     '- Low stock\n'
     '- Missing ingredients\n'
     '- Supplier delivery issues\n'
     '- Food quality issues\n'
     '- Equipment problems\n'
     '- Cleaning issues\n'
     '- Temperature issues\n'
     '- Salary issue\n'
     '- Conflict\n'
     '- Any task that cannot be completed\n\n'
     'If the issue is urgent, staff must also inform the manager immediately.'),

    (5, 'Work Assignment Principles',
     'Each staff member must perform the tasks assigned to their position.\n\n'
     'Staff must not interfere with or overlap with another staff member\'s responsibilities.\n\n'
     'Each position has its own role to make sure work is completed efficiently and smoothly.\n\n'
     'Staff may assist with other tasks only when:\n'
     '- Requested by a manager; or\n'
     '- Assigned to help when they are available.\n\n'
     'Staff are not allowed to leave their assigned position without permission.'),

    (6, 'Movement During Work Hours',
     'If staff need to leave their workstation or go outside, they must inform the correct person:\n'
     '- BOH staff must inform Ms. Phung.\n'
     '- FOH staff must inform Do Nguyen.\n\n'
     'Bathroom breaks are limited to a maximum of 5 minutes (over 5 minutes only in special cases).\n\n'
     'Staff must return to their workstation as soon as possible.'),

    (7, 'Morning Breakfast Policy',
     'Staff may have a light snack in the morning.\n\n'
     'Morning breakfast must be from shop food only.\n\n'
     'Staff are not allowed to take or eat supplier food, including but not limited to:\n'
     '- Supplier pastry food\n'
     '- Supplier cakes\n'
     '- Supplier snacks\n'
     '- Any food delivered by suppliers for sale or restaurant use\n\n'
     'Supplier food is not staff food unless approved by the manager.'),

    (8, 'Lunch Meal Policy',
     'Each staff member is allowed one lunch meal per day.\n\n'
     'Staff must inform the FOH lead or manager before taking their lunch meal.\n\n'
     'Staff are not allowed to take lunch breaks at the same time if there is no replacement for their position (FOH).\n\n'
     'Any food or drinks outside the provided staff meal must be paid for.\n\n'
     'Staff must not take extra food without approval.'),

    (9, 'Break and Time Deduction Policy',
     'Each shift will have 15 minutes deducted for meal time.\n\n'
     'This 15-minute deduction includes the time used for staff food breaks, including morning snack time and lunch meal time.\n\n'
     'This deduction is separate from any paid rest break entitlements required under applicable workplace laws, award rules, or employment conditions.\n\n'
     'Staff must take breaks responsibly and must not leave their position unattended.'),

    (10, 'Uniform Policy',
     'All staff must wear the correct uniform during working hours.\n\n'
     'Staff must:\n'
     '- Wear the MCQ uniform shirt.\n'
     '- Wear the MCQ cap.\n'
     '- Wear suitable work shoes.\n'
     '- Maintain personal hygiene.\n'
     '- Keep hair tidy and suitable for food handling.\n'
     '- Keep their workstation clean and organized.\n\n'
     'Staff who do not wear the correct uniform may be reminded or warned by the manager.'),

    (11, 'Phone Usage Policy',
     'Personal phone use is not allowed during working hours.\n\n'
     'Phones may only be used for work-related purposes.\n\n'
     'Staff may use the restaurant iPad for:\n'
     '- Checklist tasks\n'
     '- Temperature records\n'
     '- Issue reporting\n'
     '- Work-related communication\n\n'
     'Staff must inform the manager before using a personal phone for work purposes.'),

    (12, 'Code of Conduct',
     '**Respect**\n'
     'Staff must communicate with colleagues, managers, suppliers, and customers in a polite, respectful, and cooperative manner.\n\n'
     '**No Abuse**\n'
     'The following behaviours are strictly prohibited:\n'
     '- Swearing\n'
     '- Using offensive language\n'
     '- Insulting colleagues or customers\n'
     '- Arguing aggressively\n'
     '- Disrespecting managers or staff\n\n'
     '**Teamwork**\n'
     'All staff must:\n'
     '- Support each other when needed.\n'
     '- Follow their assigned position.\n'
     '- Maintain a positive and cooperative working environment.\n'
     '- Avoid unnecessary conflict during work hours.'),

    (13, 'General Work Rules',
     'During working hours:\n'
     '- No personal activities.\n'
     '- No unnecessary conversations.\n'
     '- No personal phone use.\n'
     '- Staff must stay focused on their tasks.\n\n'
     'Work areas must be kept clean and organized at all times.\n\n'
     'Checklist tasks must be completed on time.\n\n'
     'Issues must be reported through the iPad.\n\n'
     'Staff must follow food safety and hygiene standards.\n\n'
     'Staff must not take restaurant food, supplier food, or drinks without permission.'),

    (14, 'Conflict Resolution',
     'If any issues or conflicts arise:\n'
     '- Staff should discuss the matter calmly and respectfully.\n'
     '- If the issue cannot be resolved, staff must report it to the manager.\n'
     '- Staff must not argue in front of customers.\n'
     '- Staff must not involve customers in internal restaurant issues.'),

    (15, 'Warning Policy',
     '**First offense:** Verbal warning\n'
     '**Second offense:** Official warning\n'
     '**Third offense:** Final warning\n'
     '**After three warnings:** Employment may be terminated\n\n'
     'Serious misconduct may result in stronger action depending on the situation.'),

    (16, 'Important Reminder',
     'All staff are expected to:\n'
     '- Follow the daily checklist.\n'
     '- Complete iPad records on time.\n'
     '- Record temperature correctly if assigned.\n'
     '- Report issues honestly.\n'
     '- Wear the correct uniform.\n'
     '- Stay in their assigned position.\n'
     '- Respect managers, staff, customers, and suppliers.\n'
     '- Keep the restaurant clean, safe, and organized.'),
]

# ── Helpers ────────────────────────────────────────────────────────────────────

def _get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def _is_admin():
    return session.get('role') == 'admin'

def _login_required(f):
    @wraps(f)
    def d(*a, **kw):
        if not session.get('logged_in'):
            return redirect(url_for('login_page'))
        return f(*a, **kw)
    return d

def _admin_required(f):
    @wraps(f)
    def d(*a, **kw):
        if not session.get('logged_in'):
            return redirect(url_for('login_page'))
        if session.get('role') != 'admin':
            return render_template('access_denied.html'), 403
        return f(*a, **kw)
    return d

# ── DB Init ────────────────────────────────────────────────────────────────────

def init_rules_tables(db_path):
    global DB_PATH
    DB_PATH = db_path
    with _get_db() as conn:
        conn.execute('''CREATE TABLE IF NOT EXISTS restaurant_rules (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            section_num INTEGER NOT NULL,
            title       TEXT NOT NULL,
            content     TEXT NOT NULL DEFAULT '',
            sort_order  INTEGER DEFAULT 0,
            active      INTEGER DEFAULT 1,
            updated_at  TEXT DEFAULT (datetime('now','localtime')),
            updated_by  TEXT DEFAULT ''
        )''')
        if conn.execute('SELECT COUNT(*) FROM restaurant_rules').fetchone()[0] == 0:
            for i, (num, title, content) in enumerate(RULES_SEED):
                conn.execute(
                    'INSERT INTO restaurant_rules (section_num, title, content, sort_order) VALUES (?,?,?,?)',
                    (num, title, content, i))

# ── Content renderer ───────────────────────────────────────────────────────────

def render_content(content):
    """Convert plain-text rule content to HTML."""
    result = []
    paragraphs = re.split(r'\n{2,}', (content or '').strip())
    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        lines = para.split('\n')
        bullets = []
        for line in lines:
            line = line.strip()
            if not line:
                continue
            if line.startswith('- '):
                bullets.append(html_mod.escape(line[2:].strip()))
            else:
                if bullets:
                    items = ''.join(f'<li>{b}</li>' for b in bullets)
                    result.append(f'<ul class="rule-list">{items}</ul>')
                    bullets = []
                # **Text** alone → sub-header
                if re.match(r'^\*\*[^*]+\*\*\s*$', line):
                    text = html_mod.escape(line.strip('*').strip())
                    result.append(f'<div class="rule-sub-hdr">{text}</div>')
                # **Key:** value → bold key + value
                elif re.match(r'^\*\*[^*]+\*\*', line):
                    m = re.match(r'^\*\*([^*]+)\*\*:?\s*(.*)', line)
                    if m:
                        key = html_mod.escape(m.group(1).rstrip(':'))
                        val = html_mod.escape(m.group(2).strip())
                        result.append(
                            f'<p class="rule-kv"><span class="rule-k">{key}</span>'
                            f'{": " + val if val else ""}</p>')
                    else:
                        result.append(f'<p>{html_mod.escape(line)}</p>')
                else:
                    result.append(f'<p>{html_mod.escape(line)}</p>')
        if bullets:
            items = ''.join(f'<li>{b}</li>' for b in bullets)
            result.append(f'<ul class="rule-list">{items}</ul>')
    return ''.join(result)

# ── Routes ─────────────────────────────────────────────────────────────────────

@rules_bp.route('/')
@_login_required
def rules_view():
    with _get_db() as conn:
        sections = [dict(r) for r in conn.execute(
            'SELECT * FROM restaurant_rules WHERE active=1 ORDER BY sort_order, section_num').fetchall()]
    for s in sections:
        s['html'] = render_content(s['content'])
    return render_template('rules.html', sections=sections, is_admin=_is_admin())

@rules_bp.route('/print')
@_login_required
def rules_print():
    with _get_db() as conn:
        sections = [dict(r) for r in conn.execute(
            'SELECT * FROM restaurant_rules WHERE active=1 ORDER BY sort_order, section_num').fetchall()]
    for s in sections:
        s['html'] = render_content(s['content'])
    return render_template('rules_print.html', sections=sections,
                           printed_at=datetime.now().strftime('%d %B %Y'))

@rules_bp.route('/<int:rule_id>/update', methods=['POST'])
@_admin_required
def rule_update(rule_id):
    with _get_db() as conn:
        conn.execute(
            '''UPDATE restaurant_rules
               SET section_num=?, title=?, content=?,
                   updated_at=datetime('now','localtime'), updated_by=?
               WHERE id=?''',
            (int(request.form.get('section_num', 0)),
             request.form.get('title', '').strip(),
             request.form.get('content', '').strip(),
             session.get('role', 'admin'),
             rule_id))
    return redirect(url_for('rules.rules_view'))

@rules_bp.route('/add', methods=['POST'])
@_admin_required
def rule_add():
    with _get_db() as conn:
        max_order = conn.execute('SELECT MAX(sort_order) FROM restaurant_rules').fetchone()[0] or 0
        max_num   = conn.execute('SELECT MAX(section_num) FROM restaurant_rules').fetchone()[0] or 0
        conn.execute(
            'INSERT INTO restaurant_rules (section_num, title, content, sort_order) VALUES (?,?,?,?)',
            (max_num + 1,
             request.form.get('title', 'New Section').strip(),
             request.form.get('content', '').strip(),
             max_order + 1))
    return redirect(url_for('rules.rules_view'))

@rules_bp.route('/<int:rule_id>/delete', methods=['POST'])
@_admin_required
def rule_delete(rule_id):
    with _get_db() as conn:
        conn.execute('UPDATE restaurant_rules SET active=0 WHERE id=?', (rule_id,))
    return redirect(url_for('rules.rules_view'))

@rules_bp.route('/export/excel')
@_admin_required
def rules_export_excel():
    with _get_db() as conn:
        sections = [dict(r) for r in conn.execute(
            'SELECT * FROM restaurant_rules WHERE active=1 ORDER BY sort_order, section_num').fetchall()]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Restaurant Rules'

    # Palette
    RED      = 'C0392B'
    DARK     = '1A1A2E'
    GOLD     = 'F39C12'
    LIGHT_BG = 'FDF6EC'
    RULE_BG  = 'FEF9F0'
    BULLET_BG= 'FFFFFF'

    thin = Side(style='thin', color='D5D8DC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title row ──
    ws.merge_cells('A1:C1')
    tc = ws['A1']
    tc.value = 'MCQ MIRRABOOKA — RESTAURANT RULES'
    tc.font = Font(name='Calibri', bold=True, size=16, color='FFFFFF')
    tc.fill = PatternFill('solid', fgColor=RED)
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:C2')
    dc = ws['A2']
    dc.value = f'Exported: {datetime.now().strftime("%d %B %Y")}'
    dc.font = Font(name='Calibri', size=10, color='888888')
    dc.alignment = Alignment(horizontal='center')
    dc.fill = PatternFill('solid', fgColor='F8F9F9')
    ws.row_dimensions[2].height = 18

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 72

    current_row = 3

    for sec in sections:
        # Section header
        ws.merge_cells(f'A{current_row}:C{current_row}')
        hdr = ws[f'A{current_row}']
        hdr.value = f'  {sec["section_num"]:02d}.  {sec["title"].upper()}'
        hdr.font  = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        hdr.fill  = PatternFill('solid', fgColor=DARK)
        hdr.alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Parse and write content
        paragraphs = re.split(r'\n{2,}', (sec['content'] or '').strip())
        for para in paragraphs:
            lines = para.strip().split('\n')
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                if line.startswith('- '):
                    # Bullet row
                    nr = ws.cell(current_row, 1, '•')
                    nr.font = Font(name='Calibri', bold=True, size=10, color=RED)
                    nr.alignment = Alignment(horizontal='center', vertical='top')
                    nr.fill = PatternFill('solid', fgColor=BULLET_BG)

                    ws.merge_cells(f'B{current_row}:C{current_row}')
                    bc = ws[f'B{current_row}']
                    bc.value = line[2:].strip()
                    bc.font  = Font(name='Calibri', size=10)
                    bc.fill  = PatternFill('solid', fgColor=BULLET_BG)
                    bc.alignment = Alignment(wrap_text=True, vertical='top', indent=1)
                    bc.border = border
                    ws.row_dimensions[current_row].height = 15
                    current_row += 1
                elif re.match(r'^\*\*[^*]+\*\*', line):
                    m = re.match(r'^\*\*([^*]+)\*\*:?\s*(.*)', line)
                    if m:
                        ws.merge_cells(f'A{current_row}:C{current_row}')
                        sc = ws[f'A{current_row}']
                        key = m.group(1).rstrip(':')
                        val = m.group(2).strip()
                        sc.value = f'{key}: {val}' if val else key
                        sc.font  = Font(name='Calibri', bold=True, size=10, color=DARK)
                        sc.fill  = PatternFill('solid', fgColor='EBF5FB')
                        sc.alignment = Alignment(wrap_text=True, indent=1, vertical='top')
                        ws.row_dimensions[current_row].height = 15
                        current_row += 1
                else:
                    ws.merge_cells(f'A{current_row}:C{current_row}')
                    pc = ws[f'A{current_row}']
                    pc.value = line
                    pc.font  = Font(name='Calibri', size=10)
                    pc.fill  = PatternFill('solid', fgColor=RULE_BG)
                    pc.alignment = Alignment(wrap_text=True, indent=1, vertical='top')
                    ws.row_dimensions[current_row].height = 15
                    current_row += 1

        # Spacer
        ws.row_dimensions[current_row].height = 6
        current_row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'MCQ_Restaurant_Rules_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return Response(buf.read(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename="{fname}"'})
