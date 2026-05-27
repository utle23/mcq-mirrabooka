from flask import Blueprint, render_template, request, redirect, url_for, session, jsonify, send_file
import sqlite3, json, os
from io import BytesIO
from datetime import datetime, date, timedelta
from functools import wraps

try:
    import email_service
except Exception:
    email_service = None

training_bp = Blueprint('training', __name__, url_prefix='/training')
DB_PATH = None
CHECKLISTS_DATA = {}   # injected from app.py at init time

# Roles are taken from CHECKLISTS short titles + a "General" catch-all.
# A safe default list is used if init was called without CHECKLISTS.
DEFAULT_ROLES = ['Cashier', 'Banh Mi', 'Chef', 'Kitchen Hand', 'Drinks', 'General']

SHIFTS = [('morning', 'Morning'), ('afternoon', 'Afternoon'), ('full', 'Full Day')]

RATINGS = [
    ('excellent',        'Excellent',         '#2E7D32', '#E8F5E9'),
    ('good',             'Good',              '#1565C0', '#E3F2FD'),
    ('developing',       'Developing',        '#E65100', '#FFF3E0'),
    ('needs_improvement','Needs Improvement', '#C62828', '#FFEBEE'),
]

# General topics applied to all roles (foundation training).
GENERAL_TOPICS = [
    ('Food Safety & Hygiene', [
        'Hand washing procedure',
        'Gloves & PPE usage',
        'Personal hygiene standards',
        'Food temperature awareness',
    ]),
    ('Uniform & Presentation', [
        'Full uniform worn correctly (cap, MCQ shirt)',
        'Clean & neat presentation during shift',
    ]),
    ('iPad & Records', [
        'Daily checklist completion on iPad',
        'Temperature records on iPad',
        'Issue reporting on iPad',
    ]),
    ('Conduct', [
        'Phone policy during working hours',
        'Staying at assigned position',
        'Communicating issues to manager',
    ]),
]

# Extra role-specific skills (beyond what the checklist already covers).
EXTRA_ROLE_TOPICS = {
    'Cashier': [
        ('Cashier Skills', [
            'POS system operation',
            'Cash handling & counting',
            'EFTPOS / card processing',
            'End-of-day cash reconciliation',
            'Greeting & serving customers',
            'Handling complaints politely',
            'Upselling & suggesting items',
        ]),
    ],
    'Banh Mi': [
        ('Banh Mi Skills', [
            'Bread selection & preparation',
            'Filling portions & assembly',
            'Customisation handling',
            'Speed & efficiency under pressure',
            'Presentation & wrapping',
            'Change gloves for vegetarian orders',
        ]),
    ],
    'Chef': [
        ('Chef Skills', [
            'Correct cooking temperatures',
            'Cross-contamination prevention',
            'Portion control & plating',
            'Food labelling & dating',
            'FIFO stock rotation',
            'Safe equipment use (stove, fryer, grill)',
        ]),
    ],
    'Kitchen Hand': [
        ('Kitchen Hand Skills', [
            'Knife skills (slicing meat, vegetables)',
            'Roasting / grilling procedures',
            'Marinating & seasoning',
            'Cleaning kitchen equipment',
            'Empty rubbish bins',
            'Wash dishes properly',
        ]),
    ],
    'Drinks': [
        ('Drink Station Skills', [
            'Coffee / condensed milk preparation',
            'Watermelon / tropical / sugarcane juice',
            'Smoothie & shake preparation',
            'Fresh rice paper rolls',
            'Cleaning juicer & sugarcane machine',
        ]),
    ],
    'General': [],
}


def _get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def _is_admin():
    return session.get('role') == 'admin'

def _login_required(f):
    @wraps(f)
    def d(*a, **kw):
        if not session.get('logged_in'):
            return redirect(url_for('login_page'))
        return f(*a, **kw)
    return d

def _admin_required(f):
    @wraps(f)
    def d(*a, **kw):
        if not session.get('logged_in'):
            return redirect(url_for('login_page'))
        if session.get('role') != 'admin':
            return render_template('access_denied.html'), 403
        return f(*a, **kw)
    return d


def _get_roles():
    """Roles = unique full titles from CHECKLISTS + General. Title preferred (more readable)."""
    roles = []
    seen = set()
    for chk in (CHECKLISTS_DATA or {}).values():
        name = (chk.get('title') or chk.get('short') or '').strip()
        # Normalise the 'Serve Order / Drinks' case to just 'Drinks' for clarity.
        if name == 'Serve Order / Drinks':
            name = 'Drinks'
        if name and name not in seen:
            roles.append(name); seen.add(name)
    if 'General' not in seen:
        roles.append('General')
    return roles or DEFAULT_ROLES


def init_training_tables(db_path, checklists=None):
    global DB_PATH, CHECKLISTS_DATA
    DB_PATH = db_path
    CHECKLISTS_DATA = checklists or {}
    with _get_db() as conn:
        conn.execute('''CREATE TABLE IF NOT EXISTS training_sessions (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            trainee_name    TEXT NOT NULL,
            trainee_role    TEXT NOT NULL DEFAULT '',
            trainer_name    TEXT NOT NULL DEFAULT '',
            session_date    TEXT NOT NULL,
            shift           TEXT DEFAULT 'full',
            overall_rating  TEXT DEFAULT '',
            key_achievements    TEXT DEFAULT '',
            needs_improvement   TEXT DEFAULT '',
            next_session_focus  TEXT DEFAULT '',
            general_notes       TEXT DEFAULT '',
            created_at      TEXT DEFAULT (datetime('now','localtime')),
            created_by      TEXT DEFAULT '',
            updated_at      TEXT DEFAULT (datetime('now','localtime')),
            updated_by      TEXT DEFAULT ''
        )''')
        conn.execute('''CREATE TABLE IF NOT EXISTS training_session_items (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id  INTEGER NOT NULL REFERENCES training_sessions(id) ON DELETE CASCADE,
            category    TEXT DEFAULT '',
            item_name   TEXT NOT NULL,
            status      TEXT DEFAULT 'not_covered',
            notes       TEXT DEFAULT '',
            sort_order  INTEGER DEFAULT 0
        )''')
        # Topic templates — admin-editable list of categories & items per role.
        conn.execute('''CREATE TABLE IF NOT EXISTS training_topic_templates (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            role        TEXT NOT NULL,
            category    TEXT NOT NULL,
            item_name   TEXT NOT NULL,
            sort_order  INTEGER DEFAULT 0,
            active      INTEGER DEFAULT 1,
            source      TEXT DEFAULT 'custom'
        )''')

        # Seed templates if empty.
        if conn.execute('SELECT COUNT(*) as c FROM training_topic_templates').fetchone()['c'] == 0:
            _seed_topic_templates(conn)


def _seed_topic_templates(conn):
    """Populate the topic templates from GENERAL_TOPICS + CHECKLISTS + EXTRA_ROLE_TOPICS."""
    roles = _get_roles()
    # Build a {role -> [(category, item)]} mapping from CHECKLISTS based on titles.
    checklist_items = {r: [] for r in roles}
    for chk in (CHECKLISTS_DATA or {}).values():
        role_name = (chk.get('title') or chk.get('short') or '').strip()
        if role_name == 'Serve Order / Drinks':
            role_name = 'Drinks'
        if role_name not in checklist_items:
            continue
        for task in chk.get('opening', []) or []:
            checklist_items[role_name].append(('Opening Checklist Tasks', task))
        for task in chk.get('closing', []) or []:
            checklist_items[role_name].append(('Closing Checklist Tasks', task))

    order = 0
    for role in roles:
        # 1. General topics first (every role).
        for cat, items in GENERAL_TOPICS:
            for item in items:
                conn.execute('''INSERT INTO training_topic_templates
                    (role, category, item_name, sort_order, active, source)
                    VALUES (?,?,?,?,1,?)''',
                    (role, cat, item, order, 'seed_general'))
                order += 1
        # 2. Role-specific checklist tasks.
        for cat, item in checklist_items.get(role, []):
            conn.execute('''INSERT INTO training_topic_templates
                (role, category, item_name, sort_order, active, source)
                VALUES (?,?,?,?,1,?)''',
                (role, cat, item, order, 'seed_checklist'))
            order += 1
        # 3. Extra role-specific skills.
        for cat, items in EXTRA_ROLE_TOPICS.get(role, []):
            for item in items:
                conn.execute('''INSERT INTO training_topic_templates
                    (role, category, item_name, sort_order, active, source)
                    VALUES (?,?,?,?,1,?)''',
                    (role, cat, item, order, 'seed_extra'))
                order += 1


def _build_topics(role):
    """Return [{category, items: [..]}] for a role, sourced from the editable templates."""
    result = []
    with _get_db() as conn:
        rows = conn.execute('''
            SELECT category, item_name FROM training_topic_templates
            WHERE role=? AND active=1 ORDER BY sort_order, id''', (role,)).fetchall()
    cat_map = {}
    for r in rows:
        cat = r['category'] or 'General'
        if cat not in cat_map:
            cat_map[cat] = []
            result.append({'category': cat, 'items': cat_map[cat]})
        cat_map[cat].append(r['item_name'])
    return result


def _get_staff_list():
    with _get_db() as conn:
        return [r['name'] for r in conn.execute(
            'SELECT name FROM staff_members WHERE active=1 ORDER BY name').fetchall()]


def _save_items(conn, session_id, form):
    conn.execute('DELETE FROM training_session_items WHERE session_id=?', (session_id,))
    categories = form.getlist('topic_category[]')
    names      = form.getlist('topic_name[]')
    statuses   = form.getlist('topic_status[]')
    notes_list = form.getlist('topic_notes[]')
    for i, (cat, name, status, note) in enumerate(zip(categories, names, statuses, notes_list)):
        if name.strip():
            conn.execute('''INSERT INTO training_session_items
                (session_id, category, item_name, status, notes, sort_order)
                VALUES (?,?,?,?,?,?)''',
                (session_id, cat, name.strip(), status or 'not_covered', note, i))


# ── Routes ─────────────────────────────────────────────────────────────────────

@training_bp.route('/')
@_login_required
def training_list():
    filter_trainee = request.args.get('trainee', '')
    filter_role    = request.args.get('role', '')
    roles = _get_roles()

    with _get_db() as conn:
        q = 'SELECT * FROM training_sessions WHERE 1=1'
        params = []
        if filter_trainee:
            q += ' AND trainee_name=?'; params.append(filter_trainee)
        if filter_role:
            q += ' AND trainee_role=?'; params.append(filter_role)
        q += ' ORDER BY session_date DESC, id DESC'
        sessions = [dict(r) for r in conn.execute(q, params).fetchall()]

        for s in sessions:
            rows = conn.execute(
                'SELECT status, COUNT(*) c FROM training_session_items WHERE session_id=? GROUP BY status',
                (s['id'],)).fetchall()
            s['cnt_achieved']      = sum(r['c'] for r in rows if r['status'] == 'achieved')
            s['cnt_needs_practice'] = sum(r['c'] for r in rows if r['status'] == 'needs_practice')
            s['cnt_total']          = sum(r['c'] for r in rows)

        trainees = [r['trainee_name'] for r in conn.execute(
            'SELECT DISTINCT trainee_name FROM training_sessions ORDER BY trainee_name').fetchall()]

    return render_template('training_list.html',
        sessions=sessions, trainees=trainees, roles=roles,
        filter_trainee=filter_trainee, filter_role=filter_role,
        ratings={r[0]: (r[1], r[2], r[3]) for r in RATINGS},
        is_admin=_is_admin())


@training_bp.route('/new', methods=['GET', 'POST'])
@_login_required
def training_new():
    roles = _get_roles()
    if request.method == 'POST':
        with _get_db() as conn:
            cur = conn.execute('''INSERT INTO training_sessions
                (trainee_name, trainee_role, trainer_name, session_date, shift,
                 overall_rating, key_achievements, needs_improvement,
                 next_session_focus, general_notes, created_by)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                (request.form.get('trainee_name','').strip(),
                 request.form.get('trainee_role','').strip(),
                 request.form.get('trainer_name','').strip(),
                 request.form.get('session_date', datetime.now().strftime('%Y-%m-%d')),
                 request.form.get('shift','full'),
                 request.form.get('overall_rating',''),
                 request.form.get('key_achievements','').strip(),
                 request.form.get('needs_improvement','').strip(),
                 request.form.get('next_session_focus','').strip(),
                 request.form.get('general_notes','').strip(),
                 session.get('role','')))
            sid = cur.lastrowid
            _save_items(conn, sid, request.form)

        # Email notification (fire-and-forget).
        if email_service:
            statuses = request.form.getlist('topic_status[]')
            ach_count = sum(1 for s in statuses if s == 'achieved')
            np_count = sum(1 for s in statuses if s == 'needs_practice')
            rating_label = next((r[1] for r in RATINGS if r[0] == request.form.get('overall_rating','')), '-')
            email_service.send_notification(
                'training',
                subject=f'Training: {request.form.get("trainee_name","-")} ({request.form.get("trainee_role","-")})',
                lines=[
                    f'Trainee: {request.form.get("trainee_name","-")}',
                    f'Role: {request.form.get("trainee_role","-")}',
                    f'Trainer: {request.form.get("trainer_name","-")}',
                    f'Date: {request.form.get("session_date", datetime.now().strftime("%Y-%m-%d"))}',
                    f'Shift: {request.form.get("shift","full")}',
                    f'Overall rating: {rating_label}',
                    f'Topics achieved: {ach_count}',
                    f'Needs practice: {np_count}',
                    f'Key achievements: {request.form.get("key_achievements","").strip() or "-"}',
                    f'Next session focus: {request.form.get("next_session_focus","").strip() or "-"}',
                ],
                link_path=f'/training/{sid}',
                actor=request.form.get('trainer_name','').strip() or session.get('role',''),
            )
        return redirect(url_for('training.training_detail', session_id=sid))

    topics_json = {role: _build_topics(role) for role in roles}
    return render_template('training_form.html',
        staff_list=_get_staff_list(), roles=roles, shifts=SHIFTS, ratings=RATINGS,
        topics_json=json.dumps(topics_json), edit_mode=False, s=None, items=[],
        today=datetime.now().strftime('%Y-%m-%d'), is_admin=_is_admin())


@training_bp.route('/<int:session_id>')
@_login_required
def training_detail(session_id):
    with _get_db() as conn:
        s = conn.execute('SELECT * FROM training_sessions WHERE id=?', (session_id,)).fetchone()
        if not s:
            return redirect(url_for('training.training_list'))
        s = dict(s)
        items = [dict(r) for r in conn.execute(
            'SELECT * FROM training_session_items WHERE session_id=? ORDER BY sort_order',
            (session_id,)).fetchall()]

    by_cat = {}
    for item in items:
        by_cat.setdefault(item['category'] or 'General', []).append(item)

    rating_info = {r[0]: (r[1], r[2], r[3]) for r in RATINGS}
    cnt = {'achieved': 0, 'needs_practice': 0, 'not_covered': 0}
    for item in items:
        cnt[item['status']] = cnt.get(item['status'], 0) + 1

    return render_template('training_detail.html',
        s=s, items=items, by_cat=by_cat,
        rating_info=rating_info, cnt=cnt, total=len(items),
        is_admin=_is_admin())


@training_bp.route('/<int:session_id>/edit', methods=['GET', 'POST'])
@_login_required
def training_edit(session_id):
    roles = _get_roles()
    with _get_db() as conn:
        s = conn.execute('SELECT * FROM training_sessions WHERE id=?', (session_id,)).fetchone()
        if not s:
            return redirect(url_for('training.training_list'))

        if request.method == 'POST':
            conn.execute('''UPDATE training_sessions SET
                trainee_name=?, trainee_role=?, trainer_name=?, session_date=?, shift=?,
                overall_rating=?, key_achievements=?, needs_improvement=?,
                next_session_focus=?, general_notes=?,
                updated_at=datetime('now','localtime'), updated_by=?
                WHERE id=?''',
                (request.form.get('trainee_name','').strip(),
                 request.form.get('trainee_role','').strip(),
                 request.form.get('trainer_name','').strip(),
                 request.form.get('session_date',''),
                 request.form.get('shift','full'),
                 request.form.get('overall_rating',''),
                 request.form.get('key_achievements','').strip(),
                 request.form.get('needs_improvement','').strip(),
                 request.form.get('next_session_focus','').strip(),
                 request.form.get('general_notes','').strip(),
                 session.get('role',''), session_id))
            _save_items(conn, session_id, request.form)
            return redirect(url_for('training.training_detail', session_id=session_id))

        s = dict(s)
        items = [dict(r) for r in conn.execute(
            'SELECT * FROM training_session_items WHERE session_id=? ORDER BY sort_order',
            (session_id,)).fetchall()]
        staff_list = _get_staff_list()

    topics_json = {role: _build_topics(role) for role in roles}
    return render_template('training_form.html',
        staff_list=staff_list, roles=roles, shifts=SHIFTS, ratings=RATINGS,
        topics_json=json.dumps(topics_json), edit_mode=True, s=s, items=items,
        today=datetime.now().strftime('%Y-%m-%d'), is_admin=_is_admin())


@training_bp.route('/<int:session_id>/delete', methods=['POST'])
@_admin_required
def training_delete(session_id):
    with _get_db() as conn:
        conn.execute('DELETE FROM training_session_items WHERE session_id=?', (session_id,))
        conn.execute('DELETE FROM training_sessions WHERE id=?', (session_id,))
    return redirect(url_for('training.training_list'))


@training_bp.route('/bulk-delete', methods=['POST'])
@_admin_required
def training_bulk_delete():
    ids = [int(x) for x in request.form.getlist('ids[]') if x.isdigit()]
    with _get_db() as conn:
        for sid in ids:
            conn.execute('DELETE FROM training_session_items WHERE session_id=?', (sid,))
            conn.execute('DELETE FROM training_sessions WHERE id=?', (sid,))
    return jsonify({'ok': True, 'deleted': len(ids)})


# ── Admin: Topic Templates Management ─────────────────────────────────────────

@training_bp.route('/topics')
@_admin_required
def topics_manage():
    roles = _get_roles()
    selected_role = request.args.get('role', roles[0] if roles else '')
    with _get_db() as conn:
        rows = [dict(r) for r in conn.execute('''
            SELECT * FROM training_topic_templates
            WHERE role=? ORDER BY sort_order, id''', (selected_role,)).fetchall()]
    # Group by category for display.
    by_cat = {}
    for r in rows:
        by_cat.setdefault(r['category'] or 'General', []).append(r)
    return render_template('training_topics.html',
        roles=roles, selected_role=selected_role,
        by_cat=by_cat, total=len(rows), is_admin=True)


@training_bp.route('/topics/add', methods=['POST'])
@_admin_required
def topics_add():
    role = request.form.get('role', '').strip()
    category = request.form.get('category', '').strip() or 'General'
    item_name = request.form.get('item_name', '').strip()
    if not (role and item_name):
        return redirect(url_for('training.topics_manage', role=role))
    with _get_db() as conn:
        max_ord = conn.execute(
            'SELECT COALESCE(MAX(sort_order),0)+1 as n FROM training_topic_templates WHERE role=?',
            (role,)).fetchone()['n']
        conn.execute('''INSERT INTO training_topic_templates
            (role, category, item_name, sort_order, active, source)
            VALUES (?,?,?,?,1,'custom')''',
            (role, category, item_name, max_ord))
    return redirect(url_for('training.topics_manage', role=role))


@training_bp.route('/topics/<int:tid>/edit', methods=['POST'])
@_admin_required
def topics_edit(tid):
    category = request.form.get('category', '').strip() or 'General'
    item_name = request.form.get('item_name', '').strip()
    role_redirect = request.form.get('role', '')
    if not item_name:
        return redirect(url_for('training.topics_manage', role=role_redirect))
    with _get_db() as conn:
        conn.execute('UPDATE training_topic_templates SET category=?, item_name=? WHERE id=?',
                     (category, item_name, tid))
    return redirect(url_for('training.topics_manage', role=role_redirect))


@training_bp.route('/topics/<int:tid>/toggle', methods=['POST'])
@_admin_required
def topics_toggle(tid):
    role_redirect = request.form.get('role', '')
    with _get_db() as conn:
        row = conn.execute('SELECT active FROM training_topic_templates WHERE id=?', (tid,)).fetchone()
        if row:
            conn.execute('UPDATE training_topic_templates SET active=? WHERE id=?',
                         (0 if row['active'] else 1, tid))
    return redirect(url_for('training.topics_manage', role=role_redirect))


@training_bp.route('/topics/<int:tid>/delete', methods=['POST'])
@_admin_required
def topics_delete(tid):
    role_redirect = request.form.get('role', '')
    with _get_db() as conn:
        conn.execute('DELETE FROM training_topic_templates WHERE id=?', (tid,))
    return redirect(url_for('training.topics_manage', role=role_redirect))


@training_bp.route('/topics/reseed', methods=['POST'])
@_admin_required
def topics_reseed():
    """Re-import any checklist tasks that aren't in the templates yet (non-destructive)."""
    role_redirect = request.form.get('role', '')
    with _get_db() as conn:
        existing = {(r['role'], r['category'], r['item_name']) for r in conn.execute(
            'SELECT role, category, item_name FROM training_topic_templates').fetchall()}
        order_per_role = {r['role']: r['n'] for r in conn.execute(
            'SELECT role, COALESCE(MAX(sort_order),0)+1 as n FROM training_topic_templates GROUP BY role').fetchall()}
        for chk in (CHECKLISTS_DATA or {}).values():
            role_name = (chk.get('title') or chk.get('short') or '').strip()
            if role_name == 'Serve Order / Drinks':
                role_name = 'Drinks'
            if not role_name:
                continue
            for task in chk.get('opening', []) or []:
                key = (role_name, 'Opening Checklist Tasks', task)
                if key not in existing:
                    o = order_per_role.get(role_name, 1)
                    conn.execute('''INSERT INTO training_topic_templates
                        (role, category, item_name, sort_order, active, source)
                        VALUES (?,?,?,?,1,'seed_checklist')''',
                        (role_name, 'Opening Checklist Tasks', task, o))
                    order_per_role[role_name] = o + 1
                    existing.add(key)
            for task in chk.get('closing', []) or []:
                key = (role_name, 'Closing Checklist Tasks', task)
                if key not in existing:
                    o = order_per_role.get(role_name, 1)
                    conn.execute('''INSERT INTO training_topic_templates
                        (role, category, item_name, sort_order, active, source)
                        VALUES (?,?,?,?,1,'seed_checklist')''',
                        (role_name, 'Closing Checklist Tasks', task, o))
                    order_per_role[role_name] = o + 1
                    existing.add(key)
    return redirect(url_for('training.topics_manage', role=role_redirect))


# ── Admin: Stats & Exports ────────────────────────────────────────────────────

def _collect_stats(date_from='', date_to='', role=''):
    """Aggregate sessions / topics by trainee. Returns (sessions, summary_by_trainee, by_role)."""
    with _get_db() as conn:
        q = 'SELECT * FROM training_sessions WHERE 1=1'
        params = []
        if date_from:
            q += ' AND session_date >= ?'; params.append(date_from)
        if date_to:
            q += ' AND session_date <= ?'; params.append(date_to)
        if role:
            q += ' AND trainee_role = ?'; params.append(role)
        q += ' ORDER BY session_date DESC, id DESC'
        sessions = [dict(r) for r in conn.execute(q, params).fetchall()]
        # Aggregate per trainee.
        per_trainee = {}
        all_items = []
        for s in sessions:
            items = [dict(r) for r in conn.execute(
                'SELECT * FROM training_session_items WHERE session_id=?',
                (s['id'],)).fetchall()]
            s['items'] = items
            for it in items:
                it['_trainee'] = s['trainee_name']
                it['_role']    = s['trainee_role']
                it['_date']    = s['session_date']
                all_items.append(it)

            key = s['trainee_name']
            t = per_trainee.setdefault(key, {
                'trainee_name': s['trainee_name'],
                'role': s['trainee_role'],
                'sessions': 0,
                'achieved': 0, 'needs_practice': 0, 'not_covered': 0,
                'total_topics': 0,
                'first_date': s['session_date'], 'last_date': s['session_date'],
                'ratings': {},
                'trainers': set(),
                'achieved_items': set(),
                'practice_items': set(),
            })
            t['sessions'] += 1
            t['first_date'] = min(t['first_date'], s['session_date'])
            t['last_date']  = max(t['last_date'],  s['session_date'])
            if s.get('trainer_name'): t['trainers'].add(s['trainer_name'])
            if s.get('overall_rating'):
                t['ratings'][s['overall_rating']] = t['ratings'].get(s['overall_rating'], 0) + 1
            for it in items:
                st = it['status']
                if st in t: t[st] += 1
                t['total_topics'] += 1
                if st == 'achieved': t['achieved_items'].add(it['item_name'])
                elif st == 'needs_practice': t['practice_items'].add(it['item_name'])

        # Convert sets to sorted lists for serialization.
        for t in per_trainee.values():
            t['trainers']        = sorted(t['trainers'])
            t['achieved_items']  = sorted(t['achieved_items'])
            t['practice_items']  = sorted(t['practice_items'])
            t['coverage_pct']    = round(t['achieved'] / max(t['total_topics'], 1) * 100)

        # Aggregate per role.
        by_role = {}
        for s in sessions:
            r = s['trainee_role'] or 'Unknown'
            br = by_role.setdefault(r, {'role': r, 'sessions': 0, 'trainees': set(),
                                         'achieved': 0, 'needs_practice': 0, 'not_covered': 0})
            br['sessions'] += 1
            br['trainees'].add(s['trainee_name'])
            for it in s['items']:
                if it['status'] in br: br[it['status']] += 1
        for br in by_role.values():
            br['trainees_count'] = len(br['trainees'])
            br['trainees']       = sorted(br['trainees'])

    return sessions, sorted(per_trainee.values(), key=lambda x: x['trainee_name']), by_role


@training_bp.route('/stats')
@_admin_required
def training_stats():
    today = date.today()
    default_from = (today - timedelta(days=90)).isoformat()
    date_from = request.args.get('from', default_from)
    date_to   = request.args.get('to', today.isoformat())
    role      = request.args.get('role', '')
    roles = _get_roles()
    sessions, per_trainee, by_role = _collect_stats(date_from, date_to, role)
    total_sessions = len(sessions)
    total_trainees = len(per_trainee)
    total_topics_achieved = sum(t['achieved'] for t in per_trainee)
    total_topics_practice = sum(t['needs_practice'] for t in per_trainee)
    rating_info = {r[0]: (r[1], r[2], r[3]) for r in RATINGS}
    return render_template('training_stats.html',
        sessions=sessions, per_trainee=per_trainee, by_role=by_role,
        date_from=date_from, date_to=date_to, role=role, roles=roles,
        total_sessions=total_sessions, total_trainees=total_trainees,
        total_topics_achieved=total_topics_achieved,
        total_topics_practice=total_topics_practice,
        rating_info=rating_info, is_admin=True)


def _rating_label(val):
    return next((r[1] for r in RATINGS if r[0] == val), val or '—')


@training_bp.route('/stats/export/excel')
@_admin_required
def training_stats_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    date_from = request.args.get('from', '')
    date_to   = request.args.get('to', '')
    role      = request.args.get('role', '')
    sessions, per_trainee, by_role = _collect_stats(date_from, date_to, role)

    wb = Workbook()

    # ── Sheet 1: Summary by trainee ──
    ws = wb.active
    ws.title = 'Per Trainee'
    title = 'MCQ MIRRABOOKA — TRAINING STATISTICS'
    sub = f'Period: {date_from or "all"} → {date_to or "all"}' + (f'   |   Role: {role}' if role else '')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14, color='C0392B')
    ws.merge_cells('A1:I1'); ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'] = sub
    ws['A2'].font = Font(italic=True, color='555555')
    ws.merge_cells('A2:I2'); ws['A2'].alignment = Alignment(horizontal='center')
    ws.append([])

    hdr = ['Trainee', 'Role', 'Sessions', 'First Date', 'Last Date',
           'Topics Achieved', 'Needs Practice', 'Not Covered', 'Coverage %']
    ws.append(hdr)
    hr = ws.max_row
    for col in range(1, len(hdr)+1):
        c = ws.cell(row=hr, column=col)
        c.fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    for t in per_trainee:
        ws.append([t['trainee_name'], t['role'], t['sessions'],
                   t['first_date'], t['last_date'],
                   t['achieved'], t['needs_practice'], t['not_covered'],
                   f"{t['coverage_pct']}%"])

    for col, w in zip('ABCDEFGHI', [26, 16, 10, 12, 12, 16, 16, 14, 12]):
        ws.column_dimensions[col].width = w

    # ── Sheet 2: Per role ──
    ws2 = wb.create_sheet('Per Role')
    ws2.append(['Role', 'Trainees', 'Sessions', 'Achieved', 'Needs Practice', 'Not Covered'])
    hr = ws2.max_row
    for col in range(1, 7):
        c = ws2.cell(row=hr, column=col)
        c.fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
        c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center')
    for br in by_role.values():
        ws2.append([br['role'], br['trainees_count'], br['sessions'],
                    br['achieved'], br['needs_practice'], br['not_covered']])
    for col, w in zip('ABCDEF', [18, 12, 12, 14, 18, 14]):
        ws2.column_dimensions[col].width = w

    # ── Sheet 3: All session items (detailed) ──
    ws3 = wb.create_sheet('All Topics')
    headers = ['Date', 'Trainee', 'Role', 'Trainer', 'Category', 'Topic', 'Status', 'Notes']
    ws3.append(headers)
    hr = ws3.max_row
    for col in range(1, len(headers)+1):
        c = ws3.cell(row=hr, column=col)
        c.fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
        c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center')
    status_color = {'achieved': 'C8E6C9', 'needs_practice': 'FFE0B2', 'not_covered': 'F5F5F5'}
    for s in sessions:
        for it in s['items']:
            ws3.append([s['session_date'], s['trainee_name'], s['trainee_role'],
                        s['trainer_name'], it['category'], it['item_name'],
                        it['status'].replace('_', ' ').title(), it['notes'] or ''])
            row_num = ws3.max_row
            ws3.cell(row=row_num, column=7).fill = PatternFill(
                start_color=status_color.get(it['status'], 'FFFFFF'),
                end_color=status_color.get(it['status'], 'FFFFFF'),
                fill_type='solid')
    for col, w in zip('ABCDEFGH', [12, 24, 14, 20, 22, 40, 16, 28]):
        ws3.column_dimensions[col].width = w

    # ── Sheet 4: Sessions list ──
    ws4 = wb.create_sheet('Sessions')
    ws4.append(['Date', 'Trainee', 'Role', 'Trainer', 'Shift', 'Overall Rating',
                'Achieved', 'Needs Practice', 'Not Covered',
                'Key Achievements', 'Needs Improvement', 'Next Session Focus', 'Notes'])
    hr = ws4.max_row
    for col in range(1, 14):
        c = ws4.cell(row=hr, column=col)
        c.fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
        c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    for s in sessions:
        cnt = {'achieved': 0, 'needs_practice': 0, 'not_covered': 0}
        for it in s['items']:
            cnt[it['status']] = cnt.get(it['status'], 0) + 1
        ws4.append([s['session_date'], s['trainee_name'], s['trainee_role'],
                    s['trainer_name'], s['shift'], _rating_label(s['overall_rating']),
                    cnt['achieved'], cnt['needs_practice'], cnt['not_covered'],
                    s['key_achievements'], s['needs_improvement'],
                    s['next_session_focus'], s['general_notes']])
    for col, w in zip('ABCDEFGHIJKLM', [12, 22, 14, 20, 10, 14, 10, 14, 12, 28, 28, 28, 28]):
        ws4.column_dimensions[col].width = w

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f'MCQ_Training_Stats_{date_from}_to_{date_to}.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


@training_bp.route('/stats/export/pdf')
@_admin_required
def training_stats_pdf():
    from html import escape
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, PageBreak

    date_from = request.args.get('from', '')
    date_to   = request.args.get('to', '')
    role      = request.args.get('role', '')
    sessions, per_trainee, by_role = _collect_stats(date_from, date_to, role)

    # Try to use the app's font registrar.
    try:
        from app import register_pdf_fonts
        font_name, bold_font = register_pdf_fonts()
    except Exception:
        font_name, bold_font = 'Helvetica', 'Helvetica-Bold'

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=12*mm, bottomMargin=12*mm)
    base = getSampleStyleSheet()
    styles = {
        'title': ParagraphStyle('title', parent=base['Title'], fontName=bold_font, fontSize=16,
                                leading=20, textColor=colors.HexColor('#C0392B'), alignment=TA_CENTER),
        'sub':   ParagraphStyle('sub', parent=base['Normal'], fontName=font_name, fontSize=9,
                                leading=12, textColor=colors.HexColor('#555555'), alignment=TA_CENTER),
        'h2':    ParagraphStyle('h2', parent=base['Heading2'], fontName=bold_font, fontSize=12,
                                leading=16, textColor=colors.HexColor('#1A1A2E')),
        'body':  ParagraphStyle('body', parent=base['BodyText'], fontName=font_name, fontSize=8.5,
                                leading=11, textColor=colors.HexColor('#222222')),
        'small': ParagraphStyle('small', parent=base['Normal'], fontName=font_name, fontSize=7.5,
                                leading=10, textColor=colors.HexColor('#444444')),
    }

    story = [
        Paragraph('MCQ MIRRABOOKA — Training Statistics', styles['title']),
        Paragraph(f"Period: {date_from or 'all'} → {date_to or 'all'}"
                  + (f"   |   Role: {escape(role)}" if role else ''),
                  styles['sub']),
        Spacer(1, 4*mm),
    ]

    # ── Overall summary ──
    total_sessions = len(sessions)
    total_trainees = len(per_trainee)
    total_achieved = sum(t['achieved'] for t in per_trainee)
    total_practice = sum(t['needs_practice'] for t in per_trainee)
    summary = [
        ['Total Sessions', str(total_sessions), 'Trainees', str(total_trainees)],
        ['Topics Achieved', str(total_achieved), 'Needs Practice', str(total_practice)],
    ]
    sm = Table(summary, colWidths=[40*mm, 40*mm, 40*mm, 40*mm])
    sm.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F4F6F9')),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTNAME', (0, 0), (0, -1), bold_font),
        ('FONTNAME', (2, 0), (2, -1), bold_font),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(sm)
    story.append(Spacer(1, 5*mm))

    # ── Per role ──
    story.append(Paragraph('Summary by Role', styles['h2']))
    role_rows = [['Role', 'Trainees', 'Sessions', 'Achieved', 'Needs Practice', 'Not Covered']]
    for br in by_role.values():
        role_rows.append([br['role'], str(br['trainees_count']), str(br['sessions']),
                          str(br['achieved']), str(br['needs_practice']), str(br['not_covered'])])
    role_table = Table(role_rows, colWidths=[55*mm, 30*mm, 30*mm, 30*mm, 35*mm, 30*mm], repeatRows=1)
    role_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A1A2E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), bold_font),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#DDDDDD')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FBFCFC')]),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(role_table)
    story.append(Spacer(1, 6*mm))

    # ── Per trainee summary ──
    story.append(Paragraph('Summary by Trainee', styles['h2']))
    pt_rows = [['Trainee', 'Role', 'Sessions', 'First', 'Last',
                'Achieved', 'Practice', 'Coverage %']]
    for t in per_trainee:
        pt_rows.append([t['trainee_name'], t['role'], str(t['sessions']),
                        t['first_date'], t['last_date'],
                        str(t['achieved']), str(t['needs_practice']),
                        f"{t['coverage_pct']}%"])
    pt_table = Table(pt_rows, colWidths=[55*mm, 30*mm, 22*mm, 25*mm, 25*mm, 25*mm, 25*mm, 25*mm], repeatRows=1)
    pt_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A1A2E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), bold_font),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#DDDDDD')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FBFCFC')]),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(pt_table)

    # ── Per-trainee detail pages ──
    for t in per_trainee:
        story.append(PageBreak())
        story.append(Paragraph(f"{escape(t['trainee_name'])} — {escape(t['role'] or '')}", styles['h2']))
        story.append(Paragraph(
            f"Sessions: {t['sessions']}   |   Period: {t['first_date']} → {t['last_date']}   |   "
            f"Coverage: {t['coverage_pct']}%   |   Trainers: {escape(', '.join(t['trainers']) or '-')}",
            styles['sub']))
        story.append(Spacer(1, 3*mm))

        if t['achieved_items']:
            story.append(Paragraph('Topics already achieved (no need to retrain):', styles['body']))
            ach = [[i+1, escape(name)] for i, name in enumerate(t['achieved_items'])]
            ach_table = Table([['#', 'Topic']] + ach, colWidths=[10*mm, 250*mm], repeatRows=1)
            ach_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), bold_font),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#DDDDDD')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F1F8E9'), colors.white]),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            story.append(ach_table)
            story.append(Spacer(1, 3*mm))

        if t['practice_items']:
            story.append(Paragraph('Topics still need more practice:', styles['body']))
            pr = [[i+1, escape(name)] for i, name in enumerate(t['practice_items'])]
            pr_table = Table([['#', 'Topic']] + pr, colWidths=[10*mm, 250*mm], repeatRows=1)
            pr_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E65100')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), bold_font),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#DDDDDD')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#FFF8E1'), colors.white]),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            story.append(pr_table)

    def footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont(font_name, 8)
        canvas.setFillColor(colors.HexColor('#666666'))
        canvas.drawString(10*mm, 7*mm, 'MCQ Mirrabooka Cafe - Training Statistics')
        canvas.drawRightString(landscape(A4)[0] - 10*mm, 7*mm,
                               f'Generated {datetime.now().strftime("%Y-%m-%d %H:%M")}   Page {doc_obj.page}')
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buf.seek(0)
    fname = f'MCQ_Training_Stats_{date_from}_to_{date_to}.pdf'
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=fname)
