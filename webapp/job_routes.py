from flask import Blueprint, render_template, request, session, jsonify, redirect, url_for, send_file
import sqlite3
from functools import wraps
from datetime import date, timedelta
from io import BytesIO
import os
import re
from html import escape

try:
    import email_service
except Exception:
    email_service = None

jobs = Blueprint('jobs', __name__, url_prefix='/jobs')
DB_PATH = None

DAYS     = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
DAY_FULL = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

STAFF = [
    'DANG, THI LAN UY', 'DOAN, THI NI', 'NGUYEN, PHU TAN',
    'HUYNH, ANH TRI', 'Thang Nguyen', 'MA, THANH PHUNG',
    'NGUYEN, THI NHAI', 'NGUYEN, VAN PHI LONG', 'DO, NGUYEN',
    'NGUYEN, THI NGOC PHUC', 'Ho Quynh', 'NGUYEN, HANG SANG',
    'VU, TRAN DO CAO',
]

ROLES = [
    {
        'key':   'take_order',
        'title': 'Cashier',
        'color': '#2196F3',
        'bg':    '#E3F2FD',
        'icon':  'fas fa-cash-register',
        'slots': 1,
        'opening': [
            'Open till & check cash balance ($350)',
            'Uniform check (hat, shirt, apron)',
            'Counter area ready for service, clean surrounding counter area',
            'Check the POS system - tills and EFTPOS all working',
            'Prepare vegetables & garnish for pho and bun bo hue',
            'Label beverages and label price and name for pastry',
            'Record temperature of fried pastry food (Pastry Temperature Record)',
            'Arrange the line-up for customers to take order',
            'Complete checklists (Food Temperature Records)',
            'Receive bread & pastry invoices from supplier',
        ],
        'closing': [
            'Take orders & resolve service issues (during shift)',
            'Clean the counter area',
            'Clean display glass cabinets',
            'Clean Cold Unit 2 Soft Drink Fridge',
            'Clean Cold Unit 3 Rice Paper Roll Fridge',
            'Refill drinking water',
            'Refill spoons, forks, chopsticks & sauces',
            'Buy bean sprouts, mint & veg for pho and bun bo hue for tomorrow',
            'Check income & till balance before closing',
            'Clean front floor',
            'Clean air fryer',
            'Check banh mi packaging bag order with Huu Anh',
            'Refill receipt roll',
            'Refill carry bags, paper bags & cup holders for takeaway',
            'Collect invoices & send to office every Friday (give to Van Anh)',
        ],
        'extra_sections': [
            {
                'key': 'orders_supplies',
                'title': 'Orders & Supplies Management',
                'color': '#6D4C41',
                'icon': 'fas fa-boxes-packing',
                'tasks': [
                    'Order the packaging & cleaning product',
                    'Order Banh mi packaging',
                    'Received catering order from office staff by Whatsapp',
                    'Contact customer for order information then contact with restaurant to make order for customer',
                ],
            },
        ],
    },
    {
        'key':   'banh_mi',
        'title': 'Banh Mi',
        'color': '#FF9800',
        'bg':    '#FFF3E0',
        'icon':  'fas fa-bread-slice',
        'slots': 2,
        # Person 1: all tasks that are NOT in person2_opening, plus all closing
        'person1_opening': [
            'Set up & display banh mi fridge bar (including banh mi, red pork meat, cha lua, cha gan, ngò, pickles, sliced chilli, cucumber)',
            'Check & test quality of each item (report to Manager if food not good to sell)',
            'Display plain bread in the basket',
            'Change gloves for vegetarian orders',
            'Order coriander, cucumber & radish from supermarket',
            'Complete Banh Mi Food Temperature Record',
        ],
        # Person 2: morning/opening only
        'person2_opening': [
            'Cut bread, spread pate & butter',
            'Set up roast pork station',
            'Chop roast pork & chicken',
            'Pre-make roast pork banh mi (all prep done by 8:00 AM)',
            'Customer order service (prepare banh mi per requests)',
        ],
        # Closing = Person 1 only
        'closing': [
            'Prepare next-day veg: coriander, cucumber, soy sauce follow prep timetable',
            'Prepare food in the tray for tomorrow setup (include coriander, pickle, cucumber, sauce: soy, gravy, mayo)',
            'Defrost pate (move from freezer to fridge for tomorrow)',
            'Clean up banh mi station',
            'Sanitise the roast pork table & chopping board',
            'Clean Cold Unit 4 Banh Mi Fridge (end of day)',
            'Wash dishes, final dish check at 4:30 PM',
        ],
        # Standard opening/closing keys unused for banh_mi; handled separately in template
        'opening': [],
    },
    {
        'key':   'chef',
        'title': 'Chef',
        'color': '#F44336',
        'bg':    '#FFEBEE',
        'icon':  'fas fa-kitchen-set',
        'slots': 1,
        'opening': [
            'Turn on lights and ventilation/exhaust fan',
            'Turn on gas & all kitchen equipment',
            'Cook rice',
            'Make savoury sticky rice',
            'Grill the pork chop',
            'Marinate chicken & pork followed food preparation timetable',
            'Pho soup station set up',
            'Set up rice noodle (pho noodle, BBH noodle, dry noodles)',
            'Set up chef fridge bar station (coriander, mint, tomato, cucumber, salad onion, spring onion, raw beef, cooked beef, beef ball, pork, etc.)',
            'Fry fried pastry items (fried spring roll, samosa, banana, etc.)',
            'Fry eggs for service',
            'Packaging items filled',
            'Customer order service',
            'Complete Chef Section Food Temperature Record',
        ],
        'closing': [
            'Marinate chicken & pork for next day followed food preparation timetable',
            'Prepare fish sauce and soy sauce for tomorrow',
            'Prep & store next-day food properly',
            'Clean kitchen areas & equipment',
            'Spray & wash kitchen floor',
            'Clean back floor',
            'Clean grilled machine',
            'Clean oven',
            'Clean stove',
            'Clean gas burners',
            'Clean deep fryer',
            'Clean Cold Unit 6 Soup & Rice Fridge',
            'Clean Cold Unit 7 Noodle & Soup Bar',
            'Clean Cold Unit 8 Food Prep Fridge',
            'Clean cold food fridge',
            'Clean the display fridge bar',
            'All kitchen equipment turned OFF',
            'Gas turned OFF',
            'Water turned OFF',
            'Kitchen lights OFF',
            'Door locked',
        ],
    },
    {
        'key':   'grill_beef',
        'title': 'Kitchen Hand',
        'color': '#4CAF50',
        'bg':    '#E8F5E9',
        'icon':  'fas fa-utensils',
        'slots': 1,
        'opening': [
            'Roast pork (10-15/day, 15-20 weekends)',
            'Grill chicken (2 trays/day, 3-4 weekends)',
            'Stir-fry beef / pork / tofu',
            'Slice beef ~5 kg for pho & banh mi',
            'Slice pork & char siu meat',
            'Beef soup Tue/Thu/Sat — Char siu Mon/Wed/Fri',
        ],
        'closing': [
            'Check & refill seasonings',
            'Prepare food for next day when time allows',
            'Clean the preparation area',
            'Empty rubbish bins',
            'Wash dishes',
            'Scrub & clean back floor',
            'Clean Freezer 1',
            'Clean Freezer 2',
        ],
    },
    {
        'key':   'serve_order',
        'title': 'Serve Order / Drinks',
        'color': '#00BCD4',
        'bg':    '#E0F7FA',
        'icon':  'fas fa-glass-water',
        'slots': 1,
        'opening': [
            'Prepare black iced coffee base (at least 2 jars)',
            'Prepare watermelon / tropical / sugarcane juice',
            'Prepare smoothie base mix',
            'Wash sugarcane',
            'Prepare fresh rice paper rolls for the day',
            'Make at least 1 tropical juice, 1 jar watermelon juice, 1 jar sugarcane juice, 4 jars black iced coffee',
            'Check coffee, condensed milk & fruit stock',
        ],
        'closing': [
            'Peel & prepare fruit for juice (next day): orange, apple, watermelon, etc.',
            'Refill fruit containers, lids clean & covered',
            'Set up & check quality, clean fruit juice in the fridge',
            'Check enough coffee / condensed milk',
            'Clean fruit juicer',
            'Clean sugarcane juicer',
            'Change fruit display (Mon, Wed, Fri, Sun)',
            'Wash all drink containers',
            'Check & refill cups, lids, straws',
            'Wash banh mi trays',
            'Bag returned pastry, bring to counter',
            'Clean microwave',
            'Clean pastry tray',
            'Clean Cold Unit 1 Fruit Juice Fridge',
            'Clean Cold Unit 5 Coffee Fridge',
            'Clean front floor',
            'Work area tidy, lights off',
        ],
    },
]

ROLES_BY_KEY = {r['key']: r for r in ROLES}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _get_db():
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    return conn

def _is_admin():
    return session.get('role') == 'admin'

def _login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

def _admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in') or not _is_admin():
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

def _week_start(d=None):
    if d is None:
        d = date.today()
    return d - timedelta(days=d.weekday())

def _clean_color(value, fallback):
    value = (value or '').strip()
    if len(value) == 7 and value.startswith('#'):
        try:
            int(value[1:], 16)
            return value.upper()
        except ValueError:
            pass
    return fallback

def _default_sections(role):
    if role['key'] == 'banh_mi':
        sections = [
            {
                'key': 'person1_opening',
                'title': 'Person 1 - Opening',
                'color': '#4CAF50',
                'icon': 'fas fa-user',
                'tasks': role.get('person1_opening', []),
            },
            {
                'key': 'person2_opening',
                'title': 'Person 2 - Morning Tasks',
                'color': '#FF9800',
                'icon': 'fas fa-user',
                'tasks': role.get('person2_opening', []),
            },
            {
                'key': 'closing',
                'title': 'Person 1 - Closing',
                'color': '#F44336',
                'icon': 'fas fa-moon',
                'tasks': role.get('closing', []),
            },
        ]
    else:
        sections = [
            {
                'key': 'opening',
                'title': 'Opening Tasks',
                'color': role.get('color', '#1565C0'),
                'icon': 'fas fa-sun',
                'tasks': role.get('opening', []),
            },
            {
                'key': 'closing',
                'title': 'Closing Tasks',
                'color': '#F44336',
                'icon': 'fas fa-moon',
                'tasks': role.get('closing', []),
            },
        ]
    sections.extend(role.get('extra_sections', []))
    return sections

def _seed_job_templates(conn):
    for sort_order, role in enumerate(ROLES):
        conn.execute('''INSERT OR IGNORE INTO job_role_templates
            (role_key,title,color,bg,icon,slots,sort_order)
            VALUES (?,?,?,?,?,?,?)''',
            (role['key'], role['title'], role['color'], role.get('bg', ''),
             role.get('icon', 'fas fa-id-badge'), role.get('slots', 1), sort_order))

        for section_order, section in enumerate(_default_sections(role)):
            conn.execute('''INSERT OR IGNORE INTO job_description_sections
                (role_key,section_key,title,color,icon,sort_order)
                VALUES (?,?,?,?,?,?)''',
                (role['key'], section['key'], section['title'],
                 section.get('color', role['color']), section.get('icon', 'fas fa-list-check'),
                 section_order))
            existing = conn.execute('''SELECT COUNT(*) as c
                FROM job_description_tasks
                WHERE role_key=? AND section_key=?''',
                (role['key'], section['key'])).fetchone()[0]
            if existing == 0:
                for task_order, task_name in enumerate(section.get('tasks', [])):
                    conn.execute('''INSERT INTO job_description_tasks
                        (role_key,section_key,task_order,task_name)
                        VALUES (?,?,?,?)''',
                        (role['key'], section['key'], task_order, task_name))

def _load_job_roles(conn):
    rows = conn.execute('''
        SELECT * FROM job_role_templates
        ORDER BY sort_order, title
    ''').fetchall()
    roles = []
    for row in rows:
        role = dict(row)
        role['key'] = role['role_key']
        role['slots'] = int(role.get('slots') or 1)
        role['sections'] = []

        section_rows = conn.execute('''
            SELECT * FROM job_description_sections
            WHERE role_key=?
            ORDER BY sort_order, id
        ''', (role['key'],)).fetchall()
        for section_row in section_rows:
            section = dict(section_row)
            section['key'] = section['section_key']
            task_rows = conn.execute('''
                SELECT id, task_order, task_name
                FROM job_description_tasks
                WHERE role_key=? AND section_key=?
                ORDER BY task_order, id
            ''', (role['key'], section['key'])).fetchall()
            section['tasks'] = [dict(t) for t in task_rows]
            role['sections'].append(section)
        roles.append(role)
    return roles

def _section_exists(conn, role_key, section_key):
    return conn.execute('''
        SELECT 1 FROM job_description_sections
        WHERE role_key=? AND section_key=?
    ''', (role_key, section_key)).fetchone() is not None


def _role_exists(conn, role_key):
    return conn.execute(
        'SELECT 1 FROM job_role_templates WHERE role_key=?',
        (role_key,)).fetchone() is not None


def _slugify(value, fallback='item'):
    s = re.sub(r'[^a-z0-9]+', '_', (value or '').lower()).strip('_')
    return s or fallback

def _hex_for_excel(value, fallback='7B1FA2'):
    return _clean_color(value, f'#{fallback}').replace('#', '').upper()

def _sheet_name(value, used):
    name = re.sub(r'[\[\]\:\*\?\/\\]', ' ', value or 'Role').strip()[:31] or 'Role'
    base = name
    n = 2
    while name in used:
        suffix = f' {n}'
        name = (base[:31 - len(suffix)] + suffix).strip()
        n += 1
    used.add(name)
    return name

def _role_task_count(role):
    return sum(len(section.get('tasks', [])) for section in role.get('sections', []))

def _register_pdf_fonts():
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except Exception:
        return 'Helvetica', 'Helvetica-Bold'

    regular_candidates = [
        '/System/Library/Fonts/Supplemental/Arial Unicode.ttf',
        '/System/Library/Fonts/Supplemental/Arial.ttf',
        '/Library/Fonts/Arial Unicode.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
    ]
    bold_candidates = [
        '/System/Library/Fonts/Supplemental/Arial Bold.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf',
    ]

    regular = next((p for p in regular_candidates if os.path.exists(p)), None)
    bold = next((p for p in bold_candidates if os.path.exists(p)), None)
    if not regular:
        return 'Helvetica', 'Helvetica-Bold'

    try:
        pdfmetrics.registerFont(TTFont('MCQSans', regular))
        if bold:
            pdfmetrics.registerFont(TTFont('MCQSans-Bold', bold))
            return 'MCQSans', 'MCQSans-Bold'
        return 'MCQSans', 'MCQSans'
    except Exception:
        return 'Helvetica', 'Helvetica-Bold'


# ── DB init ────────────────────────────────────────────────────────────────────

def init_job_tables(db_path):
    global DB_PATH
    DB_PATH = db_path
    with sqlite3.connect(db_path, timeout=30) as conn:
        conn.row_factory = sqlite3.Row
        conn.execute('''CREATE TABLE IF NOT EXISTS job_assignments (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            role_key   TEXT NOT NULL,
            week_start TEXT NOT NULL,
            day        TEXT NOT NULL,
            slot       INTEGER NOT NULL DEFAULT 1,
            staff_name TEXT DEFAULT '',
            start_time TEXT DEFAULT '',
            end_time   TEXT DEFAULT ''
        )''')
        conn.execute('''CREATE UNIQUE INDEX IF NOT EXISTS uidx_job
            ON job_assignments(role_key, week_start, day, slot)''')
        conn.execute('''CREATE TABLE IF NOT EXISTS job_role_templates (
            role_key   TEXT PRIMARY KEY,
            title      TEXT NOT NULL,
            color      TEXT NOT NULL DEFAULT '#7B1FA2',
            bg         TEXT DEFAULT '',
            icon       TEXT DEFAULT 'fas fa-id-badge',
            slots      INTEGER NOT NULL DEFAULT 1,
            sort_order INTEGER NOT NULL DEFAULT 0
        )''')
        conn.execute('''CREATE TABLE IF NOT EXISTS job_description_sections (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            role_key    TEXT NOT NULL,
            section_key TEXT NOT NULL,
            title       TEXT NOT NULL,
            color       TEXT NOT NULL DEFAULT '#7B1FA2',
            icon        TEXT DEFAULT 'fas fa-list-check',
            sort_order  INTEGER NOT NULL DEFAULT 0,
            UNIQUE(role_key, section_key)
        )''')
        conn.execute('''CREATE TABLE IF NOT EXISTS job_description_tasks (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            role_key    TEXT NOT NULL,
            section_key TEXT NOT NULL,
            task_order  INTEGER NOT NULL DEFAULT 0,
            task_name   TEXT NOT NULL
        )''')
        # Migrate legacy 'ordering' + 'catering_order' sections into 'orders_supplies'
        # BEFORE seeding so the seed step sees the merged section already populated and skips it.
        _migrate_merge_catering_sections(conn)
        _seed_job_templates(conn)
        conn.execute('''
            UPDATE job_role_templates
            SET title='Chef', icon='fas fa-kitchen-set'
            WHERE role_key='chef'
              AND title IN ('Chef / Kitchen', 'Chef/Kitchen', 'Chef')
        ''')
        conn.commit()


def _migrate_merge_catering_sections(conn):
    """Merge legacy 'ordering' + 'catering_order' sections (Cashier) into 'orders_supplies'."""
    has_old = conn.execute('''
        SELECT 1 FROM job_description_sections
        WHERE role_key='take_order' AND section_key IN ('ordering','catering_order')
        LIMIT 1''').fetchone()
    if not has_old:
        return

    new_row = conn.execute('''
        SELECT id FROM job_description_sections
        WHERE role_key='take_order' AND section_key='orders_supplies' ''').fetchone()
    if not new_row:
        next_order = conn.execute('''
            SELECT COALESCE(MAX(sort_order), -1) + 1 as n
            FROM job_description_sections WHERE role_key='take_order' ''').fetchone()['n']
        conn.execute('''INSERT INTO job_description_sections
            (role_key, section_key, title, color, icon, sort_order)
            VALUES ('take_order','orders_supplies',
                    'Orders & Supplies Management', '#6D4C41', 'fas fa-boxes-packing', ?)''',
            (next_order,))

    # Move all tasks from the old sections into the new one (preserve order).
    base_order = conn.execute('''
        SELECT COALESCE(MAX(task_order), -1) + 1 as n
        FROM job_description_tasks
        WHERE role_key='take_order' AND section_key='orders_supplies' ''').fetchone()['n']
    old_tasks = conn.execute('''
        SELECT id FROM job_description_tasks
        WHERE role_key='take_order' AND section_key IN ('ordering','catering_order')
        ORDER BY section_key, task_order, id''').fetchall()
    for offset, row in enumerate(old_tasks):
        conn.execute('''UPDATE job_description_tasks
            SET section_key='orders_supplies', task_order=?
            WHERE id=?''', (base_order + offset, row['id']))

    # Drop the now-empty old sections.
    conn.execute('''DELETE FROM job_description_sections
        WHERE role_key='take_order' AND section_key IN ('ordering','catering_order')''')


# ── Routes ─────────────────────────────────────────────────────────────────────

@jobs.route('/')
@_login_required
def job_schedule():
    ws_str = request.args.get('week', '')
    try:
        ws = date.fromisoformat(ws_str)
        # snap to Monday
        ws = ws - timedelta(days=ws.weekday())
    except ValueError:
        ws = _week_start()

    with _get_db() as conn:
        rows = conn.execute(
            'SELECT * FROM job_assignments WHERE week_start=?',
            (ws.isoformat(),)
        ).fetchall()
        roles = _load_job_roles(conn)

    # Build nested dict: assignments[role_key][day][slot_str]
    asgn = {}
    for r in rows:
        asgn.setdefault(r['role_key'], {}).setdefault(r['day'], {})[str(r['slot'])] = {
            'staff_name': r['staff_name'] or '',
            'start_time': r['start_time'] or '',
            'end_time':   r['end_time']   or '',
        }

    today_day = DAYS[date.today().weekday()] if _week_start() == ws else None

    day_dates = [(ws + timedelta(days=i)).strftime('%d/%m') for i in range(7)]

    return render_template('job_schedule.html',
        roles      = roles,
        days       = DAYS,
        day_full   = DAY_FULL,
        day_dates  = day_dates,
        week_start = ws.isoformat(),
        week_label = f"{ws.strftime('%d %b')} – {(ws + timedelta(days=6)).strftime('%d %b %Y')}",
        prev_week  = (ws - timedelta(weeks=1)).isoformat(),
        next_week  = (ws + timedelta(weeks=1)).isoformat(),
        asgn       = asgn,
        staff      = STAFF,
        today_day  = today_day,
        is_admin   = _is_admin(),
    )


@jobs.route('/assign', methods=['POST'])
@_admin_required
def assign():
    role_key   = request.form.get('role_key', '').strip()
    week_start = request.form.get('week_start', '').strip()
    day        = request.form.get('day', '').strip()
    slot       = request.form.get('slot', '1').strip()
    staff_name = request.form.get('staff_name', '').strip()
    start_time = request.form.get('start_time', '').strip()
    end_time   = request.form.get('end_time', '').strip()

    if day not in DAYS:
        return jsonify({'error': 'invalid'}), 400
    try:
        slot = int(slot)
    except ValueError:
        return jsonify({'error': 'invalid slot'}), 400

    with _get_db() as conn:
        if not _role_exists(conn, role_key):
            return jsonify({'error': 'invalid role'}), 400
        conn.execute('''
            INSERT INTO job_assignments (role_key, week_start, day, slot, staff_name, start_time, end_time)
            VALUES (?,?,?,?,?,?,?)
            ON CONFLICT(role_key, week_start, day, slot) DO UPDATE SET
                staff_name = excluded.staff_name,
                start_time = excluded.start_time,
                end_time   = excluded.end_time
        ''', (role_key, week_start, day, slot, staff_name, start_time, end_time))
        role_row = conn.execute('SELECT title FROM job_role_templates WHERE role_key=?', (role_key,)).fetchone()
        role_title = role_row['title'] if role_row else role_key

    if email_service and staff_name:
        email_service.send_notification(
            'jobs',
            subject=f'Job assignment updated: {role_title} — {day} ({week_start})',
            lines=[
                f'Role: {role_title}',
                f'Week: {week_start}',
                f'Day: {day}',
                f'Slot: {slot}',
                f'Staff: {staff_name}',
                f'Time: {start_time or "-"} → {end_time or "-"}',
            ],
            link_path=f'/jobs/?week={week_start}',
            actor=session.get('role',''),
        )

    return jsonify({'ok': True, 'staff_name': staff_name, 'start_time': start_time, 'end_time': end_time})


@jobs.route('/export/excel')
@jobs.route('/export/<role_key>/excel')
@_login_required
def export_job_excel(role_key=None):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    with _get_db() as conn:
        roles = _load_job_roles(conn)
    if role_key:
        roles = [r for r in roles if r['key'] == role_key]
        if not roles:
            return redirect(url_for('jobs.job_schedule'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Job Descriptions'
    ws.freeze_panes = 'A7'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    brand_fill = PatternFill('solid', fgColor='1B3A2D')
    soft_fill = PatternFill('solid', fgColor='F0F4F2')
    header_fill = PatternFill('solid', fgColor='D8F3DC')
    white_font = Font(bold=True, color='FFFFFF', size=16)
    title_font = Font(bold=True, color='1B3A2D', size=13)
    small_font = Font(color='666666', size=9)
    label_font = Font(bold=True, color='1B3A2D', size=10)
    body_font = Font(color='222222', size=10)
    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:D1')
    ws['A1'] = 'MCQ MIRRABOOKA CAFE'
    ws['A1'].font = white_font
    ws['A1'].fill = brand_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:D2')
    ws['A2'] = 'JOB DESCRIPTION PACK - PRINT TEMPLATE'
    ws['A2'].font = title_font
    ws['A2'].fill = header_fill
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 22

    ws['A4'] = 'Generated'
    ws['B4'] = date.today().strftime('%d/%m/%Y')
    ws['C4'] = 'Total roles'
    ws['D4'] = len(roles)
    for cell in ws[4]:
        cell.fill = soft_fill
        cell.border = border
        cell.font = label_font if cell.column in (1, 3) else small_font

    ws.append([])
    ws.append(['Role', 'Section', '#', 'Task'])
    for cell in ws[6]:
        cell.fill = brand_fill
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    row_idx = 7
    for role in roles:
        role_fill = PatternFill('solid', fgColor=_hex_for_excel(role.get('color')))
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
        cell = ws.cell(row=row_idx, column=1)
        cell.value = f"{role['title']}  |  {len(role['sections'])} sections  |  {_role_task_count(role)} tasks"
        cell.fill = role_fill
        cell.font = Font(bold=True, color='FFFFFF', size=12)
        cell.alignment = Alignment(vertical='center')
        ws.row_dimensions[row_idx].height = 22
        row_idx += 1

        for section in role['sections']:
            section_fill = PatternFill('solid', fgColor=_hex_for_excel(section.get('color'), '7B1FA2'))
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
            cell = ws.cell(row=row_idx, column=1)
            cell.value = section['title']
            cell.fill = section_fill
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.alignment = Alignment(vertical='center')
            ws.row_dimensions[row_idx].height = 20
            row_idx += 1

            for i, task in enumerate(section['tasks'], 1):
                ws.cell(row_idx, 1).value = role['title']
                ws.cell(row_idx, 2).value = section['title']
                ws.cell(row_idx, 3).value = i
                ws.cell(row_idx, 4).value = task['task_name']
                for col in range(1, 5):
                    c = ws.cell(row_idx, col)
                    c.border = border
                    c.font = body_font
                    c.alignment = Alignment(vertical='top', wrap_text=True)
                ws.cell(row_idx, 3).alignment = Alignment(horizontal='center', vertical='top')
                ws.row_dimensions[row_idx].height = max(22, min(68, 16 + (len(task['task_name']) // 58) * 14))
                row_idx += 1
        row_idx += 1

    widths = [24, 28, 6, 82]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    used_names = {ws.title}
    for role in roles:
        rs = wb.create_sheet(_sheet_name(role['title'], used_names))
        rs.sheet_properties.pageSetUpPr.fitToPage = True
        rs.page_setup.fitToWidth = 1
        rs.page_setup.fitToHeight = 0
        rs.page_margins.left = 0.35
        rs.page_margins.right = 0.35
        rs.merge_cells('A1:C1')
        rs['A1'] = role['title']
        rs['A1'].fill = PatternFill('solid', fgColor=_hex_for_excel(role.get('color')))
        rs['A1'].font = Font(bold=True, color='FFFFFF', size=16)
        rs['A1'].alignment = Alignment(horizontal='center')
        rs.row_dimensions[1].height = 30
        r = 3
        for section in role['sections']:
            rs.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            rs.cell(r, 1).value = section['title']
            rs.cell(r, 1).fill = PatternFill('solid', fgColor=_hex_for_excel(section.get('color'), '7B1FA2'))
            rs.cell(r, 1).font = Font(bold=True, color='FFFFFF', size=12)
            r += 1
            rs.append(['#', 'Task', 'Staff notes'])
            for c in rs[r]:
                c.fill = header_fill
                c.font = label_font
                c.border = border
                c.alignment = Alignment(horizontal='center')
            r += 1
            # Add a Done? column for printing
            rs.cell(r-1, 1).value = '#'
            rs.cell(r-1, 2).value = 'TASK'
            rs.cell(r-1, 3).value = 'DONE?'
            for c in rs[r-1]:
                c.fill = PatternFill('solid', fgColor='1B3A2D')
                c.font = Font(bold=True, color='FFFFFF', size=11)
                c.border = border
                c.alignment = Alignment(horizontal='center')

            big_task_font = Font(color='1A1A1A', size=12, bold=True)
            for i, task in enumerate(section['tasks'], 1):
                rs.cell(r, 1).value = i
                rs.cell(r, 2).value = task['task_name']
                rs.cell(r, 3).value = ''   # empty box for ticking
                for col in range(1, 4):
                    c = rs.cell(r, col)
                    c.border = border
                    c.alignment = Alignment(vertical='center', wrap_text=True)
                rs.cell(r, 1).font = Font(bold=True, color='1B3A2D', size=12)
                rs.cell(r, 1).alignment = Alignment(horizontal='center', vertical='center')
                rs.cell(r, 2).font = big_task_font
                rs.cell(r, 3).alignment = Alignment(horizontal='center', vertical='center')
                rs.cell(r, 3).fill = PatternFill('solid', fgColor='FFFFFF')
                # Taller rows so there is space to tick by hand
                rs.row_dimensions[r].height = max(32, min(95, 24 + (len(task['task_name']) // 38) * 18))
                r += 1
            r += 1
        for col, width in enumerate([8, 78, 14], 1):
            rs.column_dimensions[get_column_letter(col)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    export_name = roles[0]['key'] if role_key and len(roles) == 1 else 'all'
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'MCQ_Job_Descriptions_{export_name}_{date.today().isoformat()}.xlsx',
    )


@jobs.route('/export/pdf')
@jobs.route('/export/<role_key>/pdf')
@_login_required
def export_job_pdf(role_key=None):
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except Exception:
        return 'PDF export requires reportlab. Install with: pip install reportlab', 500

    with _get_db() as conn:
        roles = _load_job_roles(conn)
    if role_key:
        roles = [r for r in roles if r['key'] == role_key]
        if not roles:
            return redirect(url_for('jobs.job_schedule'))

    font_name, bold_font = _register_pdf_fonts()
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title='MCQ Job Descriptions',
    )

    base = getSampleStyleSheet()
    styles = {
        'cover_title': ParagraphStyle('cover_title', parent=base['Title'], fontName=bold_font,
                                      fontSize=26, leading=30, textColor=colors.HexColor('#1B3A2D'),
                                      alignment=TA_CENTER, spaceAfter=8),
        'cover_sub': ParagraphStyle('cover_sub', parent=base['Normal'], fontName=font_name,
                                    fontSize=11, leading=15, textColor=colors.HexColor('#555555'),
                                    alignment=TA_CENTER, spaceAfter=18),
        'role': ParagraphStyle('role', parent=base['Heading1'], fontName=bold_font,
                               fontSize=20, leading=24, textColor=colors.white, spaceAfter=0),
        'section': ParagraphStyle('section', parent=base['Heading2'], fontName=bold_font,
                                  fontSize=13, leading=16, textColor=colors.white, spaceAfter=0),
        # Bigger, bolder task text — designed for a printed sheet hanging on the wall
        'task': ParagraphStyle('task', parent=base['BodyText'], fontName=bold_font,
                               fontSize=12, leading=15, textColor=colors.HexColor('#1A1A1A')),
        'num': ParagraphStyle('num', parent=base['BodyText'], fontName=bold_font,
                              fontSize=12, leading=15, alignment=TA_CENTER,
                              textColor=colors.HexColor('#1B3A2D')),
        'small': ParagraphStyle('small', parent=base['Normal'], fontName=font_name,
                                fontSize=8.2, leading=10, textColor=colors.HexColor('#666666')),
    }

    story = []
    story.append(Spacer(1, 16 * mm))
    story.append(Paragraph('MCQ MIRRABOOKA CAFE', styles['cover_title']))
    story.append(Paragraph('Job Description Pack for New Staff', styles['cover_sub']))
    summary_data = [['Role', 'Sections', 'Tasks']]
    for role in roles:
        summary_data.append([role['title'], str(len(role['sections'])), str(_role_task_count(role))])
    summary = Table(summary_data, colWidths=[105 * mm, 28 * mm, 28 * mm], repeatRows=1)
    summary.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B3A2D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), bold_font),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#DDDDDD')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7FAF8')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
    ]))
    story.append(summary)
    story.append(Spacer(1, 7 * mm))
    story.append(Paragraph(f'Generated: {date.today().strftime("%d/%m/%Y")}', styles['cover_sub']))

    for role in roles:
        story.append(PageBreak())
        role_color = colors.HexColor(_clean_color(role.get('color'), '#7B1FA2'))
        role_header = Table([[Paragraph(escape(role['title']), styles['role']),
                              Paragraph(f"{len(role['sections'])} sections · {_role_task_count(role)} tasks<br/>"
                                        f"Date: ______________ &nbsp; Staff: ______________",
                                        styles['section'])]],
                            colWidths=[110 * mm, 65 * mm])
        role_header.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), role_color),
            ('BOX', (0, 0), (-1, -1), 0, role_color),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('LEFTPADDING', (0, 0), (-1, -1), 14),
            ('RIGHTPADDING', (0, 0), (-1, -1), 14),
        ]))
        story.append(role_header)
        story.append(Spacer(1, 4 * mm))

        for section in role['sections']:
            section_color = colors.HexColor(_clean_color(section.get('color'), role.get('color', '#7B1FA2')))
            sec_header = Table(
                [[Paragraph(escape(section['title']).upper(), styles['section']),
                  Paragraph(f"{len(section.get('tasks', []))} task(s)", styles['section'])]],
                colWidths=[140 * mm, 35 * mm])
            sec_header.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), section_color),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 12),
                ('RIGHTPADDING', (0, 0), (-1, -1), 12),
            ]))
            story.append(sec_header)

            # Print-friendly: 3 columns — Checkbox | # | Task (bigger bold text)
            task_data = [['DONE', '#', 'TASK']]
            for i, task in enumerate(section['tasks'], 1):
                task_data.append(['', Paragraph(str(i), styles['num']),
                                  Paragraph(escape(task['task_name']), styles['task'])])
            if len(task_data) == 1:
                task_data.append(['', '-', Paragraph('No tasks yet.', styles['small'])])

            task_table = Table(task_data, colWidths=[14 * mm, 12 * mm, 149 * mm], repeatRows=1)
            task_table.setStyle(TableStyle([
                # Header row
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B3A2D')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), bold_font),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('ALIGN', (0, 0), (1, 0), 'CENTER'),
                ('ALIGN', (2, 0), (2, 0), 'LEFT'),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                # Body cells
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.6, colors.HexColor('#888888')),
                ('TOPPADDING', (0, 1), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
                ('LEFTPADDING', (2, 1), (2, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F4F8F5')]),
                # Make the checkbox column visually distinct — a thicker outline + light bg
                ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#FFFFFF')),
                ('LINEAFTER', (0, 1), (0, -1), 1.2, colors.HexColor('#1B3A2D')),
                ('LINEBEFORE', (0, 1), (0, -1), 1.2, colors.HexColor('#1B3A2D')),
            ]))
            story.append(task_table)
            story.append(Spacer(1, 5 * mm))

        # Sign-off footer per role page — for staff/manager initials at end of shift
        sign_table = Table([
            [Paragraph('<b>Staff signature</b>', styles['small']),
             Paragraph('<b>Time finished</b>', styles['small']),
             Paragraph('<b>Manager check</b>', styles['small'])],
            ['', '', ''],
        ], colWidths=[60 * mm, 50 * mm, 65 * mm])
        sign_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.6, colors.HexColor('#888888')),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F0F4F2')),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 1), (-1, 1), 22),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 22),
        ]))
        story.append(Spacer(1, 4 * mm))
        story.append(sign_table)

    def footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont(font_name, 8)
        canvas.setFillColor(colors.HexColor('#666666'))
        canvas.drawString(12 * mm, 8 * mm, 'MCQ Mirrabooka Cafe - Job Description Pack')
        canvas.drawRightString(A4[0] - 12 * mm, 8 * mm, f'Page {doc_obj.page}')
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buf.seek(0)
    export_name = roles[0]['key'] if role_key and len(roles) == 1 else 'all'
    return send_file(
        buf,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'MCQ_Job_Descriptions_{export_name}_{date.today().isoformat()}.pdf',
    )


@jobs.route('/role/update', methods=['POST'])
@_admin_required
def update_role():
    role_key = request.form.get('role_key', '').strip()
    title    = request.form.get('title', '').strip()
    color    = request.form.get('color', '').strip()
    icon     = request.form.get('icon', '').strip()
    slots    = request.form.get('slots', '').strip()

    if not role_key or not title:
        return jsonify({'error': 'invalid'}), 400

    with _get_db() as conn:
        row = conn.execute(
            'SELECT color, icon, slots FROM job_role_templates WHERE role_key=?',
            (role_key,)).fetchone()
        if not row:
            return jsonify({'error': 'role not found'}), 404
        color = _clean_color(color, row['color'] or '#7B1FA2')
        icon = icon or row['icon'] or 'fas fa-id-badge'
        try:
            slots_int = int(slots) if slots else int(row['slots'] or 1)
        except ValueError:
            slots_int = int(row['slots'] or 1)
        slots_int = max(1, min(4, slots_int))
        conn.execute('''
            UPDATE job_role_templates
            SET title=?, color=?, icon=?, slots=?
            WHERE role_key=?
        ''', (title, color, icon, slots_int, role_key))
    return jsonify({'ok': True, 'title': title, 'color': color, 'icon': icon, 'slots': slots_int})


@jobs.route('/role/add', methods=['POST'])
@_admin_required
def add_role():
    title = request.form.get('title', '').strip()
    color = _clean_color(request.form.get('color', ''), '#7B1FA2')
    icon  = request.form.get('icon', '').strip() or 'fas fa-id-badge'
    try:
        slots = int(request.form.get('slots', '1') or '1')
    except ValueError:
        slots = 1
    slots = max(1, min(4, slots))

    if not title:
        return jsonify({'error': 'title required'}), 400

    with _get_db() as conn:
        # Generate a unique role_key from title.
        base = _slugify(title, 'role')
        role_key = base
        n = 2
        while _role_exists(conn, role_key):
            role_key = f'{base}_{n}'; n += 1

        next_order = conn.execute(
            'SELECT COALESCE(MAX(sort_order), -1) + 1 as n FROM job_role_templates'
        ).fetchone()['n']

        conn.execute('''INSERT INTO job_role_templates
            (role_key, title, color, bg, icon, slots, sort_order)
            VALUES (?,?,?,?,?,?,?)''',
            (role_key, title, color, '', icon, slots, next_order))

        # Seed default Opening + Closing sections so admins have somewhere to add tasks.
        for sec_order, (skey, stitle, sicon) in enumerate([
            ('opening', 'Opening Tasks', 'fas fa-sun'),
            ('closing', 'Closing Tasks', 'fas fa-moon'),
        ]):
            conn.execute('''INSERT INTO job_description_sections
                (role_key, section_key, title, color, icon, sort_order)
                VALUES (?,?,?,?,?,?)''',
                (role_key, skey, stitle, color, sicon, sec_order))

    return jsonify({'ok': True, 'role_key': role_key, 'title': title,
                    'color': color, 'icon': icon, 'slots': slots})


@jobs.route('/role/delete', methods=['POST'])
@_admin_required
def delete_role():
    role_key = request.form.get('role_key', '').strip()
    if not role_key:
        return jsonify({'error': 'invalid'}), 400
    with _get_db() as conn:
        if not _role_exists(conn, role_key):
            return jsonify({'error': 'role not found'}), 404
        conn.execute('DELETE FROM job_description_tasks WHERE role_key=?', (role_key,))
        conn.execute('DELETE FROM job_description_sections WHERE role_key=?', (role_key,))
        conn.execute('DELETE FROM job_role_templates WHERE role_key=?', (role_key,))
        conn.execute('DELETE FROM job_assignments WHERE role_key=?', (role_key,))
    return jsonify({'ok': True, 'role_key': role_key})


@jobs.route('/section/update', methods=['POST'])
@_admin_required
def update_section():
    role_key    = request.form.get('role_key', '').strip()
    section_key = request.form.get('section_key', '').strip()
    title       = request.form.get('title', '').strip()
    color       = request.form.get('color', '').strip()
    icon        = request.form.get('icon', '').strip()

    if not role_key or not section_key or not title:
        return jsonify({'error': 'invalid'}), 400

    with _get_db() as conn:
        if not _role_exists(conn, role_key):
            return jsonify({'error': 'role not found'}), 404
        row = conn.execute('''
            SELECT color, icon FROM job_description_sections
            WHERE role_key=? AND section_key=?
        ''', (role_key, section_key)).fetchone()
        if not row:
            return jsonify({'error': 'section not found'}), 404
        color = _clean_color(color, row['color'] or '#7B1FA2')
        icon = icon or row['icon'] or 'fas fa-list-check'
        conn.execute('''
            UPDATE job_description_sections
            SET title=?, color=?, icon=?
            WHERE role_key=? AND section_key=?
        ''', (title, color, icon, role_key, section_key))
    return jsonify({'ok': True, 'title': title, 'color': color, 'icon': icon})


@jobs.route('/section/add', methods=['POST'])
@_admin_required
def add_section():
    role_key = request.form.get('role_key', '').strip()
    title    = request.form.get('title', '').strip()
    color    = _clean_color(request.form.get('color', ''), '#7B1FA2')
    icon     = request.form.get('icon', '').strip() or 'fas fa-list-check'

    if not role_key or not title:
        return jsonify({'error': 'invalid'}), 400
    with _get_db() as conn:
        if not _role_exists(conn, role_key):
            return jsonify({'error': 'role not found'}), 404
        # Unique section_key per role.
        base = _slugify(title, 'section')
        section_key = base
        n = 2
        while _section_exists(conn, role_key, section_key):
            section_key = f'{base}_{n}'; n += 1
        next_order = conn.execute('''
            SELECT COALESCE(MAX(sort_order), -1) + 1 as n
            FROM job_description_sections WHERE role_key=?
        ''', (role_key,)).fetchone()['n']
        conn.execute('''INSERT INTO job_description_sections
            (role_key, section_key, title, color, icon, sort_order)
            VALUES (?,?,?,?,?,?)''',
            (role_key, section_key, title, color, icon, next_order))
    return jsonify({'ok': True, 'role_key': role_key, 'section_key': section_key,
                    'title': title, 'color': color, 'icon': icon})


@jobs.route('/section/delete', methods=['POST'])
@_admin_required
def delete_section():
    role_key    = request.form.get('role_key', '').strip()
    section_key = request.form.get('section_key', '').strip()
    if not role_key or not section_key:
        return jsonify({'error': 'invalid'}), 400
    with _get_db() as conn:
        if not _section_exists(conn, role_key, section_key):
            return jsonify({'error': 'section not found'}), 404
        conn.execute('DELETE FROM job_description_tasks WHERE role_key=? AND section_key=?',
                     (role_key, section_key))
        conn.execute('DELETE FROM job_description_sections WHERE role_key=? AND section_key=?',
                     (role_key, section_key))
    return jsonify({'ok': True})


@jobs.route('/task/update', methods=['POST'])
@_admin_required
def update_task():
    task_id = request.form.get('task_id', '').strip()
    name    = request.form.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Task name is required'}), 400
    try:
        task_id = int(task_id)
    except (TypeError, ValueError):
        return jsonify({'error': 'invalid task'}), 400

    with _get_db() as conn:
        cur = conn.execute('''
            UPDATE job_description_tasks SET task_name=? WHERE id=?
        ''', (name, task_id))
        if cur.rowcount == 0:
            return jsonify({'error': 'task not found'}), 404
    return jsonify({'ok': True, 'id': task_id, 'name': name})


@jobs.route('/task/add', methods=['POST'])
@_admin_required
def add_task():
    role_key    = request.form.get('role_key', '').strip()
    section_key = request.form.get('section_key', '').strip()
    name        = request.form.get('name', '').strip()
    if not role_key or not section_key or not name:
        return jsonify({'error': 'invalid'}), 400

    with _get_db() as conn:
        if not _section_exists(conn, role_key, section_key):
            return jsonify({'error': 'section not found'}), 404
        next_order = conn.execute('''
            SELECT COALESCE(MAX(task_order), -1) + 1 as next_order
            FROM job_description_tasks
            WHERE role_key=? AND section_key=?
        ''', (role_key, section_key)).fetchone()['next_order']
        cur = conn.execute('''
            INSERT INTO job_description_tasks
                (role_key,section_key,task_order,task_name)
            VALUES (?,?,?,?)
        ''', (role_key, section_key, next_order, name))
        task_id = cur.lastrowid
    return jsonify({'ok': True, 'id': task_id, 'name': name, 'order': next_order})


@jobs.route('/task/delete', methods=['POST'])
@_admin_required
def delete_task():
    task_id = request.form.get('task_id', '').strip()
    try:
        task_id = int(task_id)
    except (TypeError, ValueError):
        return jsonify({'error': 'invalid task'}), 400

    with _get_db() as conn:
        cur = conn.execute('DELETE FROM job_description_tasks WHERE id=?', (task_id,))
        if cur.rowcount == 0:
            return jsonify({'error': 'task not found'}), 404
    return jsonify({'ok': True, 'id': task_id})
