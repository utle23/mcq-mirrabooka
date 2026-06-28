from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file, send_from_directory, flash, abort
from functools import wraps
import sqlite3, os, json, calendar, uuid
from datetime import datetime, date, timedelta, timezone
from io import BytesIO
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = 'mcq-mirrabooka-2024-secure-key'

# ── Session security ─────────────────────────────────────────────────────────
# Defence against shared-device drift: 30-minute idle timeout AND a hard
# 8-hour absolute cap from login (covers one full shift, no longer).
SESSION_IDLE_TIMEOUT     = timedelta(minutes=30)
SESSION_ABSOLUTE_TIMEOUT = timedelta(hours=8)
app.config['PERMANENT_SESSION_LIFETIME'] = SESSION_ABSOLUTE_TIMEOUT

from prep_routes      import prep      as prep_bp,      init_prep_tables
from pastry_routes    import pastry    as pastry_bp,    init_pastry_tables
from inventory_routes import inventory as inventory_bp, init_inventory_tables
from job_routes       import jobs      as jobs_bp,      init_job_tables
from rules_routes     import rules_bp,                  init_rules_tables
from training_routes  import training_bp,               init_training_tables
from whatsapp_share   import whatsapp_bp,                init_whatsapp
from packaging_routes import packaging_bp,               init_packaging
from orders_routes    import orders     as orders_bp,    init_order_tables
from equipment_routes import equipment  as equipment_bp, init_equipment_tables
from structure_routes import structure  as structure_bp, init_structure_tables
from webauthn_routes  import webauthn_bp,                 init_webauthn
from food_pricing_routes import food_pricing as food_pricing_bp, init_food_pricing_tables
from food_safety_routes import defrost_bp, delivery_bp, init_food_safety_tables
from branch_seed import seed_subiaco_branch
import email_service
app.register_blueprint(prep_bp)
app.register_blueprint(pastry_bp)
app.register_blueprint(inventory_bp)
app.register_blueprint(jobs_bp)
app.register_blueprint(rules_bp)
app.register_blueprint(training_bp)
app.register_blueprint(whatsapp_bp)
app.register_blueprint(packaging_bp)
app.register_blueprint(orders_bp)
app.register_blueprint(equipment_bp)
app.register_blueprint(structure_bp)
app.register_blueprint(webauthn_bp)
app.register_blueprint(food_pricing_bp)
app.register_blueprint(defrost_bp)
app.register_blueprint(delivery_bp)
DB_PATH      = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'mcq_restaurant.db')
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_EXT = {'jpg', 'jpeg', 'png', 'gif', 'webp', 'heic', 'heif'}
PHOTOS_REQUIRED = 4

# ─── Data Definitions ──────────────────────────────────────────────────────────

CHECKLISTS = {
    'take_order': {
        'title': 'Cashier',
        'short': 'Cashier',
        'color': '#2196F3',
        'badge': 'primary',
        'opening': [
            'Open till & check cash balance ($350)',
            'Uniform check (hat, shirt, apron)',
            'Counter area ready for service, clean surrounding counter area',
            'Check the POS system - tills and EFTPOS all working',
            'Prepare vegetables & garnish for pho and bun bo hue',
            'Label beverages and label price and name for pastry',
            'Record temperature of fried pastry food (Pastry Temperature Record)',
            'Arrange the line-up for customers to take order',
            'Complete checklists (Food Temperature Records)',
            'Receive bread & pastry invoices from supplier',
        ],
        'closing': [
            'Take orders & resolve service issues (during shift)',
            'Clean the counter area',
            'Clean display glass cabinets',
            'Clean Cold Unit 2 Soft Drink Fridge',
            'Clean Cold Unit 3 Rice Paper Roll Fridge',
            'Refill drinking water',
            'Refill spoons, forks, chopsticks & sauces',
            'Buy bean sprouts, mint & veg for pho and bun bo hue for tomorrow',
            'Check income & till balance before closing',
            'Clean front floor',
            'Clean air fryer',
            'Check banh mi packaging bag order with Huu Anh',
            'Refill receipt roll',
            'Refill carry bags, paper bags & cup holders for takeaway',
            'Collect invoices & send to office every Friday (give to Van Anh)',
        ],
    },
    'banh_mi': {
        'title': 'Banh Mi',
        'short': 'Banh Mi',
        'color': '#FF9800',
        'badge': 'warning',
        'opening': [
            'Set up & display banh mi fridge bar (including banh mi, red pork meat, cha lua, cha gan, ngò, pickles, sliced chilli, cucumber)',
            'Check & test quality of each item (report to Manager if food not good to sell)',
            'Cut bread, spread pate & butter',
            'Display plain bread in the basket',
            'Set up roast pork station',
            'Chop roast pork & chicken',
            'Pre-make roast pork banh mi (all prep done by 8:00 AM)',
            'Change gloves for vegetarian orders',
            'Order coriander, cucumber & radish from supermarket',
            'Complete Banh Mi Food Temperature Record',
            'Customer order service (prepare banh mi per requests)',
        ],
        'closing': [
            'Prepare next-day veg: coriander, cucumber, soy sauce follow prep timetable',
            'Prepare food in the tray for tomorrow setup (include coriander, pickle, cucumber, sauce: soy, gravy, mayo)',
            'Defrost pate (move from freezer to fridge for tomorrow)',
            'Clean up banh mi station',
            'Sanitise the roast pork table & chopping board',
            'Clean Cold Unit 4 Banh Mi Fridge (end of day)',
            'Wash dishes, final dish check at 4:30 PM',
        ],
    },
    'chef': {
        'title': 'Chef',
        'short': 'Chef',
        'color': '#F44336',
        'badge': 'danger',
        'opening': [
            'Turn on lights and ventilation/exhaust fan',
            'Turn on gas & all kitchen equipment',
            'Cook rice',
            'Make savoury sticky rice',
            'Grill the pork chop',
            'Marinate chicken & pork followed food preparation timetable',
            'Pho soup station set up',
            'Set up rice noodle (pho noodle, BBH noodle, dry noodles)',
            'Set up chef fridge bar station (coriander, mint, tomato, cucumber, salad onion, spring onion, raw beef, cooked beef, beef ball, pork, etc.)',
            'Fry fried pastry items (fried spring roll, samosa, banana, etc.)',
            'Fry eggs for service',
            'Packaging items filled',
            'Customer order service',
            'Complete Chef Section Food Temperature Record',
        ],
        'closing': [
            'Marinate chicken & pork for next day followed food preparation timetable',
            'Prepare fish sauce and soy sauce for tomorrow',
            'Prep & store next-day food properly',
            'Clean kitchen areas & equipment',
            'Spray & wash kitchen floor',
            'Clean back floor',
            'Clean grilled machine',
            'Clean oven',
            'Clean stove',
            'Clean gas burners',
            'Clean deep fryer',
            'Clean Cold Unit 6 Soup & Rice Fridge',
            'Clean Cold Unit 7 Noodle & Soup Bar',
            'Clean Cold Unit 8 Food Prep Fridge',
            'Clean cold food fridge',
            'Clean the display fridge bar',
            'All kitchen equipment turned OFF',
            'Gas turned OFF',
            'Water turned OFF',
            'Kitchen lights OFF',
            'Door locked',
        ],
    },
    'grill_beef': {
        'title': 'Kitchen Hand',
        'short': 'K.Hand',
        'color': '#4CAF50',
        'badge': 'success',
        'opening': [
            'Roast pork (10-15/day, 15-20 weekends)',
            'Grill chicken (2 trays/day, 3-4 weekends)',
            'Stir-fry beef / pork / tofu',
            'Slice beef ~5 kg for pho & banh mi',
            'Slice pork & char siu meat',
            'Beef soup Tue/Thu/Sat — Char siu Mon/Wed/Fri',
        ],
        'closing': [
            'Check & refill seasonings',
            'Prepare food for next day when time allows',
            'Clean the preparation area',
            'Empty rubbish bins',
            'Wash dishes',
            'Scrub & clean back floor',
            'Clean Freezer 1',
            'Clean Freezer 2',
        ],
    },
    'serve_order': {
        'title': 'Serve Order / Drinks',
        'short': 'Drinks',
        'color': '#00BCD4',
        'badge': 'info',
        'opening': [
            'Prepare black iced coffee base (at least 2 jars)',
            'Prepare watermelon / tropical / sugarcane juice',
            'Prepare smoothie base mix',
            'Wash sugarcane',
            'Prepare fresh rice paper rolls for the day',
            'Make at least 1 tropical juice, 1 jar watermelon juice, 1 jar sugarcane juice, 4 jars black iced coffee',
            'Check coffee, condensed milk & fruit stock',
        ],
        'closing': [
            'Peel & prepare fruit for juice (next day): orange, apple, watermelon, etc.',
            'Refill fruit containers, lids clean & covered',
            'Set up & check quality, clean fruit juice in the fridge',
            'Check enough coffee / condensed milk',
            'Clean fruit juicer',
            'Clean sugarcane juicer',
            'Change fruit display (Mon, Wed, Fri, Sun)',
            'Wash all drink containers',
            'Check & refill cups, lids, straws',
            'Wash banh mi trays',
            'Bag returned pastry, bring to counter',
            'Clean microwave',
            'Clean pastry tray',
            'Clean Cold Unit 1 Fruit Juice Fridge',
            'Clean Cold Unit 5 Coffee Fridge',
            'Clean front floor',
            'Work area tidy, lights off',
        ],
    },
}

TEMPERATURES = {
    'banh_mi': {
        'title': 'Banh Mi Food Temperature Record',
        'short': 'Banh Mi',
        'color': '#FF9800',
        'badge': 'warning',
        # Each food has 'kind' = 'hot' (must hold >= 60°C) or 'cold' (must hold <= 5°C).
        # Only hot foods at Banh Mi: roast pork, roast chicken, stir-fry beef/pork/tofu,
        # plus beef brisket. Everything else (deli meats, pate, veg, pickles) = cold.
        'foods': [
            {'name': 'Cooked Pork Ham (Cha Lua)',   'kind': 'cold'},
            {'name': 'Cooked Pork Meat (Thit Nguoi)','kind': 'cold'},
            {'name': 'Cooked Pork Jelly (Gio Thu)', 'kind': 'cold'},
            {'name': 'Roasted Pork',                'kind': 'hot'},
            {'name': 'Roasted Chicken',             'kind': 'hot'},
            {'name': 'Stir-Fry Beef',               'kind': 'hot'},
            {'name': 'Stir-Fry Pork',               'kind': 'hot'},
            {'name': 'Stir-Fry Tofu',               'kind': 'hot'},
            {'name': 'Beef Brisket',                'kind': 'hot'},
            {'name': 'Pate',                        'kind': 'cold'},
            {'name': 'Butter',                      'kind': 'cold'},
            {'name': 'Pickled',                     'kind': 'cold'},
            {'name': 'Coriander',                   'kind': 'cold'},
            {'name': 'Cucumber',                    'kind': 'cold'},
            {'name': 'Carrot Pickled',              'kind': 'cold'},
        ],
    },
    'chef': {
        'title': 'Chef / Food Temperature Record',
        'short': 'Chef',
        'color': '#F44336',
        'badge': 'danger',
        # Chef record mixes three holding rules:
        #   cold → must hold <= 5°C
        #   room → ambient items, safe between 15°C and 30°C
        #   hot  → must hold >= 60°C
        'foods': [
            {'name': 'Cooked Beef',         'kind': 'cold'},
            {'name': 'Raw Beef',            'kind': 'cold'},
            {'name': 'Cooked Beef Ball',    'kind': 'cold'},
            {'name': 'Cooked Pork',         'kind': 'cold'},
            {'name': 'Cooked Pork Ham',     'kind': 'cold'},
            {'name': 'Pho Noodle',          'kind': 'room'},
            {'name': 'Rice Vermicelli Bun', 'kind': 'room'},
            {'name': 'Pork Chop',           'kind': 'hot'},
            {'name': 'Egg',                 'kind': 'hot'},
            {'name': 'Pickled',             'kind': 'cold'},
            {'name': 'Cucumber',            'kind': 'cold'},
            {'name': 'Lettuce',             'kind': 'room'},
            {'name': 'Tomato',              'kind': 'cold'},
        ],
    },
    'pastry': {
        'title': 'Pastry / Hot Food Temperature Record',
        'short': 'Pastry',
        'color': '#9C27B0',
        'badge': 'secondary',
        # Pastry hot food display = hot holding only. All items must hold >= 60°C.
        'foods': [
            {'name': 'Sesame Ball',         'kind': 'hot'},
            {'name': 'Sesame Bread',        'kind': 'hot'},
            {'name': 'Fried Pork Dumpling', 'kind': 'hot'},
            {'name': 'Steam Bun',           'kind': 'hot'},
            {'name': 'Sweet Donuts',        'kind': 'hot'},
            {'name': 'Chinese Donut',       'kind': 'hot'},
            {'name': 'Fried Banana',        'kind': 'hot'},
            {'name': 'Pork Pastry',         'kind': 'hot'},
            {'name': 'Bird Nest Cake',      'kind': 'hot'},
            {'name': 'Curry Puff',          'kind': 'hot'},
            {'name': 'Spring Roll',         'kind': 'hot'},
            {'name': 'Banana Banh Tet',     'kind': 'hot'},
            {'name': 'Beef Samosa',         'kind': 'hot'},
            {'name': 'Banh Bao (Dumpling)', 'kind': 'hot'},
        ],
    },
}


def temp_food_kind_default(temp_type, name):
    """Look up the food's kind from the TEMPERATURES seed dict by exact name match,
    falling back to the type's prevailing kind so admin-added items aren't classed wrong."""
    for f in TEMPERATURES.get(temp_type, {}).get('foods', []):
        if f['name'].strip().lower() == (name or '').strip().lower():
            return f['kind']
    # Type-level defaults
    if temp_type == 'pastry':  return 'hot'
    if temp_type == 'chef':    return 'cold'
    return 'cold'   # banh_mi default

# Holding rules per food kind — single source of truth for headings, badges,
# colour-coding and out-of-zone alerts.
#   cold → safe ≤ 5°C   ·   room → safe 15–30°C   ·   hot → safe ≥ 60°C
TEMP_KIND_RULES = {
    'cold': {'label': 'COLD', 'icon': 'fa-snowflake',        'rule': '5°C or below (≤ 5°C)',
             'lo': None, 'hi': 5.0,  'badge': '#1565C0'},
    'room': {'label': 'ROOM', 'icon': 'fa-temperature-half', 'rule': 'room temperature, 15°C to 30°C',
             'lo': 15.0, 'hi': 30.0, 'badge': '#00897B'},
    'hot':  {'label': 'HOT',  'icon': 'fa-fire',             'rule': '60°C or above (≥ 60°C)',
             'lo': 60.0, 'hi': None, 'badge': '#C0392B'},
}

def temp_is_unsafe(kind, temp):
    """True if a temperature reading is outside the safe range for its kind."""
    if temp is None:
        return False
    r = TEMP_KIND_RULES.get(kind, TEMP_KIND_RULES['cold'])
    if r['lo'] is not None and temp < r['lo']:
        return True
    if r['hi'] is not None and temp > r['hi']:
        return True
    return False

STAFF = [
    'DANG, THI LAN UY', 'DOAN, THI NI', 'NGUYEN, PHU TAN',
    'HUYNH, ANH TRI', 'Thang Nguyen', 'MA, THANH PHUNG',
    'NGUYEN, THI NHAI', 'NGUYEN, VAN PHI LONG', 'DO, NGUYEN',
    'NGUYEN, THI NGOC PHUC', 'Ho Quynh', 'NGUYEN, HANG SANG',
    'VU, TRAN DO CAO',
]

BRANCHES = ['Mirrabooka', 'Subiaco', 'Morley']
MANAGERS = ['MA, THANH PHUNG', 'Khoi']

# ── Multi-store foundation ────────────────────────────────────────────────────
# One shared SQLite DB; each store's operational data is isolated by store_id.
# stores.id is the canonical key; branch text is display/back-compat only.
# Seed order fixes ids: mirrabooka=1, morley=2, subiaco=3.
STORES_SEED = [
    ('mirrabooka', 'MCQ Mirrabooka'),
    ('morley',     'MCQ Morley'),
    ('subiaco',    'MCQ Subiaco'),
]

# Tables that carry store_id. This includes operational records plus the
# branch-specific templates/catalogues that staff see when they log in.
STORE_SCOPED_TABLES = [
    'staff_members', 'checklist_sessions', 'temp_sessions', 'packaging_orders',
    'staff_violations', 'issue_reports', 'equipment_units', 'equipment_temp_readings',
    'prep_weekly_schedules', 'prep_weekly_tasks', 'prep_daily_status',
    'pastry_delivery', 'pastry_sales', 'orders',
    'checklist_task_templates', 'temp_food_templates', 'prep_task_templates',
    'packaging_suppliers', 'pastry_suppliers', 'pastry_items',
    # batch 3 — HR / org-chart tables
    'staff_certificates', 'staff_birthdays', 'monthly_reward_decisions',
    'monthly_reward_adjustments', 'salary_raise_reviews',
    'structure_departments', 'structure_members', 'structure_meta',
]


# Session/request-based store scoping helpers live in store_scope.py so the
# blueprints can share them without importing app.py (circular import).
from store_scope import (is_super_admin, current_store_id, selected_store_scope,
                         store_filter_clause, store_guard_clause)


def get_stores(active_only=True):
    """All stores as list of dicts (id, code, name, ...), ordered by id."""
    try:
        with get_db() as conn:
            q = 'SELECT * FROM stores'
            if active_only:
                q += ' WHERE active=1'
            q += ' ORDER BY id'
            return [dict(r) for r in conn.execute(q).fetchall()]
    except Exception:
        return []


def store_id_for_branch(branch):
    """Map a branch display name to stores.id via its code (lowercased name).
    Falls back to 1 (Mirrabooka) so logins never break during migration."""
    code = (branch or '').strip().lower()
    try:
        with get_db() as conn:
            row = conn.execute('SELECT id FROM stores WHERE code=?', (code,)).fetchone()
        if row:
            return row['id']
    except Exception:
        pass
    return 1


def get_active_staff(store_id=None):
    """Single source of truth for staff dropdowns: the live staff_members table
    (active members only) for the CURRENT store. Adding a member on the Staff
    page makes them show up in every checklist / temperature / report picker;
    deleting or deactivating them removes the name everywhere. Falls back to the
    seed STAFF list only if the table is unavailable or empty.

    Dropdowns always need a concrete store, so this uses the session store by
    default (not the super_admin all-stores view)."""
    if store_id is None:
        store_id = current_store_id()
    try:
        with get_db() as conn:
            names = [r['name'] for r in conn.execute(
                'SELECT name FROM staff_members WHERE active=1 AND store_id=? '
                'ORDER BY name', (store_id,)).fetchall()]
        return names or list(STAFF)
    except Exception:
        return list(STAFF)

ISSUE_CATEGORIES = {
    'low_stock':          {'label': 'Low Stock',                    'icon': 'fa-box-open',             'color': '#E65100'},
    'missing_ingredients':{'label': 'Missing Ingredients',          'icon': 'fa-bowl-food',            'color': '#8D6E63'},
    'supplier':           {'label': 'Supplier Delivery Issue',      'icon': 'fa-truck',                'color': '#1565C0'},
    'food_quality':       {'label': 'Food Quality Issue',           'icon': 'fa-circle-exclamation',   'color': '#C62828'},
    'equipment':          {'label': 'Equipment Problem',            'icon': 'fa-tools',                'color': '#F57C00'},
    'cleaning':           {'label': 'Cleaning Issue',               'icon': 'fa-broom',                'color': '#2E7D32'},
    'temperature':        {'label': 'Temperature Issue',            'icon': 'fa-thermometer-half',     'color': '#D84315'},
    'hr':                 {'label': 'Salary Issue',                 'icon': 'fa-money-bill-wave',      'color': '#7B1FA2'},
    'conflict':           {'label': 'Conflict',                     'icon': 'fa-user-slash',           'color': '#E53935'},
    'task_incomplete':    {'label': 'Task Cannot Be Completed',     'icon': 'fa-circle-xmark',         'color': '#546E7A'},
    'food_safety':        {'label': 'Food Safety Concern',          'icon': 'fa-exclamation-triangle', 'color': '#B71C1C'},
    'customer':           {'label': 'Customer Complaint',           'icon': 'fa-comment-dots',         'color': '#D84315'},
    'maintenance':        {'label': 'Maintenance / Facility',       'icon': 'fa-hammer',               'color': '#558B2F'},
    'supplies':           {'label': 'Supply / Stock Request',       'icon': 'fa-shopping-cart',        'color': '#0288D1'},
    'timesheet':          {'label': 'Missed Clock-in / Clock-out',  'icon': 'fa-clock',                'color': '#00796B'},
    'suggestion':         {'label': 'Idea / Suggestion',            'icon': 'fa-lightbulb',            'color': '#F9A825'},
    'other':              {'label': 'Other',                        'icon': 'fa-question-circle',      'color': '#546E7A'},
}

VIOLATION_RULES_SEED = [
    {
        'code': 'uniform',
        'title': 'Uniform not worn correctly',
        'category': 'Presentation',
        'severity': 'minor',
        'description': 'Staff member is missing required uniform items such as hat, shirt, apron, or wears unsuitable presentation for service.',
        'default_action': 'Remind staff of uniform standard. Repeat breach requires manager follow-up.',
        'color': '#1565C0',
        'icon': 'fa-shirt',
    },
    {
        'code': 'hygiene_ppe',
        'title': 'Hygiene / PPE breach',
        'category': 'Food Safety',
        'severity': 'serious',
        'description': 'Examples: no gloves when required, poor hand hygiene, hair not controlled, or unsafe handling of ready-to-eat food.',
        'default_action': 'Stop task, correct immediately, retrain on hygiene expectations.',
        'color': '#C62828',
        'icon': 'fa-hand-sparkles',
    },
    {
        'code': 'temperature_record',
        'title': 'Temperature record missed or incorrect',
        'category': 'Food Safety',
        'severity': 'serious',
        'description': 'Temperature record not completed, completed late, pre-filled, or contains unsafe/inaccurate readings.',
        'default_action': 'Manager to verify product safety and retrain staff on temperature record process.',
        'color': '#D84315',
        'icon': 'fa-temperature-half',
    },
    {
        'code': 'checklist_late',
        'title': 'Daily checklist late / incomplete',
        'category': 'Compliance',
        'severity': 'moderate',
        'description': 'Opening or closing checklist was submitted late, missing photos, or not completed properly.',
        'default_action': 'Review checklist deadline and required evidence with responsible staff.',
        'color': '#F57C00',
        'icon': 'fa-clipboard-list',
    },
    {
        'code': 'cleaning_not_done',
        'title': 'Cleaning task not completed',
        'category': 'Cleanliness',
        'severity': 'moderate',
        'description': 'Assigned cleaning task, fridge clean, floor clean, equipment clean, or closing hygiene duty was missed.',
        'default_action': 'Assign correction before close and record follow-up check.',
        'color': '#2E7D32',
        'icon': 'fa-broom',
    },
    {
        'code': 'phone_use',
        'title': 'Phone use during service',
        'category': 'Conduct',
        'severity': 'minor',
        'description': 'Personal phone use during active service or food preparation without manager approval.',
        'default_action': 'Verbal reminder. Escalate if repeated.',
        'color': '#6A1B9A',
        'icon': 'fa-mobile-screen',
    },
    {
        'code': 'customer_service',
        'title': 'Customer service issue',
        'category': 'Service',
        'severity': 'moderate',
        'description': 'Rude tone, incorrect order handling, unresolved complaint, or poor service behaviour.',
        'default_action': 'Manager to discuss incident and coach service standard.',
        'color': '#00838F',
        'icon': 'fa-comments',
    },
    {
        'code': 'attendance',
        'title': 'Late arrival / missed shift procedure',
        'category': 'Attendance',
        'severity': 'moderate',
        'description': 'Late arrival, no notice, missed clock-in/out, or shift attendance issue.',
        'default_action': 'Confirm reason, correct timesheet if needed, and monitor repeat pattern.',
        'color': '#455A64',
        'icon': 'fa-clock',
    },
    {
        'code': 'other',
        'title': 'Other',
        'category': 'Other',
        'severity': 'minor',
        'description': 'Use this when the incident does not match an existing rule. Write the exact issue before submitting.',
        'default_action': 'Manager to review details and choose the appropriate follow-up action.',
        'color': '#607D8B',
        'icon': 'fa-circle-question',
    },
]

# ─── Database ──────────────────────────────────────────────────────────────────

def get_db():
    # timeout=30 + busy_timeout: under concurrent use (several staff at once, or
    # a slow PDF holding a read while another request writes) SQLite would
    # otherwise raise "database is locked" after 5s → 500/502. WAL (set once at
    # startup) lets readers and the single writer work without blocking.
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')
    conn.execute('PRAGMA busy_timeout = 20000')
    return conn


def _update_then_insert(conn, table, key_values, data_values):
    """Upsert one row without depending on a UNIQUE index.

    Some deployed databases can briefly miss a newly-added unique index when
    startup migration runs while another worker holds SQLite's write lock.
    UPDATE-first keeps saves working in that state. SQLite serializes writers,
    and the retry handles a concurrent insert when the index is present.
    """
    where = ' AND '.join(f'{column}=?' for column in key_values)
    assignments = ', '.join(f'{column}=?' for column in data_values)
    update_values = list(data_values.values()) + list(key_values.values())
    cur = conn.execute(
        f'UPDATE {table} SET {assignments} WHERE {where}',
        update_values,
    )
    if cur.rowcount:
        return

    combined = {**key_values, **data_values}
    columns = ', '.join(combined)
    placeholders = ', '.join('?' for _ in combined)
    try:
        conn.execute(
            f'INSERT INTO {table} ({columns}) VALUES ({placeholders})',
            list(combined.values()),
        )
    except sqlite3.IntegrityError as insert_error:
        # Another request may have inserted the same natural key after our
        # UPDATE. Update that row rather than surfacing a 500 to the staff.
        retry = conn.execute(
            f'UPDATE {table} SET {assignments} WHERE {where}',
            update_values,
        )
        if not retry.rowcount:
            raise insert_error


def ensure_save_upsert_constraints(db_path):
    """Repair the two child-row keys used by checklist/temperature saves.

    The save path no longer requires these indexes to function, but keeping
    them guarantees data integrity and makes accidental duplicates impossible.
    If an older database already contains duplicates, retain the newest row.
    """
    conn = sqlite3.connect(db_path, timeout=30)
    try:
        conn.execute('PRAGMA busy_timeout = 30000')
        conn.execute('BEGIN IMMEDIATE')
        for table, order_column, index_name in (
            ('checklist_tasks', 'task_order', 'idx_chktasks_session_order'),
            ('temp_readings', 'food_order', 'idx_tempreadings_session_order'),
        ):
            exists = conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
                (table,),
            ).fetchone()
            if not exists:
                continue
            conn.execute(f'''
                DELETE FROM {table}
                WHERE id NOT IN (
                    SELECT MAX(id)
                    FROM {table}
                    GROUP BY session_id, {order_column}
                )
            ''')
            conn.execute(
                f'CREATE UNIQUE INDEX IF NOT EXISTS {index_name} '
                f'ON {table}(session_id, {order_column})'
            )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def _add_store_to_unique(conn, table, old_unique, new_unique):
    """Rebuild `table` so its natural-key UNIQUE constraint also includes
    store_id. Safe + idempotent: no-op if already migrated or the old
    constraint isn't present. Preserves all rows, columns, ids and indexes.

    SQLite can't ALTER a constraint, so we copy into a new table. The caller
    MUST pass an autocommit connection (isolation_level=None) so the
    PRAGMA foreign_keys=OFF below actually takes effect — otherwise the
    implicit DELETE that DROP TABLE performs on a parent would CASCADE into
    child tables (e.g. checklist_tasks) and wipe them."""
    row = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone()
    if not row or not row[0]:
        return
    create_sql = row[0]
    norm = create_sql.replace(' ', '')
    if new_unique.replace(' ', '') in norm:
        return  # already migrated
    if old_unique.replace(' ', '') not in norm:
        return  # constraint shape unexpected — leave it alone, do not risk data
    cols = [r[1] for r in conn.execute(f'PRAGMA table_info({table})')]
    collist = ', '.join(cols)
    # Preserve user-defined indexes (auto-indexes have NULL sql and are rebuilt
    # automatically from the UNIQUE/PK definitions).
    index_sqls = [r[0] for r in conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='index' AND tbl_name=? AND sql IS NOT NULL",
        (table,)).fetchall()]
    tmp = f'{table}__mig'
    new_create = (create_sql
                  .replace(old_unique, new_unique)
                  .replace(f'IF NOT EXISTS {table}', tmp)
                  .replace(f'EXISTS {table}', tmp))
    if tmp not in new_create:
        new_create = create_sql.replace(old_unique, new_unique).replace(
            f' {table} ', f' {tmp} ', 1).replace(f' {table}(', f' {tmp}(', 1)
    if tmp not in new_create:
        return  # couldn't safely rename target — bail without touching data

    conn.execute('PRAGMA foreign_keys = OFF')   # effective only outside a txn
    conn.execute('BEGIN')
    try:
        conn.execute(f'DROP TABLE IF EXISTS {tmp}')
        conn.execute(new_create)
        conn.execute(f'INSERT INTO {tmp} ({collist}) SELECT {collist} FROM {table}')
        conn.execute(f'DROP TABLE {table}')
        conn.execute(f'ALTER TABLE {tmp} RENAME TO {table}')
        for isql in index_sqls:   # recreate the indexes we dropped with the table
            try:
                conn.execute(isql)
            except Exception:
                pass
        conn.execute('COMMIT')
    except Exception:
        conn.execute('ROLLBACK')
        raise
    finally:
        conn.execute('PRAGMA foreign_keys = ON')


def _rebuild_structure_meta_per_store(conn):
    """structure_meta is key→value with `key` as PRIMARY KEY. Make it per store:
    PRIMARY KEY (key, store_id). Autocommit conn + FK off. Idempotent."""
    row = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='structure_meta'").fetchone()
    if not row or not row[0]:
        return
    sql = row[0]
    norm = sql.replace(' ', '')
    if 'PRIMARYKEY(key,store_id)' in norm:
        return  # already migrated
    if 'keyTEXTPRIMARYKEY' not in norm:
        return  # unexpected shape — leave it
    cols = [r[1] for r in conn.execute('PRAGMA table_info(structure_meta)')]
    if 'store_id' not in cols:
        return  # store_id column must exist first (added in the ALTER pass)
    collist = ', '.join(cols)
    conn.execute('PRAGMA foreign_keys = OFF')
    conn.execute('BEGIN')
    try:
        conn.execute('DROP TABLE IF EXISTS structure_meta__mig')
        conn.execute('''CREATE TABLE structure_meta__mig (
            key TEXT NOT NULL, value TEXT, store_id INTEGER DEFAULT 1,
            PRIMARY KEY (key, store_id))''')
        conn.execute(f'INSERT INTO structure_meta__mig ({collist}) SELECT {collist} FROM structure_meta')
        conn.execute('DROP TABLE structure_meta')
        conn.execute('ALTER TABLE structure_meta__mig RENAME TO structure_meta')
        conn.execute('COMMIT')
    except Exception:
        conn.execute('ROLLBACK')
        raise
    finally:
        conn.execute('PRAGMA foreign_keys = ON')


def _rebuild_checklist_templates_per_store(conn):
    """Make checklist task templates branch-specific.

    Earlier versions used PRIMARY KEY(chk_type, section, task_order), which
    meant Subiaco edits overwrote Mirrabooka. Rebuild to include store_id.
    """
    row = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='checklist_task_templates'"
    ).fetchone()
    if not row or not row[0]:
        return
    norm = row[0].replace(' ', '').replace('\n', '')
    if 'PRIMARYKEY(chk_type,section,task_order,store_id)' in norm:
        return
    cols = [r[1] for r in conn.execute('PRAGMA table_info(checklist_task_templates)')]
    store_expr = 'COALESCE(store_id, 1)' if 'store_id' in cols else '1'
    conn.execute('PRAGMA foreign_keys = OFF')
    conn.execute('BEGIN')
    try:
        conn.execute('DROP TABLE IF EXISTS checklist_task_templates__mig')
        conn.execute('''CREATE TABLE checklist_task_templates__mig (
            chk_type   TEXT NOT NULL,
            section    TEXT NOT NULL,
            task_order INTEGER NOT NULL,
            task_name  TEXT NOT NULL,
            store_id   INTEGER NOT NULL DEFAULT 1,
            PRIMARY KEY (chk_type, section, task_order, store_id)
        )''')
        conn.execute(f'''INSERT OR REPLACE INTO checklist_task_templates__mig
            (chk_type, section, task_order, task_name, store_id)
            SELECT chk_type, section, task_order, task_name, {store_expr}
            FROM checklist_task_templates''')
        conn.execute('DROP TABLE checklist_task_templates')
        conn.execute('ALTER TABLE checklist_task_templates__mig RENAME TO checklist_task_templates')
        conn.execute('COMMIT')
    except Exception:
        conn.execute('ROLLBACK')
        raise
    finally:
        conn.execute('PRAGMA foreign_keys = ON')


def _rebuild_temp_templates_per_store(conn):
    """Make food temperature templates branch-specific."""
    row = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='temp_food_templates'"
    ).fetchone()
    if not row or not row[0]:
        return
    norm = row[0].replace(' ', '').replace('\n', '')
    if 'PRIMARYKEY(temp_type,food_order,store_id)' in norm:
        return
    cols = [r[1] for r in conn.execute('PRAGMA table_info(temp_food_templates)')]
    store_expr = 'COALESCE(store_id, 1)' if 'store_id' in cols else '1'
    kind_expr = 'COALESCE(food_kind, "cold")' if 'food_kind' in cols else '"cold"'
    conn.execute('PRAGMA foreign_keys = OFF')
    conn.execute('BEGIN')
    try:
        conn.execute('DROP TABLE IF EXISTS temp_food_templates__mig')
        conn.execute('''CREATE TABLE temp_food_templates__mig (
            temp_type  TEXT NOT NULL,
            food_order INTEGER NOT NULL,
            food_name  TEXT NOT NULL,
            food_kind  TEXT NOT NULL DEFAULT 'cold',
            store_id   INTEGER NOT NULL DEFAULT 1,
            PRIMARY KEY (temp_type, food_order, store_id)
        )''')
        conn.execute(f'''INSERT OR REPLACE INTO temp_food_templates__mig
            (temp_type, food_order, food_name, food_kind, store_id)
            SELECT temp_type, food_order, food_name, {kind_expr}, {store_expr}
            FROM temp_food_templates''')
        conn.execute('DROP TABLE temp_food_templates')
        conn.execute('ALTER TABLE temp_food_templates__mig RENAME TO temp_food_templates')
        conn.execute('COMMIT')
    except Exception:
        conn.execute('ROLLBACK')
        raise
    finally:
        conn.execute('PRAGMA foreign_keys = ON')


def _rebuild_prep_schedules_per_store(conn):
    row = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='prep_weekly_schedules'"
    ).fetchone()
    if not row or not row[0]:
        return
    norm = row[0].replace(' ', '').replace('\n', '')
    if 'UNIQUE(week_start,store_id)' in norm:
        return
    cols = [r[1] for r in conn.execute('PRAGMA table_info(prep_weekly_schedules)')]
    store_expr = 'COALESCE(store_id, 1)' if 'store_id' in cols else '1'
    conn.execute('PRAGMA foreign_keys = OFF')
    conn.execute('BEGIN')
    try:
        conn.execute('DROP TABLE IF EXISTS prep_weekly_schedules__mig')
        conn.execute('''CREATE TABLE prep_weekly_schedules__mig (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            week_start TEXT NOT NULL,
            created_by TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            locked     INTEGER DEFAULT 0,
            locked_by  TEXT,
            locked_at  TEXT,
            notes      TEXT,
            store_id   INTEGER DEFAULT 1,
            UNIQUE(week_start, store_id)
        )''')
        conn.execute(f'''INSERT OR REPLACE INTO prep_weekly_schedules__mig
            (id, week_start, created_by, created_at, locked, locked_by, locked_at, notes, store_id)
            SELECT id, week_start, created_by, created_at, locked, locked_by, locked_at, notes, {store_expr}
            FROM prep_weekly_schedules''')
        conn.execute('DROP TABLE prep_weekly_schedules')
        conn.execute('ALTER TABLE prep_weekly_schedules__mig RENAME TO prep_weekly_schedules')
        conn.execute('COMMIT')
    except Exception:
        conn.execute('ROLLBACK')
        raise
    finally:
        conn.execute('PRAGMA foreign_keys = ON')


def migrate_multistore(db_path):
    """All multi-store schema work on a dedicated AUTOCOMMIT connection so the
    UNIQUE-constraint rebuilds can safely toggle foreign_keys. Idempotent."""
    conn = sqlite3.connect(db_path, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.isolation_level = None   # autocommit
    # Wait (don't error) if another worker/console briefly holds the DB during a
    # reload, and switch the DB to WAL once so readers never block the writer.
    try:
        conn.execute('PRAGMA busy_timeout = 30000')
        conn.execute('PRAGMA journal_mode = WAL')
    except Exception:
        pass
    try:
        # 1) stores table + seed (fixed ids: mirrabooka=1, morley=2, subiaco=3)
        conn.execute('''CREATE TABLE IF NOT EXISTS stores (
            id      INTEGER PRIMARY KEY AUTOINCREMENT,
            code    TEXT UNIQUE NOT NULL,
            name    TEXT NOT NULL,
            address TEXT DEFAULT '',
            phone   TEXT DEFAULT '',
            active  INTEGER DEFAULT 1
        )''')
        for sid, (scode, sname) in enumerate(STORES_SEED, start=1):
            conn.execute('INSERT OR IGNORE INTO stores (id,code,name) VALUES (?,?,?)',
                         (sid, scode, sname))
        # 1b) per-store login passwords (seed with the chain defaults so existing
        #     logins keep working until an owner changes them per store).
        for col in ('user_password', 'admin_password', 'kitchen_password'):
            try:
                conn.execute(f"ALTER TABLE stores ADD COLUMN {col} TEXT DEFAULT ''")
            except Exception:
                pass
        conn.execute("UPDATE stores SET user_password=?    WHERE COALESCE(user_password,'')=''",    (USER_PASSWORD,))
        conn.execute("UPDATE stores SET admin_password=?   WHERE COALESCE(admin_password,'')=''",   (ADMIN_PASSWORD,))
        conn.execute("UPDATE stores SET kitchen_password=? WHERE COALESCE(kitchen_password,'')=''", (KITCHEN_PASSWORD,))
        # 2) add store_id to operational tables + backfill existing rows to
        #    Mirrabooka (store_id=1). Legacy branch TEXT columns are left intact.
        for tbl in STORE_SCOPED_TABLES:
            try:
                conn.execute(f'ALTER TABLE {tbl} ADD COLUMN store_id INTEGER DEFAULT 1')
            except Exception:
                pass
            try:
                conn.execute(f'UPDATE {tbl} SET store_id=1 WHERE store_id IS NULL')
            except Exception:
                pass
        # 3) widen natural-key UNIQUE constraints to include store_id (only the
        #    tables whose key columns are chain-wide need it). Each rebuild is
        #    isolated: a transient failure (e.g. a brief lock during a reload) is
        #    skipped and retried next boot — it must NEVER crash app startup,
        #    because that would 502 the entire site.
        def _safe(fn, *a):
            try:
                fn(conn, *a)
            except Exception:
                pass
        _safe(_add_store_to_unique, 'checklist_sessions',
              'UNIQUE(type, section, date)', 'UNIQUE(type, section, date, store_id)')
        _safe(_add_store_to_unique, 'temp_sessions',
              'UNIQUE(type, date)', 'UNIQUE(type, date, store_id)')
        _safe(_add_store_to_unique, 'pastry_delivery',
              'UNIQUE(item_id, date)', 'UNIQUE(item_id, date, store_id)')
        _safe(_add_store_to_unique, 'pastry_sales',
              'UNIQUE(item_id, date)', 'UNIQUE(item_id, date, store_id)')
        _safe(_rebuild_prep_schedules_per_store)
        # 3b) names that were globally UNIQUE become unique PER STORE.
        _safe(_rebuild_col_unique_per_store, 'staff_members', 'name')
        _safe(_rebuild_col_unique_per_store, 'staff_birthdays', 'staff_name')
        # 3c) one reward decision per (month, award_type) PER STORE.
        _safe(_add_store_to_unique, 'monthly_reward_decisions',
              'UNIQUE(reward_month, award_type)', 'UNIQUE(reward_month, award_type, store_id)')
        # 3d) org-chart settings (structure_meta) are keyed per store.
        _safe(_rebuild_structure_meta_per_store)
        # 3e) Branch-specific compliance templates.
        _safe(_rebuild_checklist_templates_per_store)
        _safe(_rebuild_temp_templates_per_store)
        # 4) index store_id on every scoped table (after any rebuild above)
        for tbl in STORE_SCOPED_TABLES:
            try:
                conn.execute(f'CREATE INDEX IF NOT EXISTS idx_{tbl}_store_id ON {tbl}(store_id)')
            except Exception:
                pass
        for tbl in ('checklist_task_templates', 'temp_food_templates'):
            try:
                conn.execute(f'CREATE INDEX IF NOT EXISTS idx_{tbl}_store_id ON {tbl}(store_id)')
            except Exception:
                pass
    finally:
        conn.close()


def _rebuild_col_unique_per_store(conn, table, col):
    """A column with inline `<col> TEXT NOT NULL UNIQUE` should be unique PER
    STORE, not globally (so two branches can share a name). Rebuild the table so
    uniqueness becomes UNIQUE(<col>, store_id). Autocommit conn + FK off.
    Idempotent: skips if already migrated or the expected shape isn't present."""
    row = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone()
    if not row or not row[0]:
        return
    sql = row[0]
    norm = sql.replace(' ', '')
    if f'UNIQUE({col},store_id)' in norm:
        return  # already migrated
    if f'{col}TEXTNOTNULLUNIQUE' not in norm:
        return  # unexpected shape — leave data untouched
    import re
    new_inner = re.sub(rf'({col}\s+TEXT\s+NOT\s+NULL)\s+UNIQUE', r'\1', sql, count=1, flags=re.I)
    idx = new_inner.rstrip().rfind(')')
    new_inner = new_inner[:idx] + f', UNIQUE({col}, store_id)' + new_inner[idx:]
    tmp = f'{table}__mig'
    new_create = new_inner.replace(table, tmp, 1)
    if tmp not in new_create:
        return
    cols = [r[1] for r in conn.execute(f'PRAGMA table_info({table})')]
    collist = ', '.join(cols)
    conn.execute('PRAGMA foreign_keys = OFF')
    conn.execute('BEGIN')
    try:
        conn.execute(f'DROP TABLE IF EXISTS {tmp}')
        conn.execute(new_create)
        conn.execute(f'INSERT INTO {tmp} ({collist}) SELECT {collist} FROM {table}')
        conn.execute(f'DROP TABLE {table}')
        conn.execute(f'ALTER TABLE {tmp} RENAME TO {table}')
        conn.execute(f'CREATE INDEX IF NOT EXISTS idx_{table}_store_id ON {table}(store_id)')
        conn.execute('COMMIT')
    except Exception:
        conn.execute('ROLLBACK')
        raise
    finally:
        conn.execute('PRAGMA foreign_keys = ON')


def init_db():
    with get_db() as conn:
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS checklist_sessions (
                id                 INTEGER PRIMARY KEY AUTOINCREMENT,
                type               TEXT NOT NULL,
                section            TEXT NOT NULL,
                date               TEXT NOT NULL,
                day_of_week        TEXT,
                responsible        TEXT,
                submitted_by       TEXT,
                submitted_at       TEXT DEFAULT (datetime('now','localtime')),
                verified           INTEGER DEFAULT 0,
                verified_by        TEXT,
                verified_at        TEXT,
                overall_result     TEXT,
                issues_found       TEXT,
                action_responsible TEXT,
                manager_notes      TEXT,
                UNIQUE(type, section, date)
            );
            CREATE TABLE IF NOT EXISTS checklist_tasks (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id  INTEGER NOT NULL REFERENCES checklist_sessions(id) ON DELETE CASCADE,
                task_order  INTEGER NOT NULL,
                task_name   TEXT NOT NULL,
                done        INTEGER DEFAULT 0,
                done_by     TEXT,
                note        TEXT
            );
            CREATE TABLE IF NOT EXISTS temp_sessions (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                type         TEXT NOT NULL,
                date         TEXT NOT NULL,
                recorded_by  TEXT,
                checked_by   TEXT,
                submitted_at TEXT DEFAULT (datetime('now','localtime')),
                notes        TEXT,
                UNIQUE(type, date)
            );
            CREATE TABLE IF NOT EXISTS temp_readings (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id  INTEGER NOT NULL REFERENCES temp_sessions(id) ON DELETE CASCADE,
                food_order  INTEGER NOT NULL,
                food_name   TEXT NOT NULL,
                food_kind   TEXT NOT NULL DEFAULT 'cold',
                c1_time TEXT, c1_temp REAL,
                c2_time TEXT, c2_temp REAL,
                c3_time TEXT, c3_temp REAL,
                c4_time TEXT, c4_temp REAL,
                c5_time TEXT, c5_temp REAL,
                discarded   TEXT DEFAULT 'N'
            );
            CREATE TABLE IF NOT EXISTS audit_log (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                action      TEXT NOT NULL,
                record_type TEXT,
                record_id   INTEGER,
                user_name   TEXT,
                timestamp   TEXT DEFAULT (datetime('now','localtime')),
                details     TEXT
            );
            CREATE TABLE IF NOT EXISTS staff_members (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                name       TEXT NOT NULL UNIQUE,
                role       TEXT,
                active     INTEGER DEFAULT 1,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS staff_certificates (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_name    TEXT NOT NULL,
                cert_type     TEXT DEFAULT 'Food Safety Certificate',
                filename      TEXT NOT NULL,
                original_name TEXT,
                expiry_date   TEXT,
                notes         TEXT,
                uploaded_by   TEXT,
                uploaded_at   TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS checklist_photos (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id    INTEGER NOT NULL REFERENCES checklist_sessions(id) ON DELETE CASCADE,
                filename      TEXT NOT NULL,
                original_name TEXT,
                photo_number  INTEGER DEFAULT 0,
                file_size     INTEGER,
                uploaded_at   TEXT DEFAULT (datetime('now','localtime')),
                uploaded_by   TEXT
            );
            CREATE TABLE IF NOT EXISTS issue_reports (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                category     TEXT NOT NULL,
                title        TEXT NOT NULL,
                description  TEXT NOT NULL,
                reported_by  TEXT NOT NULL,
                branch       TEXT,
                date         TEXT NOT NULL,
                submitted_at TEXT DEFAULT (datetime('now','localtime')),
                priority     TEXT DEFAULT 'normal',
                status       TEXT DEFAULT 'open',
                admin_notes  TEXT,
                resolved_by  TEXT,
                resolved_at  TEXT
            );
            CREATE TABLE IF NOT EXISTS violation_rules (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                code           TEXT NOT NULL UNIQUE,
                title          TEXT NOT NULL,
                category       TEXT,
                severity       TEXT DEFAULT 'minor',
                description    TEXT,
                default_action TEXT,
                color          TEXT DEFAULT '#607D8B',
                icon           TEXT DEFAULT 'fa-triangle-exclamation',
                active         INTEGER DEFAULT 1,
                sort_order     INTEGER DEFAULT 0,
                created_at     TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS staff_violations (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                rule_id        INTEGER REFERENCES violation_rules(id) ON DELETE SET NULL,
                staff_name     TEXT NOT NULL,
                incident_date  TEXT NOT NULL,
                incident_time  TEXT NOT NULL,
                submitted_at   TEXT DEFAULT (datetime('now','localtime')),
                submitted_by   TEXT,
                branch         TEXT,
                severity       TEXT DEFAULT 'minor',
                description    TEXT NOT NULL,
                action_taken   TEXT,
                follow_up_date TEXT,
                status         TEXT DEFAULT 'open',
                resolved_by    TEXT,
                resolved_at    TEXT,
                manager_notes  TEXT
            );
            CREATE TABLE IF NOT EXISTS monthly_reward_adjustments (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                reward_month TEXT NOT NULL,
                staff_name   TEXT NOT NULL,
                points       REAL DEFAULT 0,
                reason       TEXT NOT NULL,
                created_by   TEXT,
                created_at   TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS monthly_reward_decisions (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                reward_month  TEXT NOT NULL,
                award_type    TEXT NOT NULL,
                staff_name    TEXT NOT NULL,
                reward_amount REAL DEFAULT 0,
                status        TEXT DEFAULT 'approved',
                notes         TEXT,
                approved_by   TEXT,
                approved_at   TEXT DEFAULT (datetime('now','localtime')),
                UNIQUE(reward_month, award_type)
            );
            CREATE TABLE IF NOT EXISTS salary_raise_reviews (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_name      TEXT NOT NULL,
                review_month    TEXT NOT NULL,
                current_rate    REAL DEFAULT 0,
                proposed_rate   REAL DEFAULT 0,
                effective_date  TEXT,
                status          TEXT DEFAULT 'draft',
                requested_by    TEXT,
                reviewed_by     TEXT,
                manager_notes   TEXT,
                decision_reason TEXT,
                created_at      TEXT DEFAULT (datetime('now','localtime')),
                decided_at      TEXT
            );
            CREATE TABLE IF NOT EXISTS staff_birthdays (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_name      TEXT NOT NULL UNIQUE,
                birthday        TEXT NOT NULL,
                favorite_gift   TEXT,
                gift_status     TEXT DEFAULT 'planned',
                last_given_year INTEGER,
                notes           TEXT,
                updated_at      TEXT DEFAULT (datetime('now','localtime'))
            );
        ''')
        # Checklist task name overrides (admin can rename tasks)
        conn.execute('''CREATE TABLE IF NOT EXISTS checklist_task_templates (
            chk_type   TEXT NOT NULL,
            section    TEXT NOT NULL,
            task_order INTEGER NOT NULL,
            task_name  TEXT NOT NULL,
            store_id   INTEGER NOT NULL DEFAULT 1,
            PRIMARY KEY (chk_type, section, task_order, store_id)
        )''')
        # Temperature food item templates (admin can rename/add items)
        conn.execute('''CREATE TABLE IF NOT EXISTS temp_food_templates (
            temp_type  TEXT NOT NULL,
            food_order INTEGER NOT NULL,
            food_name  TEXT NOT NULL,
            food_kind  TEXT NOT NULL DEFAULT 'cold',
            store_id   INTEGER NOT NULL DEFAULT 1,
            PRIMARY KEY (temp_type, food_order, store_id)
        )''')
        # Migrate: add food_kind for older installs that don't have it yet.
        try:
            conn.execute("ALTER TABLE temp_food_templates ADD COLUMN food_kind TEXT NOT NULL DEFAULT 'cold'")
        except sqlite3.OperationalError:
            pass
        try:
            conn.execute("ALTER TABLE checklist_task_templates ADD COLUMN store_id INTEGER NOT NULL DEFAULT 1")
        except sqlite3.OperationalError:
            pass
        try:
            conn.execute("ALTER TABLE temp_food_templates ADD COLUMN store_id INTEGER NOT NULL DEFAULT 1")
        except sqlite3.OperationalError:
            pass
        # Keep the holding type on each reading so historical records retain
        # the correct HOT / COLD / ROOM label even if a template changes later.
        try:
            conn.execute("ALTER TABLE temp_readings ADD COLUMN food_kind TEXT NOT NULL DEFAULT 'cold'")
        except sqlite3.OperationalError:
            pass
        # Migrate: per-food notes column on the reading row.
        try:
            conn.execute("ALTER TABLE temp_readings ADD COLUMN notes TEXT DEFAULT ''")
        except sqlite3.OperationalError:
            pass

        # Migrate: 'defrosting' flag for cold items being thawed (high temp is
        # then expected, so it is not treated as an out-of-zone alert).
        try:
            conn.execute("ALTER TABLE temp_readings ADD COLUMN defrosted TEXT DEFAULT 'N'")
        except sqlite3.OperationalError:
            pass
        # Seed from CHECKLISTS if empty
        if conn.execute('SELECT COUNT(*) as c FROM checklist_task_templates').fetchone()['c'] == 0:
            for chk_type, chk_data in CHECKLISTS.items():
                for section in ('opening', 'closing'):
                    for i, name in enumerate(chk_data.get(section, [])):
                        conn.execute('''INSERT OR IGNORE INTO checklist_task_templates
                            (chk_type, section, task_order, task_name, store_id) VALUES (?,?,?,?,1)''',
                            (chk_type, section, i, name))

        # Seed temperature food templates per type if missing.
        for temp_type, temp_data in TEMPERATURES.items():
            existing_foods = conn.execute(
                'SELECT COUNT(*) as c FROM temp_food_templates WHERE temp_type=?',
                (temp_type,)).fetchone()['c']
            if existing_foods == 0:
                for i, food in enumerate(temp_data.get('foods', [])):
                    conn.execute('''INSERT OR IGNORE INTO temp_food_templates
                        (temp_type, food_order, food_name, food_kind, store_id) VALUES (?,?,?,?,1)''',
                        (temp_type, i, food['name'], food['kind']))

        # Migration: classify existing food rows by kind. Runs once
        # (uses a marker row in audit_log to avoid re-running every boot).
        try:
            mig_marker = '_mig_food_kind_v1'
            already = conn.execute(
                "SELECT 1 FROM audit_log WHERE action=? LIMIT 1", (mig_marker,)
            ).fetchone()
            if not already:
                for r in conn.execute(
                    'SELECT rowid, temp_type, food_name FROM temp_food_templates').fetchall():
                    kind = temp_food_kind_default(r['temp_type'], r['food_name'])
                    conn.execute(
                        'UPDATE temp_food_templates SET food_kind=? WHERE rowid=?',
                        (kind, r['rowid']))
                # Ensure 'Beef Brisket' exists on banh_mi (added in this update).
                exists = conn.execute(
                    "SELECT 1 FROM temp_food_templates "
                    "WHERE temp_type='banh_mi' AND food_name='Beef Brisket'").fetchone()
                if not exists:
                    next_order = conn.execute(
                        "SELECT COALESCE(MAX(food_order),-1)+1 AS n "
                        "FROM temp_food_templates WHERE temp_type='banh_mi'"
                    ).fetchone()['n']
                    conn.execute(
                        "INSERT OR IGNORE INTO temp_food_templates "
                        "(temp_type,food_order,food_name,food_kind) "
                        "VALUES ('banh_mi',?,'Beef Brisket','hot')", (next_order,))
                conn.execute(
                    "INSERT INTO audit_log (action,record_type,user_name,details) "
                    "VALUES (?,?,?,?)",
                    (mig_marker, 'migration', 'system',
                     'Classified temp foods as hot/cold and ensured Beef Brisket on banh_mi.'))
        except sqlite3.OperationalError:
            pass

        # Migration v2: re-classify chef foods now that some items are 'room'
        # (Pho Noodle / Rice Vermicelli Bun / Lettuce) or 'hot' (Pork Chop / Egg).
        try:
            mig2 = '_mig_food_kind_v2'
            done = conn.execute(
                "SELECT 1 FROM audit_log WHERE action=? LIMIT 1", (mig2,)).fetchone()
            if not done:
                for r in conn.execute(
                    'SELECT rowid, temp_type, food_name FROM temp_food_templates').fetchall():
                    kind = temp_food_kind_default(r['temp_type'], r['food_name'])
                    conn.execute('UPDATE temp_food_templates SET food_kind=? WHERE rowid=?',
                                 (kind, r['rowid']))
                conn.execute(
                    "INSERT INTO audit_log (action,record_type,user_name,details) "
                    "VALUES (?,?,?,?)",
                    (mig2, 'migration', 'system',
                     'Re-synced chef food kinds: added room-temp and hot items.'))
        except sqlite3.OperationalError:
            pass

        # Backfill holding types on historical readings after the live food
        # templates have been classified. Prefer an exact template match;
        # fall back to the seeded classification for renamed/deleted rows.
        try:
            mig3 = '_mig_temp_reading_food_kind_v1'
            done = conn.execute(
                "SELECT 1 FROM audit_log WHERE action=? LIMIT 1", (mig3,)).fetchone()
            if not done:
                old_readings = conn.execute('''
                    SELECT tr.id, tr.food_order, tr.food_name,
                           ts.type AS temp_type,
                           tft.food_name AS template_name,
                           tft.food_kind AS template_kind
                    FROM temp_readings tr
                    JOIN temp_sessions ts ON ts.id=tr.session_id
                    LEFT JOIN temp_food_templates tft
                      ON tft.temp_type=ts.type AND tft.food_order=tr.food_order
                ''').fetchall()
                for r in old_readings:
                    template_matches = (
                        (r['template_name'] or '').strip().lower()
                        == (r['food_name'] or '').strip().lower()
                    )
                    kind = r['template_kind'] if template_matches else None
                    if kind not in TEMP_KIND_RULES:
                        kind = temp_food_kind_default(r['temp_type'], r['food_name'])
                    conn.execute('UPDATE temp_readings SET food_kind=? WHERE id=?',
                                 (kind, r['id']))
                conn.execute(
                    "INSERT INTO audit_log (action,record_type,user_name,details) "
                    "VALUES (?,?,?,?)",
                    (mig3, 'migration', 'system',
                     'Backfilled HOT/COLD/ROOM labels on historical temperature readings.'))
        except sqlite3.OperationalError:
            pass

        # Migrate: add new columns if they don't exist
        for col_sql in [
            'ALTER TABLE checklist_sessions ADD COLUMN general_note TEXT',
            'ALTER TABLE checklist_sessions ADD COLUMN general_done_by TEXT',
            'ALTER TABLE checklist_sessions ADD COLUMN manager_submit TEXT',
            'ALTER TABLE checklist_sessions ADD COLUMN is_late INTEGER DEFAULT 0',
            'ALTER TABLE issue_reports ADD COLUMN photo TEXT',
            'ALTER TABLE staff_members ADD COLUMN phone TEXT',
            'ALTER TABLE staff_members ADD COLUMN email TEXT',
            'ALTER TABLE staff_members ADD COLUMN emergency_contact TEXT',
            'ALTER TABLE staff_members ADD COLUMN staff_notes TEXT',
            "ALTER TABLE staff_violations ADD COLUMN warning_step TEXT DEFAULT 'Verbal Discussion'",
        ]:
            try:
                conn.execute(col_sql)
            except Exception:
                pass

        # Normalize legacy violation statuses → open / closed / cancelled.
        try:
            conn.execute("UPDATE staff_violations SET status='closed' WHERE status='resolved'")
            conn.execute("UPDATE staff_violations SET status='cancelled' WHERE status='void'")
            conn.execute("UPDATE staff_violations SET status='open' "
                         "WHERE status IN ('in_progress','counseled') OR status IS NULL OR status=''")
        except Exception:
            pass

        # Seed staff if empty
        for name in STAFF:
            conn.execute('INSERT OR IGNORE INTO staff_members (name) VALUES (?)', (name,))

        # Seed violation rule catalog.
        for i, rule in enumerate(VIOLATION_RULES_SEED):
            conn.execute('''INSERT OR IGNORE INTO violation_rules
                (code,title,category,severity,description,default_action,color,icon,sort_order)
                VALUES (?,?,?,?,?,?,?,?,?)''',
                (rule['code'], rule['title'], rule['category'], rule['severity'],
                 rule['description'], rule['default_action'], rule['color'], rule['icon'], i))

        # ── Indexes for hot query paths ─────────────────────────────────────
        # Adding these makes /analytics, /manager, /history scale to 100k+ rows
        # without slowdown. Safe to run on every boot — CREATE INDEX IF NOT EXISTS
        # is idempotent and only does work the first time.
        for ddl in [
            'CREATE INDEX IF NOT EXISTS idx_chk_date           ON checklist_sessions(date)',
            'CREATE INDEX IF NOT EXISTS idx_chk_type_section_date ON checklist_sessions(type, section, date)',
            'CREATE INDEX IF NOT EXISTS idx_chk_submitted_by   ON checklist_sessions(submitted_by)',
            'CREATE INDEX IF NOT EXISTS idx_chk_responsible    ON checklist_sessions(responsible)',
            'CREATE INDEX IF NOT EXISTS idx_chktasks_session   ON checklist_tasks(session_id)',
            'CREATE UNIQUE INDEX IF NOT EXISTS idx_chktasks_session_order ON checklist_tasks(session_id, task_order)',
            'CREATE INDEX IF NOT EXISTS idx_chkphotos_session  ON checklist_photos(session_id)',
            'CREATE INDEX IF NOT EXISTS idx_temp_date          ON temp_sessions(date)',
            'CREATE INDEX IF NOT EXISTS idx_temp_type_date     ON temp_sessions(type, date)',
            'CREATE INDEX IF NOT EXISTS idx_tempreadings_sess  ON temp_readings(session_id)',
            'CREATE UNIQUE INDEX IF NOT EXISTS idx_tempreadings_session_order ON temp_readings(session_id, food_order)',
            'CREATE INDEX IF NOT EXISTS idx_issue_date         ON issue_reports(date)',
            'CREATE INDEX IF NOT EXISTS idx_issue_status       ON issue_reports(status)',
            'CREATE INDEX IF NOT EXISTS idx_viol_date          ON staff_violations(incident_date)',
            'CREATE INDEX IF NOT EXISTS idx_viol_staff         ON staff_violations(staff_name)',
            'CREATE INDEX IF NOT EXISTS idx_viol_status        ON staff_violations(status)',
            'CREATE INDEX IF NOT EXISTS idx_audit_ts           ON audit_log(timestamp)',
        ]:
            try:
                conn.execute(ddl)
            except sqlite3.OperationalError:
                pass

        # Seed temporary birthday data for staff so the giveaway workflow is usable now.
        gift_ideas = [
            'MCQ meal voucher',
            'Coffee and pastry pack',
            'Birthday cake',
            'Gift card',
            'Team lunch shout',
            'Dessert box',
        ]
        active_staff = conn.execute('SELECT name FROM staff_members ORDER BY name').fetchall()
        for idx, staff_row in enumerate(active_staff):
            name = staff_row['name']
            if not conn.execute('SELECT 1 FROM staff_birthdays WHERE staff_name=?', (name,)).fetchone():
                seed = sum(ord(ch) for ch in name) + idx * 17
                month = seed % 12 + 1
                day = (seed * 7) % 28 + 1
                conn.execute('''INSERT INTO staff_birthdays
                    (staff_name,birthday,favorite_gift,gift_status,notes)
                    VALUES (?,?,?,?,?)''',
                    (name, f'2000-{month:02d}-{day:02d}',
                     gift_ideas[seed % len(gift_ideas)], 'planned',
                     'Temporary birthday data - update when the real birthday is confirmed.'))

def save_uploaded_photo(file_storage, dest_path, max_dim=1280, quality=82):
    """Save an uploaded image, downscaling + re-encoding as JPEG when oversized.

    Reduces phone photos (typically 3-5 MB, 4000×3000) down to ~150-300 KB
    while preserving enough detail for health-inspection evidence. Falls back
    to a plain file save if Pillow is unavailable or the image is corrupt.

    Returns the FINAL filename written (may differ from dest_path if extension changed).
    """
    try:
        from PIL import Image, ImageOps   # type: ignore
    except Exception:
        file_storage.save(dest_path)
        return os.path.basename(dest_path)

    try:
        file_storage.stream.seek(0)
    except Exception:
        pass

    img = None
    work = None
    try:
        img = Image.open(file_storage.stream)
        if img.format == 'JPEG' and max(img.size) <= max_dim:
            base, _ = os.path.splitext(dest_path)
            final_path = base + '.jpg'
            try:
                file_storage.stream.seek(0)
            except Exception:
                pass
            file_storage.save(final_path)
            return os.path.basename(final_path)
        if img.format == 'JPEG':
            try:
                img.draft('RGB', (max_dim, max_dim))
            except Exception:
                pass
        work = ImageOps.exif_transpose(img)   # respect phone rotation metadata
        if work.mode in ('RGBA', 'LA', 'P'):
            conv = work.convert('RGB')
            if work is not img:
                work.close()
            work = conv
        w, h = work.size
        if max(w, h) > max_dim:
            ratio = max_dim / float(max(w, h))
            resized = work.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
            if work is not img:
                work.close()
            work = resized
        # Always save resized photos as JPEG. Avoid optimize/progressive here:
        # those save a little space but make checklist submit noticeably slower
        # when staff upload several phone photos at once.
        base, _ = os.path.splitext(dest_path)
        final_path = base + '.jpg'
        work.save(final_path, 'JPEG', quality=quality)
        return os.path.basename(final_path)
    except Exception:
        # If decoding failed (corrupt / unsupported), fall back to original bytes.
        try:
            file_storage.stream.seek(0)
        except Exception:
            pass
        file_storage.save(dest_path)
        return os.path.basename(dest_path)
    finally:
        # Release decoded-image memory immediately so a batch of phone photos
        # (each 4000×3000) doesn't pile up and OOM-kill the worker.
        for im in (work, img):
            try:
                if im is not None:
                    im.close()
            except Exception:
                pass


def log_action(action, record_type, record_id, user_name, details=''):
    try:
        with get_db() as conn:
            log_action_conn(conn, action, record_type, record_id, user_name, details)
    except Exception:
        pass

def log_action_conn(conn, action, record_type, record_id, user_name, details=''):
    try:
        conn.execute(
            'INSERT INTO audit_log (action,record_type,record_id,user_name,details) VALUES (?,?,?,?,?)',
            (action, record_type, record_id, user_name, details)
        )
    except Exception:
        pass

def get_temp_foods(conn, temp_type, store_id=None):
    """Return list of {name, kind} dicts in display order."""
    if store_id is None:
        store_id = current_store_id()
    rows = conn.execute('''
        SELECT food_name, food_kind FROM temp_food_templates
        WHERE temp_type=? AND store_id=?
        ORDER BY food_order
    ''', (temp_type, store_id)).fetchall()
    if rows:
        return [{'name': r['food_name'],
                 'kind': r['food_kind'] or temp_food_kind_default(temp_type, r['food_name'])}
                for r in rows]
    return [dict(f) for f in TEMPERATURES.get(temp_type, {}).get('foods', [])]


def get_temp_food_names(conn, temp_type):
    """Backward-compat helper for places that only need a flat list of names."""
    return [f['name'] for f in get_temp_foods(conn, temp_type)]


def get_temp_data_for_form(conn, temp_type, store_id=None):
    temp_data = dict(TEMPERATURES[temp_type])
    temp_data['foods'] = get_temp_foods(conn, temp_type, store_id=store_id)
    return temp_data

def month_bounds(month_value=None):
    if not month_value:
        month_value = date.today().strftime('%Y-%m')
    try:
        year, month = [int(x) for x in month_value.split('-', 1)]
        start = date(year, month, 1)
    except Exception:
        start = date.today().replace(day=1)
        month_value = start.strftime('%Y-%m')
    end = date(start.year, start.month, calendar.monthrange(start.year, start.month)[1])
    return month_value, start.isoformat(), end.isoformat(), start.strftime('%B %Y')

def shift_month(month_value, offset):
    month_value, _, _, _ = month_bounds(month_value)
    year, month = [int(x) for x in month_value.split('-', 1)]
    month += offset
    while month < 1:
        month += 12
        year -= 1
    while month > 12:
        month -= 12
        year += 1
    return f'{year:04d}-{month:02d}'

def staff_matches(value, staff_lookup):
    if not value:
        return []
    raw = str(value).strip()
    if not raw:
        return []
    lowered = raw.lower()
    if lowered in staff_lookup:
        return [staff_lookup[lowered]]
    parts = [p.strip().lower() for p in raw.replace('&', ',').replace('/', ',').split(',')]
    matches = [staff_lookup[p] for p in parts if p in staff_lookup]
    if matches:
        return list(dict.fromkeys(matches))
    contained = [name for key, name in staff_lookup.items() if key and key in lowered]
    return list(dict.fromkeys(contained))

def base_staff_score(name, role=''):
    return {
        'staff_name': name,
        'role': role or '',
        'employee_score': 0.0,
        'checklist_score': 0.0,
        'checklist_points': 0.0,
        'temperature_points': 0.0,
        'report_points': 0.0,
        'adjustment_points': 0.0,
        'violation_penalty': 0.0,
        'checklist_sessions': 0,
        'checklist_done': 0,
        'checklist_total': 0,
        'checklist_late': 0,
        'checklist_verified': 0,
        'checklist_clean': 0,
        'checklist_issues': 0,
        'temp_sessions': 0,
        'issue_reports': 0,
        'resolved_reports': 0,
        'violations_total': 0,
        'minor_violations': 0,
        'moderate_violations': 0,
        'serious_violations': 0,
        'critical_violations': 0,
        'completion_rate': 0.0,
        'on_time_rate': 0.0,
        'verified_rate': 0.0,
        'clean_rate': 0.0,
        'recommendation': 'Needs more data',
    }

def calculate_monthly_reward_scores(conn, month_value=None, store_id=None):
    month_value, start_date, end_date, month_label = month_bounds(month_value)
    # Scope all scoring to one store (None = all stores, super_admin).
    sc = '' if store_id is None else ' AND store_id=?'
    sp = [] if store_id is None else [store_id]
    staff_rows = [dict(r) for r in conn.execute(
        f'SELECT name, role FROM staff_members WHERE active=1{sc} ORDER BY name',
        sp).fetchall()]
    staff_lookup = {r['name'].strip().lower(): r['name'] for r in staff_rows}
    scores = {r['name']: base_staff_score(r['name'], r.get('role') or '') for r in staff_rows}

    checklist_rows = conn.execute(f'''
        SELECT cs.*,
               (SELECT COUNT(*) FROM checklist_tasks WHERE session_id=cs.id AND done=1) as done_count,
               (SELECT COUNT(*) FROM checklist_tasks WHERE session_id=cs.id) as total_count
        FROM checklist_sessions cs
        WHERE cs.date BETWEEN ? AND ?{(' AND cs.store_id=?' if store_id is not None else '')}
    ''', [start_date, end_date] + sp).fetchall()
    for row in checklist_rows:
        d = dict(row)
        names = set()
        for field in ('responsible', 'submitted_by', 'general_done_by', 'manager_submit'):
            names.update(staff_matches(d.get(field), staff_lookup))
        if not names:
            continue
        done_count = int(d.get('done_count') or 0)
        total_count = int(d.get('total_count') or 0)
        completion = (done_count / total_count) if total_count else 0
        for name in names:
            s = scores[name]
            s['checklist_sessions'] += 1
            s['checklist_done'] += done_count
            s['checklist_total'] += total_count
            if d.get('verified'):
                s['checklist_verified'] += 1
            if d.get('is_late'):
                s['checklist_late'] += 1
            if d.get('overall_result') == 'issues_found':
                s['checklist_issues'] += 1
            # Only verified sessions earn/lose points
            if d.get('verified'):
                if d.get('overall_result') == 'issues_found':
                    points = -1   # verified but has issues
                else:
                    points = 2    # verified, clean
                if d.get('is_late'):
                    points -= 1   # late deduction
            else:
                points = 0        # not verified = no points
            s['checklist_points'] += round(points, 2)

    temp_rows = conn.execute(f'''
        SELECT * FROM temp_sessions
        WHERE date BETWEEN ? AND ?{sc}
    ''', [start_date, end_date] + sp).fetchall()
    for row in temp_rows:
        d = dict(row)
        names = set()
        for field in ('recorded_by', 'checked_by'):
            names.update(staff_matches(d.get(field), staff_lookup))
        for name in names:
            scores[name]['temp_sessions'] += 1
            scores[name]['temperature_points'] += 5

    report_rows = conn.execute(f'''
        SELECT * FROM issue_reports
        WHERE date BETWEEN ? AND ?{sc}
    ''', [start_date, end_date] + sp).fetchall()
    for row in report_rows:
        d = dict(row)
        names = staff_matches(d.get('reported_by'), staff_lookup)
        for name in names:
            scores[name]['issue_reports'] += 1
            if d.get('status') == 'resolved':
                scores[name]['resolved_reports'] += 1
                scores[name]['report_points'] += 5
            else:
                scores[name]['report_points'] += 2

    penalty_map = {'minor': 5, 'moderate': 12, 'serious': 25, 'critical': 40}
    violation_rows = conn.execute(f'''
        SELECT * FROM staff_violations
        WHERE incident_date BETWEEN ? AND ?{sc}
    ''', [start_date, end_date] + sp).fetchall()
    for row in violation_rows:
        d = dict(row)
        name = staff_lookup.get((d.get('staff_name') or '').strip().lower())
        if not name:
            continue
        severity = d.get('severity') or 'minor'
        s = scores[name]
        s['violations_total'] += 1
        if severity in ('minor', 'moderate', 'serious', 'critical'):
            s[f'{severity}_violations'] += 1
        s['violation_penalty'] += penalty_map.get(severity, 8)

    adjustment_rows = conn.execute('''
        SELECT * FROM monthly_reward_adjustments
        WHERE reward_month=?
    ''', (month_value,)).fetchall()
    for row in adjustment_rows:
        d = dict(row)
        name = staff_lookup.get((d.get('staff_name') or '').strip().lower())
        if name:
            scores[name]['adjustment_points'] += float(d.get('points') or 0)

    for s in scores.values():
        if s['checklist_total']:
            s['completion_rate'] = round(s['checklist_done'] / s['checklist_total'] * 100, 1)
        if s['checklist_sessions']:
            s['on_time_rate'] = round((s['checklist_sessions'] - s['checklist_late']) / s['checklist_sessions'] * 100, 1)
            s['verified_rate'] = round(s['checklist_verified'] / s['checklist_sessions'] * 100, 1)
            s['clean_rate'] = round((s['checklist_sessions'] - s['checklist_issues']) / s['checklist_sessions'] * 100, 1)
            s['checklist_score'] = round(
                s['completion_rate'] * 0.45 +
                s['on_time_rate'] * 0.25 +
                s['verified_rate'] * 0.20 +
                s['clean_rate'] * 0.10 -
                min(20, s['violation_penalty'] * 0.25), 1)
            s['checklist_score'] = max(0, min(100, s['checklist_score']))
        s['employee_score'] = round(
            s['checklist_points'] + s['temperature_points'] + s['report_points'] +
            s['adjustment_points'] - s['violation_penalty'], 1)
        if s['employee_score'] >= 30 and s['violations_total'] == 0:
            s['recommendation'] = 'Excellent'
        elif s['employee_score'] >= 15:
            s['recommendation'] = 'Strong'
        elif s['employee_score'] >= 5:
            s['recommendation'] = 'Developing'
        else:
            s['recommendation'] = 'Needs support'

    employee_rank = sorted(scores.values(), key=lambda x: (x['employee_score'], x['checklist_score'], -x['violations_total']), reverse=True)
    checklist_rank = sorted(scores.values(), key=lambda x: (x['checklist_score'], x['checklist_sessions'], -x['violations_total']), reverse=True)
    decisions = [dict(r) for r in conn.execute('''
        SELECT * FROM monthly_reward_decisions
        WHERE reward_month=?
        ORDER BY award_type
    ''', (month_value,)).fetchall()]
    adjustments = [dict(r) for r in conn.execute('''
        SELECT * FROM monthly_reward_adjustments
        WHERE reward_month=?
        ORDER BY created_at DESC, id DESC
    ''', (month_value,)).fetchall()]
    return {
        'month': month_value,
        'start_date': start_date,
        'end_date': end_date,
        'month_label': month_label,
        'scores': list(scores.values()),
        'employee_rank': employee_rank,
        'checklist_rank': checklist_rank,
        'decisions': decisions,
        'adjustments': adjustments,
    }

def raise_review_snapshot(conn, staff_name, month_value, store_id=None):
    sc = '' if store_id is None else ' AND store_id=?'
    sp = [] if store_id is None else [store_id]
    current = calculate_monthly_reward_scores(conn, month_value, store_id=store_id)
    current_score = next((s for s in current['scores'] if s['staff_name'] == staff_name), base_staff_score(staff_name))
    months = [shift_month(month_value, -i) for i in range(3)]
    monthly = []
    for m in months:
        data = calculate_monthly_reward_scores(conn, m, store_id=store_id)
        row = next((s for s in data['scores'] if s['staff_name'] == staff_name), base_staff_score(staff_name))
        monthly.append({
            'month': m,
            'employee_score': row['employee_score'],
            'checklist_score': row['checklist_score'],
            'violations_total': row['violations_total'],
        })
    avg_3m = round(sum(r['employee_score'] for r in monthly) / len(monthly), 1) if monthly else 0
    today_iso = date.today().isoformat()
    start_90 = (date.today() - timedelta(days=90)).isoformat()
    violations = [dict(r) for r in conn.execute(f'''
        SELECT sv.*, vr.title as rule_title, vr.category as rule_category, vr.color as rule_color
        FROM staff_violations sv
        LEFT JOIN violation_rules vr ON vr.id=sv.rule_id
        WHERE sv.staff_name=? AND sv.incident_date BETWEEN ? AND ?{(' AND sv.store_id=?' if store_id is not None else '')}
        ORDER BY sv.incident_date DESC, sv.submitted_at DESC
    ''', [staff_name, start_90, today_iso] + sp).fetchall()]
    submitted_reports = [dict(r) for r in conn.execute(f'''
        SELECT * FROM issue_reports
        WHERE reported_by=? AND date BETWEEN ? AND ?{sc}
        ORDER BY date DESC, submitted_at DESC
    ''', [staff_name, start_90, today_iso] + sp).fetchall()]
    like_name = f'%{staff_name}%'
    related_reports = [dict(r) for r in conn.execute(f'''
        SELECT * FROM issue_reports
        WHERE (title LIKE ? OR description LIKE ? OR COALESCE(admin_notes,'') LIKE ?)
          AND date BETWEEN ? AND ?{sc}
        ORDER BY date DESC, submitted_at DESC
    ''', [like_name, like_name, like_name, start_90, today_iso] + sp).fetchall()]
    reward_history = [dict(r) for r in conn.execute('''
        SELECT * FROM monthly_reward_decisions
        WHERE staff_name=?
        ORDER BY reward_month DESC, award_type
        LIMIT 12
    ''', (staff_name,)).fetchall()]
    serious_recent = sum(1 for v in violations if v.get('severity') in ('serious', 'critical'))
    open_recent = sum(1 for v in violations if v.get('status') in ('open', 'in_progress'))
    if avg_3m >= 25 and serious_recent == 0 and open_recent == 0:
        recommendation = 'Recommended'
        recommendation_class = 'success'
    elif avg_3m >= 12 and serious_recent == 0:
        recommendation = 'Review with manager'
        recommendation_class = 'warning'
    else:
        recommendation = 'Not ready yet'
        recommendation_class = 'danger'
    return {
        'current': current_score,
        'monthly': monthly,
        'avg_3m': avg_3m,
        'violations': violations,
        'submitted_reports': submitted_reports,
        'related_reports': related_reports,
        'reward_history': reward_history,
        'recommendation': recommendation,
        'recommendation_class': recommendation_class,
        'serious_recent': serious_recent,
        'open_recent': open_recent,
        'start_90': start_90,
        'today': today_iso,
    }

def birthday_rows(conn, store_id=None):
    if store_id is None:
        store_id = current_store_id()
    today_obj = date.today()
    current_year = today_obj.year
    rows = [dict(r) for r in conn.execute('''
        SELECT sb.*, sm.role, sm.active
        FROM staff_birthdays sb
        LEFT JOIN staff_members sm ON sm.name=sb.staff_name AND sm.store_id=sb.store_id
        WHERE sb.store_id=?
        ORDER BY sb.staff_name
    ''', (store_id,)).fetchall()]
    enriched = []
    for r in rows:
        try:
            _, month, day = [int(x) for x in r['birthday'].split('-')]
            next_date = date(current_year, month, day)
        except Exception:
            month, day = 1, 1
            next_date = date(current_year, 1, 1)
        if next_date < today_obj:
            next_date = date(current_year + 1, month, day)
        r['next_birthday'] = next_date.isoformat()
        r['days_until'] = (next_date - today_obj).days
        r['birthday_label'] = next_date.strftime('%d %b')
        r['gift_due'] = r.get('last_given_year') != current_year
        enriched.append(r)
    return sorted(enriched, key=lambda x: (x['days_until'], x['staff_name']))

def register_pdf_fonts():
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except Exception:
        return 'Helvetica', 'Helvetica-Bold'

    regular_candidates = [
        '/System/Library/Fonts/Supplemental/Arial Unicode.ttf',
        '/System/Library/Fonts/Supplemental/Arial.ttf',
        '/Library/Fonts/Arial Unicode.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
    ]
    bold_candidates = [
        '/System/Library/Fonts/Supplemental/Arial Bold.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf',
    ]

    regular = next((p for p in regular_candidates if os.path.exists(p)), None)
    bold = next((p for p in bold_candidates if os.path.exists(p)), None)
    if not regular:
        return 'Helvetica', 'Helvetica-Bold'
    try:
        pdfmetrics.registerFont(TTFont('MCQSans', regular))
        if bold:
            pdfmetrics.registerFont(TTFont('MCQSans-Bold', bold))
            return 'MCQSans', 'MCQSans-Bold'
        return 'MCQSans', 'MCQSans'
    except Exception:
        return 'Helvetica', 'Helvetica-Bold'

# ─── Auth ──────────────────────────────────────────────────────────────────────

USER_PASSWORD        = '7777'
ADMIN_PASSWORD       = '77771'
KITCHEN_PASSWORD     = '8888'
SUPER_ADMIN_PASSWORD = '999999'   # cross-store owner view (all branches)
LOCATION         = 'mirrabooka'

def is_admin():
    return session.get('role') == 'admin'

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated


def database_save_guard(record_kind):
    """Return staff to the form on a transient SQLite save failure.

    The form draft stays in localStorage until a successful redirect includes
    `saved=1`, so retrying does not require re-entering the record.
    """
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            try:
                return f(*args, **kwargs)
            except sqlite3.Error:
                app.logger.exception('Database error while saving %s', record_kind)
                flash(
                    'The record could not be saved yet. Your draft is still on this device. '
                    'Please wait a moment and tap Save again.',
                    'danger',
                )
                if record_kind == 'checklist':
                    return redirect(url_for(
                        'checklist_form',
                        chk_type=kwargs.get('chk_type', ''),
                        date=request.form.get('date', date.today().isoformat()),
                        section=request.form.get('section', 'opening'),
                        save_error=1,
                    ), code=303)
                return redirect(url_for(
                    'temperature_form',
                    temp_type=kwargs.get('temp_type', ''),
                    date=request.form.get('date', date.today().isoformat()),
                    save_error=1,
                ), code=303)
        return decorated
    return decorator


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login_page'))
        if session.get('role') not in ('admin', 'super_admin'):
            return render_template('access_denied.html'), 403
        return f(*args, **kwargs)
    return decorated


def super_admin_required(f):
    """Routes that show/operate across all stores (cross-store reports)."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login_page'))
        if session.get('role') != 'super_admin':
            return render_template('access_denied.html'), 403
        return f(*args, **kwargs)
    return decorated

# Perth is UTC+8 year-round (no daylight saving), so a fixed offset is exact and
# works even when the server (PythonAnywhere) runs on UTC.
PERTH_TZ = timezone(timedelta(hours=8))
OPENING_LOCK_HOUR = 15   # 3 PM Perth: the Opening checklist locks for everyone


def perth_now():
    return datetime.now(PERTH_TZ)


def opening_locked(for_date=None):
    """Opening locks only for TODAY's record after 3 PM Perth (for everyone).

    Past (and future) dates stay open so staff/admin can review or correct earlier
    Opening checklists at any time. `for_date` is an ISO 'YYYY-MM-DD' string; when
    omitted it means today, so the dashboard/nav 'today' behaviour is unchanged.
    """
    now = perth_now()
    today = now.date().isoformat()
    return (for_date or today) == today and now.hour >= OPENING_LOCK_HOUR


@app.context_processor
def inject_globals():
    _today = date.today()
    _week_start = (_today - timedelta(days=_today.weekday())).isoformat()
    return dict(
        checklists=CHECKLISTS,
        temperatures=TEMPERATURES,
        today=_today.isoformat(),
        week_start_nav=_week_start,
        day_name=_today.strftime('%A'),
        current_ep=request.endpoint or '',
        is_admin=session.get('role') in ('admin', 'super_admin'),
        is_super_admin=session.get('role') == 'super_admin',
        user_role=session.get('role', ''),
        photos_required=PHOTOS_REQUIRED,
        branch=session.get('branch', ''),
        store_id=session.get('store_id', 1),
        store_code=session.get('store_code', ''),
        all_stores=(get_stores() if session.get('role') == 'super_admin' else []),
        selected_store=(request.args.get('store', 'all') if session.get('role') == 'super_admin' else None),
        issue_categories=ISSUE_CATEGORIES,
        temp_kind_rules=TEMP_KIND_RULES,
        temp_is_unsafe=temp_is_unsafe,
        opening_locked=opening_locked(),
        default_section=('closing' if opening_locked() else 'opening'),
    )

# Paths that should NOT trigger or be affected by the timeout middleware:
# static assets (otherwise a background image kept loading would keep the
# session alive forever), the login page itself, and the cron endpoint
# which carries its own token-based auth.
_TIMEOUT_EXEMPT_PREFIXES = ('/static/', '/login', '/logout', '/cron/')


@app.before_request
def _enforce_session_timeout():
    """Idle (30 min) + absolute (8h) session timeouts.
    Runs before every request; redirects to /login with a friendly notice
    when a logged-in session has expired."""
    if any(request.path.startswith(p) for p in _TIMEOUT_EXEMPT_PREFIXES):
        return None
    if not session.get('logged_in'):
        return None

    now = datetime.now()

    # 1) Absolute cap from login time (one shift max)
    login_iso = session.get('login_ts')
    if login_iso:
        try:
            if now - datetime.fromisoformat(login_iso) > SESSION_ABSOLUTE_TIMEOUT:
                session.clear()
                return redirect(url_for('login_page', timeout='session'))
        except ValueError:
            pass

    # 2) Idle cap (no activity for 30 min)
    last_iso = session.get('last_activity')
    if last_iso:
        try:
            if now - datetime.fromisoformat(last_iso) > SESSION_IDLE_TIMEOUT:
                session.clear()
                return redirect(url_for('login_page', timeout='idle'))
        except ValueError:
            pass

    # Bump activity timestamp, keep session marked permanent so the cookie
    # itself respects PERMANENT_SESSION_LIFETIME.
    session['last_activity'] = now.isoformat(timespec='seconds')
    session.permanent = True
    return None


@app.route('/login', methods=['GET', 'POST'])
def login_page():
    error = None
    stores = get_stores()   # active stores (dynamic — supports any number of branches)
    if request.method == 'POST':
        pw     = request.form.get('password', '')
        code   = request.form.get('branch', '').strip()   # store CODE
        mode   = request.form.get('mode', 'user').strip().lower()
        store  = next((s for s in stores if s['code'] == code), None)
        if not store:
            error = 'Please select a valid branch.'
        else:
            # Per-store passwords (fall back to chain defaults if a store hasn't
            # set its own). The owner/super_admin password is global.
            role = None
            if pw and pw == SUPER_ADMIN_PASSWORD:
                role = 'super_admin'
            elif mode == 'admin' and pw and pw == (store.get('admin_password') or ADMIN_PASSWORD):
                role = 'admin'
            elif mode == 'user' and pw and pw == (store.get('user_password') or USER_PASSWORD):
                role = 'user'
            elif pw and pw == (store.get('kitchen_password') or KITCHEN_PASSWORD):
                role = 'kitchen'
            elif pw and pw == (store.get('admin_password') or ADMIN_PASSWORD):
                role = 'admin'
            elif pw and pw == (store.get('user_password') or USER_PASSWORD):
                role = 'user'
            if role is None:
                error = 'Incorrect password. Please try again.'
            else:
                now = datetime.now()
                display = store['name'][4:] if store['name'].startswith('MCQ ') else store['name']
                session.clear()
                session.update({
                    'logged_in':     True,
                    'role':          role,
                    'branch':        display,         # short display name
                    'store_id':      store['id'],     # canonical key for all filtering
                    'store_code':    store['code'],
                    'login_time':    now.strftime('%Y-%m-%d %H:%M'),
                    'login_ts':      now.isoformat(timespec='seconds'),
                    'last_activity': now.isoformat(timespec='seconds'),
                })
                session.permanent = True
                if role == 'kitchen':
                    return redirect(url_for('orders.kitchen'))
                return redirect(url_for('dashboard'))

    # Friendly notice when redirected here by the timeout middleware
    notice = None
    tk = request.args.get('timeout')
    if tk == 'idle':
        notice = 'You were signed out after 30 minutes of inactivity. Please sign in again.'
    elif tk == 'session':
        notice = 'Your 8-hour session expired. Please sign in again.'

    return render_template('login.html', error=error, notice=notice, stores=stores)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

# ─── Dashboard ─────────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def dashboard():
    today_str = date.today().isoformat()
    week_ago  = (date.today() - timedelta(days=7)).isoformat()

    sid = current_store_id()
    with get_db() as conn:
        chk_status = {}
        for t in CHECKLISTS:
            op = conn.execute(
                'SELECT id,verified,(SELECT COUNT(*) FROM checklist_tasks WHERE session_id=checklist_sessions.id AND done=1) as done_count,(SELECT COUNT(*) FROM checklist_tasks WHERE session_id=checklist_sessions.id) as total FROM checklist_sessions WHERE type=? AND section="opening" AND date=? AND store_id=?',
                (t, today_str, sid)).fetchone()
            cl = conn.execute(
                'SELECT id,verified,(SELECT COUNT(*) FROM checklist_tasks WHERE session_id=checklist_sessions.id AND done=1) as done_count,(SELECT COUNT(*) FROM checklist_tasks WHERE session_id=checklist_sessions.id) as total FROM checklist_sessions WHERE type=? AND section="closing" AND date=? AND store_id=?',
                (t, today_str, sid)).fetchone()
            chk_status[t] = {
                'opening': dict(op) if op else None,
                'closing': dict(cl) if cl else None,
            }

        temp_status = {}
        for t in TEMPERATURES:
            rec = conn.execute('SELECT id FROM temp_sessions WHERE type=? AND date=? AND store_id=?',
                               (t, today_str, sid)).fetchone()
            temp_status[t] = dict(rec) if rec else None

        # Equipment temperature status for today (for the "Today's Tasks" hub).
        equip_total = conn.execute(
            'SELECT COUNT(*) c FROM equipment_units WHERE active=1 AND store_id=?', (sid,)).fetchone()['c']
        try:
            equip_done = conn.execute(
                '''SELECT COUNT(*) c FROM equipment_temp_readings WHERE date=? AND store_id=?
                   AND (morning_temp IS NOT NULL OR closing_temp IS NOT NULL OR temp IS NOT NULL)''',
                (today_str, sid)).fetchone()['c']
        except Exception:
            equip_done = 0

        recent = [dict(r) for r in conn.execute(
            'SELECT * FROM audit_log ORDER BY timestamp DESC LIMIT 15').fetchall()]

        pending_count = conn.execute(
            'SELECT COUNT(*) as c FROM checklist_sessions WHERE verified=0 AND store_id=?',
            (sid,)).fetchone()['c']

        total_week_chk = conn.execute(
            'SELECT COUNT(*) as c FROM checklist_sessions WHERE date>=? AND store_id=?',
            (week_ago, sid)).fetchone()['c']
        total_week_temp = conn.execute(
            'SELECT COUNT(*) as c FROM temp_sessions WHERE date>=? AND store_id=?',
            (week_ago, sid)).fetchone()['c']

        # Alerts: missing today's records
        alerts = []
        total_expected = len(CHECKLISTS) * 2 + len(TEMPERATURES)
        submitted_today = conn.execute(
            'SELECT COUNT(*) as c FROM checklist_sessions WHERE date=? AND store_id=?',
            (today_str, sid)).fetchone()['c']
        temp_today = conn.execute(
            'SELECT COUNT(*) as c FROM temp_sessions WHERE date=? AND store_id=?',
            (today_str, sid)).fetchone()['c']
        if submitted_today < len(CHECKLISTS) * 2:
            alerts.append({'type': 'warning', 'msg': f'{len(CHECKLISTS)*2 - submitted_today} checklist section(s) not yet submitted today'})
        if temp_today < len(TEMPERATURES):
            alerts.append({'type': 'warning', 'msg': f'{len(TEMPERATURES) - temp_today} temperature record(s) not yet submitted today'})
        if pending_count > 0:
            alerts.append({'type': 'info', 'msg': f'{pending_count} checklist(s) awaiting manager verification'})

    return render_template('dashboard.html',
        chk_status=chk_status, temp_status=temp_status,
        recent=recent, pending_count=pending_count,
        total_week_chk=total_week_chk, total_week_temp=total_week_temp,
        alerts=alerts, equip_done=equip_done, equip_total=equip_total,
    )

# ─── Checklists ────────────────────────────────────────────────────────────────

@app.route('/checklist/<chk_type>')
@login_required
def checklist_form(chk_type):
    if chk_type not in CHECKLISTS:
        return redirect(url_for('dashboard'))
    chk_date = request.args.get('date', date.today().isoformat())
    # Only TODAY's Opening locks after 3 PM Perth; past dates stay open for review.
    locked = opening_locked(chk_date)
    section  = request.args.get('section') or ('closing' if locked else 'opening')
    section  = 'closing' if section not in ('opening', 'closing') else section
    if section == 'opening' and locked:
        flash('Opening checklist is locked after 3 PM. View it in History.', 'warning')
        return redirect(url_for('checklist_form', chk_type=chk_type, date=chk_date, section='closing'))
    chk_data = CHECKLISTS[chk_type]
    try:
        day_name = datetime.strptime(chk_date, '%Y-%m-%d').strftime('%A')
    except Exception:
        day_name = ''

    with get_db() as conn:
        # Load task names from DB (admin may have renamed them)
        db_tasks = conn.execute(
            'SELECT task_name FROM checklist_task_templates WHERE chk_type=? AND section=? AND store_id=? ORDER BY task_order',
            (chk_type, section, current_store_id())).fetchall()
        if db_tasks:
            tasks = [r['task_name'] for r in db_tasks]
        else:
            tasks = chk_data.get(section, [])

        existing = conn.execute(
            'SELECT * FROM checklist_sessions WHERE type=? AND section=? AND date=? AND store_id=?',
            (chk_type, section, chk_date, current_store_id())).fetchone()
        existing_tasks = []
        if existing:
            existing_tasks = [dict(r) for r in conn.execute(
                'SELECT * FROM checklist_tasks WHERE session_id=? ORDER BY task_order',
                (existing['id'],)).fetchall()]

    return render_template('checklist.html',
        chk_type=chk_type, chk_data=chk_data, section=section,
        chk_date=chk_date, day_name=day_name, tasks=tasks,
        existing=dict(existing) if existing else None,
        existing_tasks=existing_tasks, staff=get_active_staff(), managers=MANAGERS,
        opening_locked=locked,   # date-aware: overrides the global (today) lock
    )


@app.route('/admin/checklist-task/update', methods=['POST'])
@admin_required
def update_checklist_task():
    chk_type = request.form.get('chk_type', '').strip()
    section  = request.form.get('section', '').strip()
    order    = request.form.get('order', '')
    name     = request.form.get('name', '').strip()
    if not name or chk_type not in CHECKLISTS or section not in ('opening', 'closing'):
        return jsonify({'error': 'invalid'}), 400
    try:
        order = int(order)
    except (ValueError, TypeError):
        return jsonify({'error': 'invalid order'}), 400
    if order < 0:
        return jsonify({'error': 'invalid order'}), 400
    with get_db() as conn:
        sid = current_store_id()
        conn.execute('''INSERT INTO checklist_task_templates
            (chk_type, section, task_order, task_name, store_id)
            VALUES (?,?,?,?,?)
            ON CONFLICT(chk_type, section, task_order, store_id) DO UPDATE
            SET task_name=excluded.task_name''',
            (chk_type, section, order, name, sid))
    return jsonify({'ok': True, 'name': name})

@app.route('/admin/checklist-task/add', methods=['POST'])
@admin_required
def add_checklist_task():
    chk_type = request.form.get('chk_type', '').strip()
    section  = request.form.get('section', '').strip()
    name     = request.form.get('name', '').strip()
    if not name or chk_type not in CHECKLISTS or section not in ('opening', 'closing'):
        return jsonify({'error': 'invalid'}), 400
    with get_db() as conn:
        sid = current_store_id()
        next_order = conn.execute('''
            SELECT COALESCE(MAX(task_order), -1) + 1 as next_order
            FROM checklist_task_templates
            WHERE chk_type=? AND section=? AND store_id=?
        ''', (chk_type, section, sid)).fetchone()['next_order']
        conn.execute('''INSERT INTO checklist_task_templates
            (chk_type, section, task_order, task_name, store_id) VALUES (?,?,?,?,?)''',
            (chk_type, section, next_order, name, sid))
        total = conn.execute('''
            SELECT COUNT(*) as c FROM checklist_task_templates
            WHERE chk_type=? AND section=? AND store_id=?
        ''', (chk_type, section, sid)).fetchone()['c']
    return jsonify({'ok': True, 'order': total - 1, 'name': name})


@app.route('/admin/checklist-task/delete', methods=['POST'])
@admin_required
def delete_checklist_task():
    """Delete a task from a checklist template by its (chk_type, section, task_order)
    and renumber remaining task_orders so they stay contiguous."""
    chk_type = request.form.get('chk_type', '').strip()
    section  = request.form.get('section', '').strip()
    order    = request.form.get('order', '')
    if chk_type not in CHECKLISTS or section not in ('opening', 'closing'):
        return jsonify({'error': 'invalid'}), 400
    try:
        order = int(order)
    except (TypeError, ValueError):
        return jsonify({'error': 'invalid order'}), 400

    with get_db() as conn:
        sid = current_store_id()
        conn.execute('''DELETE FROM checklist_task_templates
            WHERE chk_type=? AND section=? AND task_order=? AND store_id=?''',
            (chk_type, section, order, sid))
        # Pull rows above the deleted slot down by 1 so task_order stays contiguous.
        # SQLite handles UNIQUE constraint within a single UPDATE statement atomically,
        # so direct decrement is safe here (the deleted slot is empty).
        conn.execute('''UPDATE checklist_task_templates
            SET task_order = task_order - 1
            WHERE chk_type=? AND section=? AND task_order > ? AND store_id=?''',
            (chk_type, section, order, sid))
    return jsonify({'ok': True})

@app.route('/checklist/<chk_type>/save', methods=['POST'])
@login_required
@database_save_guard('checklist')
def checklist_save(chk_type):
    if chk_type not in CHECKLISTS:
        return redirect(url_for('dashboard'))
    chk_date       = request.form.get('date', date.today().isoformat())
    section        = request.form.get('section', 'opening')
    # Server-side lock: only TODAY's Opening is locked after 3 PM Perth; past dates
    # stay editable so earlier records can be corrected.
    if section == 'opening' and opening_locked(chk_date):
        flash('Opening checklist is locked after 3 PM. View it in History.', 'warning')
        return redirect(url_for('checklist_form', chk_type=chk_type, date=chk_date, section='closing'))
    responsible    = request.form.get('responsible', '')
    # Banh Mi station can have two people responsible — combine into one field.
    responsible2   = request.form.get('responsible2', '').strip()
    if responsible2 and responsible2 != responsible.strip():
        responsible = ' & '.join([p for p in (responsible.strip(), responsible2) if p])
    submitted_by   = request.form.get('submitted_by', '')
    general_done_by = request.form.get('general_done_by', '')
    manager_submit = request.form.get('manager_submit', '')
    general_note   = request.form.get('general_note', '')
    try:
        day_name = datetime.strptime(chk_date, '%Y-%m-%d').strftime('%A')
    except Exception:
        day_name = ''

    now = datetime.now()
    deadline_h, deadline_m = (10, 30) if section == 'opening' else (18, 30)
    deadline_dt = now.replace(hour=deadline_h, minute=deadline_m, second=0, microsecond=0)
    is_late = 1 if now > deadline_dt else 0

    with get_db() as conn:
        db_tasks = conn.execute(
            'SELECT task_name FROM checklist_task_templates WHERE chk_type=? AND section=? AND store_id=? ORDER BY task_order',
            (chk_type, section, current_store_id())).fetchall()
        tasks = [r['task_name'] for r in db_tasks] if db_tasks else CHECKLISTS[chk_type][section]
        task_rows = []
        for i, task_name in enumerate(tasks):
            task_rows.append((
                i, task_name,
                1 if request.form.get(f'done_{i}') else 0,
                request.form.get(f'note_{i}', ''),
            ))

        # UPDATE-first upsert keeps this compatible with databases where a
        # startup migration was delayed by a SQLite write lock.
        store_id = current_store_id()
        existing_session = conn.execute(
            '''SELECT id FROM checklist_sessions
               WHERE type=? AND section=? AND date=? AND store_id=?''',
            (chk_type, section, chk_date, store_id),
        ).fetchone()
        if existing_session:
            sid = existing_session['id']
            conn.execute('''
                UPDATE checklist_sessions SET
                    responsible=?,
                    submitted_by=?,
                    submitted_at=datetime('now','localtime'),
                    day_of_week=?,
                    general_done_by=?,
                    manager_submit=?,
                    general_note=?,
                    is_late=?
                WHERE id=?
            ''', (responsible, submitted_by, day_name, general_done_by,
                  manager_submit, general_note, is_late, sid))
        else:
            try:
                cur = conn.execute('''
                    INSERT INTO checklist_sessions
                        (type,section,date,day_of_week,responsible,submitted_by,
                         general_done_by,manager_submit,general_note,is_late,store_id)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)
                ''', (chk_type, section, chk_date, day_name, responsible,
                      submitted_by, general_done_by, manager_submit,
                      general_note, is_late, store_id))
                sid = cur.lastrowid
            except sqlite3.IntegrityError:
                # A concurrent submit created the same session first.
                row = conn.execute(
                    '''SELECT id FROM checklist_sessions
                       WHERE type=? AND section=? AND date=? AND store_id=?''',
                    (chk_type, section, chk_date, store_id),
                ).fetchone()
                if not row:
                    raise
                sid = row['id']
                conn.execute('''
                    UPDATE checklist_sessions SET
                        responsible=?, submitted_by=?,
                        submitted_at=datetime('now','localtime'),
                        day_of_week=?, general_done_by=?, manager_submit=?,
                        general_note=?, is_late=?
                    WHERE id=?
                ''', (responsible, submitted_by, day_name, general_done_by,
                      manager_submit, general_note, is_late, sid))

        for i, task_name, done, note in task_rows:
            _update_then_insert(
                conn,
                'checklist_tasks',
                {'session_id': sid, 'task_order': i},
                {'task_name': task_name, 'done': done, 'note': note},
            )
        conn.execute('DELETE FROM checklist_tasks WHERE session_id=? AND task_order>=?',
                     (sid, len(task_rows)))

        # ── Handle photo uploads ──────────────────────────────────────────────
        new_photos = []
        for i in range(PHOTOS_REQUIRED + 6):
            f = request.files.get(f'photo_{i}')
            if f and f.filename and f.filename.strip():
                ext = os.path.splitext(f.filename)[1].lower().lstrip('.')
                if ext in ALLOWED_EXT:
                    new_photos.append((i, f))

        if new_photos:
            # Delete old photo files then re-insert
            old = conn.execute('SELECT filename FROM checklist_photos WHERE session_id=?', (sid,)).fetchall()
            for ph in old:
                try: os.remove(os.path.join(UPLOAD_FOLDER, ph['filename']))
                except: pass
            conn.execute('DELETE FROM checklist_photos WHERE session_id=?', (sid,))
            for i, f in new_photos:
                ext   = os.path.splitext(f.filename)[1].lower()
                fname = f'{sid}_{i}_{uuid.uuid4().hex[:8]}{ext}'
                dest  = os.path.join(UPLOAD_FOLDER, fname)
                # save_uploaded_photo resizes large phone photos before writing;
                # it may rewrite as .jpg, so use the returned filename.
                fname = save_uploaded_photo(f, dest)
                fsize = os.path.getsize(os.path.join(UPLOAD_FOLDER, fname))
                conn.execute(
                    'INSERT INTO checklist_photos (session_id,filename,original_name,photo_number,file_size,uploaded_by) VALUES (?,?,?,?,?,?)',
                    (sid, fname, secure_filename(f.filename), i, fsize, submitted_by))
        # ─────────────────────────────────────────────────────────────────────

        log_action_conn(conn, 'SAVE_CHECKLIST', 'checklist', sid, submitted_by,
                        f'{CHECKLISTS[chk_type]["title"]} / {section} / {chk_date}')

    done_count = sum(1 for t in task_rows if t[2])
    completion_pct = round(done_count / len(task_rows) * 100) if task_rows else 0
    email_service.send_notification(
        'checklist',
        subject=f'{CHECKLISTS[chk_type]["title"]} {section} checklist submitted ({chk_date})',
        lines=[
            f'Type: {CHECKLISTS[chk_type]["title"]}',
            f'Section: {section.title()}',
            f'Date: {chk_date} ({day_name})',
            f'Completion: {done_count} / {len(task_rows)} tasks ({completion_pct}%)',
            f'Late submission: {"Yes" if is_late else "On time"}',
            f'Photos attached: {len(new_photos)}',
            f'General note: {general_note or "-"}',
            f'Submitted by: {submitted_by or "-"}',
            f'Responsible: {responsible or "-"}',
            f'Manager on duty: {manager_submit or "-"}',
            f'General done by: {general_done_by or "-"}',
        ],
        link_path=f'/checklist/view/{sid}',
        actor=submitted_by,
    )
    return redirect(url_for('checklist_view', session_id=sid, saved=1))

@app.route('/checklist/view/<int:session_id>')
@login_required
def checklist_view(session_id):
    guard, gp = store_guard_clause()
    with get_db() as conn:
        sess = conn.execute(f'SELECT * FROM checklist_sessions WHERE id=? AND {guard}',
                            [session_id] + gp).fetchone()
        if not sess:
            return redirect(url_for('dashboard'))
        tasks = [dict(r) for r in conn.execute(
            'SELECT * FROM checklist_tasks WHERE session_id=? ORDER BY task_order',
            (session_id,)).fetchall()]
        photos = [dict(r) for r in conn.execute(
            'SELECT * FROM checklist_photos WHERE session_id=? ORDER BY photo_number',
            (session_id,)).fetchall()]
    return render_template('checklist_view.html',
        sess=dict(sess), tasks=tasks, photos=photos,
        chk_data=CHECKLISTS.get(sess['type'], {}), staff=get_active_staff(),
    )

@app.route('/checklist/verify/<int:session_id>', methods=['POST'])
@login_required
def checklist_verify(session_id):
    verified_by    = request.form.get('verified_by','')
    overall_result = request.form.get('overall_result','')
    issues_found   = request.form.get('issues_found','')
    action_resp    = request.form.get('action_responsible','')
    manager_notes  = request.form.get('manager_notes','')
    guard, gp = store_guard_clause()
    with get_db() as conn:
        conn.execute(
            f"UPDATE checklist_sessions SET verified=1,verified_by=?,verified_at=datetime('now','localtime'),overall_result=?,issues_found=?,action_responsible=?,manager_notes=? WHERE id=? AND {guard}",
            [verified_by, overall_result, issues_found, action_resp, manager_notes, session_id] + gp)
        log_action_conn(conn, 'VERIFY', 'checklist', session_id, verified_by,
                        f'Result: {overall_result}')
        sess = conn.execute('SELECT * FROM checklist_sessions WHERE id=?', (session_id,)).fetchone()

    if sess:
        sess = dict(sess)
        chk_label = CHECKLISTS.get(sess.get('type'), {}).get('title', sess.get('type', ''))
        result_label = (overall_result or '-').replace('_', ' ').title()
        email_service.send_notification(
            'checklist',
            subject=f'Checklist VERIFIED by manager: {chk_label} {sess.get("section","")} ({sess.get("date","")})',
            lines=[
                f'Type: {chk_label}',
                f'Section: {(sess.get("section") or "").title()}',
                f'Date: {sess.get("date", "")}',
                f'Status: Verified',
                f'Overall result: {result_label}',
                f'Submitted by: {sess.get("submitted_by") or "-"}',
                f'Verified by: {verified_by or "-"}',
                f'Action responsible: {action_resp or "-"}',
                f'Issues found: {issues_found or "None"}',
                f'Manager notes: {manager_notes or "-"}',
            ],
            link_path=f'/checklist/view/{session_id}',
            actor=verified_by,
        )
    return redirect(url_for('checklist_view', session_id=session_id))

# ─── Temperature ───────────────────────────────────────────────────────────────

@app.route('/temperature/<temp_type>')
@login_required
def temperature_form(temp_type):
    if temp_type not in TEMPERATURES:
        return redirect(url_for('dashboard'))
    temp_date = request.args.get('date', date.today().isoformat())
    with get_db() as conn:
        temp_data = get_temp_data_for_form(conn, temp_type, store_id=current_store_id())
        existing = conn.execute(
            'SELECT * FROM temp_sessions WHERE type=? AND date=? AND store_id=?',
            (temp_type, temp_date, current_store_id())).fetchone()
        existing_readings = []
        if existing:
            existing_readings = [dict(r) for r in conn.execute(
                'SELECT * FROM temp_readings WHERE session_id=? ORDER BY food_order',
                (existing['id'],)).fetchall()]
    return render_template('temperature.html',
        temp_type=temp_type, temp_data=temp_data, temp_date=temp_date,
        existing=dict(existing) if existing else None,
        existing_readings=existing_readings, staff=get_active_staff(),
    )

@app.route('/admin/temperature-food/update', methods=['POST'])
@admin_required
def update_temperature_food():
    temp_type = request.form.get('temp_type', '').strip()
    order     = request.form.get('order', '')
    name      = request.form.get('name', '').strip()
    if temp_type not in TEMPERATURES or not name:
        return jsonify({'error': 'invalid'}), 400
    try:
        order = int(order)
    except (TypeError, ValueError):
        return jsonify({'error': 'invalid order'}), 400
    if order < 0:
        return jsonify({'error': 'invalid order'}), 400

    with get_db() as conn:
        sid = current_store_id()
        kind = temp_food_kind_default(temp_type, name)
        conn.execute('''INSERT INTO temp_food_templates
            (temp_type, food_order, food_name, food_kind, store_id)
            VALUES (?,?,?,?,?)
            ON CONFLICT(temp_type, food_order, store_id) DO UPDATE
            SET food_name=excluded.food_name, food_kind=excluded.food_kind''',
            (temp_type, order, name, kind, sid))
    return jsonify({'ok': True, 'name': name})

@app.route('/admin/temperature-food/add', methods=['POST'])
@admin_required
def add_temperature_food():
    temp_type = request.form.get('temp_type', '').strip()
    name      = request.form.get('name', '').strip()
    if temp_type not in TEMPERATURES or not name:
        return jsonify({'error': 'invalid'}), 400
    with get_db() as conn:
        sid = current_store_id()
        next_order = conn.execute('''
            SELECT COALESCE(MAX(food_order), -1) + 1 as next_order
            FROM temp_food_templates
            WHERE temp_type=? AND store_id=?
        ''', (temp_type, sid)).fetchone()['next_order']
        kind = temp_food_kind_default(temp_type, name)
        conn.execute('''INSERT INTO temp_food_templates
            (temp_type, food_order, food_name, food_kind, store_id) VALUES (?,?,?,?,?)''',
            (temp_type, next_order, name, kind, sid))
    return jsonify({'ok': True, 'order': next_order, 'name': name})


@app.route('/admin/temperature-food/delete', methods=['POST'])
@admin_required
def delete_temperature_food():
    """Delete a food row from a temperature template by (temp_type, food_order)
    and renumber rows above it so food_order stays contiguous."""
    temp_type = request.form.get('temp_type', '').strip()
    order     = request.form.get('order', '')
    if temp_type not in TEMPERATURES:
        return jsonify({'error': 'invalid'}), 400
    try:
        order = int(order)
    except (TypeError, ValueError):
        return jsonify({'error': 'invalid order'}), 400

    with get_db() as conn:
        sid = current_store_id()
        conn.execute('''DELETE FROM temp_food_templates
            WHERE temp_type=? AND food_order=? AND store_id=?''', (temp_type, order, sid))
        conn.execute('''UPDATE temp_food_templates
            SET food_order = food_order - 1
            WHERE temp_type=? AND food_order > ? AND store_id=?''', (temp_type, order, sid))
    return jsonify({'ok': True})

@app.route('/temperature/<temp_type>/save', methods=['POST'])
@login_required
@database_save_guard('temperature')
def temperature_save(temp_type):
    if temp_type not in TEMPERATURES:
        return redirect(url_for('dashboard'))
    temp_date    = request.form.get('date', date.today().isoformat())
    recorded_by  = request.form.get('recorded_by', '')
    checked_by   = request.form.get('checked_by', '')
    notes        = request.form.get('notes', '')

    def p(v):
        try: return float(v) if v and v.strip() else None
        except: return None

    with get_db() as _c:
        foods = get_temp_foods(_c, temp_type, store_id=current_store_id())
    readings = []
    for i, food in enumerate(foods):
        readings.append({
            'food_order': i, 'food_name': food['name'], 'food_kind': food['kind'],
            'c1_time': request.form.get(f'c1_time_{i}',''), 'c1_temp': p(request.form.get(f'c1_temp_{i}','')),
            'c2_time': request.form.get(f'c2_time_{i}',''), 'c2_temp': p(request.form.get(f'c2_temp_{i}','')),
            'c3_time': request.form.get(f'c3_time_{i}',''), 'c3_temp': p(request.form.get(f'c3_temp_{i}','')),
            'c4_time': request.form.get(f'c4_time_{i}',''), 'c4_temp': p(request.form.get(f'c4_temp_{i}','')),
            'c5_time': request.form.get(f'c5_time_{i}',''), 'c5_temp': p(request.form.get(f'c5_temp_{i}','')),
            'discarded': request.form.get(f'discarded_{i}', 'N'),
            'defrosted': request.form.get(f'defrosted_{i}', 'N'),
            'notes':     request.form.get(f'food_notes_{i}', '').strip(),
        })

    store_id = current_store_id()
    with get_db() as conn:
        existing = conn.execute(
            'SELECT id FROM temp_sessions WHERE type=? AND date=? AND store_id=?',
            (temp_type, temp_date, store_id)).fetchone()
        if existing:
            sid = existing['id']
            conn.execute(
                "UPDATE temp_sessions SET recorded_by=?,checked_by=?,submitted_at=datetime('now','localtime'),notes=? WHERE id=?",
                (recorded_by, checked_by, notes, sid))
        else:
            try:
                cur = conn.execute(
                    'INSERT INTO temp_sessions (type,date,recorded_by,checked_by,notes,store_id) VALUES (?,?,?,?,?,?)',
                    (temp_type, temp_date, recorded_by, checked_by, notes, store_id))
                sid = cur.lastrowid
            except sqlite3.IntegrityError:
                row = conn.execute(
                    'SELECT id FROM temp_sessions WHERE type=? AND date=? AND store_id=?',
                    (temp_type, temp_date, store_id),
                ).fetchone()
                if not row:
                    raise
                sid = row['id']
                conn.execute(
                    "UPDATE temp_sessions SET recorded_by=?,checked_by=?,submitted_at=datetime('now','localtime'),notes=? WHERE id=?",
                    (recorded_by, checked_by, notes, sid))
        for r in readings:
            _update_then_insert(
                conn,
                'temp_readings',
                {'session_id': sid, 'food_order': r['food_order']},
                {
                    'food_name': r['food_name'],
                    'food_kind': r['food_kind'],
                    'c1_time': r['c1_time'],
                    'c1_temp': r['c1_temp'],
                    'c2_time': r['c2_time'],
                    'c2_temp': r['c2_temp'],
                    'c3_time': r['c3_time'],
                    'c3_temp': r['c3_temp'],
                    'c4_time': r['c4_time'],
                    'c4_temp': r['c4_temp'],
                    'c5_time': r['c5_time'],
                    'c5_temp': r['c5_temp'],
                    'discarded': r['discarded'],
                    'notes': r['notes'],
                    'defrosted': r['defrosted'],
                },
            )
        conn.execute('DELETE FROM temp_readings WHERE session_id=? AND food_order>=?',
                     (sid, len(readings)))
        log_action_conn(conn, 'SAVE_TEMP', 'temperature', sid, recorded_by,
                        f'{TEMPERATURES[temp_type]["title"]} / {temp_date}')

    # Out-of-zone check uses food_kind:
    #   cold → unsafe if > 5°C · room → unsafe if <15 or >30 · hot → unsafe if <60
    out_of_zone = []
    discarded_items = []
    for r in readings:
        # A cold item being defrosted is expected to rise above range — no alert.
        if (r.get('defrosted') or 'N').upper() == 'Y':
            continue
        for n in range(1, 6):
            tv = r.get(f'c{n}_temp')
            if tv is None:
                continue
            # Pastry hot display: the 3rd check is informational only — no alert.
            if temp_type == 'pastry' and n == 3:
                continue
            unsafe = temp_is_unsafe(r['food_kind'], tv)
            if unsafe:
                out_of_zone.append(
                    f'{r["food_name"]} ({r["food_kind"]}) check {n}: {tv}°C')
        if (r.get('discarded') or '').upper() == 'Y':
            discarded_items.append(r['food_name'])
    email_service.send_notification(
        'temperature',
        subject=f'{TEMPERATURES[temp_type]["title"]} submitted ({temp_date})',
        lines=[
            f'Type: {TEMPERATURES[temp_type]["title"]}',
            f'Date: {temp_date}',
            f'Recorded by: {recorded_by or "-"}',
            f'Checked by: {checked_by or "-"}',
            f'Items recorded: {len(readings)}',
            f'Out-of-zone readings: {len(out_of_zone)}' + (' — ' + '; '.join(out_of_zone[:5]) if out_of_zone else ''),
            f'Discarded items: {", ".join(discarded_items) if discarded_items else "none"}',
            f'Notes: {notes or "-"}',
        ],
        link_path=f'/temperature/view/{sid}',
        actor=recorded_by,
    )
    return redirect(url_for('temperature_view', session_id=sid, saved=1))

@app.route('/temperature/view/<int:session_id>')
@login_required
def temperature_view(session_id):
    guard, gp = store_guard_clause()
    with get_db() as conn:
        sess = conn.execute(f'SELECT * FROM temp_sessions WHERE id=? AND {guard}',
                            [session_id] + gp).fetchone()
        if not sess:
            return redirect(url_for('dashboard'))
        readings = [dict(r) for r in conn.execute(
            'SELECT * FROM temp_readings WHERE session_id=? ORDER BY food_order',
            (session_id,)).fetchall()]
    return render_template('temperature_view.html',
        sess=dict(sess), readings=readings,
        temp_data=TEMPERATURES.get(sess['type'], {}),
    )

# ─── History ───────────────────────────────────────────────────────────────────

@app.route('/history')
@login_required
def history():
    date_from    = request.args.get('date_from', (date.today()-timedelta(days=30)).isoformat())
    date_to      = request.args.get('date_to',   date.today().isoformat())
    rec_type     = request.args.get('type', 'all')
    staff_filter = request.args.get('staff', '')

    chk_scope, chk_sp = store_filter_clause('cs')
    temp_scope, temp_sp = store_filter_clause()
    with get_db() as conn:
        # Checklist records
        q = f'''SELECT cs.*,
               (SELECT COUNT(*) FROM checklist_tasks WHERE session_id=cs.id AND done=1) as done_count,
               (SELECT COUNT(*) FROM checklist_tasks WHERE session_id=cs.id) as total_count
               FROM checklist_sessions cs WHERE cs.date BETWEEN ? AND ? AND {chk_scope}'''
        p = [date_from, date_to] + chk_sp
        if staff_filter:
            q += ' AND (cs.submitted_by=? OR cs.responsible=?)'
            p += [staff_filter, staff_filter]
        q += ' ORDER BY cs.date DESC, cs.type, cs.section'
        chk_records = [dict(r) for r in conn.execute(q, p).fetchall()] if rec_type in ('all','checklist') else []

        # Temperature records
        q2 = f'SELECT * FROM temp_sessions WHERE date BETWEEN ? AND ? AND {temp_scope}'
        p2 = [date_from, date_to] + temp_sp
        if staff_filter:
            q2 += ' AND (recorded_by=? OR checked_by=?)'
            p2 += [staff_filter, staff_filter]
        q2 += ' ORDER BY date DESC, type'
        temp_records = [dict(r) for r in conn.execute(q2, p2).fetchall()] if rec_type in ('all','temperature') else []

    return render_template('history.html',
        chk_records=chk_records, temp_records=temp_records,
        date_from=date_from, date_to=date_to,
        rec_type=rec_type, staff_filter=staff_filter, staff=get_active_staff(),
    )

# ─── Manager Panel ─────────────────────────────────────────────────────────────

@app.route('/manager')
@admin_required
def manager():
    today_str = date.today().isoformat()
    # Panel is scoped to a single day; defaults to today but can be sorted/
    # navigated by date via ?date=YYYY-MM-DD.
    sel_date = (request.args.get('date', '') or '').strip() or today_str
    try:
        d = datetime.strptime(sel_date, '%Y-%m-%d').date()
    except ValueError:
        d = date.today()
        sel_date = today_str
    prev_date = (d - timedelta(days=1)).isoformat()
    next_date = (d + timedelta(days=1)).isoformat()
    scope, sp = store_filter_clause('cs')
    with get_db() as conn:
        pending = [dict(r) for r in conn.execute(f'''
            SELECT cs.*,
              (SELECT COUNT(*) FROM checklist_tasks WHERE session_id=cs.id AND done=1) as done_count,
              (SELECT COUNT(*) FROM checklist_tasks WHERE session_id=cs.id) as total_count
            FROM checklist_sessions cs WHERE cs.verified=0 AND cs.date=? AND {scope}
            ORDER BY cs.type, cs.section
        ''', [sel_date] + sp).fetchall()]
        issues = [dict(r) for r in conn.execute(f'''
            SELECT * FROM checklist_sessions cs WHERE cs.overall_result='issues_found' AND cs.date=? AND {scope}
            ORDER BY cs.type, cs.section LIMIT 50
        ''', [sel_date] + sp).fetchall()]
        verified_today = conn.execute(
            f"SELECT COUNT(*) as c FROM checklist_sessions cs WHERE cs.verified=1 AND cs.date=? AND {scope}",
            [sel_date] + sp).fetchone()['c']
        recent_log = [dict(r) for r in conn.execute(
            "SELECT * FROM audit_log WHERE substr(timestamp,1,10)=? "
            "ORDER BY timestamp DESC LIMIT 50", (sel_date,)).fetchall()]
    return render_template('manager.html',
        pending=pending, issues=issues,
        verified_today=verified_today, recent_log=recent_log, staff=get_active_staff(),
        sel_date=sel_date, today=today_str, is_today=(sel_date == today_str),
        prev_date=prev_date, next_date=next_date,
    )

# ─── Analytics ─────────────────────────────────────────────────────────────────

@app.route('/analytics')
@admin_required
def analytics():
    return render_template('analytics.html')

@app.route('/api/analytics-data')
@login_required
def analytics_data():
    days       = int(request.args.get('days', 30))
    start_date = (date.today()-timedelta(days=days-1)).isoformat()
    end_date   = date.today().isoformat()

    cscope, csp = store_filter_clause()      # bare store_id clause
    with get_db() as conn:
        daily_chk = conn.execute(f'''
            SELECT date, COUNT(*) as total,
                   SUM(CASE WHEN verified=1 THEN 1 ELSE 0 END) as verified
            FROM checklist_sessions WHERE date BETWEEN ? AND ? AND {cscope}
            GROUP BY date ORDER BY date
        ''', [start_date, end_date] + csp).fetchall()

        daily_temp = conn.execute(f'''
            SELECT date, COUNT(*) as total FROM temp_sessions
            WHERE date BETWEEN ? AND ? AND {cscope}
            GROUP BY date ORDER BY date
        ''', [start_date, end_date] + csp).fetchall()

        staff_act = conn.execute(f'''
            SELECT submitted_by as name, COUNT(*) as cnt
            FROM checklist_sessions WHERE date BETWEEN ? AND ? AND submitted_by!='' AND {cscope}
            GROUP BY submitted_by ORDER BY cnt DESC LIMIT 13
        ''', [start_date, end_date] + csp).fetchall()

        type_comp = conn.execute(f'''
            SELECT type, COUNT(*) as total,
                   SUM(CASE WHEN verified=1 THEN 1 ELSE 0 END) as verified
            FROM checklist_sessions WHERE date BETWEEN ? AND ? AND {cscope}
            GROUP BY type
        ''', [start_date, end_date] + csp).fetchall()

    labels = []
    cur = date.today()-timedelta(days=days-1)
    while cur <= date.today():
        labels.append(cur.isoformat())
        cur += timedelta(days=1)

    chk_map  = {r['date']: dict(r) for r in daily_chk}
    temp_map = {r['date']: dict(r) for r in daily_temp}

    return jsonify({
        'labels': labels,
        'daily_chk':      [chk_map.get(d,{}).get('total',0)    for d in labels],
        'daily_verified': [chk_map.get(d,{}).get('verified',0) for d in labels],
        'daily_temp':     [temp_map.get(d,{}).get('total',0)   for d in labels],
        'staff_names':  [r['name']  for r in staff_act],
        'staff_counts': [r['cnt']   for r in staff_act],
        'type_labels':    [CHECKLISTS.get(r['type'],{}).get('short',r['type']) for r in type_comp],
        'type_totals':    [r['total']    for r in type_comp],
        'type_verified':  [r['verified'] for r in type_comp],
    })

# ─── Cross-store reports (super_admin) ───────────────────────────────────────

@app.route('/admin/stores')
@super_admin_required
def stores_report():
    """Owner overview: key operational + revenue metrics per store, side by side.
    Foundation for cross-store reporting; date range defaults to this month."""
    date_from = request.args.get('date_from', date.today().replace(day=1).isoformat())
    date_to   = request.args.get('date_to', date.today().isoformat())
    with get_db() as conn:
        rows = [dict(r) for r in conn.execute('''
            SELECT s.id, s.code, s.name,
              (SELECT COUNT(*) FROM checklist_sessions cs
                 WHERE cs.store_id=s.id AND cs.date BETWEEN ? AND ?) AS chk_total,
              (SELECT COUNT(*) FROM checklist_sessions cs
                 WHERE cs.store_id=s.id AND cs.date BETWEEN ? AND ? AND cs.verified=1) AS chk_verified,
              (SELECT COUNT(*) FROM temp_sessions ts
                 WHERE ts.store_id=s.id AND ts.date BETWEEN ? AND ?) AS temp_total,
              (SELECT COUNT(*) FROM issue_reports ir
                 WHERE ir.store_id=s.id AND ir.status != 'resolved') AS issues_open,
              (SELECT COUNT(*) FROM staff_violations sv
                 WHERE sv.store_id=s.id AND sv.status='open') AS violations_open,
              (SELECT COALESCE(SUM(o.total),0) FROM orders o
                 WHERE o.store_id=s.id AND o.status != 'cancelled'
                   AND substr(o.created_at,1,10) BETWEEN ? AND ?) AS revenue,
              (SELECT COUNT(*) FROM staff_members sm
                 WHERE sm.store_id=s.id AND sm.active=1) AS staff_count
            FROM stores s WHERE s.active=1 ORDER BY s.id
        ''', [date_from, date_to, date_from, date_to, date_from, date_to,
              date_from, date_to]).fetchall()]
    for r in rows:
        r['verify_rate'] = round(r['chk_verified'] / r['chk_total'] * 100, 1) if r['chk_total'] else 0
    totals = {
        'chk_total':       sum(r['chk_total'] for r in rows),
        'chk_verified':    sum(r['chk_verified'] for r in rows),
        'temp_total':      sum(r['temp_total'] for r in rows),
        'issues_open':     sum(r['issues_open'] for r in rows),
        'violations_open': sum(r['violations_open'] for r in rows),
        'revenue':         sum(r['revenue'] for r in rows),
        'staff_count':     sum(r['staff_count'] for r in rows),
    }
    return render_template('stores_report.html',
        rows=rows, totals=totals, date_from=date_from, date_to=date_to,
        all_store_rows=get_stores(active_only=False))


@app.route('/admin/stores/add', methods=['POST'])
@super_admin_required
def stores_add():
    """Create a new branch. code is the login key (lowercase, no spaces)."""
    name = request.form.get('name', '').strip()
    code = (request.form.get('code', '').strip().lower()
            or ''.join(ch for ch in name.lower() if ch.isalnum()))
    if not name or not code:
        flash('Store name and code are required.', 'danger')
        return redirect(url_for('stores_report'))
    with get_db() as conn:
        exists = conn.execute('SELECT 1 FROM stores WHERE code=?', (code,)).fetchone()
        if exists:
            flash(f'A store with code "{code}" already exists.', 'danger')
            return redirect(url_for('stores_report'))
        conn.execute('''INSERT INTO stores
            (code,name,address,phone,active,user_password,admin_password,kitchen_password)
            VALUES (?,?,?,?,1,?,?,?)''',
            (code, name,
             request.form.get('address', '').strip(),
             request.form.get('phone', '').strip(),
             request.form.get('user_password', '').strip() or USER_PASSWORD,
             request.form.get('admin_password', '').strip() or ADMIN_PASSWORD,
             request.form.get('kitchen_password', '').strip() or KITCHEN_PASSWORD))
    flash(f'Store "{name}" added.', 'success')
    return redirect(url_for('stores_report'))


@app.route('/admin/stores/<int:store_id>/update', methods=['POST'])
@super_admin_required
def stores_update(store_id):
    """Edit a branch's details, per-store passwords, and active flag.
    Blank password fields keep the existing value (not cleared)."""
    with get_db() as conn:
        cur = conn.execute('SELECT * FROM stores WHERE id=?', (store_id,)).fetchone()
        if not cur:
            flash('Store not found.', 'danger')
            return redirect(url_for('stores_report'))
        cur = dict(cur)
        conn.execute('''UPDATE stores
            SET name=?, address=?, phone=?, active=?,
                user_password=?, admin_password=?, kitchen_password=?
            WHERE id=?''',
            (request.form.get('name', '').strip() or cur['name'],
             request.form.get('address', '').strip(),
             request.form.get('phone', '').strip(),
             1 if request.form.get('active') else 0,
             request.form.get('user_password', '').strip() or cur.get('user_password') or USER_PASSWORD,
             request.form.get('admin_password', '').strip() or cur.get('admin_password') or ADMIN_PASSWORD,
             request.form.get('kitchen_password', '').strip() or cur.get('kitchen_password') or KITCHEN_PASSWORD,
             store_id))
    flash('Store updated.', 'success')
    return redirect(url_for('stores_report'))


# ─── HR Rewards & Reviews ─────────────────────────────────────────────────────

@app.route('/admin/monthly-rewards')
@admin_required
def monthly_rewards():
    reward_month = request.args.get('month', date.today().strftime('%Y-%m'))
    scope_id = selected_store_scope()   # int store, or None=all (super_admin)
    clause, cp = store_filter_clause()
    with get_db() as conn:
        data = calculate_monthly_reward_scores(conn, reward_month, store_id=scope_id)
        staff = [dict(r) for r in conn.execute(
            f'SELECT * FROM staff_members WHERE active=1 AND {clause} ORDER BY name',
            cp).fetchall()]
        strike_map = {s['staff_name']: s for s in staff_strike_standings(conn, store_id=scope_id)}
    decision_map = {d['award_type']: d for d in data['decisions']}
    return render_template('monthly_rewards.html',
        reward=data, staff=staff, decision_map=decision_map,
        reward_month=data['month'],
        strike_map=strike_map, strike_threshold=STRIKE_THRESHOLD,
        strike_window_days=STRIKE_WINDOW_DAYS)

@app.route('/admin/monthly-rewards/decision', methods=['POST'])
@admin_required
def monthly_reward_decision():
    reward_month = request.form.get('reward_month', date.today().strftime('%Y-%m')).strip()
    award_type = request.form.get('award_type', '').strip()
    staff_name = request.form.get('staff_name', '').strip()
    reward_amount = float(request.form.get('reward_amount', 0) or 0)
    notes = request.form.get('notes', '').strip()
    if award_type in ('employee_month', 'best_checklist') and staff_name:
        with get_db() as conn:
            conn.execute('''INSERT INTO monthly_reward_decisions
                (reward_month,award_type,staff_name,reward_amount,status,notes,approved_by)
                VALUES (?,?,?,?,?,?,?)
                ON CONFLICT(reward_month, award_type) DO UPDATE SET
                    staff_name=excluded.staff_name,
                    reward_amount=excluded.reward_amount,
                    status=excluded.status,
                    notes=excluded.notes,
                    approved_by=excluded.approved_by,
                    approved_at=datetime('now','localtime')''',
                (reward_month, award_type, staff_name, reward_amount, 'approved',
                 notes, request.form.get('approved_by', '').strip() or 'Admin'))
    return redirect(url_for('monthly_rewards', month=reward_month))

@app.route('/admin/monthly-rewards/decision/<int:decision_id>/delete', methods=['POST'])
@admin_required
def monthly_reward_decision_delete(decision_id):
    with get_db() as conn:
        row = conn.execute('SELECT reward_month FROM monthly_reward_decisions WHERE id=?', (decision_id,)).fetchone()
        reward_month = row['reward_month'] if row else date.today().strftime('%Y-%m')
        conn.execute('DELETE FROM monthly_reward_decisions WHERE id=?', (decision_id,))
    return redirect(url_for('monthly_rewards', month=reward_month))

@app.route('/admin/monthly-rewards/adjustment', methods=['POST'])
@admin_required
def monthly_reward_adjustment():
    reward_month = request.form.get('reward_month', date.today().strftime('%Y-%m')).strip()
    staff_name = request.form.get('staff_name', '').strip()
    reason = request.form.get('reason', '').strip()
    try:
        points = float(request.form.get('points', 0) or 0)
    except ValueError:
        points = 0
    if staff_name and reason and points:
        with get_db() as conn:
            conn.execute('''INSERT INTO monthly_reward_adjustments
                (reward_month,staff_name,points,reason,created_by)
                VALUES (?,?,?,?,?)''',
                (reward_month, staff_name, points, reason,
                 request.form.get('created_by', '').strip() or 'Admin'))
    return redirect(url_for('monthly_rewards', month=reward_month))

@app.route('/admin/monthly-rewards/adjustment/<int:adjustment_id>/update', methods=['POST'])
@admin_required
def monthly_reward_adjustment_update(adjustment_id):
    reward_month = request.form.get('reward_month', date.today().strftime('%Y-%m')).strip()
    try:
        points = float(request.form.get('points', 0) or 0)
    except ValueError:
        points = 0
    with get_db() as conn:
        current = conn.execute(
            'SELECT reward_month FROM monthly_reward_adjustments WHERE id=?',
            (adjustment_id,)).fetchone()
        if current:
            reward_month = current['reward_month']
            conn.execute('''UPDATE monthly_reward_adjustments
                SET staff_name=?, points=?, reason=?, created_by=?
                WHERE id=?''',
                (request.form.get('staff_name', '').strip(),
                 points,
                 request.form.get('reason', '').strip(),
                 request.form.get('created_by', '').strip(),
                 adjustment_id))
    return redirect(url_for('monthly_rewards', month=reward_month))

@app.route('/admin/monthly-rewards/adjustment/<int:adjustment_id>/delete', methods=['POST'])
@admin_required
def monthly_reward_adjustment_delete(adjustment_id):
    with get_db() as conn:
        row = conn.execute('SELECT reward_month FROM monthly_reward_adjustments WHERE id=?', (adjustment_id,)).fetchone()
        reward_month = row['reward_month'] if row else date.today().strftime('%Y-%m')
        conn.execute('DELETE FROM monthly_reward_adjustments WHERE id=?', (adjustment_id,))
    return redirect(url_for('monthly_rewards', month=reward_month))

@app.route('/admin/raise-reviews', methods=['GET', 'POST'])
@admin_required
def raise_reviews():
    if request.method == 'POST':
        staff_name = request.form.get('staff_name', '').strip()
        review_month = request.form.get('review_month', date.today().strftime('%Y-%m')).strip()
        status = request.form.get('status', 'draft').strip()
        if status not in ('draft', 'recommended', 'approved', 'on_hold', 'rejected'):
            status = 'draft'
        if staff_name:
            decided_at = datetime.now().strftime('%Y-%m-%d %H:%M') if status in ('approved', 'rejected') else None
            with get_db() as conn:
                conn.execute('''INSERT INTO salary_raise_reviews
                    (staff_name,review_month,current_rate,proposed_rate,effective_date,status,
                     requested_by,reviewed_by,manager_notes,decision_reason,decided_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                    (staff_name, review_month,
                     float(request.form.get('current_rate', 0) or 0),
                     float(request.form.get('proposed_rate', 0) or 0),
                     request.form.get('effective_date', '').strip(),
                     status,
                     request.form.get('requested_by', '').strip() or 'Admin',
                     request.form.get('reviewed_by', '').strip(),
                     request.form.get('manager_notes', '').strip(),
                     request.form.get('decision_reason', '').strip(),
                     decided_at))
            return redirect(url_for('raise_reviews', staff=staff_name, month=review_month))

    selected_month = request.args.get('month', date.today().strftime('%Y-%m'))
    scope_id = selected_store_scope()
    clause, cp = store_filter_clause()
    with get_db() as conn:
        staff = [dict(r) for r in conn.execute(
            f'SELECT * FROM staff_members WHERE active=1 AND {clause} ORDER BY name',
            cp).fetchall()]
        selected_staff = request.args.get('staff') or (staff[0]['name'] if staff else '')
        snapshot = raise_review_snapshot(conn, selected_staff, selected_month, store_id=scope_id) if selected_staff else None
        reviews = [dict(r) for r in conn.execute('''
            SELECT * FROM salary_raise_reviews
            ORDER BY created_at DESC, id DESC
            LIMIT 40
        ''').fetchall()]
        strike_map = {s['staff_name']: s for s in staff_strike_standings(conn, store_id=scope_id)}
    selected_month, _, _, month_label = month_bounds(selected_month)
    return render_template('raise_reviews.html',
        staff=staff, selected_staff=selected_staff, selected_month=selected_month,
        month_label=month_label, snapshot=snapshot, reviews=reviews,
        strike_map=strike_map, strike_threshold=STRIKE_THRESHOLD,
        strike_window_days=STRIKE_WINDOW_DAYS)

@app.route('/admin/raise-reviews/<int:review_id>/update', methods=['POST'])
@admin_required
def raise_review_update(review_id):
    status = request.form.get('status', 'draft').strip()
    if status not in ('draft', 'recommended', 'approved', 'on_hold', 'rejected'):
        status = 'draft'
    decided_at = datetime.now().strftime('%Y-%m-%d %H:%M') if status in ('approved', 'rejected') else None
    with get_db() as conn:
        current = conn.execute('SELECT staff_name, review_month FROM salary_raise_reviews WHERE id=?', (review_id,)).fetchone()
        if current:
            staff_name = request.form.get('staff_name', current['staff_name']).strip() or current['staff_name']
            review_month = request.form.get('review_month', current['review_month']).strip() or current['review_month']
            conn.execute('''UPDATE salary_raise_reviews
                SET staff_name=?, review_month=?, current_rate=?, proposed_rate=?,
                    effective_date=?, status=?, requested_by=?, reviewed_by=?,
                    manager_notes=?, decision_reason=?, decided_at=?
                WHERE id=?''',
                (staff_name, review_month,
                 float(request.form.get('current_rate', 0) or 0),
                 float(request.form.get('proposed_rate', 0) or 0),
                 request.form.get('effective_date', '').strip(),
                 status,
                 request.form.get('requested_by', '').strip(),
                 request.form.get('reviewed_by', '').strip(),
                 request.form.get('manager_notes', '').strip(),
                 request.form.get('decision_reason', '').strip(),
                 decided_at, review_id))
            return redirect(url_for('raise_reviews', staff=staff_name, month=review_month))
    return redirect(url_for('raise_reviews'))

@app.route('/admin/raise-reviews/<int:review_id>/delete', methods=['POST'])
@admin_required
def raise_review_delete(review_id):
    with get_db() as conn:
        current = conn.execute('SELECT staff_name, review_month FROM salary_raise_reviews WHERE id=?', (review_id,)).fetchone()
        conn.execute('DELETE FROM salary_raise_reviews WHERE id=?', (review_id,))
    if current:
        return redirect(url_for('raise_reviews', staff=current['staff_name'], month=current['review_month']))
    return redirect(url_for('raise_reviews'))

@app.route('/admin/birthdays', methods=['GET', 'POST'])
@admin_required
def birthday_giveaways():
    if request.method == 'POST':
        staff_name = request.form.get('staff_name', '').strip()
        birthday = request.form.get('birthday', '').strip()
        if staff_name and birthday:
            try:
                datetime.strptime(birthday, '%Y-%m-%d')
            except ValueError:
                birthday = ''
        if staff_name and birthday:
            last_year = request.form.get('last_given_year', '').strip()
            last_year_val = int(last_year) if last_year.isdigit() else None
            sid = current_store_id()
            with get_db() as conn:
                conn.execute('''INSERT INTO staff_birthdays
                    (staff_name,birthday,favorite_gift,gift_status,last_given_year,notes,store_id)
                    VALUES (?,?,?,?,?,?,?)
                    ON CONFLICT(staff_name, store_id) DO UPDATE SET
                        birthday=excluded.birthday,
                        favorite_gift=excluded.favorite_gift,
                        gift_status=excluded.gift_status,
                        last_given_year=excluded.last_given_year,
                        notes=excluded.notes,
                        updated_at=datetime('now','localtime')''',
                    (staff_name, birthday,
                     request.form.get('favorite_gift', '').strip(),
                     request.form.get('gift_status', 'planned').strip() or 'planned',
                     last_year_val,
                     request.form.get('notes', '').strip(), sid))
        return redirect(url_for('birthday_giveaways'))

    sid = current_store_id()
    with get_db() as conn:
        # Keep birthday rows in sync when new staff are added after initial setup.
        existing = {r['staff_name'] for r in conn.execute(
            'SELECT staff_name FROM staff_birthdays WHERE store_id=?', (sid,)).fetchall()}
        staff_rows = [dict(r) for r in conn.execute(
            'SELECT * FROM staff_members WHERE active=1 AND store_id=? ORDER BY name', (sid,)).fetchall()]
        gift_ideas = ['MCQ meal voucher', 'Coffee and pastry pack', 'Birthday cake', 'Gift card', 'Team lunch shout', 'Dessert box']
        for idx, staff_row in enumerate(staff_rows):
            if staff_row['name'] not in existing:
                seed = sum(ord(ch) for ch in staff_row['name']) + idx * 17
                conn.execute('''INSERT INTO staff_birthdays
                    (staff_name,birthday,favorite_gift,gift_status,notes,store_id)
                    VALUES (?,?,?,?,?,?)''',
                    (staff_row['name'], f'2000-{seed % 12 + 1:02d}-{(seed * 7) % 28 + 1:02d}',
                     gift_ideas[seed % len(gift_ideas)], 'planned',
                     'Temporary birthday data - update when the real birthday is confirmed.', sid))
        rows = birthday_rows(conn)
    today_obj = date.today()
    next_30 = [r for r in rows if r['days_until'] <= 30]
    this_month = [r for r in rows if int((r['birthday'] or '2000-01-01').split('-')[1]) == today_obj.month]
    given_this_year = [r for r in rows if r.get('last_given_year') == today_obj.year]
    return render_template('birthday_giveaways.html',
        birthdays=rows, next_birthday=rows[0] if rows else None,
        next_30=next_30, this_month=this_month,
        given_this_year=given_this_year, current_year=today_obj.year)

# ─── Export ────────────────────────────────────────────────────────────────────

@app.route('/export/checklist/<int:session_id>/excel')
@admin_required
def export_checklist_excel(session_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    guard, gp = store_guard_clause()
    with get_db() as conn:
        sess  = conn.execute(f'SELECT * FROM checklist_sessions WHERE id=? AND {guard}',
                             [session_id] + gp).fetchone()
        if not sess: return redirect(url_for('dashboard'))
        tasks = [dict(r) for r in conn.execute(
            'SELECT * FROM checklist_tasks WHERE session_id=? ORDER BY task_order', (session_id,)).fetchall()]
        photos = [dict(r) for r in conn.execute(
            'SELECT * FROM checklist_photos WHERE session_id=? ORDER BY photo_number',
            (session_id,)).fetchall()]
    sess     = dict(sess)
    chk_data = CHECKLISTS.get(sess['type'], {})

    wb = Workbook()
    ws = wb.active
    ws.title = 'Checklist'
    thin  = Side(style='thin',   color='CCCCCC')
    thick = Side(style='medium', color='1B4332')
    border     = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
    border_top = Border(left=thick, right=thick, top=thick, bottom=thick)

    deadline = '10:30 AM' if sess['section'] == 'opening' else '6:30 PM'
    done_count = sum(1 for t in tasks if t['done'])
    total_count = len(tasks)
    pct = round(done_count / total_count * 100) if total_count else 0

    # ── Title block ──────────────────────────────────────────────────────────
    ws['A1'] = 'MCQ MIRRABOOKA CAFE'
    ws['A1'].font = Font(name='Calibri', bold=True, size=18, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='1B4332', end_color='1B4332', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:E1')
    ws.row_dimensions[1].height = 30

    ws['A2'] = f'{chk_data.get("title","").upper()}  —  {sess["section"].upper()} CHECKLIST'
    ws['A2'].font = Font(name='Calibri', bold=True, size=13, color='1B4332')
    ws['A2'].fill = PatternFill(start_color='D8F3DC', end_color='D8F3DC', fill_type='solid')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:E2')
    ws.row_dimensions[2].height = 22

    # ── Info block ───────────────────────────────────────────────────────────
    info_fill = PatternFill(start_color='F0FFF4', end_color='F0FFF4', fill_type='solid')
    label_font = Font(bold=True, color='2D6A4F', size=10)
    value_font = Font(color='333333', size=10)

    def add_info_row(label, value):
        ws.append(['', label, value])
        r = ws.max_row
        ws.cell(r, 2).font = label_font
        ws.cell(r, 3).font = value_font
        ws.cell(r, 2).fill = info_fill
        ws.cell(r, 3).fill = info_fill
        ws.merge_cells(f'C{r}:E{r}')

    ws.append([''])
    add_info_row('Date:',         f'{sess["date"]}  ({sess.get("day_of_week","") or ""})')
    add_info_row('Section:',      sess['section'].title())
    add_info_row('Deadline:',     deadline)
    add_info_row('Submitted at:', sess.get('submitted_at','') or '—')
    add_info_row('Responsible:',  sess.get('responsible','') or '—')
    add_info_row('Submitted by:', sess.get('submitted_by','') or '—')
    if sess.get('general_note'):
        add_info_row('General Note:', sess.get('general_note',''))
    add_info_row('Completion:',   f'{done_count}/{total_count} tasks done ({pct}%)')
    if sess.get('is_late'):
        add_info_row('⚠ Submission:', 'LATE — Submitted after deadline')

    # ── Task table ───────────────────────────────────────────────────────────
    ws.append([''])
    headers = ['#', 'Task Description', 'Status', 'Note']
    ws.append(headers)
    hr = ws.max_row
    col_widths = [4, 58, 14, 35]
    hdr_colors = ['1B4332', '1B4332', '1B4332', '1B4332']
    for col, (h, clr) in enumerate(zip(headers, hdr_colors), 1):
        cell = ws.cell(row=hr, column=col)
        cell.value = h
        cell.fill = PatternFill(start_color=clr, end_color=clr, fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[hr].height = 20

    for t in tasks:
        done_label = '✓  Done' if t['done'] else '✗  Not Done'
        ws.append([t['task_order']+1, t['task_name'], done_label, t.get('note','') or ''])
        row = ws.max_row
        fill_c = 'E8F5E9' if t['done'] else 'FFEBEE'
        stat_c = '2E7D32' if t['done'] else 'C62828'
        for col in range(1, 5):
            c = ws.cell(row=row, column=col)
            c.border = border
            c.font   = Font(size=10)
            c.fill   = PatternFill(start_color=fill_c, end_color=fill_c, fill_type='solid')
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3).font  = Font(bold=True, color=stat_c, size=10)
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws.row_dimensions[row].height = 18

    # ── Summary row ──────────────────────────────────────────────────────────
    ws.append(['', f'TOTAL: {done_count} done, {total_count - done_count} not done  ({pct}% complete)', '', ''])
    sr = ws.max_row
    ws.merge_cells(f'B{sr}:E{sr}')
    sum_clr = '1B4332' if pct == 100 else ('F57F17' if pct >= 50 else 'B71C1C')
    ws.cell(sr, 2).font = Font(bold=True, color='FFFFFF', size=11)
    ws.cell(sr, 2).fill = PatternFill(start_color=sum_clr, end_color=sum_clr, fill_type='solid')
    ws.cell(sr, 2).alignment = Alignment(horizontal='center')

    # ── Manager Verification ─────────────────────────────────────────────────
    ws.append([''])
    ws.append(['', 'MANAGER VERIFICATION', '', ''])
    vr = ws.max_row
    ws.merge_cells(f'B{vr}:E{vr}')
    ws.cell(vr, 2).font = Font(bold=True, size=12, color='1B4332')
    ws.cell(vr, 2).fill = PatternFill(start_color='D8F3DC', end_color='D8F3DC', fill_type='solid')

    def add_verify_row(label, value):
        ws.append(['', label, value])
        r = ws.max_row
        ws.cell(r, 2).font = Font(bold=True, size=10)
        ws.cell(r, 3).font = Font(size=10)
        ws.merge_cells(f'C{r}:E{r}')

    v_status = '✓ Verified' if sess.get('verified') else '⏳ Pending verification'
    add_verify_row('Status:',           v_status)
    add_verify_row('Verified by:',      sess.get('verified_by','') or '—')
    add_verify_row('Verified at:',      sess.get('verified_at','') or '—')
    add_verify_row('Overall result:',   (sess.get('overall_result','') or '—').replace('_',' ').title())
    add_verify_row('Issues found:',     sess.get('issues_found','') or '—')
    add_verify_row('Action:',           sess.get('action_responsible','') or '—')
    add_verify_row('Manager notes:',    sess.get('manager_notes','') or '—')

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 58
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 35

    # ── Photo Evidence sheet ─────────────────────────────────────────────────
    if photos:
        try:
            from openpyxl.drawing.image import Image as XLImage
            from PIL import Image as PILImage
            ph = wb.create_sheet('Photo Evidence')
            ph['A1'] = 'MCQ MIRRABOOKA CAFE  —  PHOTO EVIDENCE'
            ph['A1'].font = Font(bold=True, color='FFFFFF', size=14)
            ph['A1'].fill = PatternFill('solid', fgColor='1B4332')
            ph['A1'].alignment = Alignment(horizontal='center')
            ph.merge_cells('A1:D1')
            ph.row_dimensions[1].height = 28

            ph['A2'] = f'{chk_data.get("title","")} — {sess["section"].title()} — {sess["date"]}'
            ph['A2'].font = Font(italic=True, color='555555', size=10)
            ph['A2'].alignment = Alignment(horizontal='center')
            ph.merge_cells('A2:D2')

            row = 4
            for p in photos:
                src = os.path.join(UPLOAD_FOLDER, p['filename'])
                if not os.path.exists(src):
                    ph.cell(row, 1).value = f"Photo {p['photo_number']+1}: (file missing)"
                    ph.cell(row, 1).font = Font(italic=True, color='C62828')
                    row += 2
                    continue
                # Make a thumbnail to keep the xlsx small + render predictably
                try:
                    thumb_path = os.path.join(UPLOAD_FOLDER, f'_xls_thumb_{p["id"]}.jpg')
                    with PILImage.open(src) as im:
                        im = im.convert('RGB')
                        im.thumbnail((480, 360), PILImage.LANCZOS)
                        im.save(thumb_path, 'JPEG', quality=85)
                    label = f"Photo {p['photo_number']+1} — {p.get('original_name','') or p['filename']}"
                    ph.cell(row, 1).value = label
                    ph.cell(row, 1).font = Font(bold=True, color='1B4332', size=11)
                    ph.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
                    img = XLImage(thumb_path)
                    ph.add_image(img, f'A{row + 1}')
                    # Approx 18 rows tall for a 360px image at default row height
                    row += 22
                except Exception:
                    ph.cell(row, 1).value = f"Photo {p['photo_number']+1}: (could not embed)"
                    row += 2
            for col, width in zip('ABCD', [40, 20, 20, 20]):
                ph.column_dimensions[col].width = width
        except Exception:
            # If Pillow missing or any other failure, skip the photo sheet
            # rather than break the whole export.
            pass

    buf = BytesIO()
    wb.save(buf); buf.seek(0)

    # Clean up thumbnail files created above
    for p in photos:
        thumb = os.path.join(UPLOAD_FOLDER, f'_xls_thumb_{p["id"]}.jpg')
        if os.path.exists(thumb):
            try: os.remove(thumb)
            except OSError: pass

    fname = f'MCQ_Checklist_{sess["type"]}_{sess["section"]}_{sess["date"]}.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)

@app.route('/export/checklist/<int:session_id>/pdf')
@admin_required
def export_checklist_pdf(session_id):
    from html import escape
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    guard, gp = store_guard_clause()
    with get_db() as conn:
        sess = conn.execute(f'SELECT * FROM checklist_sessions WHERE id=? AND {guard}',
                            [session_id] + gp).fetchone()
        if not sess:
            return redirect(url_for('dashboard'))
        tasks = [dict(r) for r in conn.execute(
            'SELECT * FROM checklist_tasks WHERE session_id=? ORDER BY task_order',
            (session_id,)).fetchall()]
        photos = [dict(r) for r in conn.execute(
            'SELECT * FROM checklist_photos WHERE session_id=? ORDER BY photo_number',
            (session_id,)).fetchall()]
        photo_count = len(photos)

    sess = dict(sess)
    chk_data = CHECKLISTS.get(sess['type'], {})
    done_count = sum(1 for t in tasks if t['done'])
    total_count = len(tasks)
    pct = round(done_count / total_count * 100) if total_count else 0
    deadline = '10:30 AM' if sess['section'] == 'opening' else '6:30 PM'
    font_name, bold_font = register_pdf_fonts()

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=12*mm, bottomMargin=12*mm)
    base = getSampleStyleSheet()
    styles = {
        'title': ParagraphStyle('title', parent=base['Title'], fontName=bold_font, fontSize=17,
                                leading=22, textColor=colors.HexColor('#1B3A2D'), alignment=TA_CENTER),
        'sub': ParagraphStyle('sub', parent=base['Normal'], fontName=font_name, fontSize=9,
                              leading=12, textColor=colors.HexColor('#555555'), alignment=TA_CENTER),
        'body': ParagraphStyle('body', parent=base['BodyText'], fontName=font_name, fontSize=8.8,
                               leading=11, textColor=colors.HexColor('#222222')),
        'small': ParagraphStyle('small', parent=base['Normal'], fontName=font_name, fontSize=8,
                                leading=10, textColor=colors.HexColor('#555555')),
    }

    story = [
        Paragraph('MCQ MIRRABOOKA CAFE', styles['title']),
        Paragraph(f"{chk_data.get('title','Checklist')} - {sess['section'].title()} Checklist", styles['sub']),
        Spacer(1, 5*mm),
    ]
    info = [
        ['Date', sess['date'], 'Day', sess.get('day_of_week') or '-'],
        ['Deadline', deadline, 'Submitted at', sess.get('submitted_at') or '-'],
        ['Responsible', sess.get('responsible') or '-', 'Submitted by', sess.get('submitted_by') or '-'],
        ['Completion', f'{done_count}/{total_count} ({pct}%)', 'Photos', str(photo_count)],
        ['Late', 'Yes' if sess.get('is_late') else 'No', 'Verified', 'Yes' if sess.get('verified') else 'No'],
    ]
    info_table = Table(info, colWidths=[28*mm, 56*mm, 28*mm, 56*mm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F7FAF8')),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTNAME', (0, 0), (0, -1), bold_font),
        ('FONTNAME', (2, 0), (2, -1), bold_font),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#DDDDDD')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(info_table)
    if sess.get('general_note'):
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph(f"<b>General Note:</b> {escape(sess.get('general_note') or '')}", styles['small']))
    story.append(Spacer(1, 5*mm))

    task_rows = [['#', 'Task', 'Done', 'Note']]
    for t in tasks:
        task_rows.append([
            str(t['task_order'] + 1),
            Paragraph(escape(t['task_name']), styles['body']),
            'Yes' if t['done'] else 'No',
            Paragraph(escape(t.get('note') or ''), styles['body']),
        ])
    task_table = Table(task_rows, colWidths=[9*mm, 97*mm, 17*mm, 46*mm], repeatRows=1)
    task_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B3A2D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), bold_font),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (2, 0), (2, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#DDDDDD')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FBFCFC')]),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(task_table)
    story.append(Spacer(1, 5*mm))

    verify_rows = [
        ['Verified by', sess.get('verified_by') or '-',
         'Verified at', sess.get('verified_at') or '-'],
        ['Result', (sess.get('overall_result') or '-').replace('_', ' ').title(),
         'Action', sess.get('action_responsible') or '-'],
        ['Manager notes', sess.get('manager_notes') or '-', '', ''],
    ]
    verify_table = Table(verify_rows, colWidths=[28*mm, 56*mm, 28*mm, 56*mm])
    verify_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F0F4F2')),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTNAME', (0, 0), (0, -1), bold_font),
        ('FONTNAME', (2, 0), (2, -1), bold_font),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#DDDDDD')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(verify_table)

    # ── Embed photos at the bottom of the PDF ────────────────────────────────
    if photos:
        from reportlab.platypus import Image, PageBreak
        story.append(PageBreak())
        story.append(Paragraph('Photo Evidence', styles['title']))
        story.append(Paragraph(f"{len(photos)} photo(s) attached to this checklist", styles['sub']))
        story.append(Spacer(1, 5*mm))

        # 2 photos per row, 80mm wide each
        photo_cells = []
        for p in photos:
            path = os.path.join(UPLOAD_FOLDER, p['filename'])
            if not os.path.exists(path):
                photo_cells.append(Paragraph(
                    f"<i>(missing: {escape(p['filename'])})</i>", styles['small']))
                continue
            try:
                img = Image(path)
                # Maintain aspect ratio, target width 80mm
                target_w = 80 * mm
                ratio = target_w / float(img.imageWidth) if img.imageWidth else 1
                img.drawWidth  = target_w
                img.drawHeight = img.imageHeight * ratio
                # Clamp height so a giant portrait photo doesn't blow up the page
                max_h = 110 * mm
                if img.drawHeight > max_h:
                    img.drawHeight = max_h
                    img.drawWidth  = img.imageWidth * (max_h / float(img.imageHeight))
                caption = (
                    f"<b>Photo {p['photo_number'] + 1}</b><br/>"
                    f"<font color='#666'>{escape(p.get('original_name') or '')}</font>"
                )
                cell = [img, Spacer(1, 1*mm), Paragraph(caption, styles['small'])]
                photo_cells.append(cell)
            except Exception as e:
                photo_cells.append(Paragraph(
                    f"<i>(could not embed: {escape(str(e))})</i>", styles['small']))

        # Lay out as a 2-col table
        rows_of_photos = []
        for i in range(0, len(photo_cells), 2):
            row = photo_cells[i:i + 2]
            if len(row) == 1:
                row.append('')
            rows_of_photos.append(row)
        if rows_of_photos:
            photo_table = Table(rows_of_photos, colWidths=[90 * mm, 90 * mm])
            photo_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 2),
                ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            ]))
            story.append(photo_table)

    def footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont(font_name, 8)
        canvas.setFillColor(colors.HexColor('#666666'))
        canvas.drawString(12*mm, 7*mm, 'MCQ Mirrabooka Cafe - Checklist Record')
        canvas.drawRightString(A4[0] - 12*mm, 7*mm, f'Page {doc_obj.page}')
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buf.seek(0)
    fname = f'MCQ_Checklist_{sess["type"]}_{sess["section"]}_{sess["date"]}.pdf'
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=fname)

@app.route('/export/temperature/<int:session_id>/excel')
@admin_required
def export_temperature_excel(session_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    guard, gp = store_guard_clause()
    with get_db() as conn:
        sess = conn.execute(f'SELECT * FROM temp_sessions WHERE id=? AND {guard}',
                            [session_id] + gp).fetchone()
        if not sess: return redirect(url_for('dashboard'))
        readings = [dict(r) for r in conn.execute(
            'SELECT * FROM temp_readings WHERE session_id=? ORDER BY food_order', (session_id,)).fetchall()]
    sess      = dict(sess)
    temp_data = TEMPERATURES.get(sess['type'], {})

    wb = Workbook()
    ws = wb.active
    ws.title = 'Temperature Record'

    ws['A1'] = 'MCQ MIRRABOOKA CAFE — FOOD TEMPERATURE RECORD'
    ws['A1'].font = Font(name='Calibri', bold=True, size=16, color='1B4332')
    ws.merge_cells('A1:M1')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = f'{temp_data.get("title","")}   |   Date: {sess["date"]}'
    ws.merge_cells('A2:M2')
    ws['A3'] = f'Recorded by: {sess.get("recorded_by","") or ""}   |   Checked by: {sess.get("checked_by","") or ""}'
    ws.merge_cells('A3:M3')
    ws['A4'] = 'SAFE HOLDING: COLD ≤ 5°C | ROOM 15–30°C | HOT ≥ 60°C.'
    ws['A4'].font = Font(color='CC0000', italic=True)
    ws.merge_cells('A4:M4')
    ws.append([''])

    headers = ['Food Item',
               'Check 1\nTime','Check 1\nTemp°C',
               'Check 2\nTime','Check 2\nTemp°C',
               'Check 3\nTime','Check 3\nTemp°C',
               'Check 4\nTime','Check 4\nTemp°C',
               'Check 5\nTime','Check 5\nTemp°C',
               'Discarded\n(Y/N)']
    ws.append(headers)
    hr = ws.max_row
    for col in range(1, len(headers)+1):
        c = ws.cell(row=hr, column=col)
        c.fill = PatternFill(start_color='1B4332', end_color='1B4332', fill_type='solid')
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    for r in readings:
        kind = r.get('food_kind') or 'cold'
        row_data = [f'[{TEMP_KIND_RULES.get(kind, TEMP_KIND_RULES["cold"])["label"]}] {r["food_name"]}']
        for n in range(1, 6):
            row_data.append(r.get(f'c{n}_time') or '')
            row_data.append(r.get(f'c{n}_temp') if r.get(f'c{n}_temp') is not None else '')
        row_data.append(r.get('discarded','N') or 'N')
        ws.append(row_data)
        row_num = ws.max_row
        for n in range(1, 6):
            temp_col = (n-1)*2 + 3
            tv = r.get(f'c{n}_temp')
            if tv is not None:
                cell = ws.cell(row=row_num, column=temp_col)
                if sess['type'] == 'pastry' and n == 3:
                    cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
                elif temp_is_unsafe(kind, tv):
                    cell.fill = PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF')
                else:
                    cell.fill = PatternFill(start_color='B2FFB2', end_color='B2FFB2', fill_type='solid')

    ws.column_dimensions['A'].width = 30
    for col in range(2, 13):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.column_dimensions[get_column_letter(12)].width = 14

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f'MCQ_Temperature_{sess["type"]}_{sess["date"]}.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)

@app.route('/export/temperature/<int:session_id>/pdf')
@admin_required
def export_temperature_pdf(session_id):
    from html import escape
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    guard, gp = store_guard_clause()
    with get_db() as conn:
        sess = conn.execute(f'SELECT * FROM temp_sessions WHERE id=? AND {guard}',
                            [session_id] + gp).fetchone()
        if not sess:
            return redirect(url_for('dashboard'))
        readings = [dict(r) for r in conn.execute(
            'SELECT * FROM temp_readings WHERE session_id=? ORDER BY food_order',
            (session_id,)).fetchall()]

    sess = dict(sess)
    temp_data = TEMPERATURES.get(sess['type'], {})
    font_name, bold_font = register_pdf_fonts()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=9*mm, rightMargin=9*mm,
                            topMargin=10*mm, bottomMargin=10*mm)
    base = getSampleStyleSheet()
    styles = {
        'title': ParagraphStyle('title', parent=base['Title'], fontName=bold_font, fontSize=16,
                                leading=20, textColor=colors.HexColor('#1B3A2D'), alignment=TA_CENTER),
        'sub': ParagraphStyle('sub', parent=base['Normal'], fontName=font_name, fontSize=8.5,
                              leading=11, textColor=colors.HexColor('#555555'), alignment=TA_CENTER),
        'food': ParagraphStyle('food', parent=base['BodyText'], fontName=font_name, fontSize=7.5,
                               leading=9.5, textColor=colors.HexColor('#222222')),
    }

    story = [
        Paragraph('MCQ MIRRABOOKA CAFE - FOOD TEMPERATURE RECORD', styles['title']),
        Paragraph(f"{temp_data.get('title','Temperature Record')} | Date: {sess['date']} | "
                  f"Recorded by: {escape(sess.get('recorded_by') or '-')} | "
                  f"Checked by: {escape(sess.get('checked_by') or '-')}", styles['sub']),
        Paragraph('SAFE HOLDING: COLD ≤ 5°C | ROOM 15–30°C | HOT ≥ 60°C',
                  styles['sub']),
        Spacer(1, 4*mm),
    ]

    headers = ['Food Item', 'C1 Time', 'C1 °C', 'C2 Time', 'C2 °C', 'C3 Time', 'C3 °C',
               'C4 Time', 'C4 °C', 'C5 Time', 'C5 °C', 'Discard']
    rows = [headers]
    for r in readings:
        kind = r.get('food_kind') or 'cold'
        kind_label = TEMP_KIND_RULES.get(kind, TEMP_KIND_RULES['cold'])['label']
        row = [Paragraph(f'<b>[{kind_label}]</b> {escape(r["food_name"])}', styles['food'])]
        for n in range(1, 6):
            row.append(r.get(f'c{n}_time') or '')
            temp_val = r.get(f'c{n}_temp')
            row.append(f'{temp_val:g}' if temp_val is not None else '')
        row.append(r.get('discarded') or 'N')
        rows.append(row)

    table = Table(rows, colWidths=[48*mm, 19*mm, 15*mm, 19*mm, 15*mm, 19*mm, 15*mm,
                                   19*mm, 15*mm, 19*mm, 15*mm, 16*mm], repeatRows=1)
    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B3A2D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), bold_font),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#DDDDDD')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FBFCFC')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]
    for row_idx, r in enumerate(readings, 1):
        kind = r.get('food_kind') or 'cold'
        for n in range(1, 6):
            tv = r.get(f'c{n}_temp')
            if tv is None:
                continue
            col = (n - 1) * 2 + 2
            if sess['type'] == 'pastry' and n == 3:
                table_style.append(('BACKGROUND', (col, row_idx), (col, row_idx), colors.HexColor('#FFF3CD')))
            elif temp_is_unsafe(kind, tv):
                table_style.append(('BACKGROUND', (col, row_idx), (col, row_idx), colors.HexColor('#F8D7DA')))
            else:
                table_style.append(('BACKGROUND', (col, row_idx), (col, row_idx), colors.HexColor('#D8F3DC')))
        if (r.get('discarded') or 'N') == 'Y':
            table_style.append(('BACKGROUND', (11, row_idx), (11, row_idx), colors.HexColor('#FFE8A1')))
    table.setStyle(TableStyle(table_style))
    story.append(table)
    if sess.get('notes'):
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph(f"Notes: {escape(sess.get('notes') or '')}", styles['sub']))

    def footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont(font_name, 8)
        canvas.setFillColor(colors.HexColor('#666666'))
        canvas.drawString(9*mm, 6*mm, 'Green = within the item safe range | Red = outside the item safe range')
        canvas.drawRightString(landscape(A4)[0] - 9*mm, 6*mm, f'Page {doc_obj.page}')
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buf.seek(0)
    fname = f'MCQ_Temperature_{sess["type"]}_{sess["date"]}.pdf'
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=fname)

@app.route('/export/bulk/excel')
@admin_required
def export_bulk_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    date_from = request.args.get('date_from', (date.today()-timedelta(days=30)).isoformat())
    date_to   = request.args.get('date_to', date.today().isoformat())

    scope, sp = store_filter_clause('cs')
    tscope, tsp = store_filter_clause()
    with get_db() as conn:
        chk_recs = [dict(r) for r in conn.execute(f'''
            SELECT cs.*,
              (SELECT COUNT(*) FROM checklist_tasks WHERE session_id=cs.id AND done=1) as done_count,
              (SELECT COUNT(*) FROM checklist_tasks WHERE session_id=cs.id) as total_count
            FROM checklist_sessions cs WHERE date BETWEEN ? AND ? AND {scope}
            ORDER BY date DESC, type, section
        ''', [date_from, date_to] + sp).fetchall()]
        temp_recs = [dict(r) for r in conn.execute(
            f'SELECT * FROM temp_sessions WHERE date BETWEEN ? AND ? AND {tscope} ORDER BY date DESC, type',
            [date_from, date_to] + tsp).fetchall()]

    wb = Workbook()
    gfill = PatternFill(start_color='1B4332', end_color='1B4332', fill_type='solid')
    hfont = Font(bold=True, color='FFFFFF', size=11)

    # Sheet 1 – Checklist Summary
    ws1 = wb.active; ws1.title = 'Checklist Summary'
    ws1['A1'] = f'MCQ MIRRABOOKA — Checklist Summary ({date_from} to {date_to})'
    ws1['A1'].font = Font(bold=True, size=14, color='1B4332')
    ws1.merge_cells('A1:K1')
    ws1.append([''])
    hdrs = ['Date','Day','Type','Section','Responsible','Submitted By','Tasks Done','Total Tasks','Completion %','Verified','Issues']
    ws1.append(hdrs)
    for col, h in enumerate(hdrs, 1):
        c = ws1.cell(row=ws1.max_row, column=col)
        c.fill = gfill; c.font = hfont
        c.alignment = Alignment(horizontal='center')
    for r in chk_recs:
        pct = round(r['done_count']/r['total_count']*100) if r['total_count'] > 0 else 0
        ws1.append([
            r['date'], r.get('day_of_week',''),
            CHECKLISTS.get(r['type'],{}).get('title', r['type']),
            r['section'].title(), r.get('responsible','') or '', r.get('submitted_by','') or '',
            r['done_count'], r['total_count'], f'{pct}%',
            'Yes' if r['verified'] else 'No',
            (r.get('overall_result','') or '').replace('_',' ').title(),
        ])
    for col in [1,2,3,4,5,6,7,8,9,10,11]:
        ws1.column_dimensions[chr(64+col)].width = [12,10,30,12,22,22,12,12,14,10,16][col-1]

    # Sheet 2 – Temperature Summary
    ws2 = wb.create_sheet('Temperature Summary')
    ws2['A1'] = f'MCQ MIRRABOOKA — Temperature Summary ({date_from} to {date_to})'
    ws2['A1'].font = Font(bold=True, size=14, color='1B4332')
    ws2.merge_cells('A1:F1')
    ws2.append([''])
    hdrs2 = ['Date','Type','Recorded by','Checked by','Submitted At','Notes']
    ws2.append(hdrs2)
    for col, h in enumerate(hdrs2, 1):
        c = ws2.cell(row=ws2.max_row, column=col)
        c.fill = gfill; c.font = hfont
    for r in temp_recs:
        ws2.append([
            r['date'], TEMPERATURES.get(r['type'],{}).get('title', r['type']),
            r.get('recorded_by','') or '', r.get('checked_by','') or '',
            r.get('submitted_at','') or '', r.get('notes','') or '',
        ])
    for col, w in enumerate([12,35,22,22,20,30], 1):
        ws2.column_dimensions[chr(64+col)].width = w

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    fname = f'MCQ_Records_{date_from}_to_{date_to}.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)

@app.route('/export/bulk/pdf')
@admin_required
def export_bulk_pdf():
    from html import escape
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    date_from = request.args.get('date_from', (date.today()-timedelta(days=30)).isoformat())
    date_to   = request.args.get('date_to', date.today().isoformat())

    scope, sp = store_filter_clause('cs')
    tscope, tsp = store_filter_clause()
    with get_db() as conn:
        chk_recs = [dict(r) for r in conn.execute(f'''
            SELECT cs.*,
              (SELECT COUNT(*) FROM checklist_tasks WHERE session_id=cs.id AND done=1) as done_count,
              (SELECT COUNT(*) FROM checklist_tasks WHERE session_id=cs.id) as total_count
            FROM checklist_sessions cs WHERE date BETWEEN ? AND ? AND {scope}
            ORDER BY date DESC, type, section
        ''', [date_from, date_to] + sp).fetchall()]
        temp_recs = [dict(r) for r in conn.execute(f'''
            SELECT * FROM temp_sessions
            WHERE date BETWEEN ? AND ? AND {tscope}
            ORDER BY date DESC, type
        ''', [date_from, date_to] + tsp).fetchall()]

    font_name, bold_font = register_pdf_fonts()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=12*mm, bottomMargin=12*mm)
    base = getSampleStyleSheet()
    styles = {
        'title': ParagraphStyle('title', parent=base['Title'], fontName=bold_font, fontSize=18,
                                leading=23, textColor=colors.HexColor('#1B3A2D'), alignment=TA_CENTER),
        'sub': ParagraphStyle('sub', parent=base['Normal'], fontName=font_name, fontSize=9,
                              leading=12, textColor=colors.HexColor('#555555'), alignment=TA_CENTER),
        'body': ParagraphStyle('body', parent=base['BodyText'], fontName=font_name, fontSize=7.8,
                               leading=9.5, textColor=colors.HexColor('#222222')),
    }

    story = [
        Paragraph('MCQ MIRRABOOKA CAFE', styles['title']),
        Paragraph(f'Record Summary: {date_from} to {date_to}', styles['sub']),
        Spacer(1, 5*mm),
        Paragraph(f'Checklist records: {len(chk_recs)} | Temperature records: {len(temp_recs)}', styles['sub']),
        Spacer(1, 7*mm),
    ]

    chk_rows = [['Date', 'Type', 'Section', 'Submitted by', 'Progress', 'Verified']]
    for r in chk_recs:
        chk = CHECKLISTS.get(r['type'], {})
        total = r.get('total_count') or 0
        pct = round((r.get('done_count') or 0) / total * 100) if total else 0
        chk_rows.append([
            r['date'],
            Paragraph(escape(chk.get('short', r['type'])), styles['body']),
            r['section'].title(),
            Paragraph(escape(r.get('submitted_by') or '-'), styles['body']),
            f"{r.get('done_count') or 0}/{total} ({pct}%)",
            'Yes' if r.get('verified') else 'No',
        ])
    story.append(Paragraph('Daily Checklists', styles['sub']))
    chk_table = Table(chk_rows, colWidths=[24*mm, 36*mm, 24*mm, 50*mm, 27*mm, 18*mm], repeatRows=1)
    chk_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B3A2D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), bold_font),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 7.8),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#DDDDDD')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FBFCFC')]),
    ]))
    story.append(chk_table)
    story.append(PageBreak())

    temp_rows = [['Date', 'Type', 'Recorded by', 'Checked by', 'Submitted at', 'Notes']]
    for r in temp_recs:
        temp = TEMPERATURES.get(r['type'], {})
        temp_rows.append([
            r['date'],
            Paragraph(escape(temp.get('short', r['type'])), styles['body']),
            Paragraph(escape(r.get('recorded_by') or '-'), styles['body']),
            Paragraph(escape(r.get('checked_by') or '-'), styles['body']),
            r.get('submitted_at') or '-',
            Paragraph(escape(r.get('notes') or ''), styles['body']),
        ])
    story.append(Paragraph('Temperature Records', styles['sub']))
    temp_table = Table(temp_rows, colWidths=[24*mm, 28*mm, 38*mm, 38*mm, 35*mm, 28*mm], repeatRows=1)
    temp_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B3A2D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), bold_font),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 7.8),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#DDDDDD')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FBFCFC')]),
    ]))
    story.append(temp_table)

    def footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont(font_name, 8)
        canvas.setFillColor(colors.HexColor('#666666'))
        canvas.drawString(10*mm, 7*mm, 'MCQ Mirrabooka Cafe - Record Summary')
        canvas.drawRightString(A4[0] - 10*mm, 7*mm, f'Page {doc_obj.page}')
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buf.seek(0)
    fname = f'MCQ_Records_{date_from}_to_{date_to}.pdf'
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=fname)

# ─── Staff API ─────────────────────────────────────────────────────────────────

@app.route('/api/staff')
@login_required
def api_staff():
    # Pickers always need a concrete store → use the session store.
    with get_db() as conn:
        staff = [dict(r) for r in conn.execute(
            'SELECT * FROM staff_members WHERE active=1 AND store_id=? ORDER BY name',
            (current_store_id(),)).fetchall()]
    return jsonify(staff)

@app.route('/staff', methods=['GET'])
@login_required
def staff_page():
    clause, params = store_filter_clause()
    with get_db() as conn:
        staff = [dict(r) for r in conn.execute(
            f'SELECT * FROM staff_members WHERE {clause} ORDER BY active DESC, name',
            params).fetchall()]
        birthdays = birthday_rows(conn)
    birthday_map = {b['staff_name']: b for b in birthdays}
    return render_template('staff.html', staff=staff, birthdays=birthdays,
                           birthday_map=birthday_map, stores=get_stores())

@app.route('/staff/add', methods=['POST'])
@login_required
def staff_add():
    name = request.form.get('name','').strip()
    role = request.form.get('role','').strip()
    if name:
        with get_db() as conn:
            conn.execute('INSERT OR IGNORE INTO staff_members (name,role,store_id) VALUES (?,?,?)',
                         (name, role, current_store_id()))
    return redirect(url_for('staff_page'))

@app.route('/staff/toggle/<int:staff_id>', methods=['POST'])
@login_required
def staff_toggle(staff_id):
    guard, gp = store_guard_clause()
    with get_db() as conn:
        current = conn.execute(
            f'SELECT active FROM staff_members WHERE id=? AND {guard}',
            [staff_id] + gp).fetchone()
        if current:
            conn.execute(f'UPDATE staff_members SET active=? WHERE id=? AND {guard}',
                         [0 if current['active'] else 1, staff_id] + gp)
    return redirect(url_for('staff_page'))

@app.route('/staff/delete/<int:staff_id>', methods=['POST'])
@admin_required
def staff_delete(staff_id):
    """Permanently remove a staff member. Past records keep the name as plain
    text, so history is preserved, but the name disappears from every live staff
    picker (all of which read staff_members)."""
    guard, gp = store_guard_clause()
    with get_db() as conn:
        row = conn.execute(f'SELECT name FROM staff_members WHERE id=? AND {guard}',
                           [staff_id] + gp).fetchone()
        conn.execute(f'DELETE FROM staff_members WHERE id=? AND {guard}', [staff_id] + gp)
    flash(f"Removed {row['name'] if row else 'staff member'} from the team.", 'success')
    return redirect(url_for('staff_page'))

@app.route('/staff/<int:staff_id>/profile')
@admin_required
def staff_profile(staff_id):
    guard, gp = store_guard_clause()
    with get_db() as conn:
        member = conn.execute(f'SELECT * FROM staff_members WHERE id=? AND {guard}',
                              [staff_id] + gp).fetchone()
        if not member:
            flash('Staff member not found.', 'danger')
            return redirect(url_for('staff_page'))
        member = dict(member)
        reward_month = date.today().strftime('%Y-%m')
        reward_data = calculate_monthly_reward_scores(conn, reward_month, store_id=member.get('store_id'))
        my_score = next((r for r in reward_data['employee_rank'] if r['staff_name'] == member['name']), None)
        all_birthdays = birthday_rows(conn, store_id=member.get('store_id'))
        birthday = next((b for b in all_birthdays if b['staff_name'] == member['name']), None)
    return render_template('staff_profile.html',
        member=member, my_score=my_score,
        reward_month=reward_month, birthday=birthday)

@app.route('/staff/<int:staff_id>/profile/update', methods=['POST'])
@admin_required
def staff_profile_update(staff_id):
    guard, gp = store_guard_clause()
    with get_db() as conn:
        conn.execute(f'''UPDATE staff_members
            SET phone=?, email=?, emergency_contact=?, staff_notes=?, role=?
            WHERE id=? AND {guard}''',
            [request.form.get('phone','').strip(),
             request.form.get('email','').strip(),
             request.form.get('emergency_contact','').strip(),
             request.form.get('staff_notes','').strip(),
             request.form.get('role','').strip(),
             staff_id] + gp)
    flash('Profile updated.', 'success')
    return redirect(url_for('staff_profile', staff_id=staff_id))

# ─── Food Safety Certificates ───────────────────────────────────────────────────

CERT_EXT = {'jpg', 'jpeg', 'png', 'gif', 'webp', 'heic', 'heif', 'pdf'}

@app.route('/admin/certificates')
@admin_required
def certificates():
    today_iso = date.today().isoformat()
    staff = get_active_staff()
    scope, sp = store_filter_clause()
    with get_db() as conn:
        rows = [dict(r) for r in conn.execute(
            f'SELECT * FROM staff_certificates WHERE {scope} ORDER BY staff_name, uploaded_at DESC',
            sp).fetchall()]
    # Group certificates by staff member
    by_staff = {}
    for r in rows:
        r['is_pdf'] = (r['filename'] or '').lower().endswith('.pdf')
        r['expired'] = bool(r.get('expiry_date') and r['expiry_date'] < today_iso)
        r['expiring_soon'] = bool(
            r.get('expiry_date') and not r['expired']
            and r['expiry_date'] <= (date.today() + timedelta(days=30)).isoformat())
        by_staff.setdefault(r['staff_name'], []).append(r)
    return render_template('certificates.html',
        staff=staff, by_staff=by_staff, today=today_iso,
        with_cert=len(by_staff), total_staff=len(staff))

@app.route('/admin/certificates/upload', methods=['POST'])
@admin_required
def certificate_upload():
    staff_name = request.form.get('staff_name', '').strip()
    cert_type  = request.form.get('cert_type', '').strip() or 'Food Safety Certificate'
    expiry     = request.form.get('expiry_date', '').strip()
    notes      = request.form.get('notes', '').strip()
    f = request.files.get('certificate')
    if not staff_name:
        flash('Please choose a staff member.', 'warning')
        return redirect(url_for('certificates'))
    if not f or not f.filename:
        flash('Please choose a certificate file to upload.', 'warning')
        return redirect(url_for('certificates'))
    ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
    if ext not in CERT_EXT:
        flash('Unsupported file type. Use an image (JPG/PNG) or PDF.', 'danger')
        return redirect(url_for('certificates'))
    fname = f"cert_{secure_filename(staff_name)}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
    dest = os.path.join(UPLOAD_FOLDER, fname)
    try:
        if ext == 'pdf':
            f.save(dest)
        else:
            # Re-encodes images as JPEG and may change the extension — keep the
            # actual filename it writes so the file can be served back.
            fname = save_uploaded_photo(f, dest, max_dim=1800, quality=85)
    except Exception:
        f.save(dest)
    with get_db() as conn:
        conn.execute('''INSERT INTO staff_certificates
            (staff_name, cert_type, filename, original_name, expiry_date, notes, uploaded_by, store_id)
            VALUES (?,?,?,?,?,?,?,?)''',
            (staff_name, cert_type, fname, secure_filename(f.filename),
             expiry, notes, session.get('role', 'admin'), current_store_id()))
    flash(f'Certificate uploaded for {staff_name}.', 'success')
    return redirect(url_for('certificates'))

@app.route('/admin/certificate/<int:cert_id>/file')
@login_required
def certificate_file(cert_id):
    guard, gp = store_guard_clause()
    with get_db() as conn:
        row = conn.execute(f'SELECT filename FROM staff_certificates WHERE id=? AND {guard}',
                           [cert_id] + gp).fetchone()
    if not row:
        abort(404)
    return send_from_directory(UPLOAD_FOLDER, os.path.basename(row['filename']))

@app.route('/admin/certificate/<int:cert_id>/delete', methods=['POST'])
@admin_required
def certificate_delete(cert_id):
    guard, gp = store_guard_clause()
    with get_db() as conn:
        row = conn.execute(f'SELECT filename FROM staff_certificates WHERE id=? AND {guard}',
                           [cert_id] + gp).fetchone()
        if row:
            try:
                os.remove(os.path.join(UPLOAD_FOLDER, row['filename']))
            except OSError:
                pass
            conn.execute(f'DELETE FROM staff_certificates WHERE id=? AND {guard}', [cert_id] + gp)
    flash('Certificate deleted.', 'warning')
    return redirect(url_for('certificates'))

# ─── Photos ────────────────────────────────────────────────────────────────────

@app.route('/photo/<path:filename>')
@admin_required
def serve_photo(filename):
    # Photo filenames are immutable (they embed a hash), so the browser can cache
    # them hard. This stops the checklist/share pages from re-downloading dozens
    # of images on every view — a big load cut for the shared worker pool. The
    # response is marked `private` because the route is behind admin auth.
    safe = os.path.basename(filename)
    resp = send_from_directory(UPLOAD_FOLDER, safe, max_age=604800)  # 7 days
    resp.headers['Cache-Control'] = 'private, max-age=604800, immutable'
    return resp

@app.route('/admin/photos')
@admin_required
def photo_gallery():
    date_filter = request.args.get('date', date.today().isoformat())
    chk_type    = request.args.get('type', 'all')
    with get_db() as conn:
        q = '''SELECT cp.*, cs.type, cs.section, cs.date, cs.submitted_by, cs.day_of_week
               FROM checklist_photos cp
               JOIN checklist_sessions cs ON cs.id = cp.session_id
               WHERE cs.date = ?'''
        params = [date_filter]
        if chk_type != 'all':
            q += ' AND cs.type=?'; params.append(chk_type)
        q += ' ORDER BY cp.session_id, cp.photo_number'
        photos = [dict(r) for r in conn.execute(q, params).fetchall()]

        # Group by session
        sessions = {}
        for ph in photos:
            key = ph['session_id']
            if key not in sessions:
                sessions[key] = {'info': ph, 'photos': []}
            sessions[key]['photos'].append(ph)

        # Recent dates with photos
        dates_with_photos = [dict(r) for r in conn.execute('''
            SELECT DISTINCT cs.date FROM checklist_photos cp
            JOIN checklist_sessions cs ON cs.id = cp.session_id
            ORDER BY cs.date DESC LIMIT 30
        ''').fetchall()]

    return render_template('photo_gallery.html',
        sessions=sessions, date_filter=date_filter,
        chk_type=chk_type, dates_with_photos=dates_with_photos,
    )

# ─── Today's Timeline API ──────────────────────────────────────────────────────

@app.route('/api/today-timeline')
@login_required
def today_timeline():
    target = request.args.get('date', date.today().isoformat())
    with get_db() as conn:
        rows = conn.execute('''
            SELECT cs.id, cs.type, cs.section, cs.submitted_at, cs.submitted_by,
                   cs.responsible, cs.verified, cs.verified_at, cs.verified_by,
                   cs.day_of_week,
                   (SELECT COUNT(*) FROM checklist_tasks WHERE session_id=cs.id AND done=1) as done_count,
                   (SELECT COUNT(*) FROM checklist_tasks WHERE session_id=cs.id) as total_count,
                   (SELECT COUNT(*) FROM checklist_photos WHERE session_id=cs.id) as photo_count
            FROM checklist_sessions cs
            WHERE cs.date=?
            ORDER BY cs.submitted_at
        ''', (target,)).fetchall()

        temp_rows = conn.execute('''
            SELECT id, type, submitted_at, recorded_by, checked_by
            FROM temp_sessions WHERE date=?
            ORDER BY submitted_at
        ''', (target,)).fetchall()

    events = []
    for r in rows:
        r = dict(r)
        pct = round(r['done_count']/r['total_count']*100) if r['total_count'] else 0
        events.append({
            'kind': 'checklist',
            'id': r['id'],
            'type': r['type'],
            'type_label': CHECKLISTS.get(r['type'], {}).get('short', r['type']),
            'type_color': CHECKLISTS.get(r['type'], {}).get('color', '#888'),
            'section': r['section'],
            'time': r['submitted_at'][-8:-3] if r['submitted_at'] else '—',
            'by': r['submitted_by'] or r['responsible'] or '—',
            'done_pct': pct,
            'done_count': r['done_count'],
            'total_count': r['total_count'],
            'photo_count': r['photo_count'],
            'verified': bool(r['verified']),
            'verified_by': r['verified_by'] or '',
            'verified_at': r['verified_at'][-8:-3] if r['verified_at'] else '',
        })
    for r in temp_rows:
        r = dict(r)
        events.append({
            'kind': 'temperature',
            'id': r['id'],
            'type': r['type'],
            'type_label': TEMPERATURES.get(r['type'], {}).get('short', r['type']),
            'type_color': TEMPERATURES.get(r['type'], {}).get('color', '#888'),
            'section': 'temp',
            'time': r['submitted_at'][-8:-3] if r['submitted_at'] else '—',
            'by': r['recorded_by'] or '—',
            'done_pct': 100,
            'photo_count': 0,
            'verified': False,
        })

    return jsonify({'date': target, 'events': events})

# ─── Compliance report (admin) ─────────────────────────────────────────────────

@app.route('/admin/compliance')
@admin_required
def compliance_report():
    days = int(request.args.get('days', 30))
    start = (date.today()-timedelta(days=days-1)).isoformat()
    end   = date.today().isoformat()

    with get_db() as conn:
        per_staff = conn.execute('''
            SELECT submitted_by as name,
                   COUNT(*) as total,
                   SUM(CASE WHEN verified=1 THEN 1 ELSE 0 END) as verified,
                   SUM(CASE WHEN overall_result='issues_found' THEN 1 ELSE 0 END) as issues,
                   MAX(submitted_at) as last_submission
            FROM checklist_sessions
            WHERE date BETWEEN ? AND ? AND submitted_by != ''
            GROUP BY submitted_by ORDER BY total DESC
        ''', (start, end)).fetchall()

        per_type = conn.execute('''
            SELECT type,
                   SUM(CASE WHEN section='opening' THEN 1 ELSE 0 END) as openings,
                   SUM(CASE WHEN section='closing' THEN 1 ELSE 0 END) as closings,
                   SUM(CASE WHEN verified=1 THEN 1 ELSE 0 END) as verified,
                   COUNT(*) as total
            FROM checklist_sessions WHERE date BETWEEN ? AND ?
            GROUP BY type
        ''', (start, end)).fetchall()

        photo_stats = conn.execute('''
            SELECT cs.submitted_by,
                   COUNT(DISTINCT cs.id) as sessions_with_photos,
                   COUNT(cp.id) as total_photos
            FROM checklist_sessions cs
            LEFT JOIN checklist_photos cp ON cp.session_id = cs.id
            WHERE cs.date BETWEEN ? AND ? AND cs.submitted_by != ''
            GROUP BY cs.submitted_by ORDER BY total_photos DESC
        ''', (start, end)).fetchall()

    return render_template('compliance.html',
        per_staff=[dict(r) for r in per_staff],
        per_type=[dict(r) for r in per_type],
        photo_stats=[dict(r) for r in photo_stats],
        days=days, start=start, end=end,
        checklists=CHECKLISTS,
    )

# ─── Violation Rules ───────────────────────────────────────────────────────────

# ── Violation strike tracking ────────────────────────────────────────────────
# Active strikes count violations in a rolling window (resets after 6 months);
# full history is always kept. 3+ active strikes flags a staff member for review.
STRIKE_WINDOW_DAYS = 180
STRIKE_THRESHOLD   = 3
SEVERITY_ORDER = {'critical': 0, 'serious': 1, 'moderate': 2, 'minor': 3}
SEVERITY_META = {
    'critical': {'label': 'Critical', 'color': '#B71C1C'},
    'serious':  {'label': 'Serious',  'color': '#E65100'},
    'moderate': {'label': 'Moderate', 'color': '#F9A825'},
    'minor':    {'label': 'Minor',    'color': '#1565C0'},
}
# Disciplinary escalation ladder. The step is chosen MANUALLY when a violation
# is logged (not auto-escalated) — staff/manager pick where this incident sits.
WARNING_STEPS = ['Verbal Discussion', 'Written Warning', 'Final Warning', 'Termination Referral']
WARNING_STEP_META = {
    'Verbal Discussion':   {'color': '#1565C0', 'short': 'Verbal'},
    'Written Warning':     {'color': '#F9A825', 'short': 'Written'},
    'Final Warning':       {'color': '#E65100', 'short': 'Final'},
    'Termination Referral':{'color': '#B71C1C', 'short': 'Termination'},
}


def staff_active_strikes(conn, staff_name, window_days=STRIKE_WINDOW_DAYS, store_id=None):
    """Number of violations for one staff member within the rolling window.
    Scoped to one store when store_id is given."""
    cutoff = (date.today() - timedelta(days=window_days)).isoformat()
    q = 'SELECT COUNT(*) FROM staff_violations WHERE staff_name=? AND incident_date>=?'
    p = [staff_name, cutoff]
    if store_id is not None:
        q += ' AND store_id=?'; p.append(store_id)
    row = conn.execute(q, p).fetchone()
    return int(row[0] if row else 0)


def staff_strike_standings(conn, window_days=STRIKE_WINDOW_DAYS, store_id=None):
    """Per-staff strike standings: active count (rolling window) + full history,
    grouped by staff, each staff's violations sorted most-severe → least.
    Scoped to one store when store_id is given; None = all stores."""
    cutoff = (date.today() - timedelta(days=window_days)).isoformat()
    where = '' if store_id is None else 'WHERE sv.store_id=?'
    params = [] if store_id is None else [store_id]
    rows = [dict(r) for r in conn.execute(f'''
        SELECT sv.*, COALESCE(vr.title, '(rule removed)') AS rule_title,
               COALESCE(vr.color, '#607D8B') AS rule_color
        FROM staff_violations sv
        LEFT JOIN violation_rules vr ON vr.id = sv.rule_id
        {where}
        ORDER BY sv.incident_date DESC, sv.id DESC''', params).fetchall()]
    by_staff = {}
    for r in rows:
        name = (r.get('staff_name') or '').strip()
        if not name:
            continue
        sev = r.get('severity') or 'minor'
        if sev not in SEVERITY_ORDER:
            sev = 'minor'
        r['severity'] = sev
        # A strike is active only while the case is still OPEN and within window.
        _st = (r.get('status') or 'open').lower()
        r['is_active'] = (_st not in ('closed', 'cancelled', 'resolved', 'void')
                          and (r.get('incident_date') or '') >= cutoff)
        st = by_staff.setdefault(name, {
            'staff_name': name, 'active_count': 0, 'total_count': 0,
            'active_severity': {k: 0 for k in SEVERITY_ORDER},
            'total_severity':  {k: 0 for k in SEVERITY_ORDER},
            'last_incident': '', 'violations': []})
        st['violations'].append(r)
        st['total_count'] += 1
        st['total_severity'][sev] += 1
        if r['is_active']:
            st['active_count'] += 1
            st['active_severity'][sev] += 1
        if (r.get('incident_date') or '') > st['last_incident']:
            st['last_incident'] = r.get('incident_date') or ''
    standings = []
    for st in by_staff.values():
        # severity desc, then (stable) date desc as appended
        st['violations'].sort(key=lambda v: SEVERITY_ORDER.get(v.get('severity'), 3))
        a = st['active_count']
        st['tier'] = ('critical' if a >= STRIKE_THRESHOLD
                      else 'warning' if a == STRIKE_THRESHOLD - 1 else 'ok')
        st['worst_active'] = next(
            (s for s in ('critical', 'serious', 'moderate', 'minor')
             if st['active_severity'][s] > 0), None)
        standings.append(st)
    tier_rank = {'critical': 0, 'warning': 1, 'ok': 2}
    standings.sort(key=lambda s: (tier_rank[s['tier']], -s['active_count'],
                                  -s['total_count'], s['staff_name'].lower()))
    return standings


def auto_close_violations(conn, window_days=STRIKE_WINDOW_DAYS, store_id=None):
    """Per staff (within a store), once their FIRST open case is older than the
    window (6 months), close all their still-open cases — the cycle resets.
    A staff member's disciplinary cycle is per store (store_id given), so the
    same person at two stores is tracked independently."""
    today = date.today()
    q = ("SELECT staff_name, MIN(incident_date) AS first FROM staff_violations "
         "WHERE status='open'")
    p = []
    if store_id is not None:
        q += ' AND store_id=?'; p.append(store_id)
    q += ' GROUP BY staff_name'
    rows = conn.execute(q, p).fetchall()
    for r in rows:
        try:
            first = datetime.strptime(r['first'], '%Y-%m-%d').date()
        except (TypeError, ValueError):
            continue
        if (today - first).days >= window_days:
            uq = "UPDATE staff_violations SET status='closed' WHERE staff_name=? AND status='open'"
            up = [r['staff_name']]
            if store_id is not None:
                uq += ' AND store_id=?'; up.append(store_id)
            conn.execute(uq, up)


@app.route('/admin/violations', methods=['GET', 'POST'])
@admin_required
def violation_rules():
    today_iso = date.today().isoformat()
    if request.method == 'POST':
        rule_id       = request.form.get('rule_id', '').strip()
        staff_name    = request.form.get('staff_name', '').strip()
        incident_date = request.form.get('incident_date', today_iso).strip()
        incident_time = request.form.get('incident_time', datetime.now().strftime('%H:%M')).strip()
        submitted_by  = request.form.get('submitted_by', '').strip()
        description   = request.form.get('description', '').strip()
        action_taken  = request.form.get('action_taken', '').strip()
        follow_up     = request.form.get('follow_up_date', '').strip()
        warning_step  = request.form.get('warning_step', '').strip()
        if warning_step not in WARNING_STEPS:
            warning_step = WARNING_STEPS[0]

        try:
            rule_id_int = int(rule_id)
        except (TypeError, ValueError):
            rule_id_int = None

        if staff_name and rule_id_int and description:
            with get_db() as conn:
                rule = conn.execute(
                    'SELECT * FROM violation_rules WHERE id=? AND active=1',
                    (rule_id_int,)).fetchone()
                if rule:
                    severity = request.form.get('severity', '').strip() or rule['severity'] or 'minor'
                    if not action_taken:
                        action_taken = rule['default_action'] or ''
                    cur = conn.execute('''INSERT INTO staff_violations
                        (rule_id,staff_name,incident_date,incident_time,submitted_by,branch,
                         severity,description,action_taken,follow_up_date,status,warning_step,store_id)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                        (rule_id_int, staff_name, incident_date, incident_time, submitted_by,
                         session.get('branch', ''), severity, description, action_taken,
                         follow_up, 'open', warning_step, current_store_id()))
                    vid = cur.lastrowid
                    email_service.send_notification(
                        'violation',
                        subject=f'Violation logged: {rule["title"]} — {staff_name}',
                        lines=[
                            f'Staff: {staff_name}',
                            f'Rule: {rule["title"]}',
                            f'Category: {rule["category"]}',
                            f'Severity: {severity}',
                            f'Warning step: {warning_step}',
                            f'Date: {incident_date} {incident_time}',
                            f'Branch: {session.get("branch", "-")}',
                            f'Description: {description}',
                            f'Action taken: {action_taken or "-"}',
                            f'Follow-up date: {follow_up or "-"}',
                        ],
                        link_path='/admin/violations',
                        actor=submitted_by,
                    )
                    # 3-strike escalation: alert when this pushes the staff to the
                    # review threshold within the rolling window (this store only).
                    active = staff_active_strikes(conn, staff_name, store_id=current_store_id())
                    if active >= STRIKE_THRESHOLD:
                        email_service.send_notification(
                            'violation',
                            subject=f'REVIEW REQUIRED: {staff_name} has {active} violations in {STRIKE_WINDOW_DAYS} days',
                            lines=[
                                f'Staff: {staff_name}',
                                f'Active strikes (last {STRIKE_WINDOW_DAYS} days): {active}',
                                f'Latest: {rule["title"]} ({severity}) on {incident_date}',
                                f'Policy: {STRIKE_THRESHOLD}+ violations within the window → review for termination.',
                            ],
                            link_path='/admin/violations',
                            actor=submitted_by,
                        )
                        flash(f'⚠ {staff_name} now has {active} active violations (last '
                              f'{STRIKE_WINDOW_DAYS} days) — flagged for review.', 'danger')
                    else:
                        flash(f'Violation logged for {staff_name}. Active strikes: {active} '
                              f'(last {STRIKE_WINDOW_DAYS} days).', 'success')
            return redirect(url_for('violation_rules'))

    return _render_violations_page('records')


@app.route('/admin/violations/new')
@admin_required
def violation_new():
    """Create a new disciplinary case."""
    return _render_violations_page('new')


@app.route('/admin/violations/stats')
@admin_required
def violation_statistics():
    """Statistics dashboard + staff strike standings."""
    return _render_violations_page('stats')


def _render_violations_page(view):
    today_iso = date.today().isoformat()
    status_filter = request.args.get('status', 'open')
    staff_filter  = request.args.get('staff', '')
    severity_filter = request.args.get('severity', '')
    date_from = request.args.get('date_from', (date.today()-timedelta(days=30)).isoformat())
    date_to   = request.args.get('date_to', today_iso)

    scope_id = selected_store_scope()          # int store_id, or None = all stores
    vscope, vsp = store_filter_clause('sv')     # 'sv.store_id = ?' or '1=1'
    with get_db() as conn:
        # Auto-close cases once a person's first open case is >6 months old —
        # always per store so the same person at two stores is independent.
        if scope_id is None:
            for s in conn.execute('SELECT id FROM stores').fetchall():
                auto_close_violations(conn, store_id=s['id'])
        else:
            auto_close_violations(conn, store_id=scope_id)
        rules = [dict(r) for r in conn.execute('''
            SELECT * FROM violation_rules
            WHERE active=1
            ORDER BY sort_order, title
        ''').fetchall()]

        q = f'''SELECT sv.*, vr.title as rule_title, vr.category as rule_category,
                      vr.color as rule_color, vr.icon as rule_icon,
                      vr.default_action as rule_default_action
               FROM staff_violations sv
               LEFT JOIN violation_rules vr ON vr.id=sv.rule_id
               WHERE sv.incident_date BETWEEN ? AND ? AND {vscope}'''
        params = [date_from, date_to] + list(vsp)
        if status_filter != 'all':
            q += ' AND sv.status=?'; params.append(status_filter)
        if staff_filter:
            q += ' AND sv.staff_name=?'; params.append(staff_filter)
        if severity_filter:
            q += ' AND sv.severity=?'; params.append(severity_filter)
        q += ' ORDER BY sv.submitted_at DESC, sv.id DESC'
        violations = [dict(r) for r in conn.execute(q, params).fetchall()]

        counts = dict(conn.execute(
            f'SELECT status, COUNT(*) FROM staff_violations sv WHERE {vscope} GROUP BY status',
            list(vsp)).fetchall())
        severity_counts = dict(conn.execute(
            f'SELECT severity, COUNT(*) FROM staff_violations sv WHERE {vscope} GROUP BY severity',
            list(vsp)).fetchall())

        stats_where = ['sv.incident_date BETWEEN ? AND ?', vscope]
        stats_params = [date_from, date_to] + list(vsp)
        if staff_filter:
            stats_where.append('sv.staff_name=?')
            stats_params.append(staff_filter)
        if severity_filter:
            stats_where.append('sv.severity=?')
            stats_params.append(severity_filter)
        stats_sql = ' AND '.join(stats_where)

        summary_row = conn.execute(f'''
            SELECT
                COUNT(*) as total,
                COALESCE(SUM(CASE WHEN sv.status='open' THEN 1 ELSE 0 END), 0) as open_count,
                COALESCE(SUM(CASE WHEN sv.status='cancelled' THEN 1 ELSE 0 END), 0) as cancelled_count,
                COALESCE(SUM(CASE WHEN sv.status='closed' THEN 1 ELSE 0 END), 0) as closed_count,
                COALESCE(SUM(CASE WHEN sv.severity IN ('serious','critical') THEN 1 ELSE 0 END), 0) as serious_count,
                COALESCE(SUM(CASE WHEN sv.incident_date=? THEN 1 ELSE 0 END), 0) as today_count
            FROM staff_violations sv
            WHERE {stats_sql}
        ''', [today_iso] + stats_params).fetchone()
        summary = dict(summary_row)
        summary['closed_rate'] = round(
            (summary['closed_count'] / summary['total'] * 100), 1
        ) if summary['total'] else 0

        month_start = date.today().replace(day=1).isoformat()
        month_where = ['sv.incident_date BETWEEN ? AND ?', vscope]
        month_params = [month_start, today_iso] + list(vsp)
        if staff_filter:
            month_where.append('sv.staff_name=?')
            month_params.append(staff_filter)
        if severity_filter:
            month_where.append('sv.severity=?')
            month_params.append(severity_filter)
        month_sql = ' AND '.join(month_where)
        summary['month_count'] = conn.execute(f'''
            SELECT COUNT(*) as c
            FROM staff_violations sv
            WHERE {month_sql}
        ''', month_params).fetchone()['c']

        status_breakdown = [dict(r) for r in conn.execute(f'''
            SELECT sv.status, COUNT(*) as count
            FROM staff_violations sv
            WHERE {stats_sql}
            GROUP BY sv.status
            ORDER BY CASE sv.status
                WHEN 'open' THEN 1
                WHEN 'closed' THEN 2
                WHEN 'cancelled' THEN 3
                ELSE 4 END
        ''', stats_params).fetchall()]

        severity_breakdown = [dict(r) for r in conn.execute(f'''
            SELECT sv.severity, COUNT(*) as count
            FROM staff_violations sv
            WHERE {stats_sql}
            GROUP BY sv.severity
            ORDER BY CASE sv.severity
                WHEN 'minor' THEN 1
                WHEN 'moderate' THEN 2
                WHEN 'serious' THEN 3
                WHEN 'critical' THEN 4
                ELSE 5 END
        ''', stats_params).fetchall()]

        daily_trend = [dict(r) for r in conn.execute(f'''
            SELECT sv.incident_date as label, COUNT(*) as count
            FROM staff_violations sv
            WHERE {stats_sql}
            GROUP BY sv.incident_date
            ORDER BY sv.incident_date
        ''', stats_params).fetchall()]

        staff_rank = [dict(r) for r in conn.execute(f'''
            SELECT
                sv.staff_name,
                COUNT(*) as total,
                COALESCE(SUM(CASE WHEN sv.status='open' THEN 1 ELSE 0 END), 0) as open_count,
                COALESCE(SUM(CASE WHEN sv.severity IN ('serious','critical') THEN 1 ELSE 0 END), 0) as serious_count,
                MAX(sv.incident_date) as last_incident
            FROM staff_violations sv
            WHERE {stats_sql}
            GROUP BY sv.staff_name
            ORDER BY total DESC, serious_count DESC, last_incident DESC
            LIMIT 8
        ''', stats_params).fetchall()]

        rule_rank = [dict(r) for r in conn.execute(f'''
            SELECT
                COALESCE(vr.title, 'Rule removed') as rule_title,
                COALESCE(vr.category, 'Other') as category,
                COALESCE(vr.color, '#607D8B') as color,
                COUNT(*) as total,
                COALESCE(SUM(CASE WHEN sv.status='open' THEN 1 ELSE 0 END), 0) as active_count,
                COALESCE(SUM(CASE WHEN sv.severity IN ('serious','critical') THEN 1 ELSE 0 END), 0) as serious_count
            FROM staff_violations sv
            LEFT JOIN violation_rules vr ON vr.id=sv.rule_id
            WHERE {stats_sql}
            GROUP BY sv.rule_id, rule_title, category, color
            ORDER BY total DESC, serious_count DESC, active_count DESC
            LIMIT 8
        ''', stats_params).fetchall()]

        violation_stats = {
            'summary': summary,
            'status_breakdown': status_breakdown,
            'severity_breakdown': severity_breakdown,
            'daily_trend': daily_trend,
            'staff_rank': staff_rank,
            'rule_rank': rule_rank,
        }

        # Strike standings (rolling window active count + full history per staff),
        # scoped to the store in view (None = all stores for super_admin).
        strike_standings = staff_strike_standings(conn, store_id=scope_id)
        strike_watch = [s for s in strike_standings if s['active_count'] >= STRIKE_THRESHOLD - 1]

    return render_template('violation_rules.html',
        view=view,
        rules=rules, violations=violations, counts=counts,
        severity_counts=severity_counts, staff=get_active_staff(), managers=MANAGERS,
        status_filter=status_filter, staff_filter=staff_filter,
        severity_filter=severity_filter, date_from=date_from, date_to=date_to,
        violation_stats=violation_stats,
        strike_standings=strike_standings, strike_watch=strike_watch,
        strike_window_days=STRIKE_WINDOW_DAYS, strike_threshold=STRIKE_THRESHOLD,
        severity_meta=SEVERITY_META,
        warning_steps=WARNING_STEPS, warning_step_meta=WARNING_STEP_META,
        today=today_iso, now_time=datetime.now().strftime('%H:%M'),
    )

@app.route('/admin/violations/<int:violation_id>/update', methods=['POST'])
@admin_required
def update_violation(violation_id):
    status = request.form.get('status', 'open')
    if status not in ('open', 'closed', 'cancelled'):
        status = 'open'
    try:
        rule_id = int(request.form.get('rule_id', 0) or 0)
    except (TypeError, ValueError):
        rule_id = None
    staff_name = request.form.get('staff_name', '').strip()
    incident_date = request.form.get('incident_date', date.today().isoformat()).strip()
    incident_time = request.form.get('incident_time', datetime.now().strftime('%H:%M')).strip()
    submitted_by = request.form.get('submitted_by', '').strip()
    severity = request.form.get('severity', 'minor').strip()
    if severity not in ('minor', 'moderate', 'serious', 'critical'):
        severity = 'minor'
    description = request.form.get('description', '').strip()
    action_taken = request.form.get('action_taken', '').strip()
    follow_up_date = request.form.get('follow_up_date', '').strip()
    manager_notes = request.form.get('manager_notes', '').strip()
    warning_step = request.form.get('warning_step', '').strip()
    resolved_by = request.form.get('resolved_by', '').strip()
    resolved_at = datetime.now().strftime('%Y-%m-%d %H:%M') if status in ('closed', 'cancelled') else None
    guard, gp = store_guard_clause()
    with get_db() as conn:
        current = conn.execute(f'SELECT * FROM staff_violations WHERE id=? AND {guard}',
                               [violation_id] + gp).fetchone()
        if current:
            if warning_step not in WARNING_STEPS:
                warning_step = (current['warning_step'] if 'warning_step' in current.keys()
                                and current['warning_step'] else WARNING_STEPS[0])
            conn.execute(f'''UPDATE staff_violations
                SET rule_id=?, staff_name=?, incident_date=?, incident_time=?, submitted_by=?,
                    severity=?, description=?, action_taken=?, follow_up_date=?,
                    status=?, manager_notes=?, resolved_by=?, resolved_at=?, warning_step=?
                WHERE id=? AND {guard}''',
                [rule_id or current['rule_id'],
                 staff_name or current['staff_name'],
                 incident_date or current['incident_date'],
                 incident_time or current['incident_time'],
                 submitted_by,
                 severity,
                 description or current['description'],
                 action_taken,
                 follow_up_date,
                 status,
                 manager_notes,
                 resolved_by,
                 resolved_at,
                 warning_step,
                 violation_id] + gp)
    return redirect(url_for('violation_rules', status=status))

@app.route('/admin/violations/rules/add', methods=['POST'])
@admin_required
def add_violation_rule():
    title = request.form.get('title', '').strip()
    if not title:
        return redirect(url_for('violation_rules'))
    code = ''.join(ch.lower() if ch.isalnum() else '_' for ch in title).strip('_')[:40] or f'rule_{uuid.uuid4().hex[:8]}'
    with get_db() as conn:
        existing = conn.execute('SELECT 1 FROM violation_rules WHERE code=?', (code,)).fetchone()
        if existing:
            code = f'{code}_{uuid.uuid4().hex[:6]}'
        next_order = conn.execute(
            'SELECT COALESCE(MAX(sort_order), -1) + 1 FROM violation_rules').fetchone()[0]
        conn.execute('''INSERT INTO violation_rules
            (code,title,category,severity,description,default_action,color,icon,sort_order)
            VALUES (?,?,?,?,?,?,?,?,?)''',
            (code, title, request.form.get('category','Other').strip() or 'Other',
             request.form.get('severity','minor').strip() or 'minor',
             request.form.get('description','').strip(),
             request.form.get('default_action','').strip(),
             request.form.get('color','#607D8B').strip() or '#607D8B',
             request.form.get('icon','fa-triangle-exclamation').strip() or 'fa-triangle-exclamation',
             next_order))
    return redirect(url_for('violation_rules'))

@app.route('/admin/violations/rules/<int:rule_id>/update', methods=['POST'])
@admin_required
def update_violation_rule(rule_id):
    title = request.form.get('title', '').strip()
    if not title:
        return redirect(url_for('violation_rules'))
    severity = request.form.get('severity', 'minor').strip()
    if severity not in ('minor', 'moderate', 'serious', 'critical'):
        severity = 'minor'
    try:
        sort_order = int(request.form.get('sort_order', 0) or 0)
    except ValueError:
        sort_order = 0
    active = 1 if request.form.get('active') == '1' else 0
    with get_db() as conn:
        conn.execute('''UPDATE violation_rules
            SET title=?, category=?, severity=?, description=?, default_action=?,
                color=?, icon=?, active=?, sort_order=?
            WHERE id=?''',
            (title,
             request.form.get('category', 'Other').strip() or 'Other',
             severity,
             request.form.get('description', '').strip(),
             request.form.get('default_action', '').strip(),
             request.form.get('color', '#607D8B').strip() or '#607D8B',
             request.form.get('icon', 'fa-triangle-exclamation').strip() or 'fa-triangle-exclamation',
             active,
             sort_order,
             rule_id))
    return redirect(url_for('violation_rules'))

@app.route('/admin/violations/rules/<int:rule_id>/delete', methods=['POST'])
@admin_required
def delete_violation_rule(rule_id):
    with get_db() as conn:
        conn.execute('UPDATE violation_rules SET active=0 WHERE id=?', (rule_id,))
    return redirect(url_for('violation_rules'))

# ─── Issue Reports ─────────────────────────────────────────────────────────────

@app.route('/report', methods=['GET', 'POST'])
@login_required
def report_issue():
    if request.method == 'POST':
        category    = request.form.get('category', '')
        title       = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        reported_by = request.form.get('reported_by', '').strip()
        priority    = request.form.get('priority', 'normal')
        if category and title and description and reported_by:
            # Handle optional photo (resized to <=1280px JPEG)
            photo_fname = None
            f = request.files.get('photo')
            if f and f.filename and f.filename.strip():
                ext = os.path.splitext(f.filename)[1].lower().lstrip('.')
                if ext in ALLOWED_EXT:
                    initial = f'issue_{uuid.uuid4().hex[:12]}.{ext}'
                    photo_fname = save_uploaded_photo(f, os.path.join(UPLOAD_FOLDER, initial))
            with get_db() as conn:
                cur = conn.execute('''
                    INSERT INTO issue_reports (category,title,description,reported_by,branch,date,priority,photo,store_id)
                    VALUES (?,?,?,?,?,?,?,?,?)
                ''', (category, title, description, reported_by,
                      session.get('branch',''), date.today().isoformat(), priority, photo_fname,
                      current_store_id()))
                rid = cur.lastrowid
            cat_label = ISSUE_CATEGORIES.get(category, {}).get('label', category)
            email_service.send_notification(
                'issue',
                subject=f'[{priority.upper()}] Issue reported: {title}',
                lines=[
                    f'Category: {cat_label}',
                    f'Priority: {priority}',
                    f'Title: {title}',
                    f'Branch: {session.get("branch", "-")}',
                    f'Date: {date.today().isoformat()}',
                    f'Photo attached: {"yes" if photo_fname else "no"}',
                    f'Description: {description}',
                ],
                link_path='/admin/reports',
                actor=reported_by,
            )
            return redirect(url_for('report_issue', submitted=1))
    submitted = request.args.get('submitted')
    return render_template('report_issue.html', staff=get_active_staff(), submitted=submitted)

@app.route('/admin/reports')
@admin_required
def admin_reports():
    status_filter   = request.args.get('status', 'open')
    category_filter = request.args.get('category', '')
    scope, sp = store_filter_clause()
    with get_db() as conn:
        q = f'SELECT * FROM issue_reports WHERE {scope}'
        p = list(sp)
        if status_filter != 'all':
            q += ' AND status=?'; p.append(status_filter)
        if category_filter:
            q += ' AND category=?'; p.append(category_filter)
        q += ' ORDER BY submitted_at DESC'
        reports = [dict(r) for r in conn.execute(q, p).fetchall()]
        counts = dict(conn.execute(
            f'SELECT status, COUNT(*) FROM issue_reports WHERE {scope} GROUP BY status',
            list(sp)).fetchall())
    return render_template('admin_reports.html',
        reports=reports, status_filter=status_filter,
        category_filter=category_filter, counts=counts)

@app.route('/admin/reports/<int:report_id>/update', methods=['POST'])
@admin_required
def update_report(report_id):
    status      = request.form.get('status', 'open')
    admin_notes = request.form.get('admin_notes', '')
    resolved_by = request.form.get('resolved_by', '')
    resolved_at = datetime.now().strftime('%Y-%m-%d %H:%M') if status == 'resolved' else None
    guard, gp = store_guard_clause()
    with get_db() as conn:
        conn.execute(f'''
            UPDATE issue_reports SET status=?, admin_notes=?, resolved_by=?, resolved_at=?
            WHERE id=? AND {guard}
        ''', [status, admin_notes, resolved_by, resolved_at, report_id] + gp)
    return redirect(url_for('admin_reports'))

# ─── Data Management — Delete & Purge (admin only) ────────────────────────────

def _delete_checklist_session(conn, session_id):
    """Delete one checklist session, its tasks, and remove its photo files from disk.
    Refuses to touch a session that belongs to another store (unless super_admin)."""
    guard, gp = store_guard_clause()
    owns = conn.execute(f'SELECT 1 FROM checklist_sessions WHERE id=? AND {guard}',
                        [session_id] + gp).fetchone()
    if not owns:
        return
    photos = conn.execute(
        'SELECT filename FROM checklist_photos WHERE session_id=?', (session_id,)).fetchall()
    for p in photos:
        try:
            os.remove(os.path.join(UPLOAD_FOLDER, p['filename']))
        except OSError:
            pass
    conn.execute('DELETE FROM checklist_photos WHERE session_id=?', (session_id,))
    conn.execute('DELETE FROM checklist_tasks  WHERE session_id=?', (session_id,))
    conn.execute('DELETE FROM checklist_sessions WHERE id=?', (session_id,))


def _delete_temp_session(conn, session_id):
    guard, gp = store_guard_clause()
    if not conn.execute(f'SELECT 1 FROM temp_sessions WHERE id=? AND {guard}',
                        [session_id] + gp).fetchone():
        return
    conn.execute('DELETE FROM temp_readings WHERE session_id=?', (session_id,))
    conn.execute('DELETE FROM temp_sessions WHERE id=?', (session_id,))


def _delete_issue_report(conn, report_id):
    guard, gp = store_guard_clause()
    row = conn.execute(f'SELECT photo FROM issue_reports WHERE id=? AND {guard}',
                       [report_id] + gp).fetchone()
    if not row:
        return
    if row['photo']:
        try:
            os.remove(os.path.join(UPLOAD_FOLDER, row['photo']))
        except OSError:
            pass
    conn.execute('DELETE FROM issue_reports WHERE id=?', (report_id,))


@app.route('/admin/delete/checklist/<int:sid>', methods=['POST'])
@admin_required
def admin_delete_checklist(sid):
    with get_db() as conn:
        _delete_checklist_session(conn, sid)
    return redirect(request.referrer or url_for('history'))


@app.route('/admin/delete/temperature/<int:sid>', methods=['POST'])
@admin_required
def admin_delete_temperature(sid):
    with get_db() as conn:
        _delete_temp_session(conn, sid)
    return redirect(request.referrer or url_for('history'))


@app.route('/admin/delete/issue/<int:rid>', methods=['POST'])
@admin_required
def admin_delete_issue(rid):
    with get_db() as conn:
        _delete_issue_report(conn, rid)
    return redirect(request.referrer or url_for('admin_reports'))


@app.route('/admin/delete/violation/<int:vid>', methods=['POST'])
@admin_required
def admin_delete_violation(vid):
    guard, gp = store_guard_clause()
    with get_db() as conn:
        conn.execute(f'DELETE FROM staff_violations WHERE id=? AND {guard}', [vid] + gp)
    return redirect(request.referrer or url_for('violation_rules'))


@app.route('/admin/delete/checklist/bulk', methods=['POST'])
@admin_required
def admin_delete_checklist_bulk():
    ids = [int(x) for x in request.form.getlist('ids[]') if x.isdigit()]
    with get_db() as conn:
        for sid in ids:
            _delete_checklist_session(conn, sid)
    return jsonify({'ok': True, 'deleted': len(ids)})


@app.route('/admin/delete/temperature/bulk', methods=['POST'])
@admin_required
def admin_delete_temperature_bulk():
    ids = [int(x) for x in request.form.getlist('ids[]') if x.isdigit()]
    with get_db() as conn:
        for sid in ids:
            _delete_temp_session(conn, sid)
    return jsonify({'ok': True, 'deleted': len(ids)})


@app.route('/admin/delete/issue/bulk', methods=['POST'])
@admin_required
def admin_delete_issue_bulk():
    ids = [int(x) for x in request.form.getlist('ids[]') if x.isdigit()]
    with get_db() as conn:
        for rid in ids:
            _delete_issue_report(conn, rid)
    return jsonify({'ok': True, 'deleted': len(ids)})


@app.route('/admin/delete/violation/bulk', methods=['POST'])
@admin_required
def admin_delete_violation_bulk():
    ids = [int(x) for x in request.form.getlist('ids[]') if x.isdigit()]
    guard, gp = store_guard_clause()
    with get_db() as conn:
        for vid in ids:
            conn.execute(f'DELETE FROM staff_violations WHERE id=? AND {guard}', [vid] + gp)
    return jsonify({'ok': True, 'deleted': len(ids)})


# ── Data Management dashboard ────────────────────────────────────────────────

# Each entry: (key, label, table, date_column, related_cleanup_function)
PURGE_TARGETS = [
    ('checklist',   'Daily Checklists',     'checklist_sessions', 'date',         '_delete_checklist_session'),
    ('temperature', 'Temperature Records',  'temp_sessions',      'date',         '_delete_temp_session'),
    ('issue',       'Issue Reports',        'issue_reports',      'date',         '_delete_issue_report'),
    ('violation',   'Staff Violations',     'staff_violations',   'incident_date', None),
    ('training',    'Training Sessions',    'training_sessions',  'session_date', None),
    ('audit',       'Audit Log',            'audit_log',          'timestamp',    None),
    ('email_log',   'Email Notification Log','email_log',         'sent_at',      None),
]


def _table_count(conn, table):
    try:
        return conn.execute(f'SELECT COUNT(*) c FROM {table}').fetchone()['c']
    except sqlite3.OperationalError:
        return 0


def _oldest_record(conn, table, date_column):
    try:
        row = conn.execute(f'SELECT MIN({date_column}) AS oldest FROM {table}').fetchone()
        return row['oldest'] if row else None
    except sqlite3.OperationalError:
        return None


@app.route('/admin/data-management')
@admin_required
def admin_data_management():
    # SQLite db file size + table row counts
    db_size_mb = round(os.path.getsize(DB_PATH) / 1024 / 1024, 2) if os.path.exists(DB_PATH) else 0
    try:
        uploads = os.listdir(UPLOAD_FOLDER)
        upload_count = len(uploads)
        upload_size_mb = round(sum(
            os.path.getsize(os.path.join(UPLOAD_FOLDER, f)) for f in uploads
        ) / 1024 / 1024, 2)
    except OSError:
        upload_count, upload_size_mb = 0, 0

    with get_db() as conn:
        stats = []
        for key, label, table, date_col, _ in PURGE_TARGETS:
            stats.append({
                'key': key, 'label': label, 'table': table,
                'date_col': date_col,
                'count': _table_count(conn, table),
                'oldest': _oldest_record(conn, table, date_col) or '—',
            })
    return render_template('data_management.html',
        stats=stats, db_size_mb=db_size_mb,
        upload_count=upload_count, upload_size_mb=upload_size_mb)


@app.route('/admin/data-management/purge', methods=['POST'])
@admin_required
def admin_data_purge():
    """Delete records of one category older than N days. Returns JSON with delete counts."""
    target_key = request.form.get('target', '')
    try:
        days = int(request.form.get('days', '0'))
    except ValueError:
        days = 0
    if days <= 0:
        return jsonify({'error': 'Invalid number of days.'}), 400

    target = next((t for t in PURGE_TARGETS if t[0] == target_key), None)
    if not target:
        return jsonify({'error': 'Unknown target.'}), 400

    cutoff = (date.today() - timedelta(days=days)).isoformat()
    _, label, table, date_col, _ = target

    cleanup_fn = {
        'checklist':   _delete_checklist_session,
        'temperature': _delete_temp_session,
        'issue':       _delete_issue_report,
    }.get(target_key)

    deleted = 0
    with get_db() as conn:
        if cleanup_fn:
            ids = [r['id'] for r in conn.execute(
                f'SELECT id FROM {table} WHERE {date_col} < ?', (cutoff,)).fetchall()]
            for id_ in ids:
                cleanup_fn(conn, id_)
            deleted = len(ids)
        else:
            cur = conn.execute(f'DELETE FROM {table} WHERE {date_col} < ?', (cutoff,))
            deleted = cur.rowcount
    return jsonify({'ok': True, 'deleted': deleted, 'label': label, 'cutoff': cutoff})


@app.route('/admin/data-management/vacuum', methods=['POST'])
@admin_required
def admin_data_vacuum():
    """Reclaim disk space after big deletes. Locks DB briefly — admin-triggered only."""
    try:
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute('VACUUM')
        return jsonify({'ok': True, 'message': 'Database compacted.'})
    except Exception as e:
        return jsonify({'error': f'{type(e).__name__}: {e}'}), 500


@app.route('/admin/email-settings/log/clear', methods=['POST'])
@admin_required
def email_log_clear():
    with get_db() as conn:
        conn.execute('DELETE FROM email_log WHERE store_id=?', (current_store_id(),))
    return redirect(url_for('email_settings'))


# ── Daily digest ─────────────────────────────────────────────────────────────

@app.route('/admin/email-settings/digest/send-now', methods=['POST'])
@admin_required
def email_digest_send_now():
    target_date = request.form.get('date', date.today().isoformat())
    ok, msg = email_service.send_daily_digest(
        target_date, CHECKLISTS, TEMPERATURES, ISSUE_CATEGORIES,
        store_id=current_store_id())
    return jsonify({'ok': ok, 'message': msg})


@app.route('/admin/email-settings/digest/preview')
@admin_required
def email_digest_preview():
    """Render the digest HTML in the browser (no send)."""
    target_date = request.args.get('date', date.today().isoformat())
    data = email_service.collect_daily_digest(
        target_date, CHECKLISTS, TEMPERATURES, ISSUE_CATEGORIES,
        store_id=current_store_id())
    settings = email_service.get_settings(current_store_id())
    html = email_service.build_digest_html(data, base_url=settings.get('base_url') or '')
    return html


@app.route('/admin/email-settings/digest/regenerate-token', methods=['POST'])
@admin_required
def email_digest_regenerate_token():
    new_tok = email_service.regenerate_digest_token(current_store_id())
    return jsonify({'ok': True, 'token': new_tok})


@app.route('/cron/daily-digest')
def cron_daily_digest():
    """Public endpoint for external cron services to trigger the digest.
    Authenticated by a secret token passed as ?token= .
    Optional ?store=subiaco selects that branch's token, recipients and data."""
    store_code = (request.args.get('store') or '').strip().lower()
    store_id = 1
    if store_code:
        try:
            row = next((s for s in get_stores(active_only=False)
                        if (s.get('code') or '').lower() == store_code), None)
            if row:
                store_id = row['id']
        except Exception:
            store_id = 1
    expected = email_service.get_or_create_digest_token(store_id)
    given = request.args.get('token', '')
    if not given or given != expected:
        return jsonify({'error': 'forbidden'}), 403
    target_date = request.args.get('date', date.today().isoformat())
    ok, msg = email_service.send_daily_digest(
        target_date, CHECKLISTS, TEMPERATURES, ISSUE_CATEGORIES,
        store_id=store_id)
    return jsonify({'ok': ok, 'date': target_date, 'message': msg})


# ─── Email Notifications (admin) ───────────────────────────────────────────────

@app.route('/admin/email-settings', methods=['GET'])
@admin_required
def email_settings():
    settings = email_service.get_settings(current_store_id())
    recipients = email_service.list_recipients(current_store_id())
    log = email_service.get_recent_log(30, current_store_id())
    digest_token = email_service.get_or_create_digest_token(current_store_id())
    return render_template('email_settings.html',
        settings=settings, recipients=recipients, log=log,
        event_types=email_service.EVENT_TYPES,
        digest_token=digest_token,
        today=date.today().isoformat())


@app.route('/admin/email-settings/save', methods=['POST'])
@admin_required
def email_settings_save():
    email_service.update_settings(
        brevo_api_key=request.form.get('brevo_api_key', '').strip(),
        sender_email=request.form.get('sender_email', '').strip().lower(),
        from_name=request.form.get('from_name', '').strip() or 'MCQ Mirrabooka Notification',
        base_url=request.form.get('base_url', '').strip().rstrip('/'),
        enabled=1 if request.form.get('enabled') else 0,
        updated_by=session.get('role', 'admin'),
    )
    return redirect(url_for('email_settings', saved=1))


@app.route('/admin/email-settings/test', methods=['POST'])
@admin_required
def email_settings_test():
    to = request.form.get('test_to', '').strip()
    ok, msg = email_service.send_test_email(to)
    return jsonify({'ok': ok, 'message': msg})


@app.route('/admin/email-settings/recipients/add', methods=['POST'])
@admin_required
def email_recipient_add():
    try:
        events = {ev: bool(request.form.get(f'notify_{ev}')) for ev, _, _, _ in email_service.EVENT_TYPES}
        email_service.add_recipient(
            email=request.form.get('email', '').strip(),
            name=request.form.get('name', '').strip(),
            events=events,
        )
    except ValueError as e:
        return redirect(url_for('email_settings', error=str(e)))
    except sqlite3.IntegrityError:
        return redirect(url_for('email_settings', error='That email is already on the list.'))
    return redirect(url_for('email_settings'))


@app.route('/admin/email-settings/recipients/<int:rid>/update', methods=['POST'])
@admin_required
def email_recipient_update(rid):
    try:
        events = {ev: bool(request.form.get(f'notify_{ev}')) for ev, _, _, _ in email_service.EVENT_TYPES}
        email_service.update_recipient(
            rid,
            email=request.form.get('email', '').strip(),
            name=request.form.get('name', '').strip(),
            active=bool(request.form.get('active')),
            events=events,
        )
    except ValueError as e:
        return redirect(url_for('email_settings', error=str(e)))
    return redirect(url_for('email_settings'))


@app.route('/admin/email-settings/recipients/<int:rid>/toggle', methods=['POST'])
@admin_required
def email_recipient_toggle(rid):
    email_service.toggle_recipient(rid)
    return redirect(url_for('email_settings'))


@app.route('/admin/email-settings/recipients/<int:rid>/delete', methods=['POST'])
@admin_required
def email_recipient_delete(rid):
    email_service.delete_recipient(rid)
    return redirect(url_for('email_settings'))


# ─── Main ──────────────────────────────────────────────────────────────────────

# Startup must be crash-proof: if any one-time init/migration throws on the
# server (e.g. a brief DB lock during a worker reload), the WSGI app would fail
# to import and the load balancer would return 502 for EVERY request. Each step
# is idempotent, so wrapping it lets the site come up and retry next boot.
def _safe_init(fn, *args, **kwargs):
    try:
        fn(*args, **kwargs)
    except Exception as e:
        import traceback
        print(f'[startup] {getattr(fn, "__name__", fn)} failed: '
              f'{type(e).__name__}: {e}')
        traceback.print_exc()

_safe_init(init_db)
_safe_init(init_prep_tables, DB_PATH, STAFF)
_safe_init(init_pastry_tables, DB_PATH)
_safe_init(init_inventory_tables, DB_PATH)
_safe_init(init_job_tables, DB_PATH)
_safe_init(init_rules_tables, DB_PATH)
_safe_init(init_training_tables, DB_PATH, CHECKLISTS)
_safe_init(email_service.init_email_tables, DB_PATH)
_safe_init(init_whatsapp, DB_PATH, os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static'),
           UPLOAD_FOLDER, CHECKLISTS, TEMPERATURES)
_safe_init(init_packaging, DB_PATH)
_safe_init(init_order_tables, DB_PATH)
_safe_init(init_equipment_tables, DB_PATH)
_safe_init(init_structure_tables, DB_PATH, UPLOAD_FOLDER)
_safe_init(init_webauthn, DB_PATH)
_safe_init(init_food_pricing_tables, DB_PATH)
_safe_init(init_food_safety_tables, DB_PATH)

# Multi-store schema runs LAST: every operational table (orders, packaging,
# equipment, prep, pastry, ...) is created by the blueprint inits above, so they
# must all exist before we add store_id / widen UNIQUE constraints.
_safe_init(migrate_multistore, DB_PATH)
_safe_init(ensure_save_upsert_constraints, DB_PATH)
_safe_init(seed_subiaco_branch, DB_PATH,
           os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates', 'branch.xlsx'))

if __name__ == '__main__':
    print('\n' + '='*50)
    print('  MCQ Mirrabooka Restaurant Management System')
    print('  http://localhost:5050')
    print('  User password: 7777  |  Admin password: 77771')
    print('  Location: mirrabooka')
    print('='*50 + '\n')
    app.run(debug=True, port=5050, host='0.0.0.0')
