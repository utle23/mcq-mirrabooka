from flask import Blueprint, render_template, request, redirect, url_for, session, jsonify, Response
import sqlite3
import io
from datetime import date
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

inventory = Blueprint('inventory', __name__, url_prefix='/inventory')
DB_PATH = None

CATEGORIES = [
    'Beverages',
    'Confectionery',
    'Dairy & Eggs',
    'Deli',
    'Dry Goods & Beans',
    'Fresh Meat',
    'Fresh Produce & Herbs',
    'Frozen Items',
    'Household',
    'Meat & Seafood',
    'Packaging & Supplies',
    'Pastry',
    'Rice & Grain Products',
    'Rice & Noodles',
    'Sauces & Seasonings',
]

CATEGORY_ICONS = {
    'Beverages':             'fas fa-glass-water',
    'Confectionery':         'fas fa-candy-cane',
    'Dairy & Eggs':          'fas fa-egg',
    'Deli':                  'fas fa-cheese',
    'Dry Goods & Beans':     'fas fa-seedling',
    'Fresh Meat':            'fas fa-drumstick-bite',
    'Fresh Produce & Herbs': 'fas fa-leaf',
    'Frozen Items':          'fas fa-snowflake',
    'Household':             'fas fa-broom',
    'Meat & Seafood':        'fas fa-fish',
    'Packaging & Supplies':  'fas fa-box-open',
    'Pastry':                'fas fa-cookie',
    'Rice & Grain Products': 'fas fa-bowl-rice',
    'Rice & Noodles':        'fas fa-bowl-food',
    'Sauces & Seasonings':   'fas fa-bottle-droplet',
}


# (category, name, original, brand, unit_size, cost_unit, cost_carton, qty_carton, code)
ITEMS_SEED = [
    ('Beverages','Chrysanthemum Drink','Yeos Chrysanthemum Drink 300ML','Yeos','300ML',0.67,15.97,24,''),
    ('Beverages','Coca Cola','Coca Cola 24x375ML','Coca Cola','375ML',1.12,26.85,24,''),
    ('Beverages','Coca Cola No Sugar','Coca Cola No Sugar 24x375ML','Coca Cola','375ML',0.98,23.43,24,''),
    ('Beverages','Coffee','Coffee','TRUNG/N HOUSE BLEND 500G','500G',7.47,149.37,20,''),
    ('Beverages','Fanta','FANTA','','375ml',1.2,26.85,24,''),
    ('Beverages','Soybean Drink','Yeos Drink Soybean 300ML','Yeos','300ML',0.78,18.74,24,''),
    ('Beverages','Spring Water (M) 600ML*24','M Spring Water','M Spring','600ML*24',0.26,6.15,24,''),
    ('Beverages','Spring Water (M) 600ML*12','M Spring Water 600ML*12','M Spring','600ML*12',0.28,3.4,12,''),
    ('Beverages','Sprite Icebox','Sprite Icebox 24x375ML','Sprite','375ML',1.12,26.85,24,''),
    ('Beverages','Sugar Cane Drink','YEOS SUGAR CANE DRINK 300ML*24','YEOS','300ML*24',0.77,18.51,1,''),
    ('Beverages','Yeo\'s Grass Jelly Drink','YEOS GRASSJELLY','YEOS','300ML',0.67,15.97,24,''),
    ('Beverages','Yeo\'s Lychee Drink','YEOS LYCHEE','YEOS','300ML',0.67,15.97,24,''),
    ('Beverages','Yeo\'s Wintermelon Drink','YEOS WINTERMELON','YEOS','300ML',0.67,15.97,24,''),
    ('Confectionery','Jelly Powder','HAPPY/G JELLY PWDR PLAIN 280G','Happy/G','280G',2.69,80.7,30,''),
    ('Dairy & Eggs','Condensed Milk','Nestle Condensed Milk 395G','Nestle','395G',2.8,33.66,12,''),
    ('Dairy & Eggs','Egg','Egg','GOLDER EGGS 700G PER DOZEN','700G',3.08,46.2,15,''),
    ('Dairy & Eggs','Eggs','FREE RANGE EGGS 1.5KG 30Pack','FPA Australia','1.5KG',9.63,57.75,6,''),
    ('Dairy & Eggs','Fried Tofu','Winko Fried Tofu 400G','Winko','400G',3.6,3.6,1,''),
    ('Dairy & Eggs','Full Cream Milk','Country Dairy Full Cream','Country Dairy','3L',3.84,3.84,1,''),
    ('Dairy & Eggs','Hard Tofu','Winko Hard Tofu 2pcs','Winko','2pcs',2.2,55.0,25,''),
    ('Dairy & Eggs','Milk (Brownes)','BROWNES MILK EXTRA CREAMY 2L','Brownes','2L',3.47,31.23,9,''),
    ('Dairy & Eggs','Milk (Brownes)','Brownes Full Cream Milk 3L','Brownes','3L',4.5,27.0,6,''),
    ('Deli','Beef Ball','GOLDEN/B BEEF BALL 500G','H&N','500G',9.7,9.7,1,''),
    ('Deli','Beef Tendon Ball','GG VIET/H BEEF TENDON BALL 1K','GG','1KG',26.0,26.0,1,''),
    ('Dry Goods & Beans','Adzuki Bean','HN ADZUKI BEAN 500G','H&N','500G',2.6,104.0,40,''),
    ('Dry Goods & Beans','Coriander Seed','MC Coriander Seed 250G','MC','250G',2.99,29.9,10,''),
    ('Dry Goods & Beans','Coriander Seed','Katoomba Coriander Seed 250G','Katoomba','250G',2.5,50.0,20,''),
    ('Dry Goods & Beans','Dried Star Aniseed','Dried Star Aniseed 1KG','H&N','1KG',21.0,210.0,10,''),
    ('Dry Goods & Beans','Dried Tsaoko Fruit','DRIED TSAOKO FRUIT 1KG','H&N','1KG',25.0,250.0,10,''),
    ('Dry Goods & Beans','Kidney Bean','KIDNEY BEAN 500G','HN DARK RED KIDNEY BEAN 500G','500G',3.5,140.0,40,''),
    ('Dry Goods & Beans','Mung Bean','MUNGBEAN','CTF SPLIT MUNGBEAN PEELED 400G','400G',1.95,97.5,50,''),
    ('Dry Goods & Beans','Mung Bean','HN Split Mung Bean 1KG','HN','1KG',4.2,50.4,12,''),
    ('Fresh Meat','Beef Bolar Blade','BEEF BOLAR BLADE /KG','','KG',14.5,None,None,''),
    ('Fresh Meat','Beef Bolar Blade Pack','BEEF BOLAR BLADE PACK /KG','','KG',16.5,None,None,''),
    ('Fresh Meat','Beef Brisket','BEEF BRISKET /KG','','KG',13.0,None,None,''),
    ('Fresh Meat','Beef Brisket Pack','BEEF BRISKET PACK /KG','','KG',15.0,None,None,''),
    ('Fresh Meat','Beef Brisket Whole','BEEF BRISKET WHOLE /KG','','KG',13.0,None,None,''),
    ('Fresh Meat','Beef Brisket Bone','BEEF BRISKET BONE /KG','','KG',1.8,None,None,''),
    ('Fresh Meat','Beef Brisket Bone Pack','BEEF BRISKET BONE PACK /KG','','KG',1.8,None,None,''),
    ('Fresh Meat','Beef Capoff Topside','BEEF CAPOFF TOPSIDE','','',15.95,None,None,''),
    ('Fresh Meat','Capoff Topside Pack','CAPOFF TOPSIDE PACK /KG','','KG',17.95,None,None,''),
    ('Fresh Meat','Beef Chuck Roll','BEEF CHUCK ROLL /KG','','KG',13.5,None,None,''),
    ('Fresh Meat','Beef Chuck Steak','BEEF CHUCK STEAK /KG','','KG',13.5,None,None,''),
    ('Fresh Meat','Beef Chuck Tender Whole','BEEF CHUCK TENDER WHOLE/KG','','KG',14.9,None,None,''),
    ('Fresh Meat','Beef Chuck Tender','BEEF CHUCK TENDER/KG','','KG',14.9,None,None,''),
    ('Fresh Meat','Beef Chuck Tender Pack','BEEF CHUCK TENDER PACK /KG','','KG',16.9,None,None,''),
    ('Fresh Meat','Beef Dice Beef','BEEF DICE BEEF /KG','','KG',13.0,None,None,''),
    ('Fresh Meat','Beef Dolphin Shin','BEEF DOLPHIN SHIN /KG','','KG',15.99,None,None,''),
    ('Fresh Meat','Beef Dolphin Shin Pack','BEEF DOLPHIN SHIN PACK /KG','','KG',17.99,None,None,''),
    ('Fresh Meat','Beef Dolphin Shin Whole','BEEF DOLPHIN SHIN WHOLE /KG','','KG',15.99,None,None,''),
    ('Fresh Meat','Beef Eyeround Platinum Whole','BEEF EYEROUND PLATINUM WHOLE/KG','','KG',15.5,None,None,''),
    ('Fresh Meat','Beef Eyeround Platinum','BEEF EYEROUND PLATINUM/KG','','KG',15.5,None,None,''),
    ('Fresh Meat','Beef Eyeround Steak','BEEF EYEROUND STEAK /KG','','KG',15.5,None,None,''),
    ('Fresh Meat','Beef Eyeround Pack','BEEF EYEROUND PACK /KG','','KG',17.5,None,None,''),
    ('Fresh Meat','Beef Fingers','BEEF FINGERS /KG','','KG',14.5,None,None,''),
    ('Fresh Meat','Beef Fingers Pack','BEEF FINGERS PACK /KG','','KG',16.5,None,None,''),
    ('Fresh Meat','Beef Flank Steak','BEEF FLANK STEAK /KG','','KG',17.5,None,None,''),
    ('Fresh Meat','Beef Flank Steak Pack','BEEF FLANK STEAK PACK /KG','','KG',19.5,None,None,''),
    ('Fresh Meat','Beef Heart','BEEF HEART','','',None,None,None,''),
    ('Fresh Meat','Beef Heart Pack','BEEF HEART PACK /KG','','KG',5.1,None,None,''),
    ('Fresh Meat','Beef Honeycomb','BEEF HONEYCOMB /KG','','KG',10.0,None,None,''),
    ('Fresh Meat','Beef Honeycomb Tray','BEEF HONEYCOMB TRAY/KG','','KG',10.0,None,None,''),
    ('Fresh Meat','Beef Knuckle','BEEF KNUCKLE /KG','','KG',13.9,None,None,''),
    ('Fresh Meat','Beef Knuckle Pack','BEEF KNUCKLE PACK /KG','','KG',15.9,None,None,''),
    ('Fresh Meat','Beef Knuckle Whole','BEEF KNUCKLE WHOLE /KG','','KG',13.9,None,None,''),
    ('Fresh Meat','Beef Liver','BEEF LIVER','','',None,None,None,''),
    ('Fresh Meat','Beef Marrow Bone Cut','BEEF MARROW BONE CUT /KG','','KG',3.0,None,None,''),
    ('Fresh Meat','Beef Marrow Bone Cut Pack','BEEF MARROW BONE CUT PACK /KG','','KG',5.0,None,None,''),
    ('Fresh Meat','Beef Mince','BEEF MINCE /KG','','KG',7.5,None,None,''),
    ('Fresh Meat','Beef Mince Pack','BEEF MINCE PACK /KG','','KG',8.5,None,None,''),
    ('Fresh Meat','Beef Omasum Tripe','BEEF OMASUM TRIPE /KG','','KG',8.0,None,None,''),
    ('Fresh Meat','Beef Omasum Tripe Pack','BEEF OMASUM TRIPE PACK /KG','','KG',10.0,None,None,''),
    ('Fresh Meat','Beef Oyster Blade','BEEF OYSTER BLADE /KG','','KG',15.0,None,None,''),
    ('Fresh Meat','Beef Oyster Blade Pack','BEEF OYSTER BLADE PACK /KG','','KG',17.0,None,None,''),
    ('Fresh Meat','Beef Picanha Steak','BEEF PICANHA STEAK /KG','','KG',21.5,None,None,''),
    ('Fresh Meat','Beef Picanha Whole','BEEF PICANHA WHOLE /KG','','KG',21.5,None,None,''),
    ('Fresh Meat','Beef Picanha Steak Pack','BEEF PICANHA STEAK PACK /KG','','KG',23.5,None,None,''),
    ('Fresh Meat','Beef Rib','BEEF RIB /KG','','KG',19.9,None,None,''),
    ('Fresh Meat','Beef Rib Pack','BEEF RIB PACK /KG','','KG',21.9,None,None,''),
    ('Fresh Meat','Beef Rib Eye','BEEF RIB EYE /KG','','KG',19.9,None,None,''),
    ('Fresh Meat','Beef Rib Eye Pack','BEEF RIB EYE PACK /KG','','KG',21.9,None,None,''),
    ('Fresh Meat','Beef Rib Whole','BEEF RIB WHOLE/KG','','KG',19.9,None,None,''),
    ('Fresh Meat','Beef Roast Beef','BEEF ROAST BEEF /KG','','KG',13.0,None,None,''),
    ('Fresh Meat','Beef Rump Steak','BEEF RUMP STEAK /KG','','KG',11.9,None,None,''),
    ('Fresh Meat','Beef Rump Steak Pack','BEEF RUMP STEAK PACK /KG','','KG',13.9,None,None,''),
    ('Fresh Meat','Beef Scotch Fillet','BEEF SCOTCH FILLET /KG','','KG',22.9,None,None,''),
    ('Fresh Meat','Beef Scotch Fillet Pack','BEEF SCOTCH FILLET PACK /KG','','KG',24.9,None,None,''),
    ('Fresh Meat','Beef Silverside','BEEF SILVERSIDE /KG','','KG',13.9,None,None,''),
    ('Fresh Meat','Beef Silverside Pack','BEEF SILVERSIDE PACK /KG','','KG',15.9,None,None,''),
    ('Fresh Meat','Beef Silverside Whole','BEEF SILVERSIDE WHOLE /KG','','KG',13.9,None,None,''),
    ('Fresh Meat','Beef Striploin','BEEF STRIPLOIN /KG','','KG',18.9,None,None,''),
    ('Fresh Meat','Beef Striploin Pack','BEEF STRIPLOIN PACK /KG','','KG',20.9,None,None,''),
    ('Fresh Meat','Beef Tails','BEEF TAILS','','',17.9,None,None,''),
    ('Fresh Meat','Beef Tails Pack','BEEF TAILS PACK /KG','','KG',19.9,None,None,''),
    ('Fresh Meat','Beef Tbone Steak','BEEF TBONE STEAK /KG','','KG',22.5,None,None,''),
    ('Fresh Meat','Beef Tbone Steak Pack','BEEF TBONE STEAK PACK /KG','','KG',24.5,None,None,''),
    ('Fresh Meat','Beef Tenderloin','BEEF TENDERLOIN /KG','','KG',33.0,None,None,''),
    ('Fresh Meat','Beef Tenderloin Pack','BEEF TENDERLOIN PACK /KG','','KG',35.0,None,None,''),
    ('Fresh Meat','Beef Tenderloin Whole','BEEF TENDERLOIN WHOLE /KG','','KG',33.0,None,None,''),
    ('Fresh Meat','Beef Tendons','BEEF TENDONS /KG','','KG',8.5,None,None,''),
    ('Fresh Meat','Beef Tendons Tray','BEEF TENDONS TRAY /KG','','KG',8.5,None,None,''),
    ('Fresh Meat','Beef Tongue','BEEF TONGUE /KG','','KG',12.0,None,None,''),
    ('Fresh Meat','Beef Tongue Pack','BEEF TONGUE PACK /KG','','KG',14.0,None,None,''),
    ('Fresh Meat','Beef Topside Steak','BEEF TOPSIDE STEAK','','',15.95,None,None,''),
    ('Fresh Meat','Beef Tri Tip','BEEF TRI TIP /KG','','KG',14.5,None,None,''),
    ('Fresh Meat','Beef Tri Tip Pack','BEEF TRI TIP PACK /KG','','KG',16.5,None,None,''),
    ('Fresh Meat','Beef Tri Tip Whole','BEEF TRI TIP WHOLE /KG','','KG',14.5,None,None,''),
    ('Fresh Meat','Beef Wagyu Chuck Steak','BEEF WAGYU CHUCK STEAK /KG','','KG',40.0,None,None,''),
    ('Fresh Meat','Beef Wagyu Oyster Blade 8/9','BEEF WAGYU OYSTER BLADE 8/9 /KG','','KG',64.0,None,None,''),
    ('Fresh Meat','Borrello BBQ Sausage Tray','BORRELLO BBQ SAUSAGE TRAY','','TRAY',6.6,None,None,''),
    ('Fresh Meat','Borrello Chipolatas Sausage Tray','BORRELLO CHIPOLATAS SAUSAGE TRAY','','TRAY',6.6,None,None,''),
    ('Fresh Meat','Frozen Ox Tail','FROZEN OX TAIL','','',21.5,None,None,''),
    ('Fresh Meat','LOTTE Beef Wagyu Eye Round','LOTTE BEEF WAGYU EYE ROUND /KG','','KG',19.0,None,None,''),
    ('Fresh Meat','LOTTE Neck Meat Pack','LOTTE NECK MEAT PACK /KG','','KG',38.0,None,None,''),
    ('Fresh Meat','LOTTE Neck Meat Whole','LOTTE NECK MEAT WHOLE /KG','','KG',38.0,None,None,''),
    ('Fresh Meat','LOTTE Wagyu Shin Shank','LOTTE WAGYU SHIN SHANK /KG','','KG',14.8,None,None,''),
    ('Fresh Meat','Margaret/R Wagyu Shin Shank','MARGARET/R WAGYU SHIN SHANK/KG','','KG',13.5,None,None,''),
    ('Fresh Meat','LOTTE Wagyu Bolar Blade 8/9','LOTTE WAGYU BOLAR BLADE 8/9 /KG','','KG',25.0,None,None,''),
    ('Fresh Meat','LOTTE Wagyu Bolar/B 8/9 Whl','LOTTE WAGYU BOLAR/B 8/9 WHL/KG','','KG',25.0,None,None,''),
    ('Fresh Meat','LOTTE Wagyu Bolar/B 8/9Pack','LOTTE WAGYU BOLAR/B 8/9PACK/KG','','KG',27.0,None,None,''),
    ('Fresh Meat','Beef La Ribs Cut','BEEF LA RIBS CUT /KG','','KG',15.4,None,None,''),
    ('Fresh Meat','Beef La Ribs Pack','BEEF LA RIBS PACK /KG','','KG',17.4,None,None,''),
    ('Fresh Meat','Beef Wagyu Knuckle','BEEF WAGYU KNUCKLE /KG','','KG',28.0,None,None,''),
    ('Fresh Meat','Beef Wagyu Knuckle Whole','BEEF WAGYU KNUCKLE WHOLE/KG','','KG',28.0,None,None,''),
    ('Fresh Meat','Veal Chuck Tender','VEAL CHUCK TENDER /KG','','KG',14.9,None,None,''),
    ('Fresh Meat','Veal Chuck Tender Whole','VEAL CHUCK TENDER WHOLE/KG','','KG',14.9,None,None,''),
    ('Fresh Meat','Veal Chuck Roll','VEAL CHUCK ROLL /KG','','KG',14.5,None,None,''),
    ('Fresh Meat','Veal Chuck Roll Whole','VEAL CHUCK ROLL WHOLE/KG','','KG',14.5,None,None,''),
    ('Fresh Meat','Chicken Bone','CHICKEN BONE','','',0.45,None,None,''),
    ('Fresh Meat','Chicken Bone Tray','CHICKEN BONE TRAY','','TRAY',0.9,None,None,''),
    ('Fresh Meat','Chicken Breast Fillet S/OFF','CHICKEN BREAST FILLET S/OFF /KG','','KG',6.0,None,None,''),
    ('Fresh Meat','Chicken Breast Fillet S/OFF 15KG','CHICKEN BREAST FILLET S/OFF 15KG','','15KG',90.0,None,None,''),
    ('Fresh Meat','Chicken Breast Fillet SK/ON','CHICKEN BREAST FILLET SK/ON/KG','','KG',6.0,None,None,''),
    ('Fresh Meat','Chicken Breast S/OFF Tray','CHICKEN BREAST S/OFF TRAY /KG','','KG',7.5,None,None,''),
    ('Fresh Meat','Chicken Drumstick','CHICKEN DRUMSTICK /KG','','KG',4.0,None,None,''),
    ('Fresh Meat','Chicken Drumstick 15KG','CHICKEN DRUMSTICK 15KG','','15KG',60.0,None,None,''),
    ('Fresh Meat','Chicken Drumstick In Tray','CHICKEN DRUMSTICK IN TRAY /KG','','KG',5.25,None,None,''),
    ('Fresh Meat','Chicken Feet','CHICKEN FEET /KG','','KG',3.0,None,None,''),
    ('Fresh Meat','Chicken Feet In Tray','CHICKEN FEET IN TRAY /KG','','KG',4.5,None,None,''),
    ('Fresh Meat','Chicken Giblet','CHICKEN GIBLET /KG','','KG',1.75,None,None,''),
    ('Fresh Meat','Chicken Giblet In Tray','CHICKEN GIBLET IN TRAY /KG','','KG',4.5,None,None,''),
    ('Fresh Meat','Chicken Heart','CHICKEN HEART /KG','','KG',1.75,None,None,''),
    ('Fresh Meat','Chicken Heart In Tray','CHICKEN HEART IN TRAY /KG','','KG',4.5,None,None,''),
    ('Fresh Meat','Chicken Liver','CHICKEN LIVER /KG','','KG',1.75,None,None,''),
    ('Fresh Meat','Chicken Liver In Tray','CHICKEN LIVER IN TRAY /KG','','KG',4.5,None,None,''),
    ('Fresh Meat','Chicken Maryland Large','CHICKEN MARYLAND LARGE /KG','','KG',6.0,None,None,''),
    ('Fresh Meat','Chicken Maryland Fillet S/OFF','CHICKEN MARYLAND FILLET S/OFF','','',9.0,None,None,''),
    ('Fresh Meat','Chicken Maryland Fillet S/ON','CHICKEN MARYLAND FILLET S/ON','','',8.3,None,None,''),
    ('Fresh Meat','Chicken Maryland Lrg In Tray','CHICKEN MARYLAND LRG IN TRAY /KG','','KG',6.75,None,None,''),
    ('Fresh Meat','Chicken Maryland S/OFF','CHICKEN MARYLAND S/OFF /KG','','KG',6.5,None,None,''),
    ('Fresh Meat','Chicken Thigh Fillet Skin On','CHICKEN THIGH FILLET SKIN ON','','',3.8,None,None,''),
    ('Fresh Meat','Chicken Wings','CHICKEN WINGS /KG','','KG',3.95,None,None,''),
    ('Fresh Meat','Chicken Wings Tray','CHICKEN WINGS TRAY /KG','','KG',5.0,None,None,''),
    ('Fresh Meat','Chiken Feet Ctn 18KG','CHIKEN FEET CTN 18KG','','18KG',48.6,None,None,''),
    ('Fresh Meat','Frozen Chicken Drumstick','FROZEN CHICKEN DRUMSTICK','','',4.75,None,None,''),
    ('Fresh Meat','Frozen Chicken Wings','FROZEN CHICKEN WINGS','','',3.0,None,None,''),
    ('Fresh Meat','MCQ Chk M/L Fillet Asia Glz','MCQ CHK M/L FILLET ASIA GLZ/KG','','KG',10.0,None,None,''),
    ('Fresh Meat','Whole Chicken Each','WHOLE CHICKEN EACH /KG','','KG',5.1,None,None,''),
    ('Pastry','Banh Quay','Banh Quay','Rubi\'s Bakery','EA',1.5,None,None,''),
    ('Pastry','Banh Tieu','Banh Tieu','Rubi\'s Bakery','EA',1.5,None,None,''),
    ('Pastry','Banh Cam / Cong','Banh Cam / Cong','Rubi\'s Bakery','EA',1.5,None,None,''),
    ('Pastry','Banh Bao (bun)','Banh Bao (bun)','Tú Anh','EA',3.5,None,None,''),
    ('Pastry','Banh batiso','Banh batiso','Tú Anh','EA',1.5,None,None,''),
    ('Pastry','Banh Tieu','Banh Tieu','Tú Anh','EA',1.5,None,None,''),
    ('Pastry','Banh Cam','Banh Cam','Tú Anh','EA',1.5,None,None,''),
    ('Pastry','Fried Pork Dumpling','Fried Pork Dumpling','','EA',1.5,None,None,''),
    ('Pastry','Sesame Ball','Sesame Ball','','EA',1.5,None,None,''),
    ('Pastry','Banh TAI YEN','Banh TAI YEN','','EA',1.4,None,None,''),
    ('Packaging & Supplies','Quilted Brown Express (Tork Xpress Dispenser) Napkin','','','6000',0.01,49.9,6000,'NapQBES'),
    ('Packaging & Supplies','1 Pound White Fruit Bag 200x205mm (Paper Bag for Snacks)','Tui banh ngot','','1000',0.02,19.9,1000,'1WFB'),
    ('Packaging & Supplies','Grease Proof Lunch Wrap Lunch Wrap 33x40cm (Bánh Mì Wrap)','Wrap banh mi','','800',0.02,19.9,800,'LWGP33x40'),
    ('Packaging & Supplies','PET Clear Tray 5x5" (Vegetable)','','','1000',0.1,99.9,1000,'PETTray55'),
    ('Packaging & Supplies','Printed Bottle Paper Bag MCQ - Single 385x100+50mm (Bánh Mì Bag - Printed)','','','500',0.04,20.9,500,'BottleSingleMCQ'),
    ('Packaging & Supplies','Bottle Paper Bag - Single 385x100+50mm (Bánh Mì Bag - Unprinted)','','','500',0.04,18.9,500,'BottleSingle'),
    ('Packaging & Supplies','Hot Dog Tray - Kraft 210x70x36mm (Bánh Mì Tray)','','','600',0.08,49.9,600,'HotDogTrayKft'),
    ('Packaging & Supplies','Kraft Catering Box Size 2 - Medium 359x252x80mm','Hop nho_catering box','','100',0.61,60.9,100,'KCB-M'),
    ('Packaging & Supplies','Window Lid to suit Kraft Catering Box Size 2 - Medium','Nap hop nho_catering box','','100',0.49,49.0,100,'KCBWLid-M'),
    ('Packaging & Supplies','Kraft Catering Box Size 3 - Large 558x252x80mm','Hop lon_catering box','','50',0.9,44.9,50,'KCB-L'),
    ('Packaging & Supplies','Window Lid to suit Kraft Catering Box Size 3 - Large','Nap hop lon_catering box','','50',0.74,37.0,50,'KCBWLid-L'),
    ('Packaging & Supplies','Double Coated Kraft PLA Deli Tray #4 185x129x24mm (Gỏi Cuốn)','Hop goi cuon','','400',0.17,69.9,400,'KDTR-4-PLA'),
    ('Packaging & Supplies','RPET Lid for Double Coated Kraft Deli Tray #4 185x129x32mm (Lid for Gỏi Cuốn)','Nap hop goi cuon','','400',0.13,53.0,400,'KDTR-4Lid'),
    ('Packaging & Supplies','2oz / 60ml Pulp Sauce Container - White (Morley)','','','2000',0.04,76.9,2000,'PulpSC-2-W'),
    ('Packaging & Supplies','Pulp Lid to suit PulpSC-1 & 2 - White (Morley)','','','2000',0.03,60.0,2000,'PulpSCL-1/2-W'),
    ('Packaging & Supplies','50ml Sauce Container (Gỏi Cuốn Sauce) (Subiaco & Mirrabooka)','Hop nuoc mam','','3000',0.03,87.9,3000,'P200'),
    ('Packaging & Supplies','Lid for 50ml Sauce Container (Lid for Gỏi Cuốn Sauce) (Subiaco & Mirrabooka)','Nap hop nuoc mam','','3000',0.02,74.0,3000,'P200Lid'),
    ('Packaging & Supplies','500ml PLA Lined Paper Rectangular Container - Kraft 172x120x41mm (Morley)','','','300',0.16,46.9,300,'Rec500-PLA-K'),
    ('Packaging & Supplies','750ml PLA Lined Paper Rectangular Container - Kraft 172x120x56mm (Morley)','','','300',0.16,48.9,300,'Rec750-PLA-K'),
    ('Packaging & Supplies','1000ml PLA Lined Paper Rectangular Container - Kraft 172x120x75mm','Hop vuong take away','','300',0.17,51.9,300,'Rec1000-PLA-K'),
    ('Packaging & Supplies','PP Lid to suit Kraft Rectangular Deli Container 170x120mm','Nap hop vuong take away','','300',0.08,25.0,300,'RecPaper-PPLid'),
    ('Packaging & Supplies','BioCane 3 Compartments Bento Box 9"','','','200',0.27,53.9,200,'PulpBento-3C'),
    ('Packaging & Supplies','PET Lid to suit 9" (1C. 3C. 4C) Bento Box','','','200',0.23,45.0,200,'PulpBentoLid-PET'),
    ('Packaging & Supplies','16oz Aqueous Lined Paper Round Container - White 115x92x81mm (Mirrabooka)','','','500',0.13,62.9,500,'Rd16-AQ-W'),
    ('Packaging & Supplies','24oz PLA Lined Paper Round Container - White 115x87x113mm (Soup)','Hop soup trang','','500',0.15,76.9,500,'Rd24-PLA-W'),
    ('Packaging & Supplies','115mm PP Flat Lid to suit Paper Round Container','Nap hop soup trang','','500',0.06,31.0,500,'RdPPLid-115-F'),
    ('Packaging & Supplies','LargeKraft Paper Bowl 1300ml - Extra Large 185x160x68mm (Mirrabooka)','Hop bun heo quay','','300',0.26,78.9,300,'PaperBowl-Extra'),
    ('Packaging & Supplies','184mm PET Lid to suit 1100 and 1300ml Bamboo Pulp Bowl','Nap hop bun heo quay','','300',0.16,47.0,300,'BPB-PETLid184'),
    ('Packaging & Supplies','24oz (682ml) Dome Lid Show Bowl with Hinged Lid (Fruit Salad) (Subiaco)','','','150',0.35,52.9,150,'SB24DL'),
    ('Packaging & Supplies','32oz (900ml) Flat Lid Show Bowl with Hinged Lid (Fruit Salad) (Subiaco)','','','150',0.4,59.9,150,'SB32FL'),
    ('Packaging & Supplies','(300) 4 Cup Pulp Eggboard Drink Tray','Cup holder','','300',0.2,60.9,300,'4CupPulp'),
    ('Packaging & Supplies','Dispensable Single Wall Paper Cup 180ml/6oz (Morley)','','','1000',0.08,83.9,1000,'EC-DC0552'),
    ('Packaging & Supplies','EC 390ml / 12oz (90mm Rim) Watercolour Cold Paper Cup (Mirrabooka)','Ly cafe','','1000',0.11,109.9,1000,'EC-DCC390'),
    ('Packaging & Supplies','EC 500ml / 16oz (90mm Rim) Watercolour Cold Paper Cup','Ly juice','','1000',0.12,124.9,1000,'EC-DCC500'),
    ('Packaging & Supplies','BioPak (90mm Rim) BioCane Flat Lid t/s Paper Cold Cups','Nap ly','','1000',0.07,66.0,1000,'BioBCL-90C-Pulp-F'),
    ('Packaging & Supplies','360ml Clear BioCup (Dessert & Fruit Salad) (Subiaco)','','','1000',0.19,185.9,1000,'BioR-360Y'),
    ('Packaging & Supplies','500ml Clear BioCup (Dessert & Fruit Salad) (Subiaco)','500ml Clear Biocup','','1000',0.22,219.9,1000,'BioR-500Y'),
    ('Packaging & Supplies','Dome Lid without Hole for 300-700ml BioCup (Lid for Dessert & Fruit Salad) (Subiaco)','Dome Lid without Hole for 300-700ml BioCup','','1000',0.09,94.9,1000,'BioC-96D(N)'),
    ('Packaging & Supplies','Paper Black Drinking Straws 6x197mm (Subiaco)','Ong hut','','2500',0.01,36.9,2500,'DSPaperBlk'),
    ('Packaging & Supplies','(5000) Paper Jumbo Assorted Colour Stripes Drinking Straws 8x210mm','','','5000',0.02,118.9,5000,'DSPaperJumbo'),
    ('Packaging & Supplies','All Purpose Aluminium Foil 44cm x 150m (Each)','Foil_giay bac','','Each',22.9,22.9,None,'AF44/150 (Ea)'),
    ('Packaging & Supplies','Baking Paper 40cm x 120m (Each)','','','Each',26.9,26.9,None,'Bake40/120 (Ea)'),
    ('Packaging & Supplies','Prowrap Clingwrap 33cm x 600m (Each)','','','Each',20.9,20.9,None,'CW33/600 Pro (Ea)'),
    ('Packaging & Supplies','Prowrap Clingwrap 45cm x 600m (Each)','Clingwrap 45cm X 600m','','1 Roll',23.9,23.9,1,'CW45/600 Pro (Ea)'),
    ('Packaging & Supplies','Kraft Paper Bag #16 SOS 390x240+120mm','','','250',0.15,36.9,250,'PB#16'),
    ('Packaging & Supplies','Individual Wrapped Bamboo Chopstick','Dua','','3000',0.02,49.9,3000,'ChopstickBam'),
    ('Packaging & Supplies','Wooden Fork 16cm','Nia','','1000',0.02,21.9,1000,'WoodenFrk'),
    ('Packaging & Supplies','Wooden Knife 16cm','Dao','','1000',0.02,18.9,1000,'WoodenKnf'),
    ('Packaging & Supplies','Wooden Spoon 16cm','Muong','','1000',0.02,22.9,1000,'WoodenSpn'),
    ('Packaging & Supplies','BioCane Chinese Soup Spoon 14cm','Muong soup','','1000',0.05,49.9,1000,'PulpCSpn'),
    ('Packaging & Supplies','Powder Free Blue Nitrile Glove - Medium','Glove m_bao tay','','10 Box',6.99,69.9,10,'NitrileBluPF-Md'),
    ('Packaging & Supplies','Powder Free Blue Nitrile Glove - Large','Glove l_bao tay','','10 Box',6.99,69.9,10,'NitrileBluPF-Lg'),
    ('Packaging & Supplies','Blue Heavy Duty Food Service Cloth Wipes (Each)','','','1 Roll',11.9,11.9,1,'WipesBlu (Ea)'),
    ('Packaging & Supplies','82lt Extra Heavy Duty Garbage Bag','Bao rac den','','200',0.19,38.9,200,'BL82/35'),
    ('Packaging & Supplies','120lt Extra Heavy Duty Garbage Bag (Subiaco)','','','100',0.24,23.9,100,'BL120/35 (100)'),
    ('Packaging & Supplies','240lt Heavy Duty Solo Garbage Bag (Subiaco)','','','100',0.55,54.9,100,'BL240/35'),
    ('Packaging & Supplies','80m Roll Paper Towel','','','16 Roll',3.12,49.9,16,'RT80'),
    ('Packaging & Supplies','Slimline Hand Towel (230x80mm Folded)','Khan giay lau tay','','4000',0.01,44.9,4000,'HTSlimline'),
    ('Packaging & Supplies','Thermal Register Roll 80x80mm','','','30 Roll',1.59,47.7,30,'Thermal80x80'),
    ('Packaging & Supplies','Super White Bleach Commercial Strength 6%','','','5lt',2.58,12.9,5,'Bleach5'),
    ('Packaging & Supplies','Blitz - Multi Purpose (Floor) Cleaner & Degreaser','','','5lt',4.58,22.9,5,'Blitz5'),
    ('Packaging & Supplies','Jet Dry Plus Rinsing & Drying Agent','','','15lt',5.33,79.9,15,'JetDryPlus15'),
    ('Packaging & Supplies','Jet Klean Automatic Dishwasher Liquid','','','20lt',4.5,89.9,20,'JetKlean20'),
    ('Packaging & Supplies','Oven Magic Grill Cleaner','Nuoc rua lo nuong 5lit','','5lt',5.98,29.9,5,'Oven5'),
    ('Packaging & Supplies','Food Grade No Rinse Sanitiser','','','5lt',5.38,26.9,5,'Sanitiser5'),
    ('Packaging & Supplies','Surplus - Manual Dishwashing Liquid','','','5lt',3.38,16.9,5,'Surplus5'),
    ('Packaging & Supplies','Surplus - Manual Dishwashing Liquid','Nuoc rua chen 20 lit','','20lt',1.7,33.9,20,'Surplus20'),
    ('Fresh Meat','Pork Sausage','Wing/C Pork Sausage 1KG','Wing/C','1KG',19.8,198.0,10,''),
    ('Fresh Produce & Herbs','Avocado Hass bulk','Avocado Hass bulk','','EA',0.74,38.5,52,''),
    ('Fresh Produce & Herbs','Bean Sprout Organic','Bean Sprout Organic','organic 7 fresh 400g','400G',2.0,2.0,1,''),
    ('Fresh Produce & Herbs','Beetroot','Beetroot','m branch','',5.01,50.1,10,''),
    ('Fresh Produce & Herbs','Brown Onion','Onion browns 20kg','M branch','20KG',23.1,23.1,1,''),
    ('Fresh Produce & Herbs','Burpless Cucumber','Cucumber burpless 20 count','','EA',0.39,11.6,30,''),
    ('Fresh Produce & Herbs','Cabbage green','Cabbage Count 10','','EA',1.6,16.0,10,''),
    ('Fresh Produce & Herbs','Cabbage Red','Cabbage Red Each','','EA',3.35,40.2,12,''),
    ('Fresh Produce & Herbs','Carrot','Carrot bag 15KG','M branch','15KG',9.35,9.35,1,''),
    ('Fresh Produce & Herbs','Carrot','Carrot bag 1kg','M branch','1KG',0.8,16.0,20,''),
    ('Fresh Produce & Herbs','Chilli Green','Chilli Green','','KG',7.63,53.4,7,''),
    ('Fresh Produce & Herbs','Coconut water','Coconut water','MALEE UHT','1L',2.19,26.26,12,''),
    ('Fresh Produce & Herbs','Coconut water','Coconut','MALEE UHT NAMHOM','1L',2.22,26.6,12,''),
    ('Fresh Produce & Herbs','Coriander','Corriander','M BRANCH','KG',7.7,7.7,1,''),
    ('Fresh Produce & Herbs','Gralic bag','Gralic bag','','500G',1.65,33.0,20,''),
    ('Fresh Produce & Herbs','Garlic Peeled','Garlic Peeled 500G','','500G',2.2,22.0,10,''),
    ('Fresh Produce & Herbs','Ginger','GINGER FRESH /KG','','1KG',22.0,220.0,10,''),
    ('Fresh Produce & Herbs','Ginger Young','Ginger Young /KG','','1KG',11.0,110.0,10,''),
    ('Fresh Produce & Herbs','Lettuce','Lettuce','','EA',1.33,16.0,12,''),
    ('Fresh Produce & Herbs','Mint','Mint','M branch','KG',7.0,7.0,1,''),
    ('Fresh Produce & Herbs','Onion sald red','Onion sald red','','KG',5.56,55.6,10,''),
    ('Fresh Produce & Herbs','Onion Brown','Onion Brown PKT 2KG','','2KG',2.31,2.31,1,''),
    ('Fresh Produce & Herbs','Onion Red','ONION RED 2KG BAG','','2KG',3.52,3.52,1,''),
    ('Fresh Produce & Herbs','Orange Valencia','Orange Valencia Prepack 3KG','','3KG',6.01,6.01,1,''),
    ('Fresh Produce & Herbs','Pineapple (with head)','Pineapple sugar with head 12','m branch','EA',6.42,6.42,12,''),
    ('Fresh Produce & Herbs','Pineapple Sugar','Pineapple Sugar Count 10','','EA',3.3,33.0,10,''),
    ('Fresh Produce & Herbs','Radish','Raddish white chinese per kg','','KG',1.7,17.1,10,''),
    ('Fresh Produce & Herbs','Spring Onion','Spring onion','','EA',1.0,30.0,30,''),
    ('Fresh Produce & Herbs','Tomato No1','Tomato No1','','KG',33.0,3.3,10,''),
    ('Fresh Produce & Herbs','Watermelon Seedless Whole','Watermelon Seedless Whole','','EA',None,None,None,''),
    ('Frozen Items','Beef Ball','MR ABC FZ Beef Ball 500G','MR ABC','500G',9.5,380.0,40,''),
    ('Frozen Items','Beef Ball Tendon','MR ABC FZ Beef Ball Tendon 500G','MR ABC','500G',9.1,364.0,40,''),
    ('Frozen Items','Frozen Fried Banana','BAMBOO FZ FRI BANANA SESAME340','Bamboo','340G',3.32,49.73,15,''),
    ('Frozen Items','Frozen Jackfruit','FROZEN JACKFRUIT 500G','M','500G',3.01,72.24,24,''),
    ('Frozen Items','Lemongrass Mince','M FZ Lemongrass Mince 500G','M','500G',1.7,40.76,24,''),
    ('Fresh Produce & Herbs','Lime No1','Lime per kg','','kg',6.6,33.0,5,''),
    ('Fresh Produce & Herbs','Lime','LIME IN PKT /KG','','KG',3.6,3.6,1,''),
    ('Household','Cling Wrap','CAPRI CLING WRAP DIS 45CM*600M','CAPRI','45CM*600M',19.07,125.84,6,''),
    ('Household','Facial Tissue 3-Ply','M Facial Tissue 3Ply 100S','M','100S',1.08,38.72,36,''),
    ('Household','Sandwich Bags','Hercules Sandwich Bags 40PK','Hercules','40PK',1.54,20.39,12,''),
    ('Meat & Seafood','Tiger Prawn','Tasse FZ Tiger Prawn 10/15 3KG','Tasse','3KG',69.99,69.99,1,''),
    ('Meat & Seafood','Tiger Prawn','UNISEA WHL COOKED PRWN U15 Per kg','UNISEA','KG',18.7,None,None,''),
    ('Meat & Seafood','Tiger Prawn','UNISEA WHL COOKED PRWN U15 5KG','UNISEA','5KG',93.5,93.5,1,''),
    ('Packaging & Supplies','Tissue','Tissue','M facial tissue','',None,None,None,''),
    ('Sauces & Seasonings','Cooking Salt','B/GOLD COOKING SALT 2KG','B GOLD','2KG',2.58,15.48,6,''),
    ('Sauces & Seasonings','Peanut Butter Smooth','Bega Peanut Btr Smth  470G','Bega','470G',4.8,28.8,6,''),
    ('Sauces & Seasonings','Peanut Butter','BEGA PEANUT BUTTER SMOOTH 2KG','BEGA','2KG',22.99,91.96,4,''),
    ('Sauces & Seasonings','Soy Sauce','CHINSU TAM THAI TU SOY 500ML','CHINSU','500ML',1.58,37.83,24,''),
    ('Sauces & Seasonings','Sweet Chilli Sauce','PANTAI SWT CHILLI SCE 750ML','PANTAI','750ML',4.8,57.6,12,''),
    ('Sauces & Seasonings','Baking Soda','MCKEN CRAB Soda 1KG','MCKEN','1KG',3.78,22.68,6,''),
    ('Sauces & Seasonings','Chilli Sauce','Chilli sauce','HUY/F SRIRACHA','739g',6.16,73.92,12,''),
    ('Sauces & Seasonings','Coconut Cream','MAEPLOY COCONUT CREAM 560ML','Maeploy','560ML',2.91,69.96,24,''),
    ('Sauces & Seasonings','Coconut Cream','TWIN/E COCONUT CREAM 20% 400ML','Twin Elephant','400ML',2.26,54.13,24,''),
    ('Sauces & Seasonings','Cooking Oil','Oil 4L','krisla pure vegestable oil','4L',10.62,42.46,4,''),
    ('Sauces & Seasonings','Cooking Oil','Oil 4L','krisla pure canola oil','4L',12.1,48.4,4,''),
    ('Sauces & Seasonings','Canola Oil','Pure Choice Canola Oil 2L','Pure Choice','2L',5.25,31.5,6,''),
    ('Sauces & Seasonings','Cooking Oil','Oil 20L','simply vegetable oil','20L',49.35,49.35,1,''),
    ('Sauces & Seasonings','Longan in Syrup','Twinlee Longan in Syrup','Twinlee','565G',2.88,69.09,24,''),
    ('Sauces & Seasonings','Palm Sugar','CAPITAL/E PALM SUGAR 454G','Capital/E','454G',1.65,39.64,24,''),
    ('Sauces & Seasonings','Rambutan in Syrup','Twin/E Rambutan Syrup','Twin/E','565G',3.23,77.6,24,''),
    ('Sauces & Seasonings','Fungus','Dried wood Fungus Whl/Stri','','80G',1.25,125.0,100,''),
    ('Sauces & Seasonings','Rock Sugar','Rock sugar (đường phèn)','HN ROCK SUGAR 454G','454G',1.54,77.0,50,''),
    ('Sauces & Seasonings','Toddy Palm Slice','Twin/e Toddy Palm Slice','Twin/E','565G',3.23,77.6,24,''),
    ('Sauces & Seasonings','Vinegar WHITE','Vinegar malt WHITE','Anchor','15L',24.35,24.35,1,''),
    ('Sauces & Seasonings','Water Chestnut Whole','TAF/WD Water Chestnut WHL 567G','TAF/WD','567G',1.45,34.8,24,''),
    ('Sauces & Seasonings','White Sugar','B/Gold White Sugar 2KG','B/Gold','2KG',3.32,19.9,6,''),
    ('Sauces & Seasonings','White Sugar','B/Gold White Sugar 3KG','B/Gold','3KG',4.87,19.48,4,''),
    ('Rice & Grain Products','Broken Rice','Broken rice (gạo tấm) 20KG','DRAGON A1','20KG',36.0,36.0,1,''),
    ('Rice & Grain Products','Glutinous Rice','QRICE Thai White Glut Rice 10KG','QRICE','10KG',23.32,23.32,1,''),
    ('Rice & Grain Products','Jasmine Rice','Golden P. Jasmine Rice 10KG','Golden Phoenix','10KG',25.0,25.0,None,''),
    ('Rice & Grain Products','Jasmine Rice','Lion Rice Jasmine 10KG','Lion','10KG',30.58,30.58,1,''),
    ('Rice & Grain Products','Jasmine Rice','Lion Rice Jasmine 20KG','Lion','20KG',61.38,61.38,1,''),
    ('Rice & Grain Products','Jasmine Rice','QRICE Thai Jasmine Rice 10KG','QRICE','10KG',21.4,21.4,1,''),
    ('Rice & Noodles','Longkou Vermicelli','Longkou Vermicelli 250G','Longkou','250G',1.38,68.75,50,''),
    ('Rice & Noodles','Rice Noodle','Rice noodle (bun tuoi) 908G','M RICE VER BUNTUOI','908G',68.59,68.59,1,''),
    ('Rice & Noodles','Rice Paper','Rice paper roll','M square rice paper 22CM 340G','340G',1.91,83.83,44,''),
    ('Rice & Noodles','Rice Stick Pho','M RICE STICK PHO 3MM 400G (BANH PHO)','M','400G',1.62,48.74,30,''),
    ('Rice & Noodles','Rice Stick Pho','M RICE STICK PHO 3MM 400G (BANH PHO)','M','400G*30',48.74,48.74,1,''),
    ('Rice & Noodles','Rice Vermicelli','FL JIANGXI R/VERM XL 400*30 (BUNBOHUE)','FL JIANGXI','400*30',40.5,40.5,1,''),
    ('Sauces & Seasonings','Char Siu Sauce','LKK Char Siu Sauce GF 2250G','Lee Kum Kee','2250G',17.0,102.0,6,''),
    ('Sauces & Seasonings','Chilli Crush','Maharani Chilli Crush 500G','Maharani','500G',7.23,86.7,12,''),
    ('Sauces & Seasonings','Chilli Crushed','MC Chilli Crushed 1KG','Maharani','1KG',12.54,75.24,6,''),
    ('Sauces & Seasonings','Cinnamon Quill','HN Dried Cinnamon QuilS 1KG','HN','1KG',13.0,130.0,10,''),
    ('Sauces & Seasonings','Clove Whole','HN Whole Clove 25G','HN','25G',1.6,16.0,10,''),
    ('Sauces & Seasonings','Clove Whole','Katoomba Clove Whole 100G','Katoomba','100G',3.0,75.0,25,''),
    ('Sauces & Seasonings','Cooking Salt','B/Gold Cooking Salt 1KG','B/Gold','1KG',1.92,23.06,12,''),
    ('Sauces & Seasonings','Cooking Salt','Saxa Cooking Salt 10KG','Saxa','10KG',12.5,12.5,1,''),
    ('Sauces & Seasonings','Coriander Seed','HN Dried Coriander Seeds 50G','HN','50G',1.4,14.0,10,''),
    ('Sauces & Seasonings','Fennel Seed','MC FENNEL SEED 500G','MC','500G',4.31,103.34,24,''),
    ('Sauces & Seasonings','Fish Sauce','Three Crab Fish Sauce 682ML','Three Crab','682ML',10.78,129.36,12,''),
    ('Sauces & Seasonings','Fish Sauce Golden','Trachang Fish Sauce Golden 725ML','Trachang','725ML',2.34,28.08,12,''),
    ('Sauces & Seasonings','Fish Sauce Red','Trachang Fish Sauce Red 4.5L','Trachang','4.5L',13.5,40.5,3,''),
    ('Sauces & Seasonings','Five Spice Powder','HN Five Spice Powder 80G','HN','80G',1.6,16.0,10,''),
    ('Sauces & Seasonings','Iodised Salt','Chinsu Tamthaitu Iodised Salt 650G','Chinsu','650G',1.31,15.7,12,''),
    ('Sauces & Seasonings','Licorice Root','HN Dried Licorice Root 50G','HN','50G',1.8,18.0,10,''),
    ('Sauces & Seasonings','MSG','Ajinomoto MSG 1KG','Ajinomoto','1KG',7.42,148.37,20,''),
    ('Sauces & Seasonings','Mayonnaise','Kewpie Mayonnaise 1kg','Kewpie','1KG',9.35,13.99,112,''),
    ('Sauces & Seasonings','Onion Powder','MC Onion Powder 1KG','MC','1KG',5.51,33.08,6,''),
    ('Sauces & Seasonings','Onion Powder','MC Onion Powder 500G','MC','500G',3.35,40.23,12,''),
    ('Sauces & Seasonings','Pepper Black Whole','Maharani Pepper Blk Whole 1KG','Maharani','1KG',17.74,248.32,14,''),
    ('Sauces & Seasonings','Rock Salt','Saxa Salt Rock 10KG','Saxa','10KG',12.99,12.99,1,''),
    ('Sauces & Seasonings','Soy Sauce','LKK Premium Soy Sauce 1.75L','LKK','1.75L',6.6,39.6,6,''),
    ('Sauces & Seasonings','Soy Sauce Dark','Maggi Soy Sauce Dark 700ML','Maggi','700ML',2.53,30.37,12,''),
    ('Sauces & Seasonings','Soy Sauce Light','Maggi Soya Sauce Light VN 700ML','Maggi','700ML',2.1,25.24,12,''),
    ('Sauces & Seasonings','Sweet Soy Sauce','ABC Sweet Soy Sauce 620ML','ABC','620ML',3.9,23.43,6,''),
    ('Sauces & Seasonings','Vinegar White','B/Gold Vinegar White 2L','B/Gold','2L',1.96,11.77,6,'')
]


# ── Helpers ────────────────────────────────────────────────────────────────────

def _get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def _is_admin():
    return session.get('role') == 'admin'

def _login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

def _admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login_page'))
        if not _is_admin():
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated


def _add_carton_margin(item: dict) -> dict:
    """Carton margin compares carton cost against buying the same qty as units."""
    try:
        cost_unit = float(item.get('cost_unit') or 0)
        cost_carton = float(item.get('cost_carton') or 0)
        qty_carton = int(item.get('qty_carton') or 0)
    except (TypeError, ValueError):
        cost_unit = cost_carton = 0
        qty_carton = 0

    unit_total = cost_unit * qty_carton
    if unit_total > 0 and cost_carton > 0:
        margin = unit_total - cost_carton
        item['carton_margin'] = margin
        item['carton_margin_abs'] = abs(margin)
        item['carton_margin_pct'] = (margin / unit_total) * 100
        item['carton_unit_cost'] = cost_carton / qty_carton if qty_carton else None
    else:
        item['carton_margin'] = None
        item['carton_margin_abs'] = None
        item['carton_margin_pct'] = None
        item['carton_unit_cost'] = None
    return item


# ── DB Init ────────────────────────────────────────────────────────────────────

def init_inventory_tables(db_path):
    global DB_PATH
    DB_PATH = db_path
    with sqlite3.connect(db_path) as conn:
        conn.execute('''CREATE TABLE IF NOT EXISTS inventory_items (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            category    TEXT NOT NULL,
            name        TEXT NOT NULL,
            original    TEXT,
            brand       TEXT,
            unit_size   TEXT,
            cost_unit   REAL DEFAULT 0,
            cost_carton REAL DEFAULT 0,
            qty_carton  INTEGER DEFAULT 1,
            code        TEXT,
            active      INTEGER DEFAULT 1
        )''')
        conn.commit()

        # Force re-seed if data is stale (< 330 items or > 0 with wrong count)
        count = conn.execute('SELECT COUNT(*) FROM inventory_items WHERE active=1').fetchone()[0]
        if count < 330:
            conn.execute('DELETE FROM inventory_items')
            conn.executemany(
                '''INSERT INTO inventory_items
                   (category, name, original, brand, unit_size, cost_unit, cost_carton, qty_carton, code)
                   VALUES (?,?,?,?,?,?,?,?,?)''',
                ITEMS_SEED
            )
            conn.commit()


# ── Routes ─────────────────────────────────────────────────────────────────────

@inventory.route('/')
@_login_required
def inventory_list():
    category = request.args.get('category', 'All')
    search   = request.args.get('q', '').strip()

    with _get_db() as conn:
        if category and category != 'All':
            rows = conn.execute(
                '''SELECT * FROM inventory_items
                   WHERE active=1 AND category=?
                   ORDER BY name''',
                (category,)
            ).fetchall()
        else:
            rows = conn.execute(
                '''SELECT * FROM inventory_items
                   WHERE active=1
                   ORDER BY category, name'''
            ).fetchall()

        # Category counts
        cat_counts_raw = conn.execute(
            'SELECT category, COUNT(*) as cnt FROM inventory_items WHERE active=1 GROUP BY category'
        ).fetchall()
        cat_counts = {r['category']: r['cnt'] for r in cat_counts_raw}
        total = sum(cat_counts.values())

    items = [_add_carton_margin(dict(r)) for r in rows]

    # Client-side search applied server-side too for non-JS fallback
    if search:
        sl = search.lower()
        items = [i for i in items if sl in i['name'].lower()
                 or sl in (i['original'] or '').lower()
                 or sl in (i['brand'] or '').lower()]

    return render_template('inventory.html',
        items=items,
        categories=CATEGORIES,
        category_icons=CATEGORY_ICONS,
        cat_counts=cat_counts,
        total=total,
        selected_cat=category,
        search=search,
        is_admin=_is_admin())


@inventory.route('/item/<int:item_id>/edit', methods=['POST'])
@_admin_required
def edit_item(item_id):
    name        = request.form.get('name', '').strip()
    original    = request.form.get('original', '').strip()
    brand       = request.form.get('brand', '').strip()
    unit_size   = request.form.get('unit_size', '').strip()
    category    = request.form.get('category', '').strip()
    code        = request.form.get('code', '').strip()
    try:
        cost_unit   = float(request.form.get('cost_unit', 0) or 0)
        cost_carton = float(request.form.get('cost_carton', 0) or 0)
        qty_carton  = int(request.form.get('qty_carton', 1) or 1)
    except (ValueError, TypeError):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'error': 'Invalid number'}), 400
        return redirect(url_for('inventory.inventory_list'))

    if not name:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'error': 'Name required'}), 400
        return redirect(url_for('inventory.inventory_list'))

    with _get_db() as conn:
        conn.execute(
            '''UPDATE inventory_items SET
               name=?, original=?, brand=?, unit_size=?, category=?,
               cost_unit=?, cost_carton=?, qty_carton=?, code=?
               WHERE id=?''',
            (name, original, brand, unit_size, category,
             cost_unit, cost_carton, qty_carton, code, item_id)
        )

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'ok': True})
    return redirect(url_for('inventory.inventory_list'))


@inventory.route('/item/add', methods=['POST'])
@_admin_required
def add_item():
    name        = request.form.get('name', '').strip()
    original    = request.form.get('original', '').strip()
    brand       = request.form.get('brand', '').strip()
    unit_size   = request.form.get('unit_size', '').strip()
    category    = request.form.get('category', '').strip()
    code        = request.form.get('code', '').strip()
    try:
        cost_unit   = float(request.form.get('cost_unit', 0) or 0)
        cost_carton = float(request.form.get('cost_carton', 0) or 0)
        qty_carton  = int(request.form.get('qty_carton', 1) or 1)
    except (ValueError, TypeError):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'error': 'Invalid number'}), 400
        return redirect(url_for('inventory.inventory_list'))

    if not name:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'error': 'Name required'}), 400
        return redirect(url_for('inventory.inventory_list'))

    with _get_db() as conn:
        conn.execute(
            '''INSERT INTO inventory_items
               (category, name, original, brand, unit_size, cost_unit, cost_carton, qty_carton, code)
               VALUES (?,?,?,?,?,?,?,?,?)''',
            (category, name, original, brand, unit_size, cost_unit, cost_carton, qty_carton, code)
        )

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'ok': True})
    return redirect(url_for('inventory.inventory_list'))


@inventory.route('/item/<int:item_id>/toggle', methods=['POST'])
@_admin_required
def toggle_item(item_id):
    with _get_db() as conn:
        row = conn.execute('SELECT active FROM inventory_items WHERE id=?', (item_id,)).fetchone()
        if not row:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'error': 'Not found'}), 404
            return redirect(url_for('inventory.inventory_list'))
        new_active = 0 if row['active'] else 1
        conn.execute('UPDATE inventory_items SET active=? WHERE id=?', (new_active, item_id))

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'ok': True, 'active': new_active})
    return redirect(url_for('inventory.inventory_list'))


@inventory.route('/export')
@_login_required
def export_excel():
    category = request.args.get('category', '').strip()
    search   = request.args.get('q', '').strip()

    with _get_db() as conn:
        if category and category != 'All':
            rows = conn.execute(
                'SELECT * FROM inventory_items WHERE active=1 AND category=? ORDER BY category, name',
                (category,)
            ).fetchall()
        else:
            rows = conn.execute(
                'SELECT * FROM inventory_items WHERE active=1 ORDER BY category, name'
            ).fetchall()

    items = [_add_carton_margin(dict(r)) for r in rows]
    if search:
        sl = search.lower()
        items = [i for i in items if sl in i['name'].lower()
                 or sl in (i['original'] or '').lower()
                 or sl in (i['brand'] or '').lower()]

    # ── Build workbook ──────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Item List'

    # Colour palette
    GREEN_DARK  = '1B4332'
    GREEN_MID   = '2D6A4F'
    GREEN_LIGHT = 'E8F5E9'
    ORANGE_LIGHT= 'FFF3E0'
    HEADER_TXT  = 'FFFFFF'
    CAT_TXT     = 'FFFFFF'

    thin = Side(style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title row ──
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = f'MCQ Mirrabooka — Item List   (exported {date.today().strftime("%d %b %Y")})'
    title_cell.font = Font(bold=True, size=13, color=HEADER_TXT)
    title_cell.fill = PatternFill('solid', fgColor=GREEN_DARK)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    # ── Column headers (row 2) ──
    headers = ['#', 'Item Name', 'Original / Full Description', 'Brand',
               'Unit Size', 'Cost / Unit ($)', 'Cost / Carton ($)', 'Qty / Carton',
               'Margin / Carton', 'Code']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color=HEADER_TXT, size=10)
        c.fill = PatternFill('solid', fgColor=GREEN_MID)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
    ws.row_dimensions[2].height = 22

    # ── Column widths ──
    col_widths = [5, 28, 42, 18, 12, 16, 18, 13, 18, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Data rows ──
    cur_cat  = None
    data_row = 3
    seq      = 0

    for item in items:
        # Category header row when category changes
        if item['category'] != cur_cat:
            cur_cat = item['category']
            icon = CATEGORY_ICONS.get(cur_cat, '')
            ws.merge_cells(f'A{data_row}:J{data_row}')
            cat_cell = ws[f'A{data_row}']
            cat_cell.value = f'  {cur_cat}'
            cat_cell.font = Font(bold=True, color=CAT_TXT, size=10)
            cat_cell.fill = PatternFill('solid', fgColor=GREEN_MID)
            cat_cell.alignment = Alignment(vertical='center')
            ws.row_dimensions[data_row].height = 18
            data_row += 1
            seq = 0

        seq += 1
        row_fill = PatternFill('solid', fgColor='FFFFFF') if seq % 2 == 0 else PatternFill('solid', fgColor='F9FFFE')

        values = [
            seq,
            item['name'],
            item['original'] or '',
            item['brand'] or '',
            item['unit_size'] or '',
            item['cost_unit'] or None,
            item['cost_carton'] or None,
            item['qty_carton'] or None,
            (f"{item['carton_margin']:.2f} ({item['carton_margin_pct']:.1f}%)"
             if item.get('carton_margin') is not None else ''),
            item['code'] or '',
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=data_row, column=col, value=val)
            c.fill = row_fill
            c.border = border
            c.alignment = Alignment(vertical='center')
            if col == 2:
                c.font = Font(bold=True, color='1B3A2D', size=9)
            elif col in (3, 4, 5, 10):
                c.font = Font(color='666666', size=9)
            elif col in (6, 7) and val is not None:
                c.font = Font(bold=True, color='1565C0', size=9)
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right', vertical='center')
            elif col == 8 and val is not None:
                c.alignment = Alignment(horizontal='center', vertical='center')
            elif col == 9 and val:
                c.font = Font(bold=True, color='2E7D32' if item['carton_margin'] >= 0 else 'C62828', size=9)
                c.alignment = Alignment(horizontal='right', vertical='center')
            else:
                c.font = Font(size=9)
        ws.row_dimensions[data_row].height = 16
        data_row += 1

    # ── Total row ──
    ws.merge_cells(f'A{data_row}:F{data_row}')
    tot = ws[f'A{data_row}']
    tot.value = f'Total: {len(items)} items'
    tot.font = Font(bold=True, size=10, color=GREEN_DARK)
    tot.alignment = Alignment(horizontal='right', vertical='center')
    tot.fill = PatternFill('solid', fgColor=GREEN_LIGHT)

    # ── Freeze panes below header ──
    ws.freeze_panes = 'A3'

    # ── Output ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname_cat = f'_{category}' if category and category != 'All' else ''
    fname = f'MCQ_ItemList{fname_cat}_{date.today().isoformat()}.xlsx'

    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'}
    )
