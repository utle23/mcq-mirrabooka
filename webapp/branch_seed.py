"""One-time branch data importer.

Loads the Subiaco setup workbook into store_id=3 without touching Mirrabooka.
The importer is idempotent through an audit_log marker so app restarts do not
overwrite later admin edits made in the web UI.
"""
from __future__ import annotations

import os
import re
import sqlite3
from typing import Any

from openpyxl import load_workbook

from packaging_routes import JACCUS_PRICE_DATA, SUBIACO_PACKAGING_DELIVERY_DAYS


SUBIACO_STORE_ID = 3
MARKER = 'subiaco_branch_seed_v1'
PRICE_MARKER = 'subiaco_packaging_prices_from_pdf_v1'
CONTACT_FIX_MARKER = 'subiaco_contact_fix_v1'

# ── Morley (store_id=2) Cashier checklist ─────────────────────────────────────
# Morley has no branch workbook; its Cashier (take_order) Opening + Closing
# tasks are hardcoded from the store's handwritten operations notes. Seeded once
# (guarded by an audit_log marker) so later admin edits in the web UI survive
# restarts. Bump the marker version if these lists change and must re-seed.
MORLEY_STORE_ID = 2
MORLEY_CHECKLIST_MARKER = 'morley_cashier_checklist_v1'

MORLEY_CASHIER_OPENING = [
    'Enter door code (3 or 4); log in to POS as Cashier; turn on the coffee machine & hot-water machine',
    'Uniform check; wear gloves when handling/filling food',
    'Set up the Noodle bar — fill items from the freezer & fridge (odourless items & small trays first, then medium, then mixed sizes)',
    'Peel & prep vegetables for all trays',
    'Organise the pastry display & set out latte/name labels; if an item is out of stock, buy it, scan the barcode & photograph the receipt',
    'Make tropical fruit juice (apple + orange, orange peeled) — 1 jug weekdays / 2 jugs weekend; if it runs out early, make to order',
    'Prepare fruit & peel pomelo for the afternoon service',
    'Make sugarcane juice — 1 jug (bring sugarcane from storage each morning); if it runs out early, make to order',
    'Make milk coffee (750 ml black coffee + 1 can condensed milk, then 650 ml fresh milk) — 4–5 jugs weekdays / 6–7 jugs weekend',
    'Prepare 2 jugs of black coffee for iced-black-coffee orders',
    'Brew hot tea (¼ cup steeped tea leaves into the thermos, topped with hot water) ready to serve',
    'Arrange the customer order line-up; counter clean & ready for service',
]

MORLEY_CASHIER_CLOSING = [
    'Take orders, serve customers & resolve service issues during the shift (review menu & promos: banh mi + coffee $10, big meal combo $17)',
    'Collect & clear the fried-food counter; turn off one machine',
    'Clean & turn off all machines — outdoor coffee machine, juicer, tea urn, sugarcane machine, and all juice/coffee/cane jugs',
    'Clean the juice-counter trays & wrap food',
    'Clean the noodle counter & cooking area — wrap all trays; sort meat/fish cake/cheese → freezer, vegetables → cold fridge',
    'Close the curtains & clean/tidy the surrounding area',
    'Dining area, fried-food counter & cashier — sweep & put chairs up; refill spoons, chopsticks, straws & paper if low; wipe the front of the fried-food counter & cashier',
    'Refill drinking water in the fridge in front of the counter',
    'Check coffee is ready for tomorrow (2 shifts / 2 black coffee)',
    'Refill cup lids & cups',
    'Check stock — if low, buy at the supermarket (scan the barcode) or record it for the restaurant',
    'Turn off & check all equipment — gas, cabinets/shelves & machines',
    'Take photos & send the closing report to the restaurant + cf group: dining area, coffee counter, outdoor drinks fridge, fried food, banh mi, both kitchen sides, bins',
]

# ── Noodle Bar checklist (Morley + Subiaco) ──────────────────────────────────
# From the Noodle Bar staff timetable: everything before 3 PM is the Opening
# section, the 3:00 PM onward stock/pack-down routine is the Closing section.
NOODLE_BAR_MARKER = 'noodle_bar_checklist_v1'
NOODLE_BAR_STORE_IDS = (MORLEY_STORE_ID, SUBIACO_STORE_ID)

NOODLE_BAR_OPENING = [
    '10:00 AM — Arrive on time',
    '10:00 AM — Check & clean the display / topping fridge area if needed',
    '10:05 AM — Check all toppings and sauces',
    '10:05 AM — Inform the kitchen if anything needs to be prepared or refilled',
    '10:10 AM — Dry the bowls and make sure they are ready for service',
    '10:30 AM–2:00 PM — Serve customers & refill toppings when needed',
    '10:30 AM–2:00 PM — Clean tables & chairs; clear used bowls, plates and cutlery',
    '2:30–3:00 PM — Break time',
]

NOODLE_BAR_CLOSING = [
    '3:00–4:00 PM — Check stock levels',
    '3:00–4:00 PM — Fill in the stock/prep list for items running low that need preparing for tomorrow',
    '3:00–4:00 PM — Prepare stock for the required items',
    '3:00–4:00 PM — Refill noodles',
    '4:00–4:15 PM — Put toppings away properly',
    '4:00–4:15 PM — Clean the topping fridge / display area thoroughly',
]

# Subiaco's Jaccus packaging contact (from branch.xlsx). A historic global
# UPDATE in init_packaging clobbered this with Mirrabooka's "Khoi: 0449819235"
# on every startup; _fix_subiaco_contact restores it once.
SUBIACO_PACKAGING_CONTACT = 'Kenny Ho — 0406552462 — mcqsubiaco@mcqinternational.com'
MIRRABOOKA_PACKAGING_CONTACT = 'Khoi: 0449819235'

DAYS = {
    'mon': 'mon', 'monday': 'mon',
    'tue': 'tue', 'tues': 'tue', 'tuesday': 'tue',
    'wed': 'wed', 'wednesday': 'wed',
    'thu': 'thu', 'thur': 'thu', 'thurs': 'thu', 'thursday': 'thu',
    'fri': 'fri', 'friday': 'fri',
    'sat': 'sat', 'saturday': 'sat',
    'sun': 'sun', 'sunday': 'sun',
}

VALID_CHECKLISTS = {'take_order', 'banh_mi', 'chef', 'grill_beef', 'serve_order', 'noodle_bar'}
VALID_TEMP_TYPES = {'banh_mi', 'chef', 'pastry'}
VALID_KINDS = {'cold', 'room', 'hot', 'freezer'}

PREP_STATION_COLORS = {
    'banh mi station': '#FF9800',
    'pho / kitchen station': '#F44336',
    'drink station': '#00BCD4',
    'chef / general prep': '#4CAF50',
}


def _s(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def _phone(value: Any) -> str:
    text = _s(value)
    if not text:
        return ''
    if re.fullmatch(r'\d+', text) and len(text) == 9 and text.startswith('4'):
        return '0' + text
    return text


def _num(value: Any, default: float = 0.0) -> float:
    if value is None or value == '':
        return default
    if isinstance(value, (int, float)):
        return float(value)
    m = re.search(r'-?\d+(?:\.\d+)?', str(value))
    return float(m.group(0)) if m else default


def _qty(value: Any) -> int:
    try:
        return max(0, int(round(_num(value, 0))))
    except Exception:
        return 0


def _bool(value: Any) -> int:
    return 1 if _s(value).lower() in {'yes', 'y', 'true', '1', 'on'} else 0


def _kind(value: Any, default: str = 'cold') -> str:
    k = _s(value).lower()
    return k if k in VALID_KINDS else default


def _days(value: Any) -> str:
    raw = _s(value).lower()
    if not raw:
        return ''
    out: list[str] = []
    for part in re.split(r'[,/; ]+', raw):
        d = DAYS.get(part.strip().lower())
        if d and d not in out:
            out.append(d)
    return ','.join(out)


def _rows(ws, header_label: str) -> list[list[Any]]:
    start = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if _s(row[0]).lower() == header_label.lower():
            start = idx + 1
            break
    if start is None:
        return []
    data = []
    for row in ws.iter_rows(min_row=start, values_only=True):
        vals = list(row)
        if not any(_s(v) for v in vals):
            continue
        data.append(vals)
    return data


def _prep_station_id(conn: sqlite3.Connection, station_name: str) -> int:
    name = station_name.strip()
    row = conn.execute(
        'SELECT id FROM prep_stations WHERE lower(name_en)=lower(?) AND active=1',
        (name,)).fetchone()
    if row:
        return row['id']
    next_order = conn.execute(
        'SELECT COALESCE(MAX(sort_order), -1) + 1 AS n FROM prep_stations'
    ).fetchone()['n']
    color = PREP_STATION_COLORS.get(name.lower(), '#607D8B')
    cur = conn.execute('''INSERT INTO prep_stations
        (name_en, name_vi, color, sort_order, active)
        VALUES (?, '', ?, ?, 1)''', (name, color, next_order))
    return cur.lastrowid


def _seed_store(conn: sqlite3.Connection) -> None:
    conn.execute('''UPDATE stores
        SET name=?, address=?, phone=?, active=1,
            user_password=?, admin_password=?, kitchen_password=?
        WHERE id=?''',
        ('Subiaco',
         '4A Seddon St, Subiaco WA 6008',
         '0406552462',
         '8888', '8888', 'banhmivietnam',
         SUBIACO_STORE_ID))


def _seed_staff(conn: sqlite3.Connection, wb) -> None:
    for row in _rows(wb['2. Staff Members'], 'Full name'):
        name = _s(row[0])
        if not name:
            continue
        conn.execute('''INSERT INTO staff_members
            (name, role, phone, email, emergency_contact, active, store_id)
            VALUES (?,?,?,?,?,1,?)
            ON CONFLICT(name, store_id) DO UPDATE SET
                role=excluded.role,
                phone=excluded.phone,
                email=excluded.email,
                emergency_contact=excluded.emergency_contact,
                active=1''',
            (name, _s(row[1]), _phone(row[2]), _s(row[3]).strip(),
             _phone(row[4]) if len(row) > 4 else '', SUBIACO_STORE_ID))


def _seed_checklists(conn: sqlite3.Connection, wb) -> None:
    conn.execute('DELETE FROM checklist_task_templates WHERE store_id=?', (SUBIACO_STORE_ID,))
    order: dict[tuple[str, str], int] = {}
    for row in _rows(wb['4. Daily Checklists'], 'Station (chk_type)'):
        chk_type, section, task = _s(row[0]), _s(row[1]).lower(), _s(row[2])
        if chk_type not in VALID_CHECKLISTS or section not in {'opening', 'closing'} or not task:
            continue
        key = (chk_type, section)
        idx = order.get(key, 0)
        conn.execute('''INSERT INTO checklist_task_templates
            (chk_type, section, task_order, task_name, store_id)
            VALUES (?,?,?,?,?)''', (chk_type, section, idx, task, SUBIACO_STORE_ID))
        order[key] = idx + 1


def _seed_temperatures(conn: sqlite3.Connection, wb) -> None:
    conn.execute('DELETE FROM temp_food_templates WHERE store_id=?', (SUBIACO_STORE_ID,))
    order: dict[str, int] = {}
    for row in _rows(wb['5. Food Temperature Items'], 'Record (banh_mi/chef/pastry)'):
        temp_type, food = _s(row[0]), _s(row[1])
        if temp_type not in VALID_TEMP_TYPES or not food:
            continue
        idx = order.get(temp_type, 0)
        conn.execute('''INSERT INTO temp_food_templates
            (temp_type, food_order, food_name, food_kind, store_id)
            VALUES (?,?,?,?,?)''',
            (temp_type, idx, food, _kind(row[2] if len(row) > 2 else None), SUBIACO_STORE_ID))
        order[temp_type] = idx + 1


def _seed_equipment(conn: sqlite3.Connection, wb) -> None:
    conn.execute('DELETE FROM equipment_temp_readings WHERE store_id=?', (SUBIACO_STORE_ID,))
    conn.execute('DELETE FROM equipment_units WHERE store_id=?', (SUBIACO_STORE_ID,))
    for idx, row in enumerate(_rows(wb['6. Equipment (Fridges)'], 'Equipment name')):
        name = _s(row[0])
        if not name:
            continue
        conn.execute('''INSERT INTO equipment_units
            (name, kind, sort_order, active, store_id)
            VALUES (?,?,?,1,?)''',
            (name, _kind(row[1] if len(row) > 1 else None), idx, SUBIACO_STORE_ID))


def _seed_prep(conn: sqlite3.Connection, wb) -> None:
    conn.execute('DELETE FROM prep_task_templates WHERE store_id=?', (SUBIACO_STORE_ID,))
    orders: dict[int, int] = {}
    for row in _rows(wb['7. Prep Timetable'], 'Station'):
        station, task = _s(row[0]), _s(row[1])
        if not station or not task:
            continue
        station_id = _prep_station_id(conn, station)
        idx = orders.get(station_id, 0)
        conn.execute('''INSERT INTO prep_task_templates
            (task_name_en, task_name_vi, station_id, default_time, active_days,
             default_assignee, is_supplier, supplier_name, sort_order, active, store_id)
            VALUES (?, '', ?, ?, ?, '', 0, '', ?, 1, ?)''',
            (task, station_id, _s(row[3]) if len(row) > 3 else '',
             _days(row[2] if len(row) > 2 else None), idx, SUBIACO_STORE_ID))
        orders[station_id] = idx + 1


def _seed_packaging(conn: sqlite3.Connection, wb) -> None:
    old_suppliers = [r['id'] for r in conn.execute(
        'SELECT id FROM packaging_suppliers WHERE store_id=?', (SUBIACO_STORE_ID,)).fetchall()]
    for sid in old_suppliers:
        conn.execute('DELETE FROM packaging_items WHERE supplier_id=?', (sid,))
    conn.execute('DELETE FROM packaging_suppliers WHERE store_id=?', (SUBIACO_STORE_ID,))

    fields = {_s(r[0]).lower(): _s(r[1]) for r in _rows(wb['8. Packaging Supplier'], 'Field') if _s(r[0])}
    cur = conn.execute('''INSERT INTO packaging_suppliers
        (name, email, phone, cc_emails, delivery_days, cafe_name, cafe_address,
         cafe_contacts, notes, active, sort_order, store_id)
        VALUES (?,?,?,?,?,?,?,?,'',1,0,?)''',
        (fields.get('supplier name', 'Jaccus Trading'),
         fields.get('order email', ''),
         fields.get('phone', ''),
         fields.get('cc email', ''),
         SUBIACO_PACKAGING_DELIVERY_DAYS,
         'Saigon Alley by MCQ Street Food — SUBIACO',
         fields.get('deliver-to address', '4A Seddon St, Subiaco WA 6008'),
         SUBIACO_PACKAGING_CONTACT,
         SUBIACO_STORE_ID))
    supplier_id = cur.lastrowid
    for idx, row in enumerate(_rows(wb['9. Packaging Items'], 'Product code')):
        code, en, vi = _s(row[0]), _s(row[1]), _s(row[2])
        if not (code or en):
            continue
        unit = _s(row[3] if len(row) > 3 else '') or 'carton'
        qty_val = row[4] if len(row) > 4 else None
        if isinstance(row[3] if len(row) > 3 else None, (int, float)) and not qty_val:
            qty_val, unit = row[3], 'carton'
        unit_measure, unit_price = JACCUS_PRICE_DATA.get(code, ('', 0))
        conn.execute('''INSERT INTO packaging_items
            (supplier_id, product_code, name_en, name_vi, unit, unit_measure,
             unit_price, default_qty, sort_order, active)
            VALUES (?,?,?,?,?,?,?,?,?,1)''',
            (supplier_id, code, en, vi, unit, unit_measure, unit_price, _qty(qty_val), idx))


def _sync_packaging_prices(conn: sqlite3.Connection) -> None:
    """Apply the PDF price list to Subiaco's existing Jaccus catalogue once."""
    exists = conn.execute('SELECT 1 FROM audit_log WHERE action=? LIMIT 1', (PRICE_MARKER,)).fetchone()
    if exists:
        return
    updated = 0
    for product_code, (unit_measure, unit_price) in JACCUS_PRICE_DATA.items():
        cur = conn.execute('''UPDATE packaging_items
            SET unit_measure=?, unit_price=?
            WHERE product_code=? AND active=1 AND supplier_id IN (
                SELECT id FROM packaging_suppliers
                WHERE store_id=? AND lower(name)=lower('Jaccus Trading') AND active=1
            )''', (unit_measure, unit_price, product_code, SUBIACO_STORE_ID))
        updated += cur.rowcount or 0
    conn.execute('''INSERT INTO audit_log(action, record_type, user_name, details)
        VALUES (?, 'migration', 'system', ?)''',
        (PRICE_MARKER, f'Applied Jaccus PDF price list to {updated} Subiaco packaging item(s).'))


def _fix_subiaco_contact(conn: sqlite3.Connection) -> None:
    """Restore Subiaco's Jaccus packaging contact once.

    Earlier builds clobbered every branch's contact with Mirrabooka's on each
    startup. Now that init_packaging only touches store_id=1, repair the live
    Subiaco row — but only if it is still blank or holds the clobbered Mirrabooka
    value, so genuine admin edits are never overwritten. Guarded by an audit
    marker so it runs at most once.
    """
    if conn.execute('SELECT 1 FROM audit_log WHERE action=? LIMIT 1', (CONTACT_FIX_MARKER,)).fetchone():
        return
    cur = conn.execute('''UPDATE packaging_suppliers
        SET cafe_contacts=?
        WHERE store_id=? AND lower(name)=lower('Jaccus Trading') AND active=1
          AND (COALESCE(cafe_contacts,'')='' OR cafe_contacts=?)''',
        (SUBIACO_PACKAGING_CONTACT, SUBIACO_STORE_ID, MIRRABOOKA_PACKAGING_CONTACT))
    conn.execute('''INSERT INTO audit_log(action, record_type, user_name, details)
        VALUES (?, 'migration', 'system', ?)''',
        (CONTACT_FIX_MARKER, f'Restored Subiaco Jaccus contact on {cur.rowcount or 0} row(s).'))


def _fix_subiaco_delivery_days(conn: sqlite3.Connection) -> None:
    """Subiaco packaging orders must default to Tue/Thu supplier delivery."""
    conn.execute('''UPDATE packaging_suppliers
        SET delivery_days=?
        WHERE store_id=? AND active=1''',
        (SUBIACO_PACKAGING_DELIVERY_DAYS, SUBIACO_STORE_ID))


def _seed_pastry(conn: sqlite3.Connection, wb) -> None:
    conn.execute('DELETE FROM pastry_items WHERE store_id=?', (SUBIACO_STORE_ID,))
    conn.execute('DELETE FROM pastry_suppliers WHERE store_id=?', (SUBIACO_STORE_ID,))
    suppliers: dict[str, int] = {}

    def supplier_id(name: str) -> int | None:
        name = name.strip()
        if not name:
            return None
        key = name.lower()
        if key in suppliers:
            return suppliers[key]
        cur = conn.execute('''INSERT INTO pastry_suppliers (name, phone, active, store_id)
            VALUES (?, '', 1, ?)''', (name, SUBIACO_STORE_ID))
        suppliers[key] = cur.lastrowid
        return cur.lastrowid

    for idx, row in enumerate(_rows(wb['10. Pastry Items'], 'Name (EN)')):
        name_en, name_vi, sup = _s(row[0]), _s(row[1]), _s(row[2])
        if not name_en:
            continue
        conn.execute('''INSERT INTO pastry_items
            (name_en, name_vi, supplier_id, selling_price, cost_price, delivery_days,
             returnable, on_order_only, active, sort_order, store_id)
            VALUES (?,?,?,?,?,?,?,?,1,?,?)''',
            (name_en, name_vi, supplier_id(sup), _num(row[3] if len(row) > 3 else None),
             _num(row[4] if len(row) > 4 else None), _days(row[5] if len(row) > 5 else None),
             _bool(row[6] if len(row) > 6 else None), _bool(row[7] if len(row) > 7 else None),
             idx, SUBIACO_STORE_ID))


def _seed_structure(conn: sqlite3.Connection, wb) -> None:
    conn.execute('DELETE FROM structure_members WHERE store_id=?', (SUBIACO_STORE_ID,))
    conn.execute('DELETE FROM structure_departments WHERE store_id=?', (SUBIACO_STORE_ID,))
    conn.execute('DELETE FROM structure_meta WHERE store_id=?', (SUBIACO_STORE_ID,))
    conn.execute('''INSERT INTO structure_meta(key, value, store_id) VALUES
        ('manager_name', 'Kenny', ?),
        ('manager_title', 'MANAGER', ?),
        ('live_updated_by', 'subiaco-seed', ?),
        ('live_updated_at', datetime('now','localtime'), ?)''',
        (SUBIACO_STORE_ID, SUBIACO_STORE_ID, SUBIACO_STORE_ID, SUBIACO_STORE_ID))

    colors = ['#2F83C2', '#CF850D', '#D75435', '#1AA17E', '#96394E', '#704C31']
    departments: dict[str, int] = {}
    for row in _rows(wb['3. Staff Structure'], 'Department'):
        dept, person, level = _s(row[0]).upper(), _s(row[1]), _s(row[2]).upper()
        if not dept or dept in {'MANAGER', 'SUPERVISOR'}:
            continue
        if dept not in departments:
            cur = conn.execute('''INSERT INTO structure_departments
                (name, color, lead_name, lead_badge, sort_order, active, store_id)
                VALUES (?,?,?,?,?,1,?)''',
                (dept, colors[len(departments) % len(colors)], person,
                 level or 'LEAD', len(departments), SUBIACO_STORE_ID))
            departments[dept] = cur.lastrowid
        elif person:
            dept_id = departments[dept]
            sort_order = conn.execute('''SELECT COALESCE(MAX(sort_order), -1) + 1 AS n
                FROM structure_members WHERE department_id=?''', (dept_id,)).fetchone()['n']
            conn.execute('''INSERT INTO structure_members
                (department_id, level_label, staff_name, sort_order, active, store_id)
                VALUES (?,?,?,?,1,?)''',
                (dept_id, level, person, sort_order, SUBIACO_STORE_ID))


def _seed_email(conn: sqlite3.Connection) -> None:
    conn.execute('INSERT OR IGNORE INTO email_settings (store_id, from_name) VALUES (?, ?)',
                 (SUBIACO_STORE_ID, 'MCQ Subiaco Notification'))
    conn.execute('''UPDATE email_settings
        SET from_name=CASE WHEN COALESCE(from_name,'')='' THEN 'MCQ Subiaco Notification' ELSE from_name END
        WHERE store_id=?''', (SUBIACO_STORE_ID,))
    conn.execute('''INSERT INTO email_recipients
        (email, name, active, notify_checklist, notify_temperature, notify_violation,
         notify_issue, notify_prep, notify_training, notify_pastry, notify_jobs, store_id)
        VALUES ('mcqsubiaco@mcqinternational.com','Subiaco Manager',1,1,1,1,1,1,1,1,1,?)
        ON CONFLICT(email, store_id) DO UPDATE SET active=1, name=excluded.name''',
        (SUBIACO_STORE_ID,))


def _annotate_cert_notes(conn: sqlite3.Connection, wb) -> None:
    cert_names = [_s(r[0]) for r in _rows(wb['11. Food Safety Certs'], 'Staff name') if _s(r[0])]
    for name in cert_names:
        conn.execute('''UPDATE staff_members
            SET staff_notes=TRIM(COALESCE(staff_notes,'') || ' Food safety certificate listed in Subiaco workbook.')
            WHERE lower(name)=lower(?) AND store_id=? AND COALESCE(staff_notes,'') NOT LIKE '%Subiaco workbook%' ''',
            (name, SUBIACO_STORE_ID))


def seed_subiaco_branch(db_path: str, workbook_path: str | None = None) -> None:
    workbook_path = workbook_path or os.path.join(os.path.dirname(__file__), 'templates', 'branch.xlsx')
    if not os.path.exists(workbook_path):
        return
    conn = sqlite3.connect(db_path, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')
    try:
        _seed_store(conn)
        exists = conn.execute('SELECT 1 FROM audit_log WHERE action=? LIMIT 1', (MARKER,)).fetchone()
        if exists:
            _sync_packaging_prices(conn)
            _fix_subiaco_contact(conn)
            _fix_subiaco_delivery_days(conn)
            conn.commit()
            return
        wb = load_workbook(workbook_path, data_only=True)
        _seed_staff(conn, wb)
        _seed_checklists(conn, wb)
        _seed_temperatures(conn, wb)
        _seed_equipment(conn, wb)
        _seed_prep(conn, wb)
        _seed_packaging(conn, wb)
        _sync_packaging_prices(conn)
        _fix_subiaco_contact(conn)
        _fix_subiaco_delivery_days(conn)
        _seed_pastry(conn, wb)
        _seed_structure(conn, wb)
        _seed_email(conn)
        _annotate_cert_notes(conn, wb)
        conn.execute('''INSERT INTO audit_log(action, record_type, user_name, details)
            VALUES (?, 'migration', 'system', ?)''',
            (MARKER, f'Seeded Subiaco branch data from {os.path.basename(workbook_path)}.'))
        conn.commit()
    finally:
        conn.close()


def seed_morley_branch(db_path: str) -> None:
    """Seed Morley's own Cashier (take_order) Opening + Closing checklist once.

    Morley (store_id=2) has no branch workbook, so the tasks are hardcoded above.
    Only store 2's take_order templates are touched — no passwords, no other data.
    Guarded by an audit_log marker so app restarts never overwrite later admin
    edits made through the web UI.
    """
    conn = sqlite3.connect(db_path, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')
    try:
        if conn.execute('SELECT 1 FROM audit_log WHERE action=? LIMIT 1',
                        (MORLEY_CHECKLIST_MARKER,)).fetchone():
            return
        conn.execute('''DELETE FROM checklist_task_templates
            WHERE store_id=? AND chk_type=?''', (MORLEY_STORE_ID, 'take_order'))
        for section, tasks in (('opening', MORLEY_CASHIER_OPENING),
                               ('closing', MORLEY_CASHIER_CLOSING)):
            for idx, task in enumerate(tasks):
                conn.execute('''INSERT INTO checklist_task_templates
                    (chk_type, section, task_order, task_name, store_id)
                    VALUES ('take_order', ?, ?, ?, ?)''',
                    (section, idx, task, MORLEY_STORE_ID))
        conn.execute('''INSERT INTO audit_log(action, record_type, user_name, details)
            VALUES (?, 'migration', 'system', ?)''',
            (MORLEY_CHECKLIST_MARKER,
             f'Seeded Morley Cashier checklist: {len(MORLEY_CASHIER_OPENING)} opening / '
             f'{len(MORLEY_CASHIER_CLOSING)} closing task(s).'))
        conn.commit()
    finally:
        conn.close()


def seed_noodle_bar_checklists(db_path: str) -> None:
    """Seed the Noodle Bar checklist templates for Morley + Subiaco once.

    Only the noodle_bar rows of stores 2 and 3 are touched. Guarded by an
    audit_log marker so restarts never overwrite later admin edits; bump the
    marker version if the lists change and must re-seed.
    """
    conn = sqlite3.connect(db_path, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')
    try:
        if conn.execute('SELECT 1 FROM audit_log WHERE action=? LIMIT 1',
                        (NOODLE_BAR_MARKER,)).fetchone():
            return
        for store_id in NOODLE_BAR_STORE_IDS:
            conn.execute('''DELETE FROM checklist_task_templates
                WHERE store_id=? AND chk_type=?''', (store_id, 'noodle_bar'))
            for section, tasks in (('opening', NOODLE_BAR_OPENING),
                                   ('closing', NOODLE_BAR_CLOSING)):
                for idx, task in enumerate(tasks):
                    conn.execute('''INSERT INTO checklist_task_templates
                        (chk_type, section, task_order, task_name, store_id)
                        VALUES ('noodle_bar', ?, ?, ?, ?)''',
                        (section, idx, task, store_id))
        conn.execute('''INSERT INTO audit_log(action, record_type, user_name, details)
            VALUES (?, 'migration', 'system', ?)''',
            (NOODLE_BAR_MARKER,
             f'Seeded Noodle Bar checklist for stores {NOODLE_BAR_STORE_IDS}: '
             f'{len(NOODLE_BAR_OPENING)} opening / {len(NOODLE_BAR_CLOSING)} closing task(s).'))
        conn.commit()
    finally:
        conn.close()
